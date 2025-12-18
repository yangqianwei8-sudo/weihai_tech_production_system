"""
诉讼管理模块数据导出视图
"""
import logging
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.utils import timezone
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from backend.apps.system_management.services import get_user_permission_codes
from backend.core.views import _permission_granted
from backend.apps.litigation_management.models import LitigationCase, LitigationExpense

logger = logging.getLogger(__name__)


def _get_case_export_queryset(request, permission_codes):
    """获取案件导出数据集"""
    cases = LitigationCase.objects.select_related(
        'project', 'client', 'contract', 'registered_by', 
        'registered_department', 'case_manager'
    ).all()
    
    # 权限过滤
    if not _permission_granted('litigation_management.case.view_all', permission_codes):
        cases = cases.filter(Q(case_manager=request.user) | Q(registered_by=request.user))
    
    # 应用筛选条件（与case_list相同的筛选逻辑）
    case_type = request.GET.get('type', '')
    if case_type:
        cases = cases.filter(case_type=case_type)
    
    case_nature = request.GET.get('nature', '')
    if case_nature:
        cases = cases.filter(case_nature=case_nature)
    
    status = request.GET.get('status', '')
    if status:
        cases = cases.filter(status=status)
    
    priority = request.GET.get('priority', '')
    if priority:
        cases = cases.filter(priority=priority)
    
    search = request.GET.get('search', '')
    if search:
        cases = cases.filter(
            Q(case_name__icontains=search) |
            Q(case_number__icontains=search) |
            Q(description__icontains=search) |
            Q(project__project_number__icontains=search) |
            Q(client__name__icontains=search) |
            Q(contract__contract_number__icontains=search)
        )
    
    date_from = request.GET.get('date_from', '')
    if date_from:
        cases = cases.filter(registration_date__gte=date_from)
    
    date_to = request.GET.get('date_to', '')
    if date_to:
        cases = cases.filter(registration_date__lte=date_to)
    
    return cases.order_by('-registration_date', '-created_at')


@login_required
def case_list_export(request):
    """案件列表导出（Excel格式）"""
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('litigation_management.case.view', permission_codes):
        messages.error(request, '您没有权限导出案件列表')
        return redirect('litigation_pages:case_list')
    
    try:
        # 获取数据
        cases = _get_case_export_queryset(request, permission_codes)
        
        # 创建Excel工作簿
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "案件列表"
        
        # 设置表头
        headers = [
            '案件编号', '案件名称', '案件类型', '案件性质', '案件状态', '优先级',
            '诉讼标的额', '争议金额', '登记日期', '立案日期', '开庭日期', 
            '判决日期', '结案日期', '关联项目', '关联客户', '关联合同',
            '案件负责人', '登记人', '登记部门', '案件描述'
        ]
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入数据
        case_type_dict = dict(LitigationCase.CASE_TYPE_CHOICES)
        case_nature_dict = dict(LitigationCase.CASE_NATURE_CHOICES)
        status_dict = dict(LitigationCase.STATUS_CHOICES)
        priority_dict = dict(LitigationCase.PRIORITY_CHOICES)
        
        for row_idx, case in enumerate(cases, 2):
            row = [
                case.case_number,
                case.case_name,
                case_type_dict.get(case.case_type, case.case_type),
                case_nature_dict.get(case.case_nature, case.case_nature),
                status_dict.get(case.status, case.status),
                priority_dict.get(case.priority, case.priority),
                float(case.litigation_amount) if case.litigation_amount else '',
                float(case.dispute_amount) if case.dispute_amount else '',
                case.registration_date.strftime('%Y-%m-%d') if case.registration_date else '',
                case.filing_date.strftime('%Y-%m-%d') if case.filing_date else '',
                case.trial_date.strftime('%Y-%m-%d') if case.trial_date else '',
                case.judgment_date.strftime('%Y-%m-%d') if case.judgment_date else '',
                case.closing_date.strftime('%Y-%m-%d') if case.closing_date else '',
                case.project.project_number if case.project else '',
                case.client.name if case.client else '',
                case.contract.contract_number if case.contract else '',
                case.case_manager.username if case.case_manager else '',
                case.registered_by.username if case.registered_by else '',
                case.registered_department.name if case.registered_department else '',
                case.description[:100] if case.description else '',  # 限制描述长度
            ]
            
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if col_idx in [7, 8]:  # 金额列右对齐
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 设置列宽
        column_widths = {
            'A': 18,  # 案件编号
            'B': 30,  # 案件名称
            'C': 12,  # 案件类型
            'D': 10,  # 案件性质
            'E': 12,  # 案件状态
            'F': 8,   # 优先级
            'G': 15,  # 诉讼标的额
            'H': 15,  # 争议金额
            'I': 12,  # 登记日期
            'J': 12,  # 立案日期
            'K': 12,  # 开庭日期
            'L': 12,  # 判决日期
            'M': 12,  # 结案日期
            'N': 18,  # 关联项目
            'O': 20,  # 关联客户
            'P': 18,  # 关联合同
            'Q': 12,  # 案件负责人
            'R': 12,  # 登记人
            'S': 15,  # 登记部门
            'T': 40,  # 案件描述
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # 设置行高
        worksheet.row_dimensions[1].height = 25  # 表头行高
        
        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'案件列表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        logger.info(f'用户 {request.user.username} 导出了案件列表，共 {cases.count()} 条记录')
        return response
        
    except Exception as e:
        logger.error(f'导出案件列表失败: {str(e)}', exc_info=True)
        messages.error(request, f'导出失败：{str(e)}')
        return redirect('litigation_pages:case_list')


def _get_expense_export_queryset(request, permission_codes, case_id=None):
    """获取费用导出数据集"""
    if case_id:
        expenses = LitigationExpense.objects.filter(case_id=case_id)
    else:
        expenses = LitigationExpense.objects.select_related('case', 'project', 'created_by').all()
        
        # 权限过滤
        if not _permission_granted('litigation_management.expense.view', permission_codes):
            # 只能查看自己负责的案件的费用
            user_cases = LitigationCase.objects.filter(
                Q(case_manager=request.user) | Q(registered_by=request.user)
            ).values_list('id', flat=True)
            expenses = expenses.filter(case_id__in=user_cases)
    
    # 应用筛选条件
    expense_type = request.GET.get('type', '')
    if expense_type:
        expenses = expenses.filter(expense_type=expense_type)
    
    payment_status = request.GET.get('payment_status', '')
    if payment_status:
        expenses = expenses.filter(payment_status=payment_status)
    
    date_from = request.GET.get('date_from', '')
    if date_from:
        expenses = expenses.filter(expense_date__gte=date_from)
    
    date_to = request.GET.get('date_to', '')
    if date_to:
        expenses = expenses.filter(expense_date__lte=date_to)
    
    return expenses.order_by('-expense_date', '-created_at')


@login_required
def expense_list_export(request, case_id=None):
    """费用明细导出（Excel格式）"""
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('litigation_management.expense.view', permission_codes):
        messages.error(request, '您没有权限导出费用明细')
        if case_id:
            return redirect('litigation_pages:expense_list', case_id=case_id)
        return redirect('litigation_pages:case_list')
    
    try:
        # 获取数据
        expenses = _get_expense_export_queryset(request, permission_codes, case_id)
        
        # 创建Excel工作簿
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "费用明细"
        
        # 设置表头
        headers = [
            '案件编号', '案件名称', '费用名称', '费用类型', '费用金额',
            '费用日期', '支付方式', '支付状态', '关联项目', '报销状态',
            '登记人', '登记时间', '费用说明'
        ]
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入数据
        expense_type_dict = dict(LitigationExpense.EXPENSE_TYPE_CHOICES)
        payment_method_dict = dict(LitigationExpense.PAYMENT_METHOD_CHOICES)
        payment_status_dict = dict(LitigationExpense.PAYMENT_STATUS_CHOICES)
        reimbursement_status_dict = dict(LitigationExpense.REIMBURSEMENT_STATUS_CHOICES)
        
        total_amount = Decimal('0')
        for row_idx, expense in enumerate(expenses, 2):
            if expense.amount:
                total_amount += expense.amount
            
            row = [
                expense.case.case_number,
                expense.case.case_name,
                expense.expense_name,
                expense_type_dict.get(expense.expense_type, expense.expense_type),
                float(expense.amount) if expense.amount else '',
                expense.expense_date.strftime('%Y-%m-%d') if expense.expense_date else '',
                payment_method_dict.get(expense.payment_method, expense.payment_method) if expense.payment_method else '',
                payment_status_dict.get(expense.payment_status, expense.payment_status),
                expense.project.project_number if expense.project else '',
                reimbursement_status_dict.get(expense.reimbursement_status, expense.reimbursement_status) if expense.reimbursement_status else '',
                expense.created_by.username if expense.created_by else '',
                expense.created_at.strftime('%Y-%m-%d %H:%M') if expense.created_at else '',
                expense.description[:100] if expense.description else '',
            ]
            
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if col_idx == 5:  # 金额列右对齐
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 添加合计行
        if expenses.exists():
            total_row = row_idx + 1
            worksheet.cell(row=total_row, column=4, value='合计').font = Font(bold=True)
            worksheet.cell(row=total_row, column=5, value=float(total_amount)).font = Font(bold=True)
            worksheet.cell(row=total_row, column=5).alignment = Alignment(horizontal="right", vertical="center")
            for col_idx in range(1, len(headers) + 1):
                worksheet.cell(row=total_row, column=col_idx).border = border
        
        # 设置列宽
        column_widths = {
            'A': 18,  # 案件编号
            'B': 30,  # 案件名称
            'C': 25,  # 费用名称
            'D': 12,  # 费用类型
            'E': 15,  # 费用金额
            'F': 12,  # 费用日期
            'G': 12,  # 支付方式
            'H': 12,  # 支付状态
            'I': 18,  # 关联项目
            'J': 12,  # 报销状态
            'K': 12,  # 登记人
            'L': 18,  # 登记时间
            'M': 40,  # 费用说明
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # 设置行高
        worksheet.row_dimensions[1].height = 25  # 表头行高
        
        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'费用明细_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        logger.info(f'用户 {request.user.username} 导出了费用明细，共 {expenses.count()} 条记录')
        return response
        
    except Exception as e:
        logger.error(f'导出费用明细失败: {str(e)}', exc_info=True)
        messages.error(request, f'导出失败：{str(e)}')
        if case_id:
            return redirect('litigation_pages:expense_list', case_id=case_id)
        return redirect('litigation_pages:case_list')


@login_required
def statistics_export(request):
    """统计报表导出（Excel格式）"""
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('litigation_management.statistics.view', permission_codes):
        messages.error(request, '您没有权限导出统计报表')
        return redirect('litigation_pages:case_list')
    
    try:
        # 获取筛选参数
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # 获取案件列表
        cases = LitigationCase.objects.all()
        
        # 权限过滤
        if not _permission_granted('litigation_management.case.view_all', permission_codes):
            cases = cases.filter(Q(case_manager=request.user) | Q(registered_by=request.user))
        
        # 时间筛选
        if date_from:
            cases = cases.filter(registration_date__gte=date_from)
        if date_to:
            cases = cases.filter(registration_date__lte=date_to)
        
        # 创建Excel工作簿
        workbook = Workbook()
        
        # 1. 案件统计表
        worksheet1 = workbook.active
        worksheet1.title = "案件统计"
        
        headers1 = ['统计项', '数量', '占比']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers1, 1):
            cell = worksheet1.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 案件总数
        total_cases = cases.count()
        row_idx = 2
        worksheet1.cell(row=row_idx, column=1, value='案件总数').border = border
        worksheet1.cell(row=row_idx, column=2, value=total_cases).border = border
        worksheet1.cell(row=row_idx, column=3, value='100%').border = border
        row_idx += 1
        
        # 按类型统计
        case_type_dict = dict(LitigationCase.CASE_TYPE_CHOICES)
        stats_by_type = cases.values('case_type').annotate(count=Count('id'))
        worksheet1.cell(row=row_idx, column=1, value='案件类型统计').font = Font(bold=True)
        row_idx += 1
        for stat in stats_by_type:
            case_type = stat['case_type']
            count = stat['count']
            percentage = f"{(count / total_cases * 100):.2f}%" if total_cases > 0 else "0%"
            worksheet1.cell(row=row_idx, column=1, value=case_type_dict.get(case_type, case_type)).border = border
            worksheet1.cell(row=row_idx, column=2, value=count).border = border
            worksheet1.cell(row=row_idx, column=3, value=percentage).border = border
            row_idx += 1
        
        row_idx += 1
        
        # 按状态统计
        status_dict = dict(LitigationCase.STATUS_CHOICES)
        stats_by_status = cases.values('status').annotate(count=Count('id'))
        worksheet1.cell(row=row_idx, column=1, value='案件状态统计').font = Font(bold=True)
        row_idx += 1
        for stat in stats_by_status:
            status = stat['status']
            count = stat['count']
            percentage = f"{(count / total_cases * 100):.2f}%" if total_cases > 0 else "0%"
            worksheet1.cell(row=row_idx, column=1, value=status_dict.get(status, status)).border = border
            worksheet1.cell(row=row_idx, column=2, value=count).border = border
            worksheet1.cell(row=row_idx, column=3, value=percentage).border = border
            row_idx += 1
        
        # 设置列宽
        worksheet1.column_dimensions['A'].width = 25
        worksheet1.column_dimensions['B'].width = 15
        worksheet1.column_dimensions['C'].width = 15
        worksheet1.row_dimensions[1].height = 25
        
        # 2. 费用统计表
        worksheet2 = workbook.create_sheet("费用统计")
        
        headers2 = ['费用类型', '费用金额', '费用笔数', '平均金额']
        for col_idx, header in enumerate(headers2, 1):
            cell = worksheet2.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 获取费用统计
        expenses = LitigationExpense.objects.filter(case__in=cases)
        expense_type_dict = dict(LitigationExpense.EXPENSE_TYPE_CHOICES)
        stats_by_expense_type = expenses.values('expense_type').annotate(
            total_amount=Sum('amount'),
            count=Count('id')
        )
        
        row_idx = 2
        total_expense_amount = Decimal('0')
        for stat in stats_by_expense_type:
            expense_type = stat['expense_type']
            total_amount = stat['total_amount'] or Decimal('0')
            count = stat['count']
            avg_amount = total_amount / count if count > 0 else Decimal('0')
            total_expense_amount += total_amount
            
            worksheet2.cell(row=row_idx, column=1, value=expense_type_dict.get(expense_type, expense_type)).border = border
            worksheet2.cell(row=row_idx, column=2, value=float(total_amount)).border = border
            worksheet2.cell(row=row_idx, column=3, value=count).border = border
            worksheet2.cell(row=row_idx, column=4, value=float(avg_amount)).border = border
            worksheet2.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right", vertical="center")
            worksheet2.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right", vertical="center")
            row_idx += 1
        
        # 添加合计行
        if expenses.exists():
            worksheet2.cell(row=row_idx, column=1, value='合计').font = Font(bold=True)
            worksheet2.cell(row=row_idx, column=2, value=float(total_expense_amount)).font = Font(bold=True)
            worksheet2.cell(row=row_idx, column=3, value=expenses.count()).font = Font(bold=True)
            worksheet2.cell(row=row_idx, column=4, value=float(total_expense_amount / expenses.count() if expenses.count() > 0 else Decimal('0'))).font = Font(bold=True)
            for col_idx in range(1, 5):
                worksheet2.cell(row=row_idx, column=col_idx).border = border
                if col_idx in [2, 4]:
                    worksheet2.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="right", vertical="center")
        
        # 设置列宽
        worksheet2.column_dimensions['A'].width = 20
        worksheet2.column_dimensions['B'].width = 18
        worksheet2.column_dimensions['C'].width = 15
        worksheet2.column_dimensions['D'].width = 18
        worksheet2.row_dimensions[1].height = 25
        
        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'诉讼统计报表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        logger.info(f'用户 {request.user.username} 导出了统计报表')
        return response
        
    except Exception as e:
        logger.error(f'导出统计报表失败: {str(e)}', exc_info=True)
        messages.error(request, f'导出失败：{str(e)}')
        return redirect('litigation_pages:case_statistics')

