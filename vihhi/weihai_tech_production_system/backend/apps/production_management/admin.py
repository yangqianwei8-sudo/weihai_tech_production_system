"""
项目中心模块的Admin配置
包含专业配置：服务类型和服务专业
"""

from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse, path
from django import forms
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required

from backend.apps.production_management.models import (
    ServiceType, ServiceProfession, BusinessType, DesignStage, ResultFileType,
    BusinessContract, BusinessPaymentPlan, StructureType, DesignUnitCategory, Project,
    ComprehensiveAdjustmentCoefficient
)
from backend.core.admin_base import BaseModelAdmin, AuditAdminMixin, LinkAdminMixin


class ServiceProfessionInline(admin.TabularInline):
    """服务专业内联编辑"""
    model = ServiceProfession
    extra = 1
    fields = ('code', 'name', 'order')
    ordering = ('order',)


@admin.register(ServiceType)
class ServiceTypeAdmin(LinkAdminMixin, BaseModelAdmin):
    """服务类型管理"""
    list_display = ('name', 'code', 'order', 'profession_count')
    list_filter = ('order',)
    search_fields = ('name', 'code')
    ordering = ('order', 'id')
    fieldsets = (
        ('基本信息', {
            'fields': ('code', 'name', 'order')
        }),
    )
    inlines = [ServiceProfessionInline]
    
    def profession_count(self, obj):
        """显示服务类型下的专业数量"""
        count = obj.professions.count()
        if count > 0:
            url = f'/admin/production_management/serviceprofession/?service_type__id__exact={obj.id}'
            return self.make_link(url, f'{count} 个专业')
        return '0 个专业'
    profession_count.short_description = '专业数量'


@admin.register(ServiceProfession)
class ServiceProfessionAdmin(BaseModelAdmin):
    """服务专业管理"""
    list_display = ('name', 'code', 'service_type', 'order')
    list_filter = ('service_type',)
    search_fields = ('name', 'code', 'service_type__name')
    ordering = ('service_type__order', 'order', 'id')
    raw_id_fields = ('service_type',)
    fieldsets = (
        ('基本信息', {
            'fields': ('service_type', 'code', 'name', 'order')
        }),
    )


@admin.register(BusinessType)
class BusinessTypeAdmin(LinkAdminMixin, BaseModelAdmin):
    """项目业态管理"""
    list_display = ('name', 'code', 'order', 'is_active', 'project_count')
    list_filter = ('is_active',)
    search_fields = ('name', 'code')
    ordering = ('order', 'id')
    fieldsets = (
        ('基本信息', {
            'fields': ('code', 'name', 'order', 'is_active', 'description')
        }),
    )
    
    def project_count(self, obj):
        """显示使用该业态的项目数量"""
        count = obj.projects.count()
        if count > 0:
            url = f'/admin/production_management/project/?business_type__id__exact={obj.id}'
            return self.make_link(url, f'{count} 个项目')
        return '0 个项目'
    project_count.short_description = '项目数量'


@admin.register(DesignStage)
class DesignStageAdmin(LinkAdminMixin, BaseModelAdmin):
    """图纸阶段管理"""
    list_display = ('name', 'code', 'order', 'is_active', 'project_count')
    list_filter = ('is_active',)
    search_fields = ('name', 'code')
    ordering = ('order', 'id')
    fieldsets = (
        ('基本信息', {
            'fields': ('code', 'name', 'order', 'is_active', 'description')
        }),
    )
    
    def project_count(self, obj):
        """显示使用该图纸阶段的项目数量"""
        count = obj.projects.count()
        if count > 0:
            url = f'/admin/production_management/project/?design_stage__id__exact={obj.id}'
            return self.make_link(url, f'{count} 个项目')
        return '0 个项目'
    project_count.short_description = '项目数量'


@admin.register(StructureType)
class StructureTypeAdmin(LinkAdminMixin, BaseModelAdmin):
    """结构形式管理"""
    list_display = ('name', 'code', 'order', 'is_active', 'contract_count', 'project_count')
    list_filter = ('is_active',)
    search_fields = ('name', 'code')
    ordering = ('order', 'id')
    fieldsets = (
        ('基本信息', {
            'fields': ('code', 'name', 'order', 'is_active', 'description')
        }),
    )
    
    def contract_count(self, obj):
        """显示使用该结构形式的合同数量"""
        from backend.apps.production_management.models import BusinessContract
        count = BusinessContract.objects.filter(structure_type=obj.code).count()
        if count > 0:
            url = f'/admin/production_management/businesscontract/?structure_type={obj.code}'
            return self.make_link(url, f'{count} 个合同')
        return '0 个合同'
    contract_count.short_description = '合同数量'
    
    def project_count(self, obj):
        """显示使用该结构形式的项目数量"""
        count = Project.objects.filter(structure_type=obj.code).count()
        if count > 0:
            url = f'/admin/production_management/project/?structure_type={obj.code}'
            return self.make_link(url, f'{count} 个项目')
        return '0 个项目'
    project_count.short_description = '项目数量'


@admin.register(DesignUnitCategory)
class DesignUnitCategoryAdmin(LinkAdminMixin, BaseModelAdmin):
    """设计单位分类管理"""
    list_display = ('name', 'code', 'order', 'is_active', 'contract_count')
    list_filter = ('is_active',)
    search_fields = ('name', 'code')
    ordering = ('order', 'id')
    fieldsets = (
        ('基本信息', {
            'fields': ('code', 'name', 'order', 'is_active', 'description')
        }),
    )
    
    def contract_count(self, obj):
        """显示使用该分类的合同数量"""
        from backend.apps.production_management.models import BusinessContract
        count = BusinessContract.objects.filter(design_unit_category=obj.code).count()
        if count > 0:
            url = f'/admin/production_management/businesscontract/?design_unit_category={obj.code}'
            return self.make_link(url, f'{count} 个合同')
        return '0 个合同'
    contract_count.short_description = '合同数量'


@admin.register(ResultFileType)
class ResultFileTypeAdmin(BaseModelAdmin):
    """成果文件类型管理"""
    
    list_display = (
        'service_category_display', 'name', 'code', 'order', 
        'is_active', 'created_time'
    )
    list_filter = ('service_category', 'is_active', 'created_time')
    search_fields = ('name', 'code', 'description')
    ordering = ('service_category', 'order', 'id')
    list_editable = ('order', 'is_active')
    readonly_fields = ('created_time', 'updated_time')
    
    fieldsets = (
        ('基本信息', {
            'fields': ('service_category', 'code', 'name', 'order', 'is_active')
        }),
        ('详细信息', {
            'fields': ('description',)
        }),
        # 时间信息会自动添加
    )
    
    def service_category_display(self, obj):
        """服务类别显示"""
        return obj.get_service_category_display()
    service_category_display.short_description = '服务类别'
    service_category_display.admin_order_field = 'service_category'


class BusinessPaymentPlanInline(admin.TabularInline):
    """商务回款计划内联编辑"""
    model = BusinessPaymentPlan
    extra = 0
    fields = ('phase_name', 'planned_amount', 'planned_date', 'actual_amount', 'actual_date', 'status')
    readonly_fields = ('created_time', 'updated_time')


class BusinessContractForm(forms.ModelForm):
    """商务合同表单，用于改善服务专业字段的输入体验"""
    SERVICE_PROFESSION_CHOICES = [
        ('structure', '结构'),
        ('construction', '构造'),
        ('electrical', '电气'),
        ('plumbing', '给排水'),
        ('other_1', '其他专业1'),
        ('other_2', '其他专业2'),
        ('other_3', '其他专业3'),
        ('other_4', '其他专业4'),
    ]
    
    service_professions_display = forms.MultipleChoiceField(
        choices=SERVICE_PROFESSION_CHOICES,
        widget=forms.CheckboxSelectMultiple,
        required=False,
        label='服务专业',
        help_text='服务专业调整（T3）：结构：0.32；构造：0.48，电气：0.14，给排水：0.06，其他专业每增加一个，调整系数增加0.1，但总系数不超过1.5'
    )
    
    class Meta:
        model = BusinessContract
        fields = '__all__'
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # 从 JSONField 中读取数据并设置到 MultipleChoiceField
        if self.instance:
            service_professions = self.instance.service_professions or []
            self.fields['service_professions_display'].initial = service_professions
        # 隐藏原始的 service_professions 字段，使用 service_professions_display 代替
        if 'service_professions' in self.fields:
            self.fields['service_professions'].widget = forms.HiddenInput()
    
    def clean_basement_area(self):
        """验证地下室面积"""
        basement_area = self.cleaned_data.get('basement_area')
        if basement_area is not None and basement_area < 0:
            raise forms.ValidationError('地下室面积不能为负数')
        return basement_area
    
    def clean_total_building_area(self):
        """验证总建筑面积"""
        total_building_area = self.cleaned_data.get('total_building_area')
        if total_building_area is not None and total_building_area <= 0:
            raise forms.ValidationError('总建筑面积必须大于0')
        return total_building_area
    
    def clean(self):
        """验证地下室面积和总建筑面积的逻辑关系"""
        cleaned_data = super().clean()
        basement_area = cleaned_data.get('basement_area')
        total_building_area = cleaned_data.get('total_building_area')
        
        if basement_area is not None and total_building_area is not None:
            if basement_area > total_building_area:
                raise forms.ValidationError({
                    'basement_area': '地下室面积不能大于总建筑面积'
                })
        
        return cleaned_data
    
    def save(self, commit=True):
        instance = super().save(commit=False)
        # 将 MultipleChoiceField 的值保存到 JSONField
        if 'service_professions_display' in self.cleaned_data:
            instance.service_professions = self.cleaned_data['service_professions_display']
        if commit:
            instance.save()
        return instance


@admin.register(BusinessContract)
class BusinessContractAdmin(AuditAdminMixin, BaseModelAdmin):
    """商务合同管理"""
    form = BusinessContractForm
    list_display = (
        'contract_number', 'contract_name', 'contract_type', 'status', 
        'client', 'project', 'structure_type', 'design_unit_category', 
        'coefficient_display', 'contract_amount', 'contract_date', 'created_time'
    )
    list_display_links = ('contract_number', 'contract_name')  # 设置可点击的列
    list_filter = ('contract_type', 'status', 'structure_type', 'design_unit_category', 'service_type', 'project_type', 'drawing_stage', 'contract_date', 'created_time')
    search_fields = ('contract_number', 'contract_name', 'client__name', 'project__name')
    ordering = ('-created_time',)
    list_max_show_all = 200
    raw_id_fields = ('project', 'client', 'opportunity', 'parent_contract', 'created_by')
    readonly_fields = ('contract_number', 'project_number', 'contract_period', 'comprehensive_adjustment_coefficient', 'coefficient_details_display', 'created_time', 'updated_time')
    date_hierarchy = 'created_time'  # 使用 created_time 而不是 contract_date，因为 created_time 总是有值
    inlines = [BusinessPaymentPlanInline]
    fieldsets = (
        ('基本信息', {
            'fields': ('contract_number', 'project_number', 'contract_name', 'contract_type', 'status')
        }),
        ('关联信息', {
            'fields': ('project', 'client', 'opportunity', 'parent_contract')
        }),
        ('项目信息', {
            'fields': ('structure_type', 'design_unit_category'),
            'description': '项目相关信息，用于合同管理和维护'
        }),
        ('综合调整系数', {
            'fields': (
                'service_type', 'project_type', 'service_professions_display', 
                'drawing_stage', 'basement_area', 'total_building_area',
                'comprehensive_adjustment_coefficient', 'coefficient_details_display'
            ),
            'description': '综合调整系数 = T1*T2*T3*T4*T5*T6*T7，最大不超过2.0<br>'
                          'T1: 服务类型调整（结果优化：1.0，过程优化：1.5）<br>'
                          'T2: 项目业态调整（住宅：1.0，综合体：1.2，工业厂房：1.10，写字楼=1.15，商业=1.3，学校=1.05，医院=1.25，市政=1.4，其他=1.0）<br>'
                          'T3: 服务专业调整（结构：0.32；构造：0.48，电气：0.14，给排水：0.06，其他专业每增加一个，调整系数增加0.1，但总系数不超过1.5）<br>'
                          'T4: 设计质量调整（一类设计院：1.0，二类设计院：1.1，三类设计院：1.2，四类设计院：1.3）<br>'
                          'T5: 图纸阶段调整（施工图（未审图）：1.0，施工图（已审图）：0.6，初步方案：1.5，详细方案：1.4，初步设计：1.3，扩初设计：1.2，施工阶段：0.5，专项设计：1.0）<br>'
                          'T6: 地下面积占比调整（地下室面积/总建筑面积：1.0（小于0.20时），1.2（大于0.2时））<br>'
                          'T7: 结构类型调整（剪力墙结构1.0，框架结构：0.6，钢结构：1.2，其他：0.9）'
        }),
        ('金额信息', {
            'fields': ('contract_amount', 'tax_rate', 'contract_amount_excl_tax', 'contract_amount_tax', 
                      'settlement_amount', 'payment_amount', 'unpaid_amount')
        }),
        ('时间信息', {
            'fields': ('contract_date', 'effective_date', 'start_date', 'end_date', 'contract_period')
        }),
        ('其他信息', {
            'fields': ('description', 'notes', 'is_active', 'created_by')
        }),
        # 系统时间信息会自动添加
    )
    
    def get_queryset(self, request):
        """获取查询集，优化查询性能"""
        qs = super().get_queryset(request)
        # 使用 select_related 优化关联查询
        return qs.select_related('client', 'project', 'created_by')
    
    def save_model(self, request, obj, form, change):
        """保存时自动设置创建人"""
        if not change:  # 新建时
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    def service_professions_display(self, obj):
        """显示服务专业"""
        if obj.service_professions:
            profession_names = {
                'structure': '结构',
                'construction': '构造',
                'electrical': '电气',
                'plumbing': '给排水',
                'other_1': '其他专业1',
                'other_2': '其他专业2',
                'other_3': '其他专业3',
                'other_4': '其他专业4',
            }
            names = [profession_names.get(p, p) for p in obj.service_professions]
            return ', '.join(names)
        return '-'
    service_professions_display.short_description = '服务专业'
    
    def coefficient_display(self, obj):
        """显示综合调整系数，带颜色标识"""
        if obj.comprehensive_adjustment_coefficient:
            value = float(obj.comprehensive_adjustment_coefficient)
            # 根据系数值设置颜色
            if value >= 2.0:
                badge = 'danger'
            elif value >= 1.5:
                badge = 'warning'
            elif value >= 1.0:
                badge = 'success'
            else:
                badge = 'secondary'
            
            # 先格式化数值，因为 format_html 不支持格式化代码
            value_str = f"{value:.2f}"
            return format_html(
                '<span class="badge badge-{}" style="font-size: 14px; padding: 5px 10px;">{}</span>',
                badge, value_str
            )
        return format_html('<span class="text-muted">未计算</span>')
    coefficient_display.short_description = '综合调整系数'
    coefficient_display.admin_order_field = 'comprehensive_adjustment_coefficient'
    
    def coefficient_details_display(self, obj):
        """显示综合调整系数计算明细"""
        if not obj.pk:
            return '请先保存合同后再查看计算明细'
        
        details = obj.get_adjustment_coefficient_details()
        
        html = '<div style="background: #f8f9fa; padding: 15px; border-radius: 5px; margin-top: 10px;">'
        html += '<h4 style="margin-bottom: 15px;">计算明细</h4>'
        html += '<table class="table table-sm table-bordered" style="margin-bottom: 0;">'
        html += '<thead><tr><th>系数</th><th>名称</th><th>值</th><th>说明</th></tr></thead>'
        html += '<tbody>'
        
        for key in ['T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7']:
            detail = details[key]
            html += f'<tr>'
            html += f'<td><strong>{key}</strong></td>'
            html += f'<td>{detail["name"]}</td>'
            html += f'<td><span class="badge badge-primary">{detail["value"]}</span></td>'
            html += f'<td>{detail["description"] or "未设置"}</td>'
            html += f'</tr>'
        
        html += '</tbody>'
        html += '</table>'
        
        # 显示最终结果
        final = details['final']
        html += '<div style="margin-top: 15px; padding: 10px; background: #e7f3ff; border-left: 4px solid #007bff; border-radius: 3px;">'
        html += f'<strong>计算公式：</strong>{final["formula"]}<br>'
        html += f'<strong>最终结果：</strong><span class="badge badge-success" style="font-size: 16px; padding: 5px 10px;">{final["value"]:.2f}</span>'
        if final['value'] >= 2.0:
            html += ' <span class="text-danger">（已达到上限2.0）</span>'
        html += '</div>'
        html += '</div>'
        
        return format_html(html)
    coefficient_details_display.short_description = '计算明细'


@admin.register(ComprehensiveAdjustmentCoefficient)
class ComprehensiveAdjustmentCoefficientAdmin(BusinessContractAdmin):
    """综合调整系数管理"""
    form = BusinessContractForm
    list_display = (
        'contract_number', 'contract_name', 'service_type', 'project_type', 
        'service_professions_display', 'drawing_stage', 'structure_type', 
        'design_unit_category', 'coefficient_display', 'coefficient_details_link'
    )
    list_display_links = ('contract_number', 'contract_name')
    list_filter = ('service_type', 'project_type', 'drawing_stage', 'structure_type', 'design_unit_category', 'created_time')
    search_fields = ('contract_number', 'contract_name', 'client__name', 'project__name')
    ordering = ('-created_time',)
    actions = ['recalculate_coefficients', 'export_coefficients']
    
    # 只显示综合调整系数相关的字段集
    fieldsets = (
        ('基本信息', {
            'fields': ('contract_number', 'project_number', 'contract_name', 'contract_type', 'status')
        }),
        ('关联信息', {
            'fields': ('project', 'client', 'opportunity', 'parent_contract')
        }),
        ('综合调整系数', {
            'fields': (
                'service_type', 'project_type', 'service_professions_display', 
                'drawing_stage', 'basement_area', 'total_building_area',
                'structure_type', 'design_unit_category',
                'comprehensive_adjustment_coefficient', 'coefficient_details_display'
            ),
            'description': '综合调整系数 = T1*T2*T3*T4*T5*T6*T7，最大不超过2.0<br>'
                          'T1: 服务类型调整（结果优化：1.0，过程优化：1.5）<br>'
                          'T2: 项目业态调整（住宅：1.0，综合体：1.2，工业厂房：1.10，写字楼=1.15，商业=1.3，学校=1.05，医院=1.25，市政=1.4，其他=1.0）<br>'
                          'T3: 服务专业调整（结构：0.32；构造：0.48，电气：0.14，给排水：0.06，其他专业每增加一个，调整系数增加0.1，但总系数不超过1.5）<br>'
                          'T4: 设计质量调整（一类设计院：1.0，二类设计院：1.1，三类设计院：1.2，四类设计院：1.3）<br>'
                          'T5: 图纸阶段调整（施工图（未审图）：1.0，施工图（已审图）：0.6，初步方案：1.5，详细方案：1.4，初步设计：1.3，扩初设计：1.2，施工阶段：0.5，专项设计：1.0）<br>'
                          'T6: 地下面积占比调整（地下室面积/总建筑面积：1.0（小于0.20时），1.2（大于0.2时））<br>'
                          'T7: 结构类型调整（剪力墙结构1.0，框架结构：0.6，钢结构：1.2，其他：0.9）'
        }),
    )
    
    # 禁用内联编辑
    inlines = []
    readonly_fields = ('coefficient_details_display',)
    
    def get_queryset(self, request):
        """获取查询集，优化查询性能"""
        qs = super().get_queryset(request)
        return qs.select_related('client', 'project', 'created_by')
    
    def changelist_view(self, request, extra_context=None):
        """添加统计信息到列表页面"""
        extra_context = extra_context or {}
        
        # 获取统计信息
        queryset = self.get_queryset(request)
        total_count = queryset.count()
        
        # 统计系数分布
        calculated_qs = queryset.exclude(comprehensive_adjustment_coefficient__isnull=True)
        coefficient_stats = {
            'total': total_count,
            'calculated': calculated_qs.count(),
            'not_calculated': queryset.filter(comprehensive_adjustment_coefficient__isnull=True).count(),
            'max_reached': queryset.filter(comprehensive_adjustment_coefficient=2.0).count(),
            'high_range': queryset.filter(comprehensive_adjustment_coefficient__gte=1.5, comprehensive_adjustment_coefficient__lt=2.0).count(),
            'normal_range': queryset.filter(comprehensive_adjustment_coefficient__gte=1.0, comprehensive_adjustment_coefficient__lt=1.5).count(),
            'low_range': queryset.filter(comprehensive_adjustment_coefficient__lt=1.0).exclude(comprehensive_adjustment_coefficient__isnull=True).count(),
        }
        
        # 计算平均值
        if calculated_qs.exists():
            from django.db.models import Avg
            avg_coefficient = calculated_qs.aggregate(Avg('comprehensive_adjustment_coefficient'))['comprehensive_adjustment_coefficient__avg']
            coefficient_stats['average'] = round(float(avg_coefficient), 2) if avg_coefficient else 0
        else:
            coefficient_stats['average'] = 0
        
        # 添加统计信息到页面顶部
        # 先格式化平均系数，因为 format_html 不支持格式化代码
        avg_coefficient_str = f"{coefficient_stats['average']:.2f}"
        stats_html = format_html(
            '<div style="background: #f8f9fa; padding: 15px; margin-bottom: 20px; border-radius: 5px; border-left: 4px solid #007bff;">'
            '<h4 style="margin-top: 0;">综合调整系数统计</h4>'
            '<div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px;">'
            '<div><strong>总合同数：</strong>{}</div>'
            '<div><strong>已计算：</strong><span style="color: #28a745;">{}</span></div>'
            '<div><strong>未计算：</strong><span style="color: #6c757d;">{}</span></div>'
            '<div><strong>平均系数：</strong><span style="color: #007bff; font-size: 18px; font-weight: bold;">{}</span></div>'
            '<div><strong>达到上限(2.0)：</strong><span style="color: #dc3545;">{}</span></div>'
            '<div><strong>较高(1.5-2.0)：</strong><span style="color: #ffc107;">{}</span></div>'
            '<div><strong>正常(1.0-1.5)：</strong><span style="color: #28a745;">{}</span></div>'
            '<div><strong>较低(&lt;1.0)：</strong><span style="color: #6c757d;">{}</span></div>'
            '</div>'
            '</div>',
            coefficient_stats['total'],
            coefficient_stats['calculated'],
            coefficient_stats['not_calculated'],
            avg_coefficient_str,
            coefficient_stats['max_reached'],
            coefficient_stats['high_range'],
            coefficient_stats['normal_range'],
            coefficient_stats['low_range']
        )
        
        extra_context['coefficient_stats_html'] = stats_html
        extra_context['coefficient_stats'] = coefficient_stats
        
        return super().changelist_view(request, extra_context)
    
    def coefficient_display(self, obj):
        """显示综合调整系数，带颜色标识"""
        if obj.comprehensive_adjustment_coefficient:
            value = float(obj.comprehensive_adjustment_coefficient)
            # 根据系数值设置颜色
            if value >= 2.0:
                color = '#dc3545'  # 红色 - 达到上限
                badge = 'danger'
            elif value >= 1.5:
                color = '#ffc107'  # 黄色 - 较高
                badge = 'warning'
            elif value >= 1.0:
                color = '#28a745'  # 绿色 - 正常
                badge = 'success'
            else:
                color = '#6c757d'  # 灰色 - 较低
                badge = 'secondary'
            
            # 先格式化数值，因为 format_html 不支持格式化代码
            value_str = f"{value:.2f}"
            return format_html(
                '<span class="badge badge-{}" style="font-size: 14px; padding: 5px 10px;">{}</span>',
                badge, value_str
            )
        return format_html('<span class="text-muted">未计算</span>')
    coefficient_display.short_description = '综合调整系数'
    coefficient_display.admin_order_field = 'comprehensive_adjustment_coefficient'
    
    def coefficient_details_link(self, obj):
        """显示计算明细链接"""
        if obj.comprehensive_adjustment_coefficient:
            url = reverse('admin:production_management_comprehensiveadjustmentcoefficient_change', args=[obj.pk])
            return format_html(
                '<a href="{}#coefficient_details_display" class="btn btn-sm btn-info">查看明细</a>',
                url
            )
        return '-'
    coefficient_details_link.short_description = '计算明细'
    
    def coefficient_details_display(self, obj):
        """显示综合调整系数计算明细"""
        if not obj.pk:
            return '请先保存合同后再查看计算明细'
        
        details = obj.get_adjustment_coefficient_details()
        
        html = '<div style="background: #f8f9fa; padding: 15px; border-radius: 5px; margin-top: 10px;">'
        html += '<h4 style="margin-bottom: 15px;">计算明细</h4>'
        html += '<table class="table table-sm table-bordered" style="margin-bottom: 0;">'
        html += '<thead><tr><th>系数</th><th>名称</th><th>值</th><th>说明</th></tr></thead>'
        html += '<tbody>'
        
        for key in ['T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7']:
            detail = details[key]
            html += f'<tr>'
            html += f'<td><strong>{key}</strong></td>'
            html += f'<td>{detail["name"]}</td>'
            html += f'<td><span class="badge badge-primary">{detail["value"]}</span></td>'
            html += f'<td>{detail["description"] or "未设置"}</td>'
            html += f'</tr>'
        
        html += '</tbody>'
        html += '</table>'
        
        # 显示最终结果
        final = details['final']
        html += '<div style="margin-top: 15px; padding: 10px; background: #e7f3ff; border-left: 4px solid #007bff; border-radius: 3px;">'
        html += f'<strong>计算公式：</strong>{final["formula"]}<br>'
        html += f'<strong>最终结果：</strong><span class="badge badge-success" style="font-size: 16px; padding: 5px 10px;">{final["value"]:.2f}</span>'
        if final['value'] >= 2.0:
            html += ' <span class="text-danger">（已达到上限2.0）</span>'
        html += '</div>'
        html += '</div>'
        
        return format_html(html)
    coefficient_details_display.short_description = '计算明细'
    
    def recalculate_coefficients(self, request, queryset):
        """批量重新计算综合调整系数"""
        from django.contrib import messages
        count = 0
        errors = 0
        
        for obj in queryset:
            try:
                # 重新计算并保存
                obj.comprehensive_adjustment_coefficient = obj.calculate_comprehensive_adjustment_coefficient()
                obj.save(update_fields=['comprehensive_adjustment_coefficient'])
                count += 1
            except Exception as e:
                errors += 1
                messages.error(request, f'合同 {obj.contract_number} 计算失败: {str(e)}')
        
        if count > 0:
            messages.success(request, f'成功重新计算 {count} 个合同的综合调整系数。')
        if errors > 0:
            messages.warning(request, f'{errors} 个合同计算失败，请检查数据。')
    recalculate_coefficients.short_description = '批量重新计算综合调整系数'
    
    def export_coefficients(self, request, queryset):
        """导出综合调整系数数据"""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '综合调整系数'
        
        # 设置表头
        headers = [
            '合同编号', '合同名称', '服务类型', '项目业态', '服务专业',
            '图纸阶段', '结构类型', '设计单位分类', '地下室面积(㎡)', '总建筑面积(㎡)',
            'T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7', '综合调整系数'
        ]
        worksheet.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 填充数据
        for obj in queryset.select_related('client', 'project'):
            details = obj.get_adjustment_coefficient_details()
            
            # 服务专业显示
            profession_names = {
                'structure': '结构',
                'construction': '构造',
                'electrical': '电气',
                'plumbing': '给排水',
            }
            professions_display = []
            if obj.service_professions:
                for p in obj.service_professions:
                    if p.startswith('other_'):
                        professions_display.append('其他专业')
                    else:
                        professions_display.append(profession_names.get(p, p))
            
            row = [
                obj.contract_number or '',
                obj.contract_name or '',
                obj.get_service_type_display() if obj.service_type else '',
                obj.get_project_type_display() if obj.project_type else '',
                ', '.join(professions_display) if professions_display else '',
                obj.get_drawing_stage_display() if obj.drawing_stage else '',
                obj.get_structure_type_display() if obj.structure_type else '',
                obj.get_design_unit_category_display() if obj.design_unit_category else '',
                float(obj.basement_area) if obj.basement_area else '',
                float(obj.total_building_area) if obj.total_building_area else '',
                float(details['T1']['value']),
                float(details['T2']['value']),
                float(details['T3']['value']),
                float(details['T4']['value']),
                float(details['T5']['value']),
                float(details['T6']['value']),
                float(details['T7']['value']),
                float(obj.comprehensive_adjustment_coefficient) if obj.comprehensive_adjustment_coefficient else '',
            ]
            worksheet.append(row)
        
        # 调整列宽
        column_widths = [15, 30, 12, 12, 20, 15, 12, 15, 12, 12, 8, 8, 8, 8, 8, 8, 8, 12]
        for i, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=i).column_letter].width = width
        
        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'综合调整系数_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        return response
    export_coefficients.short_description = '导出综合调整系数数据（Excel）'


@admin.register(BusinessPaymentPlan)
class BusinessPaymentPlanAdmin(BaseModelAdmin):
    """商务回款计划管理"""
    list_display = (
        'contract', 'phase_name', 'planned_amount', 'planned_date', 
        'actual_amount', 'actual_date', 'status', 'created_time'
    )
    list_filter = ('status', 'planned_date', 'actual_date', 'created_time')
    search_fields = ('phase_name', 'contract__contract_number', 'contract__contract_name')
    ordering = ('-planned_date', '-created_time')
    raw_id_fields = ('contract',)
    readonly_fields = ('created_time', 'updated_time')
    date_hierarchy = 'planned_date'
    fieldsets = (
        ('基本信息', {
            'fields': ('contract', 'phase_name', 'phase_description')
        }),
        ('计划信息', {
            'fields': ('planned_amount', 'planned_date', 'trigger_condition', 'condition_detail')
        }),
        ('实际信息', {
            'fields': ('actual_amount', 'actual_date', 'status')
        }),
        ('其他信息', {
            'fields': ('notes',)
        }),
        # 时间信息会自动添加
    )

