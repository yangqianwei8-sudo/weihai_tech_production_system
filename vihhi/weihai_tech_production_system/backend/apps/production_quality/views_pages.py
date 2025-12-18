from __future__ import annotations

import csv
import io
import json
from collections import Counter
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Sum, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone

from backend.apps.production_management.models import Project
from backend.apps.production_management.views_pages import (
    _has_permission,
    _project_ids_user_can_access,
)
from backend.apps.system_management.models import User
from backend.apps.system_management.services import get_user_permission_codes
from backend.apps.resource_standard.models import ProfessionalCategory, StandardReviewItem, ReportTemplate
from backend.core.views import _permission_granted, HOME_NAV_STRUCTURE, _build_full_top_nav

from .forms import OpinionAttachmentFormSet, OpinionBulkImportForm, OpinionForm
from backend.apps.production_quality.models import (
    Opinion,
    OpinionParticipant,
    OpinionReview,
    OpinionSavingItem,
    OpinionWorkflowLog,
    ProductionReport,
    ProductionStatistic,
)
from .services import (
    calculate_saving_amount,
    generate_opinion_number,
    infer_review_role,
    record_workflow_log,
    sync_opinion_participants,
    sync_opinion_saving_items,
)

FIELD_STEP_MAP = {
    "project": 1,
    "professional_category": 1,
    "source": 2,
    "priority": 2,
    "drawing_number": 2,
    "drawing_version": 2,
    "location_name": 2,
    "review_points": 2,
    "issue_description": 3,
    "current_practice": 3,
    "recommendation": 3,
    "issue_category": 3,
    "severity_level": 3,
    "reference_codes": 3,
    "calculation_mode": 4,
    "quantity_before": 4,
    "quantity_after": 4,
    "measure_unit": 4,
    "unit_price_before": 4,
    "unit_price_after": 4,
    "saving_amount": 4,
    "calculation_note": 4,
    "impact_scope": 5,
    "expected_complete_date": 5,
    "actual_complete_date": 5,
    "response_deadline": 5,
}

IMPACT_SCOPE_PRESETS = [
    {"label": "结构主体", "value": "结构主体"},
    {"label": "机电系统", "value": "机电系统"},
    {"label": "建筑外立面", "value": "建筑外立面"},
    {"label": "室内精装", "value": "室内精装"},
    {"label": "市政配套", "value": "市政配套"},
    {"label": "安全文明", "value": "安全文明"},
]


IMPORT_COLUMNS = [
    ("project_number", "项目编号", True),
    ("professional_code", "专业编码", True),
    ("drawing_number", "图纸编号", False),
    ("drawing_version", "图纸版本", False),
    ("location_name", "部位名称", True),
    ("issue_description", "问题描述", True),
    ("current_practice", "现行做法", False),
    ("recommendation", "优化建议", True),
    ("issue_category", "问题类别", True),
    ("severity_level", "严重等级", True),
    ("reference_codes", "引用规范", False),
    ("calculation_mode", "计算方式", False),
    ("quantity_before", "优化前工程量", False),
    ("quantity_after", "优化后工程量", False),
    ("measure_unit", "计量单位", False),
    ("unit_price_before", "优化前综合单价", False),
    ("unit_price_after", "优化后综合单价", False),
    ("saving_amount", "节省金额", False),
    ("calculation_note", "计算说明", False),
]


def _build_production_top_nav(permission_set):
    """
    生成生产管理专用的顶部导航菜单 - 已废弃
    
    注意：此函数已不再使用，系统现在统一使用 _build_full_top_nav 生成全局系统主菜单。
    保留此函数仅用于历史参考，可以安全删除。
    """
    from django.urls import reverse, NoReverseMatch
    
    # 定义生产管理功能模块（从左到右的顺序）
    production_modules = [
        {
            'label': '生产启动',
            'url_name': 'production_quality_pages:production_startup_list',
            'permission': None,
            'icon': '🚀',
        },
        {
            'label': '意见填报',
            'url_name': 'production_quality_pages:opinion_create',
            'permission': None,  # 意见填报无需权限
            'icon': '✍️',
        },
        {
            'label': '草稿管理',
            'url_name': 'production_quality_pages:opinion_drafts',
            'permission': None,  # 草稿管理无需权限（只能看自己的）
            'icon': '📝',
        },
        {
            'label': '质量审核',
            'url_name': 'production_quality_pages:opinion_review',
            'permission': 'production_quality.professional_review',
            'icon': '✅',
        },
        {
            'label': '审核列表',
            'url_name': 'production_quality_pages:opinion_review_list',
            'permission': 'production_quality.professional_review',
            'icon': '📋',
        },
        {
            'label': '意见导入',
            'url_name': 'production_quality_pages:opinion_import',
            'permission': None,  # 意见导入无需权限
            'icon': '📥',
        },
        {
            'label': '报告生成',
            'url_name': 'production_quality_pages:report_generate',
            'permission': 'production_quality.generate_report',
            'icon': '📊',
        },
        {
            'label': '生产统计',
            'url_name': 'production_quality_pages:production_stats',
            'permission': 'production_quality.view_statistics',
            'icon': '📈',
        },
    ]
    
    # 过滤有权限的模块，直接返回导航项
    nav_items = []
    for module in production_modules:
        if not module.get('permission') or _permission_granted(module['permission'], permission_set):
            try:
                url = reverse(module['url_name'])
            except NoReverseMatch:
                url = '#'
            nav_items.append({
                'label': module['label'],
                'url': url,
                'icon': module.get('icon', ''),
            })
    
    return nav_items


# 使用统一的顶部导航菜单生成函数
from backend.core.views import _build_full_top_nav


def _context(page_title, page_icon, description, summary_cards=None, sections=None, request=None):
    """构建页面上下文"""
    context = {
        "page_title": page_title,
        "page_icon": page_icon,
        "description": description,
        "summary_cards": summary_cards or [],
        "sections": sections or [],
    }
    
    if request and request.user.is_authenticated:
        permission_set = get_user_permission_codes(request.user)
        # 统一使用全局系统主菜单（与客户管理、财务管理模块保持一致）
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    
    return context

ISSUE_CATEGORY_LOOKUP = {}
for value, label in Opinion.IssueCategory.choices:
    ISSUE_CATEGORY_LOOKUP[value.lower()] = value
    ISSUE_CATEGORY_LOOKUP[label.lower()] = value
    ISSUE_CATEGORY_LOOKUP[value] = value
    ISSUE_CATEGORY_LOOKUP[label] = value

SEVERITY_LOOKUP = {}
for value, label in Opinion.SeverityLevel.choices:
    SEVERITY_LOOKUP[value.lower()] = value
    SEVERITY_LOOKUP[label.lower()] = value
    SEVERITY_LOOKUP[value] = value
    SEVERITY_LOOKUP[label] = value

CALCULATION_MODE_LOOKUP = {}
for value, label in Opinion.CalculationMode.choices:
    CALCULATION_MODE_LOOKUP[value.lower()] = value
    CALCULATION_MODE_LOOKUP[label.lower()] = value
    CALCULATION_MODE_LOOKUP[value] = value
    CALCULATION_MODE_LOOKUP[label] = value
CALCULATION_MODE_LOOKUP["自动"] = Opinion.CalculationMode.AUTO
CALCULATION_MODE_LOOKUP["手动"] = Opinion.CalculationMode.MANUAL


@login_required
def opinion_create(request):
    """新建咨询意见页面"""
    initial_project_id = request.GET.get("project")
    initial_profession_id = request.GET.get("profession")
    initial = {}
    if initial_project_id and Project.objects.filter(id=initial_project_id).exists():
        initial["project"] = initial_project_id
    if initial_profession_id and ProfessionalCategory.objects.filter(id=initial_profession_id).exists():
        initial["professional_category"] = initial_profession_id

    form = OpinionForm(
        request.POST or None,
        user=request.user,
        initial=initial if request.method == "GET" else None,
    )
    formset = OpinionAttachmentFormSet(
        request.POST or None,
        request.FILES or None,
        prefix="attachments",
    )

    if request.method == "POST":
        submit_action = request.POST.get("submit_action", "draft")
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                opinion: Opinion = form.save(commit=False)
                opinion.created_by = request.user
                if submit_action == "submit":
                    opinion.status = Opinion.OpinionStatus.SUBMITTED
                    opinion.submitted_at = timezone.now()
                else:
                    opinion.status = Opinion.OpinionStatus.DRAFT
                opinion.save()
                form.save_m2m()

                formset.instance = opinion
                formset.save()

                participants_payload_raw = request.POST.get("participants_payload", "[]")
                saving_items_payload_raw = request.POST.get("saving_items_payload", "[]")
                try:
                    participants_payload = json.loads(participants_payload_raw or "[]")
                except json.JSONDecodeError:
                    participants_payload = []
                try:
                    saving_items_payload = json.loads(saving_items_payload_raw or "[]")
                except json.JSONDecodeError:
                    saving_items_payload = []

                sync_opinion_participants(opinion, participants_payload, operator=request.user)
                sync_opinion_saving_items(opinion, saving_items_payload)
                record_workflow_log(
                    opinion=opinion,
                    action=OpinionWorkflowLog.ActionType.CREATED,
                    operator=request.user,
                    from_status=None,
                    to_status=opinion.status,
                    message="创建意见",
                )
                if submit_action == "submit":
                    record_workflow_log(
                        opinion=opinion,
                        action=OpinionWorkflowLog.ActionType.SUBMITTED,
                        operator=request.user,
                        from_status=Opinion.OpinionStatus.DRAFT,
                        to_status=Opinion.OpinionStatus.SUBMITTED,
                        message="提交审核",
                    )

            if submit_action == "submit":
                messages.success(request, "咨询意见已提交审核。")
            else:
                messages.success(request, "咨询意见草稿已保存。")
            return redirect(
                reverse("production_quality_pages:opinion_create")
                + f"?project={opinion.project_id}&profession={opinion.professional_category_id}"
            )
        else:
            messages.error(request, "请检查表单输入，确保所有必填信息已填写完整。")

    project_queryset = (
        form.fields["project"]
        .queryset.select_related("service_type")
        .prefetch_related("service_professions__service_type")
    )
    professional_categories = list(
        ProfessionalCategory.objects.all().values("id", "name", "code", "service_types")
    )

    service_type_to_categories = {}
    for category in professional_categories:
        for service_type_code in category["service_types"]:
            service_type_to_categories.setdefault(service_type_code, []).append(
                {"id": category["id"], "name": category["name"], "code": category["code"]}
            )

    project_payload = []
    for project in project_queryset:
        service_type_codes = list(
            project.service_professions.values_list("service_type__code", flat=True).distinct()
        )
        categories = []
        seen = set()
        for code in service_type_codes:
            for category in service_type_to_categories.get(code, []):
                if category["id"] not in seen:
                    seen.add(category["id"])
                    categories.append(category)
        members = []
        for team_member in project.team_members.select_related("user", "service_profession"):
            if not team_member.user_id:
                continue
            members.append(
                {
                    "id": team_member.user_id,
                    "name": team_member.user.get_full_name() or team_member.user.username,
                    "role": team_member.role,
                    "roleLabel": team_member.get_role_display(),
                    "profession": team_member.service_profession.name if team_member.service_profession else "",
                    "unit": team_member.unit,
                }
            )
        project_payload.append(
            {
                "id": project.id,
                "name": project.name,
                "number": project.project_number,
                "serviceTypes": service_type_codes,
                "categories": categories,
                "members": members,
            }
        )

    review_queryset = form.fields["review_points"].queryset
    review_point_templates = {
        item.id: {
            "section": item.section_name,
            "content": item.review_point,
        }
        for item in review_queryset
    }

    default_participants = [
        {
            "user": request.user.id,
            "user_name": request.user.get_full_name() or request.user.username,
            "role": OpinionParticipant.ParticipantRole.PROPOSER,
            "is_primary": True,
        }
    ]
    if request.method == "POST":
        participants_payload_raw = request.POST.get("participants_payload", json.dumps(default_participants, ensure_ascii=False))
        saving_items_payload_raw = request.POST.get("saving_items_payload", "[]")
    else:
        participants_payload_raw = json.dumps(default_participants, ensure_ascii=False)
        saving_items_payload_raw = "[]"

    initial_step = 1
    if request.method == "POST":
        posted_step = request.POST.get("current_step")
        if posted_step and posted_step.isdigit():
            initial_step = max(initial_step, min(int(posted_step), 6))
        for field_name in form.errors:
            initial_step = max(initial_step, FIELD_STEP_MAP.get(field_name, initial_step))
        if formset.non_form_errors():
            initial_step = max(initial_step, 6)
        for attachment_form in formset.forms:
            if attachment_form.errors:
                initial_step = max(initial_step, 6)
                break

    context = {
        "form": form,
        "formset": formset,
        "initial_step": initial_step,
        "review_point_count": StandardReviewItem.objects.count(),
        "project_options": project_payload,
        "professional_categories": professional_categories,
        "review_point_templates": review_point_templates,
        "impact_scope_options": IMPACT_SCOPE_PRESETS,
        "source_choices": list(Opinion.OpinionSource.choices),
        "priority_choices": list(Opinion.PriorityLevel.choices),
        "participant_role_choices": list(OpinionParticipant.ParticipantRole.choices),
        "saving_category_choices": list(OpinionSavingItem.SavingCategory.choices),
        "initial_participants_payload": participants_payload_raw,
        "initial_saving_items_payload": saving_items_payload_raw,
        "current_user_payload": {
            "id": request.user.id,
            "name": request.user.get_full_name() or request.user.username,
        },
    }
    response = render(request, "production_quality/opinion_form.html", context)
    response.context_data = context
    return response


def _project_ids_user_can_access(user):
    if user.is_superuser:
        return Project.objects.values_list("id", flat=True)
    project_ids = set(
        Project.objects.filter(
            Q(project_manager=user) | Q(business_manager=user) | Q(created_by=user)
        ).values_list("id", flat=True)
    )
    team_ids = set(
        Project.objects.filter(team_members__user=user).values_list("id", flat=True)
    )
    return project_ids | team_ids


@login_required
def opinion_review_dashboard(request):
    """质量审核总览页面"""
    accessible_ids = _project_ids_user_can_access(request.user)
    opinions = list(
        Opinion.objects.filter(project_id__in=accessible_ids)
        .select_related("project", "professional_category", "created_by")
        .prefetch_related("review_points", "reviews")
        .order_by("-submitted_at", "-created_at")
    )

    pending_status = [
        Opinion.OpinionStatus.SUBMITTED,
        Opinion.OpinionStatus.IN_REVIEW,
        Opinion.OpinionStatus.NEEDS_UPDATE,
    ]
    approved_status = [Opinion.OpinionStatus.APPROVED]
    rejected_status = [Opinion.OpinionStatus.REJECTED]

    pending_opinions = [op for op in opinions if op.status in pending_status]
    approved_recent = [
        op
        for op in opinions
        if op.status in approved_status
        and op.reviewed_at
        and op.reviewed_at >= timezone.now() - timedelta(days=30)
    ]
    rejected_recent = [
        op
        for op in opinions
        if op.status in rejected_status
        and op.reviewed_at
        and op.reviewed_at >= timezone.now() - timedelta(days=30)
    ]

    def _group_by_professional_category(opinion_list):
        buckets = {}
        for opinion in opinion_list:
            key = opinion.professional_category_id or 0
            entry = buckets.setdefault(
                key,
                {
                    "category": opinion.professional_category.name
                    if opinion.professional_category
                    else "未分类",
                    "items": [],
                },
            )
            entry["items"].append(opinion)
        return buckets.values()

    today = timezone.now().date()
    now_dt = timezone.now()

    overdue_candidates = [
        op
        for op in pending_opinions
        if op.response_deadline and op.response_deadline < today
    ]
    overdue_list = sorted(
        [
            {
                "opinion": op,
                "days": (today - op.response_deadline).days,
                "deadline": op.response_deadline,
            }
            for op in overdue_candidates
        ],
        key=lambda item: item["deadline"],
    )

    pending_unassigned = sum(1 for op in pending_opinions if op.current_reviewer_id is None)

    summary_metrics = {
        "total_pending": len(pending_opinions),
        "total_unassigned": pending_unassigned,
        "total_overdue": len(overdue_list),
        "total_approved_month": len(approved_recent),
        "total_rejected_month": len(rejected_recent),
        "total_all": len(opinions),
    }

    cycle_values = [
        float(op.cycle_time_hours)
        for op in opinions
        if op.cycle_time_hours is not None
    ]
    avg_cycle_hours = round(sum(cycle_values) / len(cycle_values), 2) if cycle_values else None

    response_values = []
    for op in opinions:
        if op.submitted_at and op.first_response_at and op.first_response_at >= op.submitted_at:
            delta_hours = (op.first_response_at - op.submitted_at).total_seconds() / 3600
            response_values.append(delta_hours)
    avg_response_hours = round(sum(response_values) / len(response_values), 2) if response_values else None

    sla_metrics = {
        "avg_response_hours": avg_response_hours,
        "avg_cycle_hours": avg_cycle_hours,
        "pending_overdue_ratio": round(
            summary_metrics["total_overdue"] / summary_metrics["total_pending"] * 100, 1
        )
        if summary_metrics["total_pending"]
        else 0,
    }

    total_saving = sum(Decimal(op.saving_amount or 0) for op in opinions)
    recent_saving = sum(Decimal(op.saving_amount or 0) for op in approved_recent)
    financial_summary = {
        "total_saving": total_saving,
        "recent_saving": recent_saving,
    }

    status_labels = [
        (Opinion.OpinionStatus.SUBMITTED, "已提交"),
        (Opinion.OpinionStatus.IN_REVIEW, "审核中"),
        (Opinion.OpinionStatus.NEEDS_UPDATE, "需修改"),
        (Opinion.OpinionStatus.APPROVED, "已通过"),
        (Opinion.OpinionStatus.REJECTED, "已驳回"),
    ]
    status_counter = Counter(op.status for op in opinions)
    status_chart = {
        "labels": [label for _, label in status_labels],
        "data": [status_counter.get(code, 0) for code, _ in status_labels],
    }

    profession_counter = Counter(
        opinion.professional_category.name if opinion.professional_category else "未分类"
        for opinion in pending_opinions
    )
    top_professions = profession_counter.most_common(6)
    if top_professions:
        profession_chart = {
            "labels": [label for label, _ in top_professions],
            "data": [count for _, count in top_professions],
        }
    else:
        profession_chart = {"labels": ["暂无待审"], "data": [0]}

    reviewer_counter = Counter(
        opinion.current_reviewer.get_full_name() or opinion.current_reviewer.username
        for opinion in pending_opinions
        if opinion.current_reviewer
    )
    top_reviewers = reviewer_counter.most_common(6)
    if top_reviewers:
        reviewer_chart = {
            "labels": [label for label, _ in top_reviewers],
            "data": [count for _, count in top_reviewers],
        }
    else:
        reviewer_chart = {"labels": ["未分配"], "data": [len(pending_opinions)]}

    pending_cards = []
    for entry in _group_by_professional_category(pending_opinions):
        sorted_items = sorted(
            entry["items"],
            key=lambda op: op.submitted_at or op.created_at,
            reverse=True,
        )[:6]
        entry_items = []
        for opinion in sorted_items:
            base_time = opinion.submitted_at or opinion.created_at
            waiting_hours = None
            if base_time:
                waiting_hours = (now_dt - base_time).total_seconds() / 3600
            entry_items.append(
                {
                    "opinion": opinion,
                    "waiting_hours": waiting_hours,
                    "is_overdue": bool(
                        opinion.response_deadline and opinion.response_deadline < today
                    ),
                    "is_unassigned": opinion.current_reviewer_id is None,
                }
            )
        pending_cards.append(
            {
                "category": entry["category"],
                "count": len(entry["items"]),
                "items": entry_items,
            }
        )

    insights = []
    if sla_metrics["avg_cycle_hours"]:
        insights.append(
            {
                "title": "平均结案耗时",
                "content": f"{sla_metrics['avg_cycle_hours']} 小时",
                "time": "最近全部",
            }
        )
    if sla_metrics["avg_response_hours"]:
        insights.append(
            {
                "title": "平均响应速度",
                "content": f"{sla_metrics['avg_response_hours']} 小时",
                "time": "最近全部",
            }
        )
    if financial_summary["recent_saving"]:
        insights.append(
            {
                "title": "近30天节省",
                "content": f"¥{financial_summary['recent_saving']}",
                "time": "已通过意见",
            }
        )
    if overdue_list:
        overdue_top = overdue_list[0]
        op = overdue_top["opinion"]
        insights.append(
            {
                "title": "超期提醒",
                "content": f"{op.project.project_number if op.project else ''} · {op.location_name}",
                "time": f"超期 {overdue_top['days']} 天",
            }
        )
    if pending_opinions:
        oldest = min(
            pending_opinions,
            key=lambda op: op.submitted_at or op.created_at,
        )
        insights.append(
            {
                "title": "最久待审意见",
                "content": f"{oldest.project.project_number if oldest.project else ''} · {oldest.location_name}",
                "time": (oldest.submitted_at or oldest.created_at).strftime("%Y-%m-%d"),
            }
        )
    if approved_recent:
        latest = max(approved_recent, key=lambda op: op.reviewed_at)
        insights.append(
            {
                "title": "最新通过意见",
                "content": f"{latest.project.project_number if latest.project else ''} · {latest.location_name}",
                "time": latest.reviewed_at.strftime("%Y-%m-%d") if latest.reviewed_at else "",
            }
        )
    if rejected_recent:
        latest_reject = max(rejected_recent, key=lambda op: op.reviewed_at)
        insights.append(
            {
                "title": "驳回提醒",
                "content": f"{latest_reject.project.project_number if latest_reject.project else ''} · {latest_reject.location_name}",
                "time": latest_reject.reviewed_at.strftime("%Y-%m-%d") if latest_reject.reviewed_at else "",
            }
        )

    context = {
        "project": None,
        "summary": summary_metrics,
        "sla_metrics": sla_metrics,
        "financial_summary": financial_summary,
        "pending_cards": pending_cards,
        "approved_recent": approved_recent[:6],
        "rejected_recent": rejected_recent[:6],
        "insights": insights,
        "status_chart": status_chart,
        "profession_chart": profession_chart,
        "reviewer_chart": reviewer_chart,
        "overdue_list": overdue_list[:6],
        "today": today,
    }
    return render(request, "production_quality/opinion_review.html", context)


@login_required
def opinion_review_detail(request, opinion_id):
    """意见审核详情页"""
    opinion = get_object_or_404(
        Opinion.objects.select_related(
            "project",
            "professional_category",
            "created_by",
            "current_reviewer",
        ).prefetch_related(
            "review_points",
            "attachments",
            "reviews__reviewer",
            "participants__user",
            "saving_items",
            "workflow_logs__operator",
        ),
        id=opinion_id,
    )
    project_ids = _project_ids_user_can_access(request.user)
    if opinion.project_id not in project_ids:
        messages.error(request, "您无权查看该意见。")
        return redirect("production_quality_pages:opinion_review")

    attachment_entries = opinion.attachments.all()
    participants = opinion.participants.select_related("user").order_by("-is_primary", "role", "joined_at")
    saving_items = opinion.saving_items.order_by("created_at")

    workflow_logs = list(
        opinion.workflow_logs.select_related("operator").order_by("-created_at")[:30]
    )
    status_labels = dict(Opinion.OpinionStatus.choices)
    review_status_labels = dict(OpinionReview.ReviewStatus.choices)

    user_name_map: Dict[int, str] = {}

    def register_user(user_obj):
        if user_obj and user_obj.id not in user_name_map:
            user_name_map[user_obj.id] = user_obj.get_full_name() or user_obj.username

    register_user(request.user)
    register_user(opinion.created_by)
    register_user(opinion.current_reviewer)
    for participant in participants:
        register_user(participant.user)

    payload_user_ids = set()
    for log in workflow_logs:
        if log.payload:
            reviewer_id = log.payload.get("reviewer_id")
            if reviewer_id:
                payload_user_ids.add(reviewer_id)
        register_user(log.operator)

    if payload_user_ids:
        for extra_user in User.objects.filter(id__in=payload_user_ids):
            register_user(extra_user)

    history_timeline = []
    for log in workflow_logs:
        details = []
        if log.from_status or log.to_status:
            if log.from_status:
                details.append(f"原状态：{status_labels.get(log.from_status, log.from_status)}")
            if log.to_status:
                details.append(f"目标状态：{status_labels.get(log.to_status, log.to_status)}")
        if log.action == OpinionWorkflowLog.ActionType.REASSIGNED:
            reviewer_id = (log.payload or {}).get("reviewer_id")
            if reviewer_id:
                reviewer_name = user_name_map.get(reviewer_id, f"ID {reviewer_id}")
                details.append(f"指派给：{reviewer_name}")
        if log.action == OpinionWorkflowLog.ActionType.REVIEWED:
            review_status = (log.payload or {}).get("review_status")
            if review_status:
                details.append(f"审核结论：{review_status_labels.get(review_status, review_status)}")
            payload_comment = (log.payload or {}).get("comment")
            if payload_comment:
                details.append(f"批量备注：{payload_comment}")
        history_timeline.append(
            {
                "title": log.get_action_display(),
                "timestamp": log.created_at,
                "operator": user_name_map.get(log.operator_id, "系统"),
                "message": log.message,
                "details": details,
            }
        )

    if not history_timeline:
        history_timeline.append(
            {
                "title": "创建草稿",
                "timestamp": opinion.created_at,
                "operator": user_name_map.get(opinion.created_by_id, "提交人"),
                "message": "",
                "details": [],
            }
        )

    saving_total = sum(item.total_saving or Decimal("0") for item in saving_items)

    pending_status = [
        Opinion.OpinionStatus.SUBMITTED,
        Opinion.OpinionStatus.IN_REVIEW,
        Opinion.OpinionStatus.NEEDS_UPDATE,
    ]
    review_allowed = opinion.status in pending_status
    is_current_reviewer = opinion.current_reviewer_id == request.user.id
    review_requires_assignment = review_allowed and opinion.current_reviewer_id not in (
        None,
        request.user.id,
    )
    can_review = review_allowed and not review_requires_assignment
    can_assign = review_allowed
    can_revert = opinion.status in [
        Opinion.OpinionStatus.NEEDS_UPDATE,
        Opinion.OpinionStatus.REJECTED,
    ]

    available_reviewers = []
    seen_users = set()

    def append_candidate(user_obj, label=""):
        if user_obj and user_obj.id not in seen_users:
            available_reviewers.append(
                {
                    "id": user_obj.id,
                    "name": user_obj.get_full_name() or user_obj.username,
                    "role": label,
                }
            )
            seen_users.add(user_obj.id)

    if opinion.project_id:
        team_members = (
            opinion.project.team_members.select_related("user", "service_profession")
            .order_by("user__first_name", "user__last_name", "user__username")
        )
        for member in team_members:
            if member.user:
                role_label = member.get_role_display()
                append_candidate(member.user, role_label)

    append_candidate(opinion.current_reviewer, "当前审核人")
    append_candidate(request.user, "当前登录")

    sla_detail = {
        "submitted_at": opinion.submitted_at,
        "first_assigned_at": opinion.first_assigned_at,
        "first_response_at": opinion.first_response_at,
        "closed_at": opinion.closed_at,
        "cycle_time_hours": float(opinion.cycle_time_hours)
        if opinion.cycle_time_hours is not None
        else None,
        "waiting_hours": None,
        "response_hours": None,
    }
    if opinion.submitted_at:
        waiting_seconds = (timezone.now() - opinion.submitted_at).total_seconds()
        if waiting_seconds >= 0:
            sla_detail["waiting_hours"] = round(waiting_seconds / 3600, 2)
    if opinion.submitted_at and opinion.first_response_at:
        response_seconds = (opinion.first_response_at - opinion.submitted_at).total_seconds()
        if response_seconds >= 0:
            sla_detail["response_hours"] = round(response_seconds / 3600, 2)

    default_review_role = infer_review_role(opinion, request.user)
    default_review_role_label = (
        OpinionReview.ReviewRole(default_review_role).label
        if default_review_role in dict(OpinionReview.ReviewRole.choices)
        else ""
    )
    review_status_message = ""
    if not review_allowed:
        review_status_message = f"当前意见状态为 {opinion.get_status_display()}，暂不可提交新的审核意见。"
    elif review_requires_assignment:
        assignee_name = ""
        if opinion.current_reviewer:
            assignee_name = (
                opinion.current_reviewer.get_full_name()
                or opinion.current_reviewer.username
            )
        review_status_message = (
            f"当前审核人：{assignee_name}，请先指派给自己后再审核。"
        )

    api_urls = {
        "assign": reverse("production_quality:opinion-assign", kwargs={"pk": opinion.id}),
        "revert": reverse("production_quality:opinion-revert", kwargs={"pk": opinion.id}),
        "review": reverse(
            "production_quality:opinion-review-list", kwargs={"opinion_pk": opinion.id}
        ),
        "list": reverse("production_quality_pages:opinion_review_list"),
    }

    context = {
        "opinion": opinion,
        "attachments": attachment_entries,
        "history": history_timeline,
        "can_review": can_review,
        "current_user": request.user,
        "participants": participants,
        "saving_items": saving_items,
        "saving_total": saving_total,
        "available_reviewers": available_reviewers,
        "can_assign": can_assign,
        "can_assign_self": can_assign and not is_current_reviewer,
        "can_revert": can_revert,
        "is_current_reviewer": is_current_reviewer,
        "review_allowed": review_allowed,
        "review_requires_assignment": review_requires_assignment,
        "review_status_message": review_status_message,
        "default_review_role": default_review_role,
        "default_review_role_label": default_review_role_label,
        "sla_detail": sla_detail,
        "detail_api_urls": api_urls,
    }
    return render(request, "production_quality/opinion_review_detail.html", context)


@login_required
def report_generate(request):
    """报告生成页面"""
    permission_set = get_user_permission_codes(request.user)
    if not _has_permission(permission_set, "production_quality.generate_report", "production_quality.professional_review"):
        messages.error(request, "您没有访问报告生成中心的权限。")
        return redirect("home")

    accessible_ids = _project_ids_user_can_access(request.user)
    reports = (
        ProductionReport.objects.filter(project_id__in=accessible_ids)
        .select_related("project", "professional_category")
        .order_by("-generated_at", "-updated_at")[:8]
    )
    templates = ReportTemplate.objects.order_by("name")[:6]

    summary_cards = [
        {
            "label": "可生成项目",
            "value": Project.objects.filter(id__in=accessible_ids).count(),
            "hint": "当前可生成报告的项目数量。",
        },
        {
            "label": "近期生成",
            "value": reports.count(),
            "hint": "最近生成的报告份数。",
        },
        {
            "label": "可用模板",
            "value": templates.count(),
            "hint": "开放使用的报告模板数量。",
        },
        {
            "label": "待发布",
            "value": ProductionReport.objects.filter(
                project_id__in=accessible_ids, status=ProductionReport.ReportStatus.GENERATED
            ).count(),
            "hint": "等待审核发布的报告。",
        },
    ]

    template_items = [
        {
            "label": tpl.name,
            "description": tpl.description or "标准化模板，支持快速套用。",
            "url": "#",
            "icon": "📄",
        }
        for tpl in templates
    ]
    report_items = [
        {
            "label": f"{report.project.project_number if report.project else '未关联'} · {report.name}",
            "description": f"状态：{report.get_status_display()}",
            "url": "#",
            "icon": "📝",
        }
        for report in reports
    ]

    context = {
        "page_title": "报告生成中心",
        "page_icon": "🧾",
        "description": "集中管理专业报告的生成、模板选择与发布流程，提升成果输出效率。",
        "summary_cards": summary_cards,
        "sections": [
            {
                "title": "模板资源",
                "description": "选择适合的模板以快速生成专业报告。",
                "items": template_items or [
                    {
                        "label": "暂无模板",
                        "description": "当前未配置报告模板，请联系系统管理员。",
                        "url": "#",
                        "icon": "ℹ️",
                    }
                ],
            },
            {
                "title": "最近生成",
                "description": "最新生成或待发布的报告列表。",
                "items": report_items or [
                    {
                        "label": "暂无报告",
                        "description": "尚未生成任何专业报告。",
                        "url": reverse("production_quality_pages:opinion_review"),
                        "icon": "⏳",
                    }
                ],
            },
        ],
    }
    return render(request, "shared/center_dashboard.html", context)


@login_required
def production_stats(request):
    """生产统计视图"""
    permission_set = get_user_permission_codes(request.user)
    if not _has_permission(permission_set, "production_quality.view_statistics", "production_quality.professional_review"):
        messages.error(request, "您没有查看生产统计的权限。")
        return redirect("home")

    accessible_ids = _project_ids_user_can_access(request.user)
    projects_queryset = Project.objects.filter(id__in=accessible_ids).order_by("project_number")
    project_options = [
        {
            "value": "global",
            "label": "全局汇总",
        }
    ] + [
        {
            "value": str(project.id),
            "label": f"{project.project_number} · {project.name}",
        }
        for project in projects_queryset
    ]

    selected_project_param = request.GET.get("project", "global")
    selected_project = None

    statistics_queryset = (
        ProductionStatistic.objects.filter(
            Q(project__isnull=True) | Q(project_id__in=accessible_ids),
            statistic_type="quality",
        )
        .select_related("project")
        .order_by("-snapshot_date", "-id")
    )

    if selected_project_param == "global" or not selected_project_param:
        statistics_queryset = statistics_queryset.filter(project__isnull=True)
    elif selected_project_param.isdigit():
        project_id = int(selected_project_param)
        if project_id not in accessible_ids:
            messages.error(request, "您无权查看该项目的统计数据。")
            return redirect("production_quality_pages:production_stats")
        selected_project = projects_queryset.filter(id=project_id).first()
        statistics_queryset = statistics_queryset.filter(project_id=project_id)
    else:
        selected_project_param = "global"
        statistics_queryset = statistics_queryset.filter(project__isnull=True)

    export_queryset = statistics_queryset
    stats_records = list(statistics_queryset[:30])
    stats_records_reversed = list(reversed(stats_records))

    latest_snapshot = stats_records[0] if stats_records else None
    latest_payload = latest_snapshot.payload if latest_snapshot else {}

    pending_payload = latest_payload.get("pending", {})
    averages_payload = latest_payload.get("averages", {})
    financial_payload = latest_payload.get("financial", {})
    sla_payload = latest_payload.get("sla", {}) or {}
    compliance_payload = sla_payload.get("compliance", {}) or {}
    reviews_payload = latest_payload.get("reviews", {}) or {}
    reminders_payload = latest_payload.get("reminders", {}) or {}

    avg_cycle_value = averages_payload.get("cycle_time_hours")
    avg_response_value = averages_payload.get("first_response_hours")
    response_rate = compliance_payload.get("response_within_24h", {}).get("rate")
    cycle_rate = compliance_payload.get("cycle_within_7d", {}).get("rate")
    review_status_counts = reviews_payload.get("status", {}) or {}
    review_total = reviews_payload.get("total", 0) or 0
    review_approved = review_status_counts.get(OpinionReview.ReviewStatus.APPROVED, 0)
    review_rejected = review_status_counts.get(OpinionReview.ReviewStatus.REJECTED, 0)
    review_pending = review_status_counts.get(OpinionReview.ReviewStatus.PENDING, 0)
    reminder_pending_total = reminders_payload.get("pending_total", 0)

    summary_cards = [
        {
            "label": "待审核",
            "value": pending_payload.get("total", 0),
            "hint": "最新快照中的待审核意见数量。",
        },
        {
            "label": "未指派",
            "value": pending_payload.get("unassigned", 0),
            "hint": "尚未指派审核人的意见数量。",
        },
        {
            "label": "超期",
            "value": pending_payload.get("overdue", 0),
            "hint": "已超过整改截止时间的意见数量。",
        },
        {
            "label": "累计节省",
            "value": f"¥{financial_payload.get('total_saving', 0):,.0f}",
            "hint": "截至当前的节省金额总计。",
        },
        {
            "label": "平均首响",
            "value": f"{avg_response_value:.1f} h" if avg_response_value is not None else "--",
            "hint": "所有意见的平均首次响应耗时。",
        },
        {
            "label": "平均结案",
            "value": f"{avg_cycle_value:.1f} h" if avg_cycle_value is not None else "--",
            "hint": "所有意见从提交到结案的平均耗时。",
        },
        {
            "label": "24h首响达成率",
            "value": f"{response_rate:.1f}%" if response_rate is not None else "--",
            "hint": "首次响应在 24 小时内完成的意见占比。",
        },
        {
            "label": "7天结案达成率",
            "value": f"{cycle_rate:.1f}%" if cycle_rate is not None else "--",
            "hint": "结案耗时在 7 天内完成的意见占比。",
        },
        {
            "label": "待处理提醒",
            "value": reminder_pending_total,
            "hint": "尚未确认的质量提醒数量。",
        },
        {
            "label": "审核通过率",
            "value": f"{(review_approved / review_total * 100):.1f}%"
            if review_total
            else "--",
            "hint": "已完成审核中通过的占比。",
        },
    ]

    chart_payload = {
        "labels": [],
        "pending": [],
        "overdue": [],
        "saving": [],
        "cycle_hours": [],
        "response_hours": [],
        "alerts_pending": [],
        "response_rate": [],
        "cycle_rate": [],
    }

    for stat in stats_records_reversed:
        label = stat.snapshot_date.strftime("%Y-%m-%d")
        payload = stat.payload or {}
        pending = payload.get("pending", {})
        averages = payload.get("averages", {})
        financial = payload.get("financial", {})
        sla_inner = payload.get("sla", {}) or {}
        compliance_inner = sla_inner.get("compliance", {}) or {}
        reminders_inner = payload.get("reminders", {}) or {}

        chart_payload["labels"].append(label)
        chart_payload["pending"].append(pending.get("total", 0))
        chart_payload["overdue"].append(pending.get("overdue", 0))
        chart_payload["saving"].append(float(financial.get("total_saving", 0) or 0))
        cycle_value = averages.get("cycle_time_hours")
        response_value = averages.get("first_response_hours")
        chart_payload["cycle_hours"].append(float(cycle_value) if cycle_value is not None else None)
        chart_payload["response_hours"].append(float(response_value) if response_value is not None else None)
        chart_payload["alerts_pending"].append(reminders_inner.get("pending_total", 0))
        chart_payload["response_rate"].append(
            compliance_inner.get("response_within_24h", {}).get("rate")
        )
        chart_payload["cycle_rate"].append(
            compliance_inner.get("cycle_within_7d", {}).get("rate")
        )

    table_rows = []
    for stat in stats_records[:10]:
        payload = stat.payload or {}
        pending = payload.get("pending", {}) or {}
        financial = payload.get("financial", {}) or {}
        averages = payload.get("averages", {}) or {}
        sla_inner = payload.get("sla", {}) or {}
        compliance_inner = sla_inner.get("compliance", {}) or {}
        reminders_inner = payload.get("reminders", {}) or {}
        reviews_inner = payload.get("reviews", {}) or {}
        review_status_inner = reviews_inner.get("status", {}) or {}
        review_total_inner = reviews_inner.get("total", 0) or 0

        approved_count = review_status_inner.get(OpinionReview.ReviewStatus.APPROVED, 0)
        approval_rate = (
            round(approved_count / review_total_inner * 100, 1)
            if review_total_inner
            else None
        )
        table_rows.append(
            {
                "id": stat.id,
                "date": stat.snapshot_date,
                "project": stat.project,
                "pending": pending.get("total", 0),
                "overdue": pending.get("overdue", 0),
                "saving": financial.get("total_saving", 0),
                "cycle": averages.get("cycle_time_hours"),
                "response_rate": compliance_inner.get("response_within_24h", {}).get("rate"),
                "alerts": reminders_inner.get("pending_total", 0),
                "approval_rate": approval_rate,
            }
        )

    export_headers = [
        "snapshot_date",
        "project_number",
        "pending_total",
        "pending_unassigned",
        "pending_overdue",
        "avg_cycle_hours",
        "avg_response_hours",
        "total_saving",
        "recent_saving",
        "response_within_24h_rate",
        "cycle_within_7d_rate",
        "review_total",
        "review_approved",
        "review_rejected",
        "reminders_pending",
        "reminders_sent_last_7_days",
        "reminders_ack_last_7_days",
    ]
    export_rows = []
    export_records = []
    for stat in export_queryset.iterator():
        payload = stat.payload or {}
        pending = payload.get("pending", {}) or {}
        averages = payload.get("averages", {}) or {}
        financial = payload.get("financial", {}) or {}
        sla_inner = payload.get("sla", {}) or {}
        compliance_inner = sla_inner.get("compliance", {}) or {}
        reviews_inner = payload.get("reviews", {}) or {}
        reminders_inner = payload.get("reminders", {}) or {}
        review_status_inner = reviews_inner.get("status", {}) or {}
        export_rows.append([
            stat.snapshot_date.strftime("%Y-%m-%d"),
            stat.project.project_number if stat.project else "GLOBAL",
            pending.get("total", 0),
            pending.get("unassigned", 0),
            pending.get("overdue", 0),
            averages.get("cycle_time_hours", ""),
            averages.get("first_response_hours", ""),
            financial.get("total_saving", ""),
            financial.get("recent_saving", ""),
            compliance_inner.get("response_within_24h", {}).get("rate", ""),
            compliance_inner.get("cycle_within_7d", {}).get("rate", ""),
            reviews_inner.get("total", 0),
            review_status_inner.get(OpinionReview.ReviewStatus.APPROVED, 0),
            review_status_inner.get(OpinionReview.ReviewStatus.REJECTED, 0),
            reminders_inner.get("pending_total", 0),
            reminders_inner.get("sent_last_7_days", 0),
            reminders_inner.get("ack_last_7_days", 0),
        ])
        export_records.append(
            {
                "snapshot_date": stat.snapshot_date.strftime("%Y-%m-%d"),
                "project": stat.project.project_number if stat.project else "GLOBAL",
                "pending": pending,
                "averages": averages,
                "financial": financial,
                "sla": sla_inner,
                "reviews": reviews_inner,
                "reminders": reminders_inner,
            }
        )

    export_format = (request.GET.get("export") or "").lower()
    if export_format == "csv":
        response = HttpResponse(content_type="text/csv")
        filename_suffix = selected_project.project_number if selected_project else "global"
        response["Content-Disposition"] = f'attachment; filename="statistics_{filename_suffix}.csv"'
        writer = csv.writer(response)
        writer.writerow(export_headers)
        writer.writerows(export_rows)
        return response
    if export_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Statistics"
        sheet.append(export_headers)
        for row in export_rows:
            sheet.append(row)
        output_stream = io.BytesIO()
        workbook.save(output_stream)
        output_stream.seek(0)
        response = HttpResponse(
            output_stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename_suffix = selected_project.project_number if selected_project else "global"
        response["Content-Disposition"] = f'attachment; filename="statistics_{filename_suffix}.xlsx"'
        return response
    if export_format == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            messages.error(request, "服务器未安装 reportlab 库，暂无法导出 PDF。")
        else:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            table = Table([export_headers] + export_rows, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightyellow]),
            ]))
            story = [Paragraph("生产统计快照导出", styles["Heading3"]), Spacer(1, 12), table]
            doc.build(story)
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
            filename_suffix = selected_project.project_number if selected_project else "global"
            response["Content-Disposition"] = f'attachment; filename="statistics_{filename_suffix}.pdf"'
            return response
    if export_format == "json":
        return JsonResponse(
            {
                "project": selected_project.project_number if selected_project else "GLOBAL",
                "records": export_records,
            }
        )

    context = {
        "selected_project": selected_project,
        "selected_project_value": selected_project_param,
        "project_options": project_options,
        "summary_cards": summary_cards,
        "chart_payload": chart_payload,
        "latest_snapshot": latest_snapshot,
        "table_rows": table_rows,
        "has_data": bool(stats_records),
        "review_summary": {
            "total": review_total,
            "approved": review_approved,
            "rejected": review_rejected,
            "pending": review_pending,
        },
        "reminder_summary": {
            "pending_total": reminder_pending_total,
            "pending_by_type": reminders_payload.get("pending_by_type", {}),
            "sent_last_7_days": reminders_payload.get("sent_last_7_days", 0),
            "ack_last_7_days": reminders_payload.get("ack_last_7_days", 0),
        },
        "sla_compliance": {
            "response_rate": response_rate,
            "cycle_rate": cycle_rate,
        },
    }
    return render(request, "production_quality/statistics_dashboard.html", context)

def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_decimal(value: Any, errors: List[str], label: str, row_index: int) -> Decimal | None:
    if value in (None, "", "-"):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        errors.append(f"第 {row_index} 行字段“{label}”格式错误，需为数字。")
        return None


def _resolve_choice(value: Any, lookup: Dict[str, str], label: str, row_index: int) -> str | None:
    if value in (None, "", "-"):
        return None
    candidate = str(value).strip()
    resolved = lookup.get(candidate) or lookup.get(candidate.lower())
    if not resolved:
        resolved = lookup.get(candidate.upper())
    if not resolved:
        choices = " / ".join(sorted({v for v in lookup.values()}))
        raise ValueError(f"第 {row_index} 行字段“{label}”无法识别：{candidate}（支持：{choices}）")
    return resolved


def _build_import_preview(upload, user):
    upload.seek(0)
    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception as exc:
        raise ValueError("无法读取 Excel 文件，请确认文件格式是否正确。") from exc

    sheet = workbook.active
    try:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration as exc:
        raise ValueError("Excel 文件为空，缺少表头。") from exc

    headers = [(_normalize_text(cell) if cell is not None else "") for cell in header_row]
    header_map = {header: index for index, header in enumerate(headers) if header}

    missing_headers = [label for _, label, required in IMPORT_COLUMNS if required and label not in header_map]
    if missing_headers:
        raise ValueError(f"缺少必填列：{'、'.join(missing_headers)}。请使用最新版模板。")

    accessible_projects = set(_project_ids_user_can_access(user))
    rows: List[Dict[str, Any]] = []
    payload: List[Dict[str, Any]] = []
    total_saving = Decimal("0")

    for absolute_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(_is_blank(cell) for cell in row):
            continue

        row_data = {}
        for key, label, _ in IMPORT_COLUMNS:
            if label in header_map and header_map[label] < len(row):
                row_data[key] = row[header_map[label]]
            else:
                row_data[key] = None

        errors: List[str] = []

        project_number = _normalize_text(row_data["project_number"])
        if not project_number:
            errors.append(f"第 {absolute_index} 行缺少项目编号。")
        project = None
        if project_number:
            project = Project.objects.filter(project_number=project_number).first()
            if not project:
                errors.append(f"第 {absolute_index} 行未找到项目编号“{project_number}”。")
            elif project.id not in accessible_projects:
                errors.append(f"第 {absolute_index} 行项目“{project_number}”无访问权限。")

        professional_code = _normalize_text(row_data["professional_code"])
        category = None
        if not professional_code:
            errors.append(f"第 {absolute_index} 行缺少专业编码。")
        else:
            category = ProfessionalCategory.objects.filter(code=professional_code).first()
            if not category:
                errors.append(f"第 {absolute_index} 行未找到专业编码“{professional_code}”。")

        location_name = _normalize_text(row_data["location_name"])
        if not location_name:
            errors.append(f"第 {absolute_index} 行缺少部位名称。")

        issue_description = _normalize_text(row_data["issue_description"])
        if not issue_description:
            errors.append(f"第 {absolute_index} 行缺少问题描述。")

        recommendation = _normalize_text(row_data["recommendation"])
        if not recommendation:
            errors.append(f"第 {absolute_index} 行缺少优化建议。")

        try:
            issue_category = _resolve_choice(
                row_data["issue_category"], ISSUE_CATEGORY_LOOKUP, "问题类别", absolute_index
            )
        except ValueError as exc:
            errors.append(str(exc))
            issue_category = None

        try:
            severity_level = _resolve_choice(
                row_data["severity_level"], SEVERITY_LOOKUP, "严重等级", absolute_index
            )
        except ValueError as exc:
            errors.append(str(exc))
            severity_level = None

        calc_mode_value = row_data.get("calculation_mode")
        calc_mode = Opinion.CalculationMode.AUTO
        if not _is_blank(calc_mode_value):
            try:
                calc_mode = _resolve_choice(
                    calc_mode_value, CALCULATION_MODE_LOOKUP, "计算方式", absolute_index
                )
            except ValueError as exc:
                errors.append(str(exc))

        quantity_before = _parse_decimal(row_data["quantity_before"], errors, "优化前工程量", absolute_index)
        quantity_after = _parse_decimal(row_data["quantity_after"], errors, "优化后工程量", absolute_index)
        unit_price_before = _parse_decimal(row_data["unit_price_before"], errors, "优化前综合单价", absolute_index)
        unit_price_after = _parse_decimal(row_data["unit_price_after"], errors, "优化后综合单价", absolute_index)
        manual_saving = _parse_decimal(row_data["saving_amount"], errors, "节省金额", absolute_index)

        if calc_mode == Opinion.CalculationMode.AUTO:
            if any(val is None for val in (quantity_before, quantity_after, unit_price_before, unit_price_after)):
                errors.append(f"第 {absolute_index} 行为自动计算模式，需要完整填写工程量与单价。")
            saving_amount = (
                calculate_saving_amount(quantity_before, quantity_after, unit_price_before, unit_price_after)
                if not errors
                else None
            )
        else:
            if manual_saving is None:
                errors.append(f"第 {absolute_index} 行为手动模式，请填写节省金额。")
            saving_amount = manual_saving

        reference_codes = _normalize_text(row_data.get("reference_codes"))
        calculation_note = _normalize_text(row_data.get("calculation_note"))
        current_practice = _normalize_text(row_data.get("current_practice"))
        drawing_number = _normalize_text(row_data.get("drawing_number"))
        drawing_version = _normalize_text(row_data.get("drawing_version"))
        measure_unit = _normalize_text(row_data.get("measure_unit"))

        rows.append(
            {
                "index": absolute_index,
                "project_number": project_number,
                "project_name": project.name if project else "",
                "professional_code": professional_code,
                "professional_name": category.name if category else "",
                "location_name": location_name,
                "issue_category": issue_category,
                "severity_level": severity_level,
                "calculation_mode": calc_mode,
                "calculation_mode_label": Opinion.CalculationMode(calc_mode).label
                if calc_mode in dict(Opinion.CalculationMode.choices)
                else "",
                "saving_amount": saving_amount,
                "reference_codes": reference_codes,
                "calculation_note": calculation_note,
                "current_practice": current_practice,
                "drawing_number": drawing_number,
                "drawing_version": drawing_version,
                "measure_unit": measure_unit,
                "quantity_before": quantity_before,
                "quantity_after": quantity_after,
                "unit_price_before": unit_price_before,
                "unit_price_after": unit_price_after,
                "errors": errors,
                "issue_description": issue_description,
                "recommendation": recommendation,
            }
        )

        if errors:
            continue

        commit_payload = {
            "project_id": project.id,
            "professional_category_id": category.id,
            "drawing_number": drawing_number,
            "drawing_version": drawing_version,
            "location_name": location_name,
            "issue_description": issue_description,
            "current_practice": current_practice,
            "recommendation": recommendation,
            "issue_category": issue_category,
            "severity_level": severity_level,
            "reference_codes": reference_codes,
            "calculation_mode": calc_mode,
            "quantity_before": str(quantity_before) if quantity_before is not None else "",
            "quantity_after": str(quantity_after) if quantity_after is not None else "",
            "measure_unit": measure_unit,
            "unit_price_before": str(unit_price_before) if unit_price_before is not None else "",
            "unit_price_after": str(unit_price_after) if unit_price_after is not None else "",
            "saving_amount": str(saving_amount) if saving_amount is not None else "",
            "calculation_note": calculation_note,
        }
        payload.append(commit_payload)

        if saving_amount is not None:
            total_saving += saving_amount

    summary = {
        "total": len(rows),
        "success": sum(1 for row in rows if not row["errors"]),
        "failed": sum(1 for row in rows if row["errors"]),
        "total_saving": f"{total_saving:.2f}",
    }
    return rows, summary, payload


@login_required
def opinion_import_template(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "意见导入模板"

    worksheet.append([label for _, label, _ in IMPORT_COLUMNS])
    worksheet.append(
        [
            "PJT-2025-001",
            "ARCH",
            "A-101",
            "V2",
            "主体结构-三层梁板",
            "楼板 reinforcement overlap 不足，存在承载风险",
            "当前按原图施工",
            "增加钢筋搭接长度并补强",
            "错误",
            "重大",
            "GB50010-2010",
            "自动计算",
            120,
            100,
            "㎡",
            520,
            480,
            "",
            "调整钢筋搭接长度，确保满足规范要求",
        ]
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="opinion_import_template.xlsx"'
    workbook.save(response)
    return response


@login_required
def opinion_import(request):
    preview_rows: List[Dict[str, Any]] = []
    summary: Dict[str, Any] | None = None
    payload_json = ""

    if request.method == "POST" and "payload" in request.POST:
        payload = json.loads(request.POST.get("payload") or "[]")
        if not payload:
            messages.error(request, "导入数据为空，请重新上传。")
            return redirect("production_quality_pages:opinion_import")

        created_count = 0
        total_saving = Decimal("0")
        accessible_projects = set(_project_ids_user_can_access(request.user))
        try:
            with transaction.atomic():
                for item in payload:
                    project = Project.objects.filter(id=item["project_id"]).first()
                    category = ProfessionalCategory.objects.filter(id=item["professional_category_id"]).first()
                    if not project or not category:
                        raise ValueError("导入数据中的项目或专业分类已不存在，请重新导入。")
                    if project.id not in accessible_projects:
                        raise ValueError(f"项目 {project.project_number} 对当前用户不可用，请重新选择。")
                    opinion = Opinion(
                        opinion_number=generate_opinion_number(project, category),
                        project=project,
                        professional_category=category,
                        created_by=request.user,
                        status=Opinion.OpinionStatus.DRAFT,
                        drawing_number=item.get("drawing_number", ""),
                        drawing_version=item.get("drawing_version", ""),
                        location_name=item.get("location_name", ""),
                        issue_description=item.get("issue_description", ""),
                        current_practice=item.get("current_practice", ""),
                        recommendation=item.get("recommendation", ""),
                        issue_category=item.get("issue_category", Opinion.IssueCategory.ERROR),
                        severity_level=item.get("severity_level", Opinion.SeverityLevel.NORMAL),
                        reference_codes=item.get("reference_codes", ""),
                        calculation_mode=item.get("calculation_mode", Opinion.CalculationMode.AUTO),
                        measure_unit=item.get("measure_unit", ""),
                        calculation_note=item.get("calculation_note", ""),
                    )
                    for field in ("quantity_before", "quantity_after", "unit_price_before", "unit_price_after"):
                        raw_value = item.get(field) or ""
                        setattr(
                            opinion,
                            field,
                            Decimal(raw_value) if raw_value not in ("", None) else None,
                        )
                    raw_saving = item.get("saving_amount") or ""
                    if raw_saving not in ("", None):
                        opinion.saving_amount = Decimal(raw_saving)
                    else:
                        if opinion.calculation_mode == Opinion.CalculationMode.AUTO:
                            opinion.saving_amount = calculate_saving_amount(
                                opinion.quantity_before,
                                opinion.quantity_after,
                                opinion.unit_price_before,
                                opinion.unit_price_after,
                            )
                    opinion.save()
                    created_count += 1
                    if opinion.saving_amount is not None:
                        total_saving += opinion.saving_amount
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("production_quality_pages:opinion_import")

        messages.success(
            request,
            f"成功导入 {created_count} 条意见，预计节省金额合计 ¥{total_saving:.2f}。",
        )
        return redirect("production_quality_pages:opinion_import")

    form = OpinionBulkImportForm()
    if request.method == "POST" and request.FILES:
        form = OpinionBulkImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                preview_rows, summary, payload = _build_import_preview(
                    form.cleaned_data["file"], request.user
                )
            except ValueError as exc:
                form.add_error("file", str(exc))
            else:
                if not preview_rows:
                    messages.warning(request, "未检测到有效数据行。")
                else:
                    if summary["failed"]:
                        messages.warning(
                            request,
                            f"共解析 {summary['total']} 条记录，其中 {summary['failed']} 条存在错误，请修正后重新上传。",
                        )
                    if summary["success"] > 0 and summary["failed"] == 0:
                        payload_json = json.dumps(payload, ensure_ascii=False)
                    elif summary["success"] > 0:
                        messages.info(
                            request,
                            "部分记录校验通过，如需导入请剔除错误行后重新上传。",
                        )
                    else:
                        messages.success(
                            request,
                            f"共解析 {summary['total']} 条记录，校验通过，可点击“确认导入”写入系统。",
                        )
    context = {
        "form": form,
        "preview_rows": preview_rows,
        "summary": summary,
        "payload_json": payload_json,
        "import_columns": IMPORT_COLUMNS,
    }
    return render(request, "production_quality/opinion_import.html", context)


@login_required
def opinion_drafts(request):
    return render(request, "production_quality/opinion_drafts.html")


@login_required
def opinion_review_list(request):
    return render(request, "production_quality/opinion_review_list.html")

