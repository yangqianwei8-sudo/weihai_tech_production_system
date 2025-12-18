from __future__ import annotations

import csv
import io
import os
import tempfile
from unittest import mock

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import Client, TestCase
from django.urls import reverse

from backend.apps.production_management.models import Project
from backend.apps.production_quality.models import OpinionReview, ProductionStatistic

User = get_user_model()


class ExportStatisticsTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="tester", password="pass1234")
        self.project = Project.objects.create(name="导出项目", project_number="EXP-001")
        ProductionStatistic.objects.create(
            project=None,
            statistic_type="quality",
            snapshot_date="2025-01-01",
            payload={
                "pending": {"total": 2, "unassigned": 1, "overdue": 1},
                "averages": {"cycle_time_hours": 12, "first_response_hours": 4},
                "financial": {"total_saving": 1000, "recent_saving": 200},
                "sla": {
                    "compliance": {
                        "response_within_24h": {"rate": 50.0, "met": 1, "total": 2},
                        "cycle_within_7d": {"rate": 75.0, "met": 3, "total": 4},
                    },
                },
                "reviews": {
                    "total": 4,
                    "status": {
                        OpinionReview.ReviewStatus.APPROVED: 3,
                        OpinionReview.ReviewStatus.REJECTED: 1,
                    },
                },
                "reminders": {
                    "pending_total": 2,
                    "sent_last_7_days": 5,
                    "ack_last_7_days": 4,
                },
            },
        )
        ProductionStatistic.objects.create(
            project=self.project,
            statistic_type="quality",
            snapshot_date="2025-01-02",
            payload={
                "pending": {"total": 5, "unassigned": 2, "overdue": 1},
                "averages": {"cycle_time_hours": 24, "first_response_hours": 6},
                "financial": {"total_saving": 5000, "recent_saving": 800},
                "sla": {
                    "compliance": {
                        "response_within_24h": {"rate": 80.0, "met": 8, "total": 10},
                        "cycle_within_7d": {"rate": 60.0, "met": 6, "total": 10},
                    },
                },
                "reviews": {
                    "total": 6,
                    "status": {
                        OpinionReview.ReviewStatus.APPROVED: 4,
                        OpinionReview.ReviewStatus.REJECTED: 1,
                    },
                },
                "reminders": {
                    "pending_total": 1,
                    "sent_last_7_days": 3,
                    "ack_last_7_days": 2,
                },
            },
        )

    def test_export_statistics_command_csv(self):
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tmp:
            path = tmp.name
        try:
            call_command("export_statistics", output=path, format="csv")
            with open(path, newline="", encoding="utf-8") as csvfile:
                reader = csv.reader(csvfile)
                rows = list(reader)
            self.assertGreaterEqual(len(rows), 2)
            self.assertEqual(rows[0][0], "snapshot_date")
            self.assertIn("response_within_24h_rate", rows[0])
            self.assertEqual(len(rows[0]), 17)
        finally:
            os.remove(path)

    def test_export_statistics_command_xlsx(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        try:
            call_command("export_statistics", output=path, format="xlsx")
            from openpyxl import load_workbook

            workbook = load_workbook(path)
            sheet = workbook.active
            self.assertEqual(sheet.title, "Statistics")
            self.assertGreaterEqual(sheet.max_row, 2)
        finally:
            os.remove(path)

    @mock.patch("backend.apps.production_quality.views_pages.get_user_permission_codes", return_value={"production_quality.view_statistics"})
    @mock.patch("backend.apps.production_quality.views_pages._has_permission", return_value=True)
    @mock.patch("backend.apps.production_quality.views_pages._project_ids_user_can_access", return_value=set())
    def test_view_export_csv(self, _access_ids, _has_perm, _perm_codes):
        client = Client()
        client.force_login(self.user)
        url = reverse("production_quality_pages:production_stats") + "?export=csv"
        response = client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        content = response.content.decode("utf-8")
        self.assertIn("snapshot_date", content)
        self.assertIn("response_within_24h_rate", content)

    @mock.patch("backend.apps.production_quality.views_pages.get_user_permission_codes", return_value={"production_quality.view_statistics"})
    @mock.patch("backend.apps.production_quality.views_pages._has_permission", return_value=True)
    @mock.patch("backend.apps.production_quality.views_pages._project_ids_user_can_access", return_value=set())
    def test_view_export_json(self, _access_ids, _has_perm, _perm_codes):
        client = Client()
        client.force_login(self.user)
        url = reverse("production_quality_pages:production_stats") + "?export=json"
        response = client.get(url)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn("records", data)
        self.assertTrue(data["records"])
        first_record = data["records"][0]
        self.assertIn("sla", first_record)
        self.assertIn("reminders", first_record)
