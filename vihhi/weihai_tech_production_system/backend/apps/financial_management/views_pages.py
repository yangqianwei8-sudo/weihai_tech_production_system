from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Count, Sum, Q, F, Max
from django.core.paginator import Paginator
from django.urls import reverse, NoReverseMatch
from django.utils import timezone
from django.forms import inlineformset_factory
from django.http import HttpResponse, JsonResponse
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from backend.apps.system_management.services import get_user_permission_codes
from backend.core.views import HOME_NAV_STRUCTURE, _permission_granted as core_permission_granted, _build_full_top_nav
from backend.apps.financial_management.models import (
    AccountSubject, Voucher, VoucherEntry,
    Ledger, Budget, Invoice, FundFlow,
    FinancialReport, ReceivableAccount, PayableAccount,
)
from .forms import (
    AccountSubjectForm, VoucherForm, VoucherEntryForm, BudgetForm, InvoiceForm, FundFlowForm
)


def _permission_granted(required_code, user_permissions: set) -> bool:
    """æ£€æŸ¥æƒé™"""
    if not required_code:
        return True
    if '__all__' in user_permissions:
        return True
    return required_code in user_permissions


def _generate_voucher_number(voucher_date=None):
    """
    è‡ªåŠ¨ç”Ÿæˆå‡­è¯å­—å·
    
    Args:
        voucher_date: å‡­è¯æ—¥æœŸï¼Œå¦‚æœä¸ºNoneåˆ™ä½¿ç”¨å½“å‰æ—¥æœŸ
    
    Returns:
        str: ç”Ÿæˆçš„å‡­è¯å­—å·ï¼Œæ ¼å¼ä¸º VOUCHER-YYYY-NNNN
    """
    if voucher_date:
        current_year = voucher_date.year
    else:
        current_year = timezone.now().year
    
    # æŸ¥æ‰¾å½“å‰å¹´åº¦æœ€å¤§çš„å‡­è¯å·
    max_voucher = Voucher.objects.filter(
        voucher_number__startswith=f'VOUCHER-{current_year}-'
    ).aggregate(max_num=Max('voucher_number'))['max_num']
    
    if max_voucher:
        try:
            # æå–åºå·éƒ¨åˆ†
            seq = int(max_voucher.split('-')[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    else:
        seq = 1
    
    return f'VOUCHER-{current_year}-{seq:04d}'


def _update_budget_from_fund_flow(fund_flow, is_create=True, old_amount=None):
    """
    æ ¹æ®èµ„é‡‘æµæ°´è‡ªåŠ¨æ›´æ–°é¢„ç®—ä½¿ç”¨é‡‘é¢
    
    Args:
        fund_flow: èµ„é‡‘æµæ°´å¯¹è±¡
        is_create: æ˜¯å¦ä¸ºåˆ›å»ºæ“ä½œï¼ˆTrue=åˆ›å»ºï¼ŒFalse=æ›´æ–°æˆ–åˆ é™¤ï¼‰
        old_amount: æ—§é‡‘é¢ï¼ˆæ›´æ–°æˆ–åˆ é™¤æ—¶éœ€è¦ï¼‰
    """
    try:
        # åªæœ‰æ”¯å‡ºç±»å‹çš„æµæ°´æ‰æ›´æ–°é¢„ç®—
        if fund_flow.flow_type != 'expense':
            return
        
        # æŸ¥æ‰¾ç›¸å…³é¢„ç®—ï¼ˆæ ¹æ®æ—¥æœŸèŒƒå›´ã€çŠ¶æ€ï¼‰
        budgets = Budget.objects.filter(
            status__in=['approved', 'executing'],
            start_date__lte=fund_flow.flow_date,
            end_date__gte=fund_flow.flow_date,
        )
        
        # å¦‚æœæœ‰å…³è”é¡¹ç›®ï¼Œå¯ä»¥é€šè¿‡é¡¹ç›®æŸ¥æ‰¾ç›¸å…³é¢„ç®—
        # ä¼˜å…ˆåŒ¹é…å…³è”é¡¹ç›®çš„é¢„ç®—
        if fund_flow.project:
            project_budgets = budgets.filter(project=fund_flow.project)
            if project_budgets.exists():
                budgets = project_budgets
        
        # å¦‚æœèµ„é‡‘æµæ°´å…³è”äº†å‡­è¯ï¼Œå¯ä»¥é€šè¿‡å‡­è¯åˆ†å½•çš„ä¼šè®¡ç§‘ç›®åŒ¹é…é¢„ç®—ç§‘ç›®
        if fund_flow.voucher and fund_flow.voucher.entries.exists():
            # è·å–å‡­è¯ä¸­æ”¯å‡ºç›¸å…³çš„ä¼šè®¡ç§‘ç›®
            expense_subjects = set()
            for entry in fund_flow.voucher.entries.all():
                if entry.debit_amount > 0:  # å€Ÿæ–¹é‡‘é¢è¡¨ç¤ºæ”¯å‡º
                    expense_subjects.add(entry.account_subject)
            
            # å¦‚æœé¢„ç®—æŒ‡å®šäº†ä¼šè®¡ç§‘ç›®ï¼Œä¼˜å…ˆåŒ¹é…
            if expense_subjects:
                subject_budgets = budgets.filter(account_subject__in=expense_subjects)
                if subject_budgets.exists():
                    budgets = subject_budgets
        
        # æ›´æ–°é¢„ç®—ä½¿ç”¨é‡‘é¢
        if is_create:
            # åˆ›å»ºï¼šå¢åŠ ä½¿ç”¨é‡‘é¢
            amount_to_add = fund_flow.amount
            for budget in budgets:
                budget.used_amount += amount_to_add
                budget.remaining_amount = budget.budget_amount - budget.used_amount
                # å¦‚æœå‰©ä½™é‡‘é¢ä¸º0æˆ–è´Ÿæ•°ï¼Œæ›´æ–°çŠ¶æ€
                if budget.remaining_amount <= 0:
                    budget.status = 'completed'
                budget.save()
        else:
            # æ›´æ–°æˆ–åˆ é™¤ï¼šè®¡ç®—é‡‘é¢å·®å¼‚
            if old_amount is not None:
                amount_diff = fund_flow.amount - old_amount
                for budget in budgets:
                    budget.used_amount += amount_diff
                    budget.remaining_amount = budget.budget_amount - budget.used_amount
                    # å¦‚æœå‰©ä½™é‡‘é¢ä¸º0æˆ–è´Ÿæ•°ï¼Œæ›´æ–°çŠ¶æ€
                    if budget.remaining_amount <= 0:
                        budget.status = 'completed'
                    elif budget.remaining_amount > 0 and budget.status == 'completed':
                        budget.status = 'executing'
                    budget.save()
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('æ›´æ–°é¢„ç®—å¤±è´¥: %s', str(e))
        # ä¸å½±å“èµ„é‡‘æµæ°´çš„ä¿å­˜ï¼Œåªè®°å½•æ—¥å¿—


# ä½¿ç”¨ç»Ÿä¸€çš„é¡¶éƒ¨å¯¼èˆªèœå•ç”Ÿæˆå‡½æ•°
from backend.core.views import _build_full_top_nav


def _context(page_title, page_icon, description, summary_cards=None, request=None, use_financial_nav=False):
    """æ„å»ºé¡µé¢ä¸Šä¸‹æ–‡"""
    context = {
        "page_title": page_title,
        "page_icon": page_icon,
        "description": description,
        "summary_cards": summary_cards or [],
    }
    
    try:
        if request and request.user.is_authenticated:
            permission_set = get_user_permission_codes(request.user)
            # ç»Ÿä¸€ä½¿ç”¨å…¨å±€ç³»ç»Ÿä¸»èœå•ï¼ˆä¸å®¢æˆ·ç®¡ç†æ¨¡å—ä¿æŒä¸€è‡´ï¼‰
            context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
            if use_financial_nav:
                context['financial_menu'] = _build_financial_sidebar_nav(permission_set, request.path)
        else:
            context['full_top_nav'] = []
            context['financial_menu'] = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'æ„å»ºé¡µé¢ä¸Šä¸‹æ–‡é”™è¯¯: {str(e)}', exc_info=True)
        context['full_top_nav'] = []
        context['financial_menu'] = []
    
    return context


def _build_financial_top_nav(permission_set):
    """
    ç”Ÿæˆè´¢åŠ¡ç®¡ç†ä¸“ç”¨çš„é¡¶éƒ¨å¯¼èˆªèœå• - å·²åºŸå¼ƒ
    
    æ³¨æ„ï¼šæ­¤å‡½æ•°å·²ä¸å†ä½¿ç”¨ï¼Œç³»ç»Ÿç°åœ¨ç»Ÿä¸€ä½¿ç”¨ _build_full_top_nav ç”Ÿæˆå…¨å±€ç³»ç»Ÿä¸»èœå•ã€‚
    ä¿ç•™æ­¤å‡½æ•°ä»…ç”¨äºå†å²å‚è€ƒï¼Œå¯ä»¥å®‰å…¨åˆ é™¤ã€‚
    """
    # å®šä¹‰è´¢åŠ¡ç®¡ç†åŠŸèƒ½æ¨¡å—ï¼ˆä»å·¦åˆ°å³çš„é¡ºåºï¼‰
    financial_modules = [
        {
            'label': 'ä¼šè®¡ç§‘ç›®',
            'url_name': 'finance_pages:account_subject_management',
            'permission': 'financial_management.account.view',
            'icon': 'ğŸ“Š',
        },
        {
            'label': 'å‡­è¯ç®¡ç†',
            'url_name': 'finance_pages:voucher_management',
            'permission': 'financial_management.voucher.view',
            'icon': 'ğŸ“',
        },
        {
            'label': 'è´¦ç°¿ç®¡ç†',
            'url_name': 'finance_pages:ledger_management',
            'permission': 'financial_management.ledger.view',
            'icon': 'ğŸ“–',
        },
        {
            'label': 'é¢„ç®—ç®¡ç†',
            'url_name': 'finance_pages:budget_management',
            'permission': 'financial_management.budget.view',
            'icon': 'ğŸ’°',
        },
        {
            'label': 'å‘ç¥¨ç®¡ç†',
            'url_name': 'finance_pages:invoice_management',
            'permission': 'financial_management.invoice.view',
            'icon': 'ğŸ§¾',
        },
        {
            'label': 'èµ„é‡‘æµæ°´',
            'url_name': 'finance_pages:fund_flow_management',
            'permission': 'financial_management.fund_flow.view',
            'icon': 'ğŸ’³',
        },
        {
            'label': 'è´¢åŠ¡æŠ¥è¡¨',
            'url_name': 'finance_pages:report_management',
            'permission': 'financial_management.report.view',
            'icon': 'ğŸ“Š',
        },
        {
            'label': 'åº”æ”¶è´¦æ¬¾',
            'url_name': 'finance_pages:receivable_management',
            'permission': 'financial_management.receivable.view',
            'icon': 'ğŸ’°',
        },
        {
            'label': 'åº”ä»˜è´¦æ¬¾',
            'url_name': 'finance_pages:payable_management',
            'permission': 'financial_management.payable.view',
            'icon': 'ğŸ’¸',
        },
    ]
    
    # è¿‡æ»¤æœ‰æƒé™çš„æ¨¡å—
    nav_items = []
    for module in financial_modules:
        if _permission_granted(module['permission'], permission_set):
            try:
                url = reverse(module['url_name'])
            except NoReverseMatch:
                url = '#'
            nav_items.append({
                'label': module['label'],
                'url': url,
                'icon': module.get('icon', ''),
            })
    
    return nav_items


def _build_financial_sidebar_nav(permission_set, request_path=None, active_id=None):
    """ç”Ÿæˆè´¢åŠ¡ç®¡ç†æ¨¡å—çš„å·¦ä¾§èœå•å¯¼èˆªï¼ˆä½¿ç”¨è®¡åˆ’ç®¡ç†æ ¼å¼ï¼‰
    
    Args:
        permission_set: ç”¨æˆ·æƒé™é›†åˆ
        request_path: å½“å‰è¯·æ±‚è·¯å¾„ï¼Œç”¨äºåˆ¤æ–­æ¿€æ´»çŠ¶æ€
        active_id: å½“å‰æ¿€æ´»çš„èœå•é¡¹IDï¼ˆå¯é€‰ï¼‰
    
    Returns:
        list: åˆ†ç»„èœå•é¡¹åˆ—è¡¨ï¼Œæ ¼å¼ä¸è®¡åˆ’ç®¡ç†ä¸€è‡´
    """
    from django.urls import reverse, NoReverseMatch
    
    # å®šä¹‰è´¢åŠ¡ç®¡ç†èœå•ç»“æ„ï¼ˆåˆ†ç»„æ ¼å¼ï¼Œä¸è®¡åˆ’ç®¡ç†ä¸€è‡´ï¼‰
    FINANCIAL_MENU_STRUCTURE = [
        {
            'id': 'financial_basic',
            'label': 'åŸºç¡€ç®¡ç†',
            'icon': 'ğŸ“Š',
            'permission': 'financial_management.account.view',
            'children': [
                {
                    'id': 'financial_home',
                    'label': 'è´¢åŠ¡ç®¡ç†é¦–é¡µ',
                    'icon': 'ğŸ’µ',
                    'url_name': 'finance_pages:financial_home',
                    'permission': None,
                    'path_keywords': ['financial_home', 'financial'],
                },
                {
                    'id': 'account_subject',
                    'label': 'ä¼šè®¡ç§‘ç›®',
                    'icon': 'ğŸ“Š',
                    'url_name': 'finance_pages:account_subject_management',
                    'permission': 'financial_management.account.view',
                    'path_keywords': ['account', 'accounts'],
                },
                {
                    'id': 'voucher',
                    'label': 'å‡­è¯ç®¡ç†',
                    'icon': 'ğŸ“',
                    'url_name': 'finance_pages:voucher_management',
                    'permission': 'financial_management.voucher.view',
                    'path_keywords': ['voucher', 'vouchers'],
                },
            ]
        },
        {
            'id': 'financial_ledger',
            'label': 'è´¦ç°¿ç®¡ç†',
            'icon': 'ğŸ“–',
            'permission': 'financial_management.ledger.view',
            'children': [
                {
                    'id': 'ledger',
                    'label': 'æ€»è´¦',
                    'icon': 'ğŸ“–',
                    'url_name': 'finance_pages:ledger_management',
                    'permission': 'financial_management.ledger.view',
                    'path_keywords': ['ledger', 'ledgers'],
                },
                {
                    'id': 'subsidiary_ledger',
                    'label': 'æ˜ç»†è´¦',
                    'icon': 'ğŸ“‹',
                    'url_name': 'finance_pages:subsidiary_ledger',
                    'permission': 'financial_management.ledger.view',
                    'path_keywords': ['subsidiary'],
                },
                {
                    'id': 'balance_sheet',
                    'label': 'ç§‘ç›®ä½™é¢è¡¨',
                    'icon': 'ğŸ“Š',
                    'url_name': 'finance_pages:account_balance_sheet',
                    'permission': 'financial_management.ledger.view',
                    'path_keywords': ['balance-sheet'],
                },
                {
                    'id': 'trial_balance',
                    'label': 'è¯•ç®—å¹³è¡¡è¡¨',
                    'icon': 'âš–ï¸',
                    'url_name': 'finance_pages:trial_balance',
                    'permission': 'financial_management.ledger.view',
                    'path_keywords': ['trial-balance'],
                },
            ]
        },
        {
            'id': 'financial_budget',
            'label': 'é¢„ç®—ä¸èµ„é‡‘',
            'icon': 'ğŸ’°',
            'permission': 'financial_management.budget.view',
            'children': [
                {
                    'id': 'budget',
                    'label': 'é¢„ç®—ç®¡ç†',
                    'icon': 'ğŸ’°',
                    'url_name': 'finance_pages:budget_management',
                    'permission': 'financial_management.budget.view',
                    'path_keywords': ['budget', 'budgets'],
                },
                {
                    'id': 'fund_flow',
                    'label': 'èµ„é‡‘æµæ°´',
                    'icon': 'ğŸ’³',
                    'url_name': 'finance_pages:fund_flow_management',
                    'permission': 'financial_management.fund_flow.view',
                    'path_keywords': ['fund-flow', 'fund_flow'],
                },
            ]
        },
        {
            'id': 'financial_invoice',
            'label': 'å‘ç¥¨ä¸è´¦æ¬¾',
            'icon': 'ğŸ§¾',
            'permission': 'financial_management.invoice.view',
            'children': [
                {
                    'id': 'invoice',
                    'label': 'å‘ç¥¨ç®¡ç†',
                    'icon': 'ğŸ§¾',
                    'url_name': 'finance_pages:invoice_management',
                    'permission': 'financial_management.invoice.view',
                    'path_keywords': ['invoice', 'invoices'],
                },
                {
                    'id': 'receivable',
                    'label': 'åº”æ”¶è´¦æ¬¾',
                    'icon': 'ğŸ’°',
                    'url_name': 'finance_pages:receivable_management',
                    'permission': 'financial_management.receivable.view',
                    'path_keywords': ['receivable', 'receivables'],
                },
                {
                    'id': 'payable',
                    'label': 'åº”ä»˜è´¦æ¬¾',
                    'icon': 'ğŸ’¸',
                    'url_name': 'finance_pages:payable_management',
                    'permission': 'financial_management.payable.view',
                    'path_keywords': ['payable', 'payables'],
                },
            ]
        },
        {
            'id': 'financial_report',
            'label': 'è´¢åŠ¡æŠ¥è¡¨',
            'icon': 'ğŸ“ˆ',
            'permission': 'financial_management.report.view',
            'children': [
                {
                    'id': 'report',
                    'label': 'è´¢åŠ¡æŠ¥è¡¨',
                    'icon': 'ğŸ“Š',
                    'url_name': 'finance_pages:report_management',
                    'permission': 'financial_management.report.view',
                    'path_keywords': ['report', 'reports', 'balance-sheet', 'income-statement', 'cash-flow'],
                },
            ]
        },
    ]
    
    menu = []
    
    for menu_group in FINANCIAL_MENU_STRUCTURE:
        # æ£€æŸ¥çˆ¶èœå•æƒé™
        permission = menu_group.get('permission')
        if permission and not _permission_granted(permission, permission_set):
            continue
        
        # å¤„ç†å­èœå•
        children = []
        for child in menu_group.get('children', []):
            # æ£€æŸ¥å­èœå•æƒé™
            child_permission = child.get('permission')
            if child_permission and not _permission_granted(child_permission, permission_set):
                continue
            
            # è·å–URL
            url_name = child.get('url_name')
            url = '#'
            if url_name:
                try:
                    url = reverse(url_name)
                except NoReverseMatch:
                    url = '#'
            
            # åˆ¤æ–­æ˜¯å¦æ¿€æ´»
            is_active = False
            if active_id:
                is_active = child.get('id') == active_id
            elif request_path:
                for keyword in child.get('path_keywords', []):
                    if keyword in request_path:
                        is_active = True
                        break
            
            children.append({
                'id': child.get('id'),
                'label': child.get('label'),
                'icon': child.get('icon'),
                'url': url,
                'active': is_active,
            })
        
        # å¦‚æœçˆ¶èœå•æ²¡æœ‰å¯è§çš„å­èœå•ï¼Œè·³è¿‡
        if not children:
            continue
        
        # åˆ¤æ–­çˆ¶èœå•æ˜¯å¦æ¿€æ´»ï¼ˆä»»æ„å­èœå•æ¿€æ´»åˆ™çˆ¶èœå•æ¿€æ´»ï¼‰
        has_active_child = any(child.get('active') for child in children)
        
        menu.append({
            'id': menu_group.get('id'),
            'label': menu_group.get('label'),
            'icon': menu_group.get('icon'),
            'active': has_active_child,
            'expanded': has_active_child,  # å¦‚æœæœ‰æ¿€æ´»é¡¹ï¼Œé»˜è®¤å±•å¼€
            'children': children,
        })
    
    return menu


@login_required
def financial_home(request):
    """è´¢åŠ¡ç®¡ç†ä¸»é¡µ"""
    permission_codes = get_user_permission_codes(request.user)
    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    
    # æ”¶é›†ç»Ÿè®¡æ•°æ®
    stats_cards = []
    
    try:
        # ä¼šè®¡ç§‘ç›®ç»Ÿè®¡
        if _permission_granted('financial_management.account.view', permission_codes):
            try:
                total_accounts = AccountSubject.objects.filter(is_active=True).count()
                accounts_by_type = AccountSubject.objects.filter(is_active=True).values('subject_type').annotate(count=Count('id'))
                
                stats_cards.append({
                    'label': 'ä¼šè®¡ç§‘ç›®',
                    'icon': 'ğŸ“Š',
                    'value': f'{total_accounts}',
                    'subvalue': f'å¯ç”¨ç§‘ç›®',
                    'url': reverse('finance_pages:account_subject_management'),
                })
            except Exception:
                pass
        
        # å‡­è¯ç®¡ç†ç»Ÿè®¡
        if _permission_granted('financial_management.voucher.view', permission_codes):
            try:
                pending_vouchers = Voucher.objects.filter(status='submitted').count()
                this_month_vouchers = Voucher.objects.filter(voucher_date__gte=this_month_start).count()
                
                stats_cards.append({
                    'label': 'å‡­è¯ç®¡ç†',
                    'icon': 'ğŸ“',
                    'value': f'{pending_vouchers}',
                    'subvalue': f'å¾…å®¡æ ¸ Â· æœ¬æœˆ {this_month_vouchers} å¼ ',
                    'url': reverse('finance_pages:voucher_management'),
                })
            except Exception:
                pass
        
        # è´¦ç°¿ç®¡ç†ç»Ÿè®¡
        if _permission_granted('financial_management.ledger.view', permission_codes):
            try:
                current_year = today.year
                current_month = today.month
                ledger_entries = Ledger.objects.filter(
                    period_year=current_year,
                    period_month=current_month
                ).count()
                
                stats_cards.append({
                    'label': 'è´¦ç°¿ç®¡ç†',
                    'icon': 'ğŸ“–',
                    'value': f'{ledger_entries}',
                    'subvalue': f'æœ¬æœˆè´¦åŠ¡è®°å½•',
                    'url': reverse('finance_pages:ledger_management'),
                })
            except Exception:
                pass
        
        # é¢„ç®—ç®¡ç†ç»Ÿè®¡
        if _permission_granted('financial_management.budget.view', permission_codes):
            try:
                executing_budgets = Budget.objects.filter(status='executing').count()
                total_budget = Budget.objects.filter(status='executing').aggregate(
                    total=Sum('budget_amount')
                )['total'] or Decimal('0')
                
                stats_cards.append({
                    'label': 'é¢„ç®—ç®¡ç†',
                    'icon': 'ğŸ’°',
                    'value': f'{executing_budgets}',
                    'subvalue': f'æ‰§è¡Œä¸­é¢„ç®—',
                    'extra': f'æ€»é¢ Â¥{total_budget:,.2f}',
                    'url': reverse('finance_pages:budget_management'),
                })
            except Exception:
                pass
        
        # å‘ç¥¨ç®¡ç†ç»Ÿè®¡
        if _permission_granted('financial_management.invoice.view', permission_codes):
            try:
                unverified_invoices = Invoice.objects.filter(status='issued').count()
                this_month_invoices = Invoice.objects.filter(invoice_date__gte=this_month_start).count()
                
                stats_cards.append({
                    'label': 'å‘ç¥¨ç®¡ç†',
                    'icon': 'ğŸ§¾',
                    'value': f'{unverified_invoices}',
                    'subvalue': f'å¾…è®¤è¯ Â· æœ¬æœˆ {this_month_invoices} å¼ ',
                    'url': reverse('finance_pages:invoice_management'),
                })
            except Exception:
                pass
        
        # èµ„é‡‘æµæ°´ç»Ÿè®¡
        if _permission_granted('financial_management.fund_flow.view', permission_codes):
            try:
                this_month_flows = FundFlow.objects.filter(flow_date__gte=this_month_start).count()
                this_month_income = FundFlow.objects.filter(
                    flow_date__gte=this_month_start,
                    flow_type='income'
                ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
                
                stats_cards.append({
                    'label': 'èµ„é‡‘æµæ°´',
                    'icon': 'ğŸ’³',
                    'value': f'{this_month_flows}',
                    'subvalue': f'æœ¬æœˆæµæ°´',
                    'extra': f'æ”¶å…¥ Â¥{this_month_income:,.2f}',
                    'url': reverse('finance_pages:fund_flow_management'),
                })
            except Exception:
                pass
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ç»Ÿè®¡æ•°æ®å¤±è´¥: %s', str(e))
    
    context = _context(
        "è´¢åŠ¡ç®¡ç†",
        "ğŸ’µ",
        "ä¼ä¸šè´¢åŠ¡ç®¡ç†å¹³å°",
        summary_cards=stats_cards,
        request=request,
        use_financial_nav=True
    )
    return render(request, "financial_management/home.html", context)


@login_required
def financial_statistics(request):
    """è´¢åŠ¡ç»Ÿè®¡ä»ªè¡¨æ¿"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.report.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æŸ¥çœ‹è´¢åŠ¡ç»Ÿè®¡')
        return redirect('finance_pages:financial_home')
    
    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    this_year_start = today.replace(month=1, day=1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)
    
    # è·å–ç»Ÿè®¡å‚æ•°
    period = request.GET.get('period', 'month')  # month, quarter, year
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month)) if period == 'month' else None
    
    # ç¡®å®šç»Ÿè®¡æ—¶é—´èŒƒå›´
    if period == 'month' and month:
        period_start = today.replace(year=year, month=month, day=1)
        if month == 12:
            period_end = today.replace(year=year+1, month=1, day=1) - timedelta(days=1)
        else:
            period_end = today.replace(year=year, month=month+1, day=1) - timedelta(days=1)
    elif period == 'quarter':
        quarter = int(request.GET.get('quarter', (today.month - 1) // 3 + 1))
        period_start = today.replace(year=year, month=(quarter-1)*3+1, day=1)
        if quarter == 4:
            period_end = today.replace(year=year+1, month=1, day=1) - timedelta(days=1)
        else:
            period_end = today.replace(year=year, month=quarter*3+1, day=1) - timedelta(days=1)
    else:  # year
        period_start = today.replace(year=year, month=1, day=1)
        period_end = today.replace(year=year, month=12, day=31)
    
    statistics = {}
    
    try:
        # å‡­è¯ç»Ÿè®¡
        vouchers = Voucher.objects.filter(voucher_date__gte=period_start, voucher_date__lte=period_end)
        statistics['vouchers'] = {
            'total': vouchers.count(),
            'draft': vouchers.filter(status='draft').count(),
            'submitted': vouchers.filter(status='submitted').count(),
            'approved': vouchers.filter(status='approved').count(),
            'posted': vouchers.filter(status='posted').count(),
            'total_debit': vouchers.aggregate(total=Sum('total_debit'))['total'] or Decimal('0'),
            'total_credit': vouchers.aggregate(total=Sum('total_credit'))['total'] or Decimal('0'),
        }
        
        # èµ„é‡‘æµæ°´ç»Ÿè®¡
        fund_flows = FundFlow.objects.filter(flow_date__gte=period_start, flow_date__lte=period_end)
        statistics['fund_flows'] = {
            'total': fund_flows.count(),
            'income': fund_flows.filter(flow_type='income').aggregate(total=Sum('amount'))['total'] or Decimal('0'),
            'expense': fund_flows.filter(flow_type='expense').aggregate(total=Sum('amount'))['total'] or Decimal('0'),
            'transfer': fund_flows.filter(flow_type='transfer').aggregate(total=Sum('amount'))['total'] or Decimal('0'),
            'net': (fund_flows.filter(flow_type='income').aggregate(total=Sum('amount'))['total'] or Decimal('0')) - 
                   (fund_flows.filter(flow_type='expense').aggregate(total=Sum('amount'))['total'] or Decimal('0')),
        }
        
        # å‘ç¥¨ç»Ÿè®¡
        invoices = Invoice.objects.filter(invoice_date__gte=period_start, invoice_date__lte=period_end)
        statistics['invoices'] = {
            'total': invoices.count(),
            'input': invoices.filter(invoice_type='input').aggregate(total=Sum('total_amount'))['total'] or Decimal('0'),
            'output': invoices.filter(invoice_type='output').aggregate(total=Sum('total_amount'))['total'] or Decimal('0'),
            'issued': invoices.filter(status='issued').count(),
            'verified': invoices.filter(status='verified').count(),
        }
        
        # åº”æ”¶è´¦æ¬¾ç»Ÿè®¡
        receivables = ReceivableAccount.objects.filter(receivable_date__gte=period_start, receivable_date__lte=period_end)
        statistics['receivables'] = {
            'total': receivables.count(),
            'total_amount': receivables.aggregate(total=Sum('receivable_amount'))['total'] or Decimal('0'),
            'paid_amount': receivables.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0'),
            'remaining_amount': receivables.aggregate(total=Sum('remaining_amount'))['total'] or Decimal('0'),
        }
        
        # åº”ä»˜è´¦æ¬¾ç»Ÿè®¡
        payables = PayableAccount.objects.filter(payable_date__gte=period_start, payable_date__lte=period_end)
        statistics['payables'] = {
            'total': payables.count(),
            'total_amount': payables.aggregate(total=Sum('payable_amount'))['total'] or Decimal('0'),
            'paid_amount': payables.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0'),
            'remaining_amount': payables.aggregate(total=Sum('remaining_amount'))['total'] or Decimal('0'),
        }
        
        # é¢„ç®—ç»Ÿè®¡
        budgets = Budget.objects.filter(
            start_date__lte=period_end,
            end_date__gte=period_start
        )
        statistics['budgets'] = {
            'total': budgets.count(),
            'total_amount': budgets.aggregate(total=Sum('budget_amount'))['total'] or Decimal('0'),
            'used_amount': budgets.aggregate(total=Sum('used_amount'))['total'] or Decimal('0'),
            'remaining_amount': budgets.aggregate(total=Sum('remaining_amount'))['total'] or Decimal('0'),
        }
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–è´¢åŠ¡ç»Ÿè®¡å¤±è´¥: %s', str(e))
        statistics = {}
    
    context = _context(
        "è´¢åŠ¡ç»Ÿè®¡",
        "ğŸ“Š",
        f"æŸ¥çœ‹ {period_start.strftime('%Y-%m-%d')} è‡³ {period_end.strftime('%Y-%m-%d')} æœŸé—´çš„è´¢åŠ¡ç»Ÿè®¡æ•°æ®",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'statistics': statistics,
        'period': period,
        'year': year,
        'month': month,
        'period_start': period_start,
        'period_end': period_end,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
        'quarters': range(1, 5),
    })
    return render(request, "financial_management/statistics.html", context)


@login_required
def account_subject_management(request):
    """ä¼šè®¡ç§‘ç›®ç®¡ç†"""
    permission_codes = get_user_permission_codes(request.user)
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    subject_type = request.GET.get('subject_type', '')
    is_active = request.GET.get('is_active', '')
    
    # è·å–ç§‘ç›®åˆ—è¡¨
    try:
        subjects = AccountSubject.objects.select_related('parent', 'created_by').order_by('code')
        
        # åº”ç”¨ç­›é€‰æ¡ä»¶
        if search:
            subjects = subjects.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        if subject_type:
            subjects = subjects.filter(subject_type=subject_type)
        if is_active == 'true':
            subjects = subjects.filter(is_active=True)
        elif is_active == 'false':
            subjects = subjects.filter(is_active=False)
        
        # åˆ†é¡µ
        paginator = Paginator(subjects, 50)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ä¼šè®¡ç§‘ç›®åˆ—è¡¨å¤±è´¥: %s', str(e))
        page_obj = None
    
    # ç»Ÿè®¡ä¿¡æ¯
    try:
        total_subjects = AccountSubject.objects.count()
        active_subjects = AccountSubject.objects.filter(is_active=True).count()
        subjects_by_type = AccountSubject.objects.filter(is_active=True).values('subject_type').annotate(count=Count('id'))
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ç»Ÿè®¡ä¿¡æ¯å¤±è´¥: %s', str(e))
        summary_cards = []
    
    context = _context(
        "ä¼šè®¡ç§‘ç›®ç®¡ç†",
        "ğŸ“Š",
        "ç®¡ç†ä¼šè®¡ç§‘ç›®ä¿¡æ¯",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'subjects': page_obj.object_list if page_obj else [],
        'subject_type_choices': AccountSubject.TYPE_CHOICES,
        'current_search': search,
        'current_subject_type': subject_type,
        'current_is_active': is_active,
    })
    return render(request, "financial_management/account_subject_list.html", context)


@login_required
def account_subject_tree_export(request):
    """å¯¼å‡ºä¼šè®¡ç§‘ç›®æ ‘å½¢ç»“æ„"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.account.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å¯¼å‡ºä¼šè®¡ç§‘ç›®')
        return redirect('finance_pages:account_subject_management')
    
    # è·å–æ‰€æœ‰å¯ç”¨çš„ä¸€çº§ç§‘ç›®
    root_subjects = AccountSubject.objects.filter(parent=None, is_active=True).order_by('code')
    
    # åˆ›å»ºExcelå·¥ä½œç°¿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'ä¼šè®¡ç§‘ç›®æ ‘'
    
    # è®¾ç½®è¡¨å¤´
    headers = ['ç§‘ç›®ç¼–ç ', 'ç§‘ç›®åç§°', 'ç§‘ç›®ç±»å‹', 'ä½™é¢æ–¹å‘', 'çº§åˆ«', 'æ˜¯å¦å¯ç”¨', 'å¤‡æ³¨è¯´æ˜']
    worksheet.append(headers)
    
    # è®¾ç½®è¡¨å¤´æ ·å¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # é€’å½’å¯¼å‡ºç§‘ç›®æ ‘
    def export_subject_tree(subject, level=1):
        """é€’å½’å¯¼å‡ºç§‘ç›®åŠå…¶å­ç§‘ç›®"""
        indent = '  ' * (level - 1)  # æ ¹æ®çº§åˆ«æ·»åŠ ç¼©è¿›
        type_dict = dict(AccountSubject.TYPE_CHOICES)
        direction_dict = dict(AccountSubject.DIRECTION_CHOICES)
        
        row = [
            subject.code,
            indent + subject.name,  # æ·»åŠ ç¼©è¿›æ˜¾ç¤ºå±‚çº§å…³ç³»
            type_dict.get(subject.subject_type, subject.subject_type),
            direction_dict.get(subject.direction, subject.direction),
            level,
            'æ˜¯' if subject.is_active else 'å¦',
            subject.description or '',
        ]
        worksheet.append(row)
        
        # é€’å½’å¯¼å‡ºå­ç§‘ç›®
        children = AccountSubject.objects.filter(parent=subject, is_active=True).order_by('code')
        for child in children:
            export_subject_tree(child, level + 1)
    
    # å¯¼å‡ºæ‰€æœ‰ä¸€çº§ç§‘ç›®åŠå…¶å­ç§‘ç›®
    for root_subject in root_subjects:
        export_subject_tree(root_subject)
    
    # è°ƒæ•´åˆ—å®½
    column_widths = [15, 40, 15, 12, 8, 10, 30]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    # è®¾ç½®æ•°æ®è¡Œæ ·å¼ï¼ˆæ ¹æ®çº§åˆ«è®¾ç½®ä¸åŒçš„ç¼©è¿›å’Œé¢œè‰²ï¼‰
    from openpyxl.styles import PatternFill as DataFill
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        level = row[4].value  # çº§åˆ«åˆ—
        if level:
            # æ ¹æ®çº§åˆ«è®¾ç½®ä¸åŒçš„èƒŒæ™¯è‰²
            if level == 1:
                fill_color = "E8F4F8"
            elif level == 2:
                fill_color = "F0F8E8"
            else:
                fill_color = "FFF8E8"
            
            for cell in row:
                cell.fill = DataFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('ä¼šè®¡ç§‘ç›®æ ‘_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def account_subject_import_template(request):
    """ä¸‹è½½ä¼šè®¡ç§‘ç›®å¯¼å…¥æ¨¡æ¿"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.account.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™ä¸‹è½½å¯¼å…¥æ¨¡æ¿')
        return redirect('finance_pages:account_subject_management')
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'ä¼šè®¡ç§‘ç›®å¯¼å…¥æ¨¡æ¿'
    
    # è®¾ç½®è¡¨å¤´
    headers = ['ç§‘ç›®ç¼–ç ', 'ç§‘ç›®åç§°', 'ä¸Šçº§ç§‘ç›®ç¼–ç ', 'ç§‘ç›®ç±»å‹', 'ä½™é¢æ–¹å‘', 'æ˜¯å¦å¯ç”¨', 'å¤‡æ³¨è¯´æ˜']
    worksheet.append(headers)
    
    # è®¾ç½®è¡¨å¤´æ ·å¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # æ·»åŠ ç¤ºä¾‹æ•°æ®
    examples = [
        ['1001', 'åº“å­˜ç°é‡‘', '', 'asset', 'debit', 'æ˜¯', ''],
        ['1002', 'é“¶è¡Œå­˜æ¬¾', '', 'asset', 'debit', 'æ˜¯', ''],
        ['100201', 'å·¥å•†é“¶è¡Œ', '1002', 'asset', 'debit', 'æ˜¯', ''],
        ['2001', 'çŸ­æœŸå€Ÿæ¬¾', '', 'liability', 'credit', 'æ˜¯', ''],
    ]
    for example in examples:
        worksheet.append(example)
    
    # æ·»åŠ è¯´æ˜è¡Œ
    worksheet.append([])
    worksheet.append(['è¯´æ˜ï¼š'])
    worksheet.append(['1. ç§‘ç›®ç¼–ç ï¼šå¿…å¡«ï¼Œå”¯ä¸€æ ‡è¯†ï¼Œä¸èƒ½é‡å¤'])
    worksheet.append(['2. ç§‘ç›®åç§°ï¼šå¿…å¡«'])
    worksheet.append(['3. ä¸Šçº§ç§‘ç›®ç¼–ç ï¼šå¯é€‰ï¼Œå¦‚æœå¡«å†™ï¼Œç³»ç»Ÿä¼šè‡ªåŠ¨å…³è”ä¸Šçº§ç§‘ç›®'])
    worksheet.append(['4. ç§‘ç›®ç±»å‹ï¼šå¿…å¡«ï¼Œå¯é€‰å€¼ï¼šasset(èµ„äº§)ã€liability(è´Ÿå€º)ã€equity(æ‰€æœ‰è€…æƒç›Š)ã€revenue(æ”¶å…¥)ã€expense(è´¹ç”¨)ã€cost(æˆæœ¬)'])
    worksheet.append(['5. ä½™é¢æ–¹å‘ï¼šå¿…å¡«ï¼Œå¯é€‰å€¼ï¼šdebit(å€Ÿæ–¹)ã€credit(è´·æ–¹)'])
    worksheet.append(['6. æ˜¯å¦å¯ç”¨ï¼šå¿…å¡«ï¼Œå¯é€‰å€¼ï¼šæ˜¯ã€å¦'])
    
    # è°ƒæ•´åˆ—å®½
    column_widths = [15, 20, 15, 15, 12, 12, 30]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ä¼šè®¡ç§‘ç›®å¯¼å…¥æ¨¡æ¿.xlsx"'
    workbook.save(response)
    return response


@login_required
def account_subject_import(request):
    """å¯¼å…¥ä¼šè®¡ç§‘ç›®"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.account.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å¯¼å…¥ä¼šè®¡ç§‘ç›®')
        return redirect('finance_pages:account_subject_management')
    
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'è¯·é€‰æ‹©è¦å¯¼å…¥çš„æ–‡ä»¶')
            return redirect('finance_pages:account_subject_management')
        
        upload_file = request.FILES['file']
        if not upload_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'è¯·ä¸Šä¼ Excelæ–‡ä»¶ï¼ˆ.xlsæˆ–.xlsxæ ¼å¼ï¼‰')
            return redirect('finance_pages:account_subject_management')
        
        try:
            from django.db import transaction
            workbook = load_workbook(upload_file, data_only=True)
            worksheet = workbook.active
            
            # è¯»å–è¡¨å¤´
            headers = [cell.value for cell in worksheet[1]]
            header_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
            
            # æ£€æŸ¥å¿…å¡«åˆ—
            required_columns = ['ç§‘ç›®ç¼–ç ', 'ç§‘ç›®åç§°', 'ç§‘ç›®ç±»å‹', 'ä½™é¢æ–¹å‘']
            missing_columns = [col for col in required_columns if col not in header_map]
            if missing_columns:
                messages.error(request, f'ç¼ºå°‘å¿…å¡«åˆ—ï¼š{", ".join(missing_columns)}')
                return redirect('finance_pages:account_subject_management')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # è¯»å–æ•°æ®è¡Œ
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                # è·³è¿‡ç©ºè¡Œ
                if not row or not any(row):
                    continue
                
                try:
                    code = str(row[header_map['ç§‘ç›®ç¼–ç ']]).strip() if row[header_map['ç§‘ç›®ç¼–ç ']] else None
                    name = str(row[header_map['ç§‘ç›®åç§°']]).strip() if row[header_map['ç§‘ç›®åç§°']] else None
                    
                    if not code or not name:
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šç§‘ç›®ç¼–ç å’Œç§‘ç›®åç§°ä¸èƒ½ä¸ºç©º')
                        continue
                    
                    # æ£€æŸ¥æ˜¯å¦å·²å­˜åœ¨
                    if AccountSubject.objects.filter(code=code).exists():
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šç§‘ç›®ç¼–ç  {code} å·²å­˜åœ¨')
                        continue
                    
                    # è§£æå…¶ä»–å­—æ®µ
                    parent_code = str(row[header_map.get('ä¸Šçº§ç§‘ç›®ç¼–ç ', -1)]).strip() if header_map.get('ä¸Šçº§ç§‘ç›®ç¼–ç ', -1) >= 0 and row[header_map.get('ä¸Šçº§ç§‘ç›®ç¼–ç ', -1)] else None
                    subject_type = str(row[header_map['ç§‘ç›®ç±»å‹']]).strip() if row[header_map['ç§‘ç›®ç±»å‹']] else None
                    direction = str(row[header_map['ä½™é¢æ–¹å‘']]).strip() if row[header_map['ä½™é¢æ–¹å‘']] else None
                    is_active_str = str(row[header_map.get('æ˜¯å¦å¯ç”¨', -1)]).strip() if header_map.get('æ˜¯å¦å¯ç”¨', -1) >= 0 and row[header_map.get('æ˜¯å¦å¯ç”¨', -1)] else 'æ˜¯'
                    description = str(row[header_map.get('å¤‡æ³¨è¯´æ˜', -1)]).strip() if header_map.get('å¤‡æ³¨è¯´æ˜', -1) >= 0 and row[header_map.get('å¤‡æ³¨è¯´æ˜', -1)] else ''
                    
                    # éªŒè¯ç§‘ç›®ç±»å‹
                    type_map = {
                        'asset': 'asset', 'èµ„äº§': 'asset',
                        'liability': 'liability', 'è´Ÿå€º': 'liability',
                        'equity': 'equity', 'æ‰€æœ‰è€…æƒç›Š': 'equity',
                        'revenue': 'revenue', 'æ”¶å…¥': 'revenue',
                        'expense': 'expense', 'è´¹ç”¨': 'expense',
                        'cost': 'cost', 'æˆæœ¬': 'cost',
                    }
                    subject_type = type_map.get(subject_type.lower(), subject_type)
                    if subject_type not in dict(AccountSubject.TYPE_CHOICES):
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šç§‘ç›®ç±»å‹æ— æ•ˆ')
                        continue
                    
                    # éªŒè¯ä½™é¢æ–¹å‘
                    direction_map = {
                        'debit': 'debit', 'å€Ÿæ–¹': 'debit',
                        'credit': 'credit', 'è´·æ–¹': 'credit',
                    }
                    direction = direction_map.get(direction.lower(), direction)
                    if direction not in dict(AccountSubject.DIRECTION_CHOICES):
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šä½™é¢æ–¹å‘æ— æ•ˆ')
                        continue
                    
                    # è§£ææ˜¯å¦å¯ç”¨
                    is_active = is_active_str.lower() in ['æ˜¯', 'yes', 'true', '1', 'y']
                    
                    # æŸ¥æ‰¾ä¸Šçº§ç§‘ç›®
                    parent = None
                    level = 1
                    if parent_code:
                        try:
                            parent = AccountSubject.objects.get(code=parent_code)
                            level = parent.level + 1
                        except AccountSubject.DoesNotExist:
                            error_count += 1
                            errors.append(f'ç¬¬{row_idx}è¡Œï¼šä¸Šçº§ç§‘ç›®ç¼–ç  {parent_code} ä¸å­˜åœ¨')
                            continue
                    
                    # åˆ›å»ºç§‘ç›®
                    with transaction.atomic():
                        AccountSubject.objects.create(
                            code=code,
                            name=name,
                            parent=parent,
                            subject_type=subject_type,
                            direction=direction,
                            level=level,
                            is_active=is_active,
                            description=description,
                            created_by=request.user,
                        )
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'ç¬¬{row_idx}è¡Œï¼š{str(e)}')
            
            if success_count > 0:
                messages.success(request, f'æˆåŠŸå¯¼å…¥ {success_count} ä¸ªä¼šè®¡ç§‘ç›®')
            if error_count > 0:
                error_msg = f'å¯¼å…¥å¤±è´¥ {error_count} ä¸ªç§‘ç›®'
                if len(errors) <= 10:
                    error_msg += 'ï¼š' + 'ï¼›'.join(errors)
                else:
                    error_msg += f'ï¼šå‰10ä¸ªé”™è¯¯ï¼š' + 'ï¼›'.join(errors[:10])
                messages.warning(request, error_msg)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('å¯¼å…¥ä¼šè®¡ç§‘ç›®å¤±è´¥: %s', str(e))
            messages.error(request, f'å¯¼å…¥å¤±è´¥ï¼š{str(e)}')
    
    return redirect('finance_pages:account_subject_management')


@login_required
def account_subject_create(request):
    """æ–°å¢ä¼šè®¡ç§‘ç›®"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.account.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æ–°å¢ä¼šè®¡ç§‘ç›®')
        return redirect('finance_pages:account_subject_management')
    
    if request.method == 'POST':
        form = AccountSubjectForm(request.POST)
        if form.is_valid():
            account_subject = form.save(commit=False)
            account_subject.created_by = request.user
            # å¦‚æœé€‰æ‹©äº†ä¸Šçº§ç§‘ç›®ï¼Œè‡ªåŠ¨è®¡ç®—çº§åˆ«
            if account_subject.parent:
                account_subject.level = account_subject.parent.level + 1
            account_subject.save()
            messages.success(request, f'ä¼šè®¡ç§‘ç›® {account_subject.name} åˆ›å»ºæˆåŠŸï¼')
            return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject.id)
    else:
        form = AccountSubjectForm()
    
    context = _context(
        "æ–°å¢ä¼šè®¡ç§‘ç›®",
        "â•",
        "åˆ›å»ºæ–°çš„ä¼šè®¡ç§‘ç›®",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/account_subject_form.html", context)


@login_required
def account_subject_update(request, account_subject_id):
    """ç¼–è¾‘ä¼šè®¡ç§‘ç›®"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.account.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™ç¼–è¾‘ä¼šè®¡ç§‘ç›®')
        return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject_id)
    
    account_subject = get_object_or_404(AccountSubject, id=account_subject_id)
    
    if request.method == 'POST':
        form = AccountSubjectForm(request.POST, instance=account_subject)
        if form.is_valid():
            account_subject = form.save(commit=False)
            # å¦‚æœé€‰æ‹©äº†ä¸Šçº§ç§‘ç›®ï¼Œè‡ªåŠ¨è®¡ç®—çº§åˆ«
            if account_subject.parent:
                account_subject.level = account_subject.parent.level + 1
            account_subject.save()
            messages.success(request, f'ä¼šè®¡ç§‘ç›® {account_subject.name} æ›´æ–°æˆåŠŸï¼')
            return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject.id)
    else:
        form = AccountSubjectForm(instance=account_subject)
    
    context = _context(
        f"ç¼–è¾‘ä¼šè®¡ç§‘ç›® - {account_subject.name}",
        "âœï¸",
        f"ç¼–è¾‘ä¼šè®¡ç§‘ç›® {account_subject.name}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'account_subject': account_subject,
        'is_create': False,
    })
    return render(request, "financial_management/account_subject_form.html", context)


@login_required
def account_subject_delete(request, account_subject_id):
    """åˆ é™¤ä¼šè®¡ç§‘ç›®"""
    account_subject = get_object_or_404(AccountSubject, id=account_subject_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.account.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ é™¤ä¼šè®¡ç§‘ç›®')
        return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject_id)
    
    # æ£€æŸ¥æ˜¯å¦æœ‰å­ç§‘ç›®
    children_count = account_subject.children.count()
    if children_count > 0:
        messages.error(request, f'è¯¥ç§‘ç›®ä¸‹æœ‰ {children_count} ä¸ªå­ç§‘ç›®ï¼Œæ— æ³•åˆ é™¤ã€‚è¯·å…ˆåˆ é™¤æˆ–ç§»åŠ¨å­ç§‘ç›®ã€‚')
        return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject_id)
    
    # æ£€æŸ¥æ˜¯å¦è¢«ä½¿ç”¨
    voucher_entry_count = account_subject.voucher_entries.count()
    ledger_entry_count = account_subject.ledger_entries.count()
    
    if voucher_entry_count > 0 or ledger_entry_count > 0:
        messages.error(request, f'è¯¥ç§‘ç›®å·²è¢«ä½¿ç”¨ï¼ˆå‡­è¯åˆ†å½•ï¼š{voucher_entry_count}æ¡ï¼Œæ€»è´¦è®°å½•ï¼š{ledger_entry_count}æ¡ï¼‰ï¼Œæ— æ³•åˆ é™¤ã€‚')
        return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject_id)
    
    if request.method == 'POST':
        try:
            subject_name = account_subject.name
            account_subject.delete()
            messages.success(request, f'ä¼šè®¡ç§‘ç›® {subject_name} å·²åˆ é™¤')
            return redirect('finance_pages:account_subject_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('åˆ é™¤ä¼šè®¡ç§‘ç›®å¤±è´¥: %s', str(e))
            messages.error(request, f'åˆ é™¤ä¼šè®¡ç§‘ç›®å¤±è´¥ï¼š{str(e)}')
            return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject_id)
    
    context = _context(
        f"åˆ é™¤ä¼šè®¡ç§‘ç›® - {account_subject.name}",
        "ğŸ—‘ï¸",
        f"ç¡®è®¤åˆ é™¤ä¼šè®¡ç§‘ç›®ï¼š{account_subject.code} {account_subject.name}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'account_subject': account_subject,
    })
    return render(request, "financial_management/account_subject_delete.html", context)


@login_required
def budget_create(request):
    """æ–°å¢é¢„ç®—"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.budget.create', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ›å»ºé¢„ç®—')
        return redirect('finance_pages:budget_management')
    
    if request.method == 'POST':
        form = BudgetForm(request.POST)
        if form.is_valid():
            budget = form.save(commit=False)
            # è‡ªåŠ¨ç”Ÿæˆé¢„ç®—ç¼–å·
            if not budget.budget_number:
                current_year = timezone.now().year
                max_budget = Budget.objects.filter(
                    budget_number__startswith=f'BUDGET-{current_year}-'
                ).aggregate(max_num=Max('budget_number'))['max_num']
                if max_budget:
                    try:
                        seq = int(max_budget.split('-')[-1]) + 1
                    except (ValueError, IndexError):
                        seq = 1
                else:
                    seq = 1
                budget.budget_number = f'BUDGET-{current_year}-{seq:04d}'
            budget.remaining_amount = budget.budget_amount
            budget.created_by = request.user
            budget.save()
            messages.success(request, f'é¢„ç®— {budget.name} åˆ›å»ºæˆåŠŸï¼')
            return redirect('finance_pages:budget_detail', budget_id=budget.id)
    else:
        form = BudgetForm()
    
    context = _context(
        "æ–°å¢é¢„ç®—",
        "â•",
        "åˆ›å»ºæ–°çš„é¢„ç®—è®°å½•",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/budget_form.html", context)


@login_required
def budget_update(request, budget_id):
    """ç¼–è¾‘é¢„ç®—"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.budget.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™ç¼–è¾‘é¢„ç®—')
        return redirect('finance_pages:budget_detail', budget_id=budget_id)
    
    budget = get_object_or_404(Budget, id=budget_id)
    
    if request.method == 'POST':
        form = BudgetForm(request.POST, instance=budget)
        if form.is_valid():
            budget = form.save(commit=False)
            # é‡æ–°è®¡ç®—å‰©ä½™é‡‘é¢
            budget.remaining_amount = budget.budget_amount - budget.used_amount
            budget.save()
            messages.success(request, f'é¢„ç®— {budget.name} æ›´æ–°æˆåŠŸï¼')
            return redirect('finance_pages:budget_detail', budget_id=budget.id)
    else:
        form = BudgetForm(instance=budget)
    
    context = _context(
        f"ç¼–è¾‘é¢„ç®— - {budget.name}",
        "âœï¸",
        f"ç¼–è¾‘é¢„ç®— {budget.name}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'budget': budget,
        'is_create': False,
    })
    return render(request, "financial_management/budget_form.html", context)


@login_required
def invoice_create(request):
    """æ–°å¢å‘ç¥¨"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.invoice.create', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ›å»ºå‘ç¥¨')
        return redirect('finance_pages:invoice_management')
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST, request.FILES)
        if form.is_valid():
            invoice = form.save(commit=False)
            # å¦‚æœæ²¡æœ‰å¡«å†™æ€»é‡‘é¢ï¼Œè‡ªåŠ¨è®¡ç®—
            if not invoice.total_amount and invoice.amount and invoice.tax_amount:
                invoice.total_amount = invoice.amount + invoice.tax_amount
            elif not invoice.total_amount:
                invoice.total_amount = invoice.amount or Decimal('0.00')
            invoice.created_by = request.user
            invoice.save()
            messages.success(request, f'å‘ç¥¨ {invoice.invoice_number} åˆ›å»ºæˆåŠŸï¼')
            return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    else:
        form = InvoiceForm()
        # é»˜è®¤å½“å‰æ—¥æœŸ
        form.fields['invoice_date'].initial = timezone.now().date()
    
    context = _context(
        "æ–°å¢å‘ç¥¨",
        "â•",
        "åˆ›å»ºæ–°çš„å‘ç¥¨è®°å½•",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/invoice_form.html", context)


@login_required
def invoice_update(request, invoice_id):
    """ç¼–è¾‘å‘ç¥¨"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.invoice.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™ç¼–è¾‘å‘ç¥¨')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice_id)
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST, request.FILES, instance=invoice)
        if form.is_valid():
            invoice = form.save(commit=False)
            # é‡æ–°è®¡ç®—æ€»é‡‘é¢
            if invoice.amount and invoice.tax_amount:
                invoice.total_amount = invoice.amount + invoice.tax_amount
            invoice.save()
            messages.success(request, f'å‘ç¥¨ {invoice.invoice_number} æ›´æ–°æˆåŠŸï¼')
            return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    else:
        form = InvoiceForm(instance=invoice)
    
    context = _context(
        f"ç¼–è¾‘å‘ç¥¨ - {invoice.invoice_number}",
        "âœï¸",
        f"ç¼–è¾‘å‘ç¥¨ {invoice.invoice_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'invoice': invoice,
        'is_create': False,
    })
    return render(request, "financial_management/invoice_form.html", context)


@login_required
def fund_flow_create(request):
    """æ–°å¢èµ„é‡‘æµæ°´"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.fund_flow.create', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ›å»ºèµ„é‡‘æµæ°´')
        return redirect('finance_pages:fund_flow_management')
    
    if request.method == 'POST':
        form = FundFlowForm(request.POST)
        if form.is_valid():
            fund_flow = form.save(commit=False)
            # è‡ªåŠ¨ç”Ÿæˆæµæ°´å·
            if not fund_flow.flow_number:
                current_year = timezone.now().year
                max_flow = FundFlow.objects.filter(
                    flow_number__startswith=f'FLOW-{current_year}-'
                ).aggregate(max_num=Max('flow_number'))['max_num']
                if max_flow:
                    try:
                        seq = int(max_flow.split('-')[-1]) + 1
                    except (ValueError, IndexError):
                        seq = 1
                else:
                    seq = 1
                fund_flow.flow_number = f'FLOW-{current_year}-{seq:04d}'
            fund_flow.created_by = request.user
            fund_flow.save()
            
            # è‡ªåŠ¨æ›´æ–°é¢„ç®—ä½¿ç”¨é‡‘é¢
            _update_budget_from_fund_flow(fund_flow, is_create=True)
            
            messages.success(request, f'èµ„é‡‘æµæ°´ {fund_flow.flow_number} åˆ›å»ºæˆåŠŸï¼')
            return redirect('finance_pages:fund_flow_detail', fund_flow_id=fund_flow.id)
    else:
        form = FundFlowForm()
        # é»˜è®¤ä»Šå¤©
        form.fields['flow_date'].initial = timezone.now().date()
    
    context = _context(
        "æ–°å¢èµ„é‡‘æµæ°´",
        "â•",
        "åˆ›å»ºæ–°çš„èµ„é‡‘æµæ°´è®°å½•",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/fund_flow_form.html", context)


@login_required
def fund_flow_update(request, fund_flow_id):
    """ç¼–è¾‘èµ„é‡‘æµæ°´"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.fund_flow.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™ç¼–è¾‘èµ„é‡‘æµæ°´')
        return redirect('finance_pages:fund_flow_detail', fund_flow_id=fund_flow_id)
    
    fund_flow = get_object_or_404(FundFlow, id=fund_flow_id)
    
    if request.method == 'POST':
        form = FundFlowForm(request.POST, instance=fund_flow)
        if form.is_valid():
            # ä¿å­˜æ—§é‡‘é¢ç”¨äºé¢„ç®—æ›´æ–°
            old_amount = fund_flow.amount
            old_flow_type = fund_flow.flow_type
            
            fund_flow = form.save()
            
            # å¦‚æœé‡‘é¢æˆ–ç±»å‹å‘ç”Ÿå˜åŒ–ï¼Œæ›´æ–°é¢„ç®—
            if old_amount != fund_flow.amount or old_flow_type != fund_flow.flow_type:
                # å…ˆå›æ»šæ—§é‡‘é¢çš„å½±å“
                if old_flow_type == 'expense':
                    _update_budget_from_fund_flow(fund_flow, is_create=False, old_amount=old_amount)
                # å†åº”ç”¨æ–°é‡‘é¢çš„å½±å“
                _update_budget_from_fund_flow(fund_flow, is_create=True)
            
            messages.success(request, f'èµ„é‡‘æµæ°´ {fund_flow.flow_number} æ›´æ–°æˆåŠŸï¼')
            return redirect('finance_pages:fund_flow_detail', fund_flow_id=fund_flow.id)
    else:
        form = FundFlowForm(instance=fund_flow)
    
    context = _context(
        f"ç¼–è¾‘èµ„é‡‘æµæ°´ - {fund_flow.flow_number}",
        "âœï¸",
        f"ç¼–è¾‘èµ„é‡‘æµæ°´ {fund_flow.flow_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'fund_flow': fund_flow,
        'is_create': False,
    })
    return render(request, "financial_management/fund_flow_form.html", context)


# åˆ›å»ºå‡­è¯åˆ†å½•çš„å†…è”è¡¨å•é›†
VoucherEntryFormSet = inlineformset_factory(
    Voucher, VoucherEntry,
    form=VoucherEntryForm,
    extra=3,  # é»˜è®¤æ˜¾ç¤º3ä¸ªç©ºè¡Œ
    can_delete=True,
    min_num=1,  # è‡³å°‘éœ€è¦1ä¸ªåˆ†å½•
    validate_min=True,
)


@login_required
def voucher_create(request):
    """æ–°å¢è®°è´¦å‡­è¯"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.create', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ›å»ºè®°è´¦å‡­è¯')
        return redirect('finance_pages:voucher_management')
    
    if request.method == 'POST':
        form = VoucherForm(request.POST)
        formset = VoucherEntryFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            voucher = form.save(commit=False)
            # è‡ªåŠ¨ç”Ÿæˆå‡­è¯å­—å·
            if not voucher.voucher_number:
                voucher.voucher_number = _generate_voucher_number(voucher.voucher_date)
            if not voucher.preparer:
                voucher.preparer = request.user
            voucher.save()
            
            # ä¿å­˜åˆ†å½•å¹¶è®¡ç®—åˆè®¡
            entries = formset.save(commit=False)
            total_debit = Decimal('0.00')
            total_credit = Decimal('0.00')
            
            for entry in entries:
                entry.voucher = voucher
                entry.save()
                total_debit += entry.debit_amount or Decimal('0.00')
                total_credit += entry.credit_amount or Decimal('0.00')
            
            # åˆ é™¤æ ‡è®°ä¸ºåˆ é™¤çš„åˆ†å½•
            for obj in formset.deleted_objects:
                obj.delete()
            
            # æ›´æ–°åˆè®¡
            voucher.total_debit = total_debit
            voucher.total_credit = total_credit
            voucher.save()
            
            messages.success(request, f'è®°è´¦å‡­è¯ {voucher.voucher_number} åˆ›å»ºæˆåŠŸï¼')
            return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
        else:
            messages.error(request, 'è¯·æ£€æŸ¥è¡¨å•ä¸­çš„é”™è¯¯ã€‚')
    else:
        form = VoucherForm(initial={'voucher_date': timezone.now().date(), 'preparer': request.user})
        formset = VoucherEntryFormSet()
    
    # è·å–æ‰€æœ‰ä¼šè®¡ç§‘ç›®ä¾› JavaScript ä½¿ç”¨
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    
    context = _context(
        "æ–°å¢è®°è´¦å‡­è¯",
        "â•",
        "åˆ›å»ºæ–°çš„è®°è´¦å‡­è¯",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'formset': formset,
        'is_create': True,
        'account_subjects': account_subjects,
    })
    return render(request, "financial_management/voucher_form.html", context)


@login_required
def voucher_update(request, voucher_id):
    """ç¼–è¾‘è®°è´¦å‡­è¯"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™ç¼–è¾‘è®°è´¦å‡­è¯')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher.objects.prefetch_related('entries'), id=voucher_id)
    
    # å·²è¿‡è´¦çš„å‡­è¯ä¸èƒ½ç¼–è¾‘
    if voucher.status == 'posted':
        messages.error(request, 'å·²è¿‡è´¦çš„å‡­è¯ä¸èƒ½ç¼–è¾‘')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        form = VoucherForm(request.POST, instance=voucher)
        formset = VoucherEntryFormSet(request.POST, instance=voucher)
        
        if form.is_valid() and formset.is_valid():
            voucher = form.save()
            
            # ä¿å­˜åˆ†å½•å¹¶è®¡ç®—åˆè®¡
            entries = formset.save(commit=False)
            total_debit = Decimal('0.00')
            total_credit = Decimal('0.00')
            
            for entry in entries:
                entry.voucher = voucher
                entry.save()
                total_debit += entry.debit_amount or Decimal('0.00')
                total_credit += entry.credit_amount or Decimal('0.00')
            
            # åˆ é™¤æ ‡è®°ä¸ºåˆ é™¤çš„åˆ†å½•
            for obj in formset.deleted_objects:
                obj.delete()
            
            # æ›´æ–°åˆè®¡
            voucher.total_debit = total_debit
            voucher.total_credit = total_credit
            voucher.save()
            
            messages.success(request, f'è®°è´¦å‡­è¯ {voucher.voucher_number} æ›´æ–°æˆåŠŸï¼')
            return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
        else:
            messages.error(request, 'è¯·æ£€æŸ¥è¡¨å•ä¸­çš„é”™è¯¯ã€‚')
    else:
        form = VoucherForm(instance=voucher)
        formset = VoucherEntryFormSet(instance=voucher)
    
    # è·å–æ‰€æœ‰ä¼šè®¡ç§‘ç›®ä¾› JavaScript ä½¿ç”¨
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    
    context = _context(
        f"ç¼–è¾‘è®°è´¦å‡­è¯ - {voucher.voucher_number}",
        "âœï¸",
        f"ç¼–è¾‘è®°è´¦å‡­è¯ {voucher.voucher_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'formset': formset,
        'voucher': voucher,
        'is_create': False,
        'account_subjects': account_subjects,
    })
    return render(request, "financial_management/voucher_form.html", context)


@login_required
def voucher_management(request):
    """å‡­è¯ç®¡ç†"""
    permission_codes = get_user_permission_codes(request.user)
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # è·å–å‡­è¯åˆ—è¡¨
    try:
        vouchers = Voucher.objects.select_related('preparer', 'reviewer', 'posted_by').order_by('-voucher_date', '-voucher_number')
        
        # åº”ç”¨ç­›é€‰æ¡ä»¶
        if search:
            vouchers = vouchers.filter(
                Q(voucher_number__icontains=search) |
                Q(notes__icontains=search)
            )
        if status:
            vouchers = vouchers.filter(status=status)
        if date_from:
            vouchers = vouchers.filter(voucher_date__gte=date_from)
        if date_to:
            vouchers = vouchers.filter(voucher_date__lte=date_to)
        
        # åˆ†é¡µ
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(vouchers, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–å‡­è¯åˆ—è¡¨å¤±è´¥: %s', str(e))
        page_obj = None
    
    # ç»Ÿè®¡ä¿¡æ¯
    try:
        total_vouchers = Voucher.objects.count()
        pending_vouchers = Voucher.objects.filter(status='submitted').count()
        approved_vouchers = Voucher.objects.filter(status='approved').count()
        posted_vouchers = Voucher.objects.filter(status='posted').count()
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ç»Ÿè®¡ä¿¡æ¯å¤±è´¥: %s', str(e))
        summary_cards = []
    
    context = _context(
        "å‡­è¯ç®¡ç†",
        "ğŸ“",
        "ç®¡ç†è®°è´¦å‡­è¯",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'vouchers': page_obj.object_list if page_obj else [],
        'status_choices': Voucher.STATUS_CHOICES,
        'current_search': search,
        'current_status': status,
        'current_date_from': date_from,
        'current_date_to': date_to,
    })
    return render(request, "financial_management/voucher_list.html", context)


@login_required
def voucher_export(request):
    """å¯¼å‡ºå‡­è¯åˆ—è¡¨ä¸ºExcel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å¯¼å‡ºå‡­è¯')
        return redirect('finance_pages:voucher_management')
    
    # è·å–ç­›é€‰å‚æ•°ï¼ˆä¸åˆ—è¡¨é¡µä¿æŒä¸€è‡´ï¼‰
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # è·å–å‡­è¯åˆ—è¡¨
    vouchers = Voucher.objects.select_related('preparer', 'reviewer', 'posted_by').order_by('-voucher_date', '-voucher_number')
    
    if search:
        vouchers = vouchers.filter(
            Q(voucher_number__icontains=search) |
            Q(notes__icontains=search)
        )
    if status:
        vouchers = vouchers.filter(status=status)
    if date_from:
        vouchers = vouchers.filter(voucher_date__gte=date_from)
    if date_to:
        vouchers = vouchers.filter(voucher_date__lte=date_to)
    
    # åˆ›å»ºExcelå·¥ä½œç°¿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'å‡­è¯åˆ—è¡¨'
    
    # è®¾ç½®è¡¨å¤´
    headers = ['å‡­è¯å­—å·', 'å‡­è¯æ—¥æœŸ', 'å€Ÿæ–¹åˆè®¡', 'è´·æ–¹åˆè®¡', 'çŠ¶æ€', 'åˆ¶å•äºº', 'å®¡æ ¸äºº', 'å®¡æ ¸æ—¶é—´', 'è¿‡è´¦äºº', 'è¿‡è´¦æ—¶é—´', 'é™„ä»¶æ•°', 'å¤‡æ³¨']
    worksheet.append(headers)
    
    # è®¾ç½®è¡¨å¤´æ ·å¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # æ·»åŠ æ•°æ®
    status_dict = dict(Voucher.STATUS_CHOICES)
    for voucher in vouchers:
        row = [
            voucher.voucher_number,
            voucher.voucher_date.strftime('%Y-%m-%d') if voucher.voucher_date else '',
            float(voucher.total_debit),
            float(voucher.total_credit),
            status_dict.get(voucher.status, voucher.status),
            voucher.preparer.get_full_name() if voucher.preparer else '',
            voucher.reviewer.get_full_name() if voucher.reviewer else '',
            voucher.reviewed_time.strftime('%Y-%m-%d %H:%M') if voucher.reviewed_time else '',
            voucher.posted_by.get_full_name() if voucher.posted_by else '',
            voucher.posted_time.strftime('%Y-%m-%d %H:%M') if voucher.posted_time else '',
            voucher.attachment_count,
            voucher.notes or '',
        ]
        worksheet.append(row)
    
    # è°ƒæ•´åˆ—å®½
    column_widths = [18, 12, 12, 12, 10, 12, 12, 18, 12, 18, 10, 30]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    # ç”Ÿæˆå“åº”
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('å‡­è¯åˆ—è¡¨_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def voucher_batch_approve(request):
    """æ‰¹é‡å®¡æ ¸å‡­è¯"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.review', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æ‰¹é‡å®¡æ ¸å‡­è¯')
        return redirect('finance_pages:voucher_management')
    
    if request.method == 'POST':
        voucher_ids = request.POST.getlist('voucher_ids')
        action = request.POST.get('action', 'approve')
        
        if not voucher_ids:
            messages.error(request, 'è¯·é€‰æ‹©è¦æ“ä½œçš„å‡­è¯')
            return redirect('finance_pages:voucher_management')
        
        try:
            vouchers = Voucher.objects.filter(id__in=voucher_ids)
            success_count = 0
            error_count = 0
            
            for voucher in vouchers:
                # æ£€æŸ¥çŠ¶æ€
                if voucher.status not in ['submitted', 'draft']:
                    error_count += 1
                    continue
                
                # æ£€æŸ¥å€Ÿè´·æ˜¯å¦å¹³è¡¡
                if voucher.total_debit != voucher.total_credit:
                    error_count += 1
                    continue
                
                if action == 'approve':
                    voucher.status = 'approved'
                    voucher.reviewer = request.user
                    voucher.reviewed_time = timezone.now()
                    voucher.save()
                    success_count += 1
                elif action == 'reject':
                    voucher.status = 'rejected'
                    voucher.reviewer = request.user
                    voucher.reviewed_time = timezone.now()
                    voucher.save()
                    success_count += 1
            
            if success_count > 0:
                messages.success(request, f'æˆåŠŸ{action == "approve" and "å®¡æ ¸é€šè¿‡" or "æ‹’ç»"} {success_count} å¼ å‡­è¯')
            if error_count > 0:
                messages.warning(request, f'{error_count} å¼ å‡­è¯æ“ä½œå¤±è´¥ï¼ˆçŠ¶æ€ä¸ç¬¦åˆè¦æ±‚æˆ–å€Ÿè´·ä¸å¹³è¡¡ï¼‰')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('æ‰¹é‡å®¡æ ¸å‡­è¯å¤±è´¥: %s', str(e))
            messages.error(request, f'æ‰¹é‡æ“ä½œå¤±è´¥ï¼š{str(e)}')
    
    return redirect('finance_pages:voucher_management')


@login_required
def voucher_batch_post(request):
    """æ‰¹é‡è¿‡è´¦å‡­è¯"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.post', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æ‰¹é‡è¿‡è´¦å‡­è¯')
        return redirect('finance_pages:voucher_management')
    
    if request.method == 'POST':
        voucher_ids = request.POST.getlist('voucher_ids')
        
        if not voucher_ids:
            messages.error(request, 'è¯·é€‰æ‹©è¦è¿‡è´¦çš„å‡­è¯')
            return redirect('finance_pages:voucher_management')
        
        try:
            vouchers = Voucher.objects.filter(id__in=voucher_ids)
            success_count = 0
            error_count = 0
            
            for voucher in vouchers:
                # æ£€æŸ¥çŠ¶æ€
                if voucher.status != 'approved':
                    error_count += 1
                    continue
                
                # æ£€æŸ¥å€Ÿè´·æ˜¯å¦å¹³è¡¡
                if voucher.total_debit != voucher.total_credit:
                    error_count += 1
                    continue
                
                # æ‰§è¡Œè¿‡è´¦
                try:
                    from django.db import transaction
                    with transaction.atomic():
                        # æ›´æ–°å‡­è¯çŠ¶æ€
                        voucher.status = 'posted'
                        voucher.posted_by = request.user
                        voucher.posted_time = timezone.now()
                        voucher.save()
                        
                        # ç”Ÿæˆæ€»è´¦è®°å½•
                        entries = voucher.entries.select_related('account_subject').all()
                        period_year = voucher.voucher_date.year
                        period_month = voucher.voucher_date.month
                        
                        for entry in entries:
                            # è·å–æˆ–åˆ›å»ºæ€»è´¦è®°å½•
                            ledger, created = Ledger.objects.get_or_create(
                                account_subject=entry.account_subject,
                                period_year=period_year,
                                period_month=period_month,
                                period_date=voucher.voucher_date,
                                defaults={
                                    'opening_balance': Decimal('0.00'),
                                    'period_debit': Decimal('0.00'),
                                    'period_credit': Decimal('0.00'),
                                    'closing_balance': Decimal('0.00'),
                                }
                            )
                            
                            # æ›´æ–°æ€»è´¦
                            ledger.period_debit += entry.debit_amount or Decimal('0.00')
                            ledger.period_credit += entry.credit_amount or Decimal('0.00')
                            
                            # è®¡ç®—æœŸæœ«ä½™é¢
                            if entry.account_subject.direction == 'debit':
                                ledger.closing_balance = ledger.opening_balance + ledger.period_debit - ledger.period_credit
                            else:
                                ledger.closing_balance = ledger.opening_balance + ledger.period_credit - ledger.period_debit
                            
                            ledger.save()
                        
                        success_count += 1
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('è¿‡è´¦å‡­è¯å¤±è´¥: %s', str(e))
                    error_count += 1
            
            if success_count > 0:
                messages.success(request, f'æˆåŠŸè¿‡è´¦ {success_count} å¼ å‡­è¯')
            if error_count > 0:
                messages.warning(request, f'{error_count} å¼ å‡­è¯è¿‡è´¦å¤±è´¥')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('æ‰¹é‡è¿‡è´¦å‡­è¯å¤±è´¥: %s', str(e))
            messages.error(request, f'æ‰¹é‡è¿‡è´¦å¤±è´¥ï¼š{str(e)}')
    
    return redirect('finance_pages:voucher_management')


@login_required
def ledger_management(request):
    """è´¦ç°¿ç®¡ç†"""
    permission_codes = get_user_permission_codes(request.user)
    today = timezone.now().date()
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', '')
    account_subject_id = request.GET.get('account_subject_id', '')
    
    # è·å–æ€»è´¦åˆ—è¡¨
    try:
        ledgers = Ledger.objects.select_related('account_subject').order_by('-period_date', 'account_subject__code')
        
        # åº”ç”¨ç­›é€‰æ¡ä»¶
        if search:
            ledgers = ledgers.filter(
                Q(account_subject__code__icontains=search) |
                Q(account_subject__name__icontains=search)
            )
        if period_year:
            ledgers = ledgers.filter(period_year=int(period_year))
        if period_month:
            ledgers = ledgers.filter(period_month=int(period_month))
        if account_subject_id:
            ledgers = ledgers.filter(account_subject_id=int(account_subject_id))
        
        # åˆ†é¡µ
        paginator = Paginator(ledgers, 50)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–æ€»è´¦åˆ—è¡¨å¤±è´¥: %s', str(e))
        page_obj = None
    
    # ç»Ÿè®¡ä¿¡æ¯
    try:
        current_year = int(period_year) if period_year else today.year
        current_month = int(period_month) if period_month else today.month
        ledger_count = Ledger.objects.filter(
            period_year=current_year,
            period_month=current_month
        ).count()
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ç»Ÿè®¡ä¿¡æ¯å¤±è´¥: %s', str(e))
        summary_cards = []
    
    context = _context(
        "è´¦ç°¿ç®¡ç†",
        "ğŸ“–",
        "ç®¡ç†æ€»è´¦ã€æ˜ç»†è´¦ç­‰",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'ledgers': page_obj.object_list if page_obj else [],
        'current_search': search,
        'current_period_year': period_year,
        'current_period_month': period_month,
        'current_account_subject_id': account_subject_id,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/ledger_list.html", context)


@login_required
def budget_management(request):
    """é¢„ç®—ç®¡ç†"""
    permission_codes = get_user_permission_codes(request.user)
    today = timezone.now().date()
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    budget_year = request.GET.get('budget_year', '')
    
    # è·å–é¢„ç®—åˆ—è¡¨
    try:
        budgets = Budget.objects.select_related('department', 'account_subject', 'approver', 'created_by').order_by('-budget_year', '-created_time')
        
        # åº”ç”¨ç­›é€‰æ¡ä»¶
        if search:
            budgets = budgets.filter(
                Q(budget_number__icontains=search) |
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        if status:
            budgets = budgets.filter(status=status)
        if budget_year:
            budgets = budgets.filter(budget_year=int(budget_year))
        
        # åˆ†é¡µ
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(budgets, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–é¢„ç®—åˆ—è¡¨å¤±è´¥: %s', str(e))
        page_obj = None
    
    # ç»Ÿè®¡ä¿¡æ¯
    try:
        total_budgets = Budget.objects.count()
        executing_budgets = Budget.objects.filter(status='executing').count()
        total_budget_amount = Budget.objects.filter(status='executing').aggregate(
            total=Sum('budget_amount')
        )['total'] or Decimal('0')
        total_used_amount = Budget.objects.filter(status='executing').aggregate(
            total=Sum('used_amount')
        )['total'] or Decimal('0')
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ç»Ÿè®¡ä¿¡æ¯å¤±è´¥: %s', str(e))
        summary_cards = []
    
    context = _context(
        "é¢„ç®—ç®¡ç†",
        "ğŸ’°",
        "ç®¡ç†é¢„ç®—ç¼–åˆ¶å’Œæ‰§è¡Œ",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'budgets': page_obj.object_list if page_obj else [],
        'status_choices': Budget.STATUS_CHOICES,
        'current_search': search,
        'current_status': status,
        'current_budget_year': budget_year,
        'years': range(today.year - 2, today.year + 2),
    })
    return render(request, "financial_management/budget_list.html", context)


@login_required
def budget_execution_analysis(request):
    """é¢„ç®—æ‰§è¡Œæƒ…å†µåˆ†æ"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.budget.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æŸ¥çœ‹é¢„ç®—æ‰§è¡Œåˆ†æ')
        return redirect('finance_pages:budget_management')
    
    today = timezone.now().date()
    
    # è·å–ç­›é€‰å‚æ•°
    budget_year = int(request.GET.get('budget_year', today.year))
    department_id = request.GET.get('department_id', '')
    account_subject_id = request.GET.get('account_subject_id', '')
    
    # è·å–é¢„ç®—æ•°æ®
    budgets = Budget.objects.select_related('department', 'account_subject', 'approver', 'created_by').filter(
        budget_year=budget_year,
        status__in=['approved', 'executing']
    )
    
    if department_id:
        budgets = budgets.filter(department_id=department_id)
    if account_subject_id:
        budgets = budgets.filter(account_subject_id=account_subject_id)
    
    # è®¡ç®—æ‰§è¡Œæƒ…å†µ
    execution_data = []
    total_budget = Decimal('0')
    total_used = Decimal('0')
    total_remaining = Decimal('0')
    
    for budget in budgets:
        usage_rate = 0
        if budget.budget_amount > 0:
            usage_rate = (budget.used_amount / budget.budget_amount) * 100
        
        remaining_rate = 0
        if budget.budget_amount > 0:
            remaining_rate = (budget.remaining_amount / budget.budget_amount) * 100
        
        execution_data.append({
            'budget': budget,
            'usage_rate': usage_rate,
            'remaining_rate': remaining_rate,
            'is_over_budget': budget.used_amount > budget.budget_amount,
            'is_near_limit': usage_rate >= 80 and usage_rate < 100,
        })
        
        total_budget += budget.budget_amount
        total_used += budget.used_amount
        total_remaining += budget.remaining_amount
    
    # è®¡ç®—æ€»ä½“æ‰§è¡Œç‡
    overall_usage_rate = 0
    if total_budget > 0:
        overall_usage_rate = (total_used / total_budget) * 100
    
    # æŒ‰æ‰§è¡Œç‡æ’åº
    execution_data.sort(key=lambda x: x['usage_rate'], reverse=True)
    
    # ç»Ÿè®¡ä¿¡æ¯
    summary_cards = []
    
    # è·å–éƒ¨é—¨åˆ—è¡¨å’Œç§‘ç›®åˆ—è¡¨ä¾›ç­›é€‰
    try:
        from backend.apps.system_management.models import Department
        departments = Department.objects.filter(is_active=True).order_by('name')
    except Exception:
        departments = []
    
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    
    context = _context(
        f"é¢„ç®—æ‰§è¡Œæƒ…å†µåˆ†æ - {budget_year}å¹´",
        "ğŸ“Š",
        f"åˆ†æ {budget_year} å¹´åº¦é¢„ç®—æ‰§è¡Œæƒ…å†µ",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'execution_data': execution_data,
        'budget_year': budget_year,
        'total_budget': total_budget,
        'total_used': total_used,
        'total_remaining': total_remaining,
        'overall_usage_rate': overall_usage_rate,
        'departments': departments,
        'account_subjects': account_subjects,
        'current_department_id': department_id,
        'current_account_subject_id': account_subject_id,
        'years': range(today.year - 2, today.year + 2),
    })
    return render(request, "financial_management/budget_execution_analysis.html", context)


@login_required
def budget_export(request):
    """å¯¼å‡ºé¢„ç®—åˆ—è¡¨ä¸ºExcel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.budget.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å¯¼å‡ºé¢„ç®—')
        return redirect('finance_pages:budget_management')
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    budget_year = request.GET.get('budget_year', '')
    
    budgets = Budget.objects.select_related('department', 'account_subject', 'approver', 'created_by').order_by('-budget_year', '-created_time')
    
    if search:
        budgets = budgets.filter(
            Q(budget_number__icontains=search) |
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    if status:
        budgets = budgets.filter(status=status)
    if budget_year:
        budgets = budgets.filter(budget_year=int(budget_year))
    
    # åˆ›å»ºExcelå·¥ä½œç°¿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'é¢„ç®—åˆ—è¡¨'
    
    headers = ['é¢„ç®—ç¼–å·', 'é¢„ç®—åç§°', 'é¢„ç®—å¹´åº¦', 'é¢„ç®—é‡‘é¢', 'å·²ç”¨é‡‘é¢', 'å‰©ä½™é‡‘é¢', 'æ‰€å±éƒ¨é—¨', 'é¢„ç®—ç§‘ç›®', 'çŠ¶æ€', 'å®¡æ‰¹äºº', 'å®¡æ‰¹æ—¶é—´', 'å¼€å§‹æ—¥æœŸ', 'ç»“æŸæ—¥æœŸ', 'åˆ›å»ºäºº', 'åˆ›å»ºæ—¶é—´']
    worksheet.append(headers)
    
    # è®¾ç½®è¡¨å¤´æ ·å¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # æ·»åŠ æ•°æ®
    status_dict = dict(Budget.STATUS_CHOICES)
    for budget in budgets:
        row = [
            budget.budget_number,
            budget.name,
            budget.budget_year,
            float(budget.budget_amount),
            float(budget.used_amount),
            float(budget.remaining_amount),
            budget.department.name if budget.department else '',
            f"{budget.account_subject.code} {budget.account_subject.name}" if budget.account_subject else '',
            status_dict.get(budget.status, budget.status),
            budget.approver.get_full_name() if budget.approver else '',
            budget.approved_time.strftime('%Y-%m-%d %H:%M') if budget.approved_time else '',
            budget.start_date.strftime('%Y-%m-%d') if budget.start_date else '',
            budget.end_date.strftime('%Y-%m-%d') if budget.end_date else '',
            budget.created_by.get_full_name() if budget.created_by else '',
            budget.created_time.strftime('%Y-%m-%d %H:%M') if budget.created_time else '',
        ]
        worksheet.append(row)
    
    # è°ƒæ•´åˆ—å®½
    column_widths = [18, 25, 10, 12, 12, 12, 15, 20, 10, 12, 18, 12, 12, 12, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('é¢„ç®—åˆ—è¡¨_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def invoice_management(request):
    """å‘ç¥¨ç®¡ç†"""
    permission_codes = get_user_permission_codes(request.user)
    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    invoice_type = request.GET.get('invoice_type', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # è·å–å‘ç¥¨åˆ—è¡¨
    try:
        invoices = Invoice.objects.select_related('verified_by', 'created_by').order_by('-invoice_date', '-invoice_number')
        
        # åº”ç”¨ç­›é€‰æ¡ä»¶
        if search:
            invoices = invoices.filter(
                Q(invoice_number__icontains=search) |
                Q(invoice_code__icontains=search) |
                Q(customer_name__icontains=search) |
                Q(supplier_name__icontains=search)
            )
        if invoice_type:
            invoices = invoices.filter(invoice_type=invoice_type)
        if status:
            invoices = invoices.filter(status=status)
        if date_from:
            invoices = invoices.filter(invoice_date__gte=date_from)
        if date_to:
            invoices = invoices.filter(invoice_date__lte=date_to)
        
        # åˆ†é¡µ
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(invoices, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–å‘ç¥¨åˆ—è¡¨å¤±è´¥: %s', str(e))
        page_obj = None
    
    # ç»Ÿè®¡ä¿¡æ¯
    try:
        total_invoices = Invoice.objects.count()
        issued_invoices = Invoice.objects.filter(status='issued').count()
        this_month_invoices = Invoice.objects.filter(invoice_date__gte=this_month_start).count()
        this_month_amount = Invoice.objects.filter(invoice_date__gte=this_month_start).aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0')
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ç»Ÿè®¡ä¿¡æ¯å¤±è´¥: %s', str(e))
        summary_cards = []
    
    context = _context(
        "å‘ç¥¨ç®¡ç†",
        "ğŸ§¾",
        "ç®¡ç†è¿›é¡¹å’Œé”€é¡¹å‘ç¥¨",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'invoices': page_obj.object_list if page_obj else [],
        'invoice_type_choices': Invoice.TYPE_CHOICES,
        'status_choices': Invoice.STATUS_CHOICES,
        'current_search': search,
        'current_invoice_type': invoice_type,
        'current_status': status,
        'current_date_from': date_from,
        'current_date_to': date_to,
    })
    return render(request, "financial_management/invoice_list.html", context)


@login_required
def invoice_export(request):
    """å¯¼å‡ºå‘ç¥¨åˆ—è¡¨ä¸ºExcel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.invoice.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å¯¼å‡ºå‘ç¥¨')
        return redirect('finance_pages:invoice_management')
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    invoice_type = request.GET.get('invoice_type', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    invoices = Invoice.objects.select_related('project', 'created_by').order_by('-invoice_date', '-invoice_number')
    
    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search) |
            Q(seller_name__icontains=search) |
            Q(buyer_name__icontains=search) |
            Q(description__icontains=search)
        )
    if invoice_type:
        invoices = invoices.filter(invoice_type=invoice_type)
    if status:
        invoices = invoices.filter(status=status)
    if date_from:
        invoices = invoices.filter(invoice_date__gte=date_from)
    if date_to:
        invoices = invoices.filter(invoice_date__lte=date_to)
    
    # åˆ›å»ºExcelå·¥ä½œç°¿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'å‘ç¥¨åˆ—è¡¨'
    
    headers = ['å‘ç¥¨å·ç ', 'å‘ç¥¨ç±»å‹', 'å‘ç¥¨æ—¥æœŸ', 'é”€å”®æ–¹', 'è´­ä¹°æ–¹', 'é‡‘é¢', 'ç¨é¢', 'ä»·ç¨åˆè®¡', 'çŠ¶æ€', 'å…³è”é¡¹ç›®', 'å¤‡æ³¨', 'åˆ›å»ºäºº', 'åˆ›å»ºæ—¶é—´']
    worksheet.append(headers)
    
    # è®¾ç½®è¡¨å¤´æ ·å¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # æ·»åŠ æ•°æ®
    type_dict = dict(Invoice.TYPE_CHOICES)
    status_dict = dict(Invoice.STATUS_CHOICES)
    for invoice in invoices:
        row = [
            invoice.invoice_number,
            type_dict.get(invoice.invoice_type, invoice.invoice_type),
            invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else '',
            invoice.seller_name or '',
            invoice.buyer_name or '',
            float(invoice.amount) if invoice.amount else 0,
            float(invoice.tax_amount) if invoice.tax_amount else 0,
            float(invoice.total_amount) if invoice.total_amount else 0,
            status_dict.get(invoice.status, invoice.status),
            invoice.project.project_number if invoice.project else '',
            invoice.description or '',
            invoice.created_by.get_full_name() if invoice.created_by else '',
            invoice.created_time.strftime('%Y-%m-%d %H:%M') if invoice.created_time else '',
        ]
        worksheet.append(row)
    
    # è°ƒæ•´åˆ—å®½
    column_widths = [20, 10, 12, 20, 20, 12, 12, 12, 10, 15, 30, 12, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('å‘ç¥¨åˆ—è¡¨_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def fund_flow_management(request):
    """èµ„é‡‘æµæ°´"""
    permission_codes = get_user_permission_codes(request.user)
    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    flow_type = request.GET.get('flow_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # è·å–èµ„é‡‘æµæ°´åˆ—è¡¨
    try:
        fund_flows = FundFlow.objects.select_related('project', 'voucher', 'created_by').order_by('-flow_date', '-flow_number')
        
        # åº”ç”¨ç­›é€‰æ¡ä»¶
        if search:
            fund_flows = fund_flows.filter(
                Q(flow_number__icontains=search) |
                Q(account_name__icontains=search) |
                Q(counterparty__icontains=search) |
                Q(summary__icontains=search)
            )
        if flow_type:
            fund_flows = fund_flows.filter(flow_type=flow_type)
        if date_from:
            fund_flows = fund_flows.filter(flow_date__gte=date_from)
        if date_to:
            fund_flows = fund_flows.filter(flow_date__lte=date_to)
        
        # åˆ†é¡µ
        paginator = Paginator(fund_flows, 50)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–èµ„é‡‘æµæ°´åˆ—è¡¨å¤±è´¥: %s', str(e))
        page_obj = None
    
    # ç»Ÿè®¡ä¿¡æ¯
    try:
        this_month_flows = FundFlow.objects.filter(flow_date__gte=this_month_start).count()
        this_month_income = FundFlow.objects.filter(
            flow_date__gte=this_month_start,
            flow_type='income'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        this_month_expense = FundFlow.objects.filter(
            flow_date__gte=this_month_start,
            flow_type='expense'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ç»Ÿè®¡ä¿¡æ¯å¤±è´¥: %s', str(e))
        summary_cards = []
    
    context = _context(
        "èµ„é‡‘æµæ°´",
        "ğŸ’³",
        "ç®¡ç†èµ„é‡‘æµå…¥æµå‡ºè®°å½•",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'fund_flows': page_obj.object_list if page_obj else [],
        'flow_type_choices': FundFlow.TYPE_CHOICES,
        'current_search': search,
        'current_flow_type': flow_type,
        'current_date_from': date_from,
        'current_date_to': date_to,
    })
    return render(request, "financial_management/fund_flow_list.html", context)


@login_required
def fund_flow_import_template(request):
    """ä¸‹è½½èµ„é‡‘æµæ°´å¯¼å…¥æ¨¡æ¿"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.fund_flow.create', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™ä¸‹è½½å¯¼å…¥æ¨¡æ¿')
        return redirect('finance_pages:fund_flow_management')
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'èµ„é‡‘æµæ°´å¯¼å…¥æ¨¡æ¿'
    
    # è®¾ç½®è¡¨å¤´
    headers = ['å‘ç”Ÿæ—¥æœŸ', 'æµæ°´ç±»å‹', 'é‡‘é¢', 'è´¦æˆ·åç§°', 'å¯¹æ–¹å•ä½', 'æ‘˜è¦', 'å…³è”é¡¹ç›®ç¼–å·']
    worksheet.append(headers)
    
    # è®¾ç½®è¡¨å¤´æ ·å¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # æ·»åŠ ç¤ºä¾‹æ•°æ®
    examples = [
        ['2025-01-15', 'income', '10000.00', 'å·¥å•†é“¶è¡Œ', 'å®¢æˆ·A', 'é¡¹ç›®å›æ¬¾', 'PJT-2025-001'],
        ['2025-01-16', 'expense', '5000.00', 'å·¥å•†é“¶è¡Œ', 'ä¾›åº”å•†B', 'ææ–™é‡‡è´­', 'PJT-2025-002'],
    ]
    for example in examples:
        worksheet.append(example)
    
    # æ·»åŠ è¯´æ˜è¡Œ
    worksheet.append([])
    worksheet.append(['è¯´æ˜ï¼š'])
    worksheet.append(['1. å‘ç”Ÿæ—¥æœŸï¼šå¿…å¡«ï¼Œæ ¼å¼ï¼šYYYY-MM-DD'])
    worksheet.append(['2. æµæ°´ç±»å‹ï¼šå¿…å¡«ï¼Œå¯é€‰å€¼ï¼šincome(æ”¶å…¥)ã€expense(æ”¯å‡º)ã€transfer(è½¬è´¦)'])
    worksheet.append(['3. é‡‘é¢ï¼šå¿…å¡«ï¼Œæ•°å­—æ ¼å¼'])
    worksheet.append(['4. è´¦æˆ·åç§°ï¼šå¿…å¡«'])
    worksheet.append(['5. å¯¹æ–¹å•ä½ï¼šå¯é€‰'])
    worksheet.append(['6. æ‘˜è¦ï¼šå¿…å¡«'])
    worksheet.append(['7. å…³è”é¡¹ç›®ç¼–å·ï¼šå¯é€‰ï¼Œå¦‚æœå¡«å†™ï¼Œç³»ç»Ÿä¼šè‡ªåŠ¨å…³è”é¡¹ç›®'])
    
    # è°ƒæ•´åˆ—å®½
    column_widths = [15, 12, 15, 20, 20, 30, 20]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="èµ„é‡‘æµæ°´å¯¼å…¥æ¨¡æ¿.xlsx"'
    workbook.save(response)
    return response


@login_required
def fund_flow_import(request):
    """å¯¼å…¥èµ„é‡‘æµæ°´"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.fund_flow.create', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å¯¼å…¥èµ„é‡‘æµæ°´')
        return redirect('finance_pages:fund_flow_management')
    
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'è¯·é€‰æ‹©è¦å¯¼å…¥çš„æ–‡ä»¶')
            return redirect('finance_pages:fund_flow_management')
        
        upload_file = request.FILES['file']
        if not upload_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'è¯·ä¸Šä¼ Excelæ–‡ä»¶ï¼ˆ.xlsæˆ–.xlsxæ ¼å¼ï¼‰')
            return redirect('finance_pages:fund_flow_management')
        
        try:
            from django.db import transaction
            from datetime import datetime
            workbook = load_workbook(upload_file, data_only=True)
            worksheet = workbook.active
            
            # è¯»å–è¡¨å¤´
            headers = [cell.value for cell in worksheet[1]]
            header_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
            
            # æ£€æŸ¥å¿…å¡«åˆ—
            required_columns = ['å‘ç”Ÿæ—¥æœŸ', 'æµæ°´ç±»å‹', 'é‡‘é¢', 'è´¦æˆ·åç§°', 'æ‘˜è¦']
            missing_columns = [col for col in required_columns if col not in header_map]
            if missing_columns:
                messages.error(request, f'ç¼ºå°‘å¿…å¡«åˆ—ï¼š{", ".join(missing_columns)}')
                return redirect('finance_pages:fund_flow_management')
            
            success_count = 0
            error_count = 0
            errors = []
            current_year = timezone.now().year
            
            # è¯»å–æ•°æ®è¡Œ
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                # è·³è¿‡ç©ºè¡Œ
                if not row or not any(row):
                    continue
                
                try:
                    # è§£ææ—¥æœŸ
                    flow_date_str = str(row[header_map['å‘ç”Ÿæ—¥æœŸ']]).strip() if row[header_map['å‘ç”Ÿæ—¥æœŸ']] else None
                    if not flow_date_str:
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šå‘ç”Ÿæ—¥æœŸä¸èƒ½ä¸ºç©º')
                        continue
                    
                    try:
                        if isinstance(row[header_map['å‘ç”Ÿæ—¥æœŸ']], datetime):
                            flow_date = row[header_map['å‘ç”Ÿæ—¥æœŸ']].date()
                        else:
                            flow_date = datetime.strptime(flow_date_str, '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šå‘ç”Ÿæ—¥æœŸæ ¼å¼é”™è¯¯ï¼Œåº”ä¸ºYYYY-MM-DD')
                        continue
                    
                    # è§£ææµæ°´ç±»å‹
                    flow_type_str = str(row[header_map['æµæ°´ç±»å‹']]).strip() if row[header_map['æµæ°´ç±»å‹']] else None
                    type_map = {
                        'income': 'income', 'æ”¶å…¥': 'income',
                        'expense': 'expense', 'æ”¯å‡º': 'expense',
                        'transfer': 'transfer', 'è½¬è´¦': 'transfer',
                    }
                    flow_type = type_map.get(flow_type_str.lower(), flow_type_str)
                    if flow_type not in dict(FundFlow.TYPE_CHOICES):
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šæµæ°´ç±»å‹æ— æ•ˆ')
                        continue
                    
                    # è§£æé‡‘é¢
                    amount_str = str(row[header_map['é‡‘é¢']]).strip() if row[header_map['é‡‘é¢']] else None
                    if not amount_str:
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šé‡‘é¢ä¸èƒ½ä¸ºç©º')
                        continue
                    try:
                        amount = Decimal(amount_str)
                        if amount <= 0:
                            error_count += 1
                            errors.append(f'ç¬¬{row_idx}è¡Œï¼šé‡‘é¢å¿…é¡»å¤§äº0')
                            continue
                    except (ValueError, InvalidOperation):
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šé‡‘é¢æ ¼å¼é”™è¯¯')
                        continue
                    
                    # è§£æå…¶ä»–å­—æ®µ
                    account_name = str(row[header_map['è´¦æˆ·åç§°']]).strip() if row[header_map['è´¦æˆ·åç§°']] else None
                    if not account_name:
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šè´¦æˆ·åç§°ä¸èƒ½ä¸ºç©º')
                        continue
                    
                    counterparty = str(row[header_map.get('å¯¹æ–¹å•ä½', -1)]).strip() if header_map.get('å¯¹æ–¹å•ä½', -1) >= 0 and row[header_map.get('å¯¹æ–¹å•ä½', -1)] else ''
                    summary = str(row[header_map['æ‘˜è¦']]).strip() if row[header_map['æ‘˜è¦']] else None
                    if not summary:
                        error_count += 1
                        errors.append(f'ç¬¬{row_idx}è¡Œï¼šæ‘˜è¦ä¸èƒ½ä¸ºç©º')
                        continue
                    
                    project_number = str(row[header_map.get('å…³è”é¡¹ç›®ç¼–å·', -1)]).strip() if header_map.get('å…³è”é¡¹ç›®ç¼–å·', -1) >= 0 and row[header_map.get('å…³è”é¡¹ç›®ç¼–å·', -1)] else None
                    
                    # æŸ¥æ‰¾å…³è”é¡¹ç›®
                    project = None
                    if project_number:
                        try:
                            from backend.apps.production_management.models import Project
                            project = Project.objects.get(project_number=project_number)
                        except Project.DoesNotExist:
                            error_count += 1
                            errors.append(f'ç¬¬{row_idx}è¡Œï¼šé¡¹ç›®ç¼–å· {project_number} ä¸å­˜åœ¨')
                            continue
                    
                    # ç”Ÿæˆæµæ°´å·
                    max_flow = FundFlow.objects.filter(
                        flow_number__startswith=f'FLOW-{current_year}-'
                    ).order_by('-flow_number').first()
                    
                    if max_flow:
                        try:
                            seq = int(max_flow.flow_number.split('-')[-1]) + 1
                        except (ValueError, IndexError):
                            seq = 1
                    else:
                        seq = 1
                    
                    flow_number = f'FLOW-{current_year}-{seq:04d}'
                    
                    # åˆ›å»ºèµ„é‡‘æµæ°´
                    with transaction.atomic():
                        FundFlow.objects.create(
                            flow_number=flow_number,
                            flow_date=flow_date,
                            flow_type=flow_type,
                            amount=amount,
                            account_name=account_name,
                            counterparty=counterparty,
                            summary=summary,
                            project=project,
                            created_by=request.user,
                        )
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'ç¬¬{row_idx}è¡Œï¼š{str(e)}')
            
            if success_count > 0:
                messages.success(request, f'æˆåŠŸå¯¼å…¥ {success_count} æ¡èµ„é‡‘æµæ°´')
            if error_count > 0:
                error_msg = f'å¯¼å…¥å¤±è´¥ {error_count} æ¡è®°å½•'
                if len(errors) <= 10:
                    error_msg += 'ï¼š' + 'ï¼›'.join(errors)
                else:
                    error_msg += f'ï¼šå‰10ä¸ªé”™è¯¯ï¼š' + 'ï¼›'.join(errors[:10])
                messages.warning(request, error_msg)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('å¯¼å…¥èµ„é‡‘æµæ°´å¤±è´¥: %s', str(e))
            messages.error(request, f'å¯¼å…¥å¤±è´¥ï¼š{str(e)}')
    
    return redirect('finance_pages:fund_flow_management')


@login_required
def fund_flow_export(request):
    """å¯¼å‡ºèµ„é‡‘æµæ°´åˆ—è¡¨ä¸ºExcel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.fund_flow.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å¯¼å‡ºèµ„é‡‘æµæ°´')
        return redirect('finance_pages:fund_flow_management')
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    flow_type = request.GET.get('flow_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    fund_flows = FundFlow.objects.select_related('project', 'voucher', 'created_by').order_by('-flow_date', '-flow_number')
    
    if search:
        fund_flows = fund_flows.filter(
            Q(flow_number__icontains=search) |
            Q(account_name__icontains=search) |
            Q(counterparty__icontains=search) |
            Q(summary__icontains=search)
        )
    if flow_type:
        fund_flows = fund_flows.filter(flow_type=flow_type)
    if date_from:
        fund_flows = fund_flows.filter(flow_date__gte=date_from)
    if date_to:
        fund_flows = fund_flows.filter(flow_date__lte=date_to)
    
    # åˆ›å»ºExcelå·¥ä½œç°¿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'èµ„é‡‘æµæ°´'
    
    headers = ['æµæ°´å·', 'å‘ç”Ÿæ—¥æœŸ', 'æµæ°´ç±»å‹', 'é‡‘é¢', 'è´¦æˆ·åç§°', 'å¯¹æ–¹å•ä½', 'æ‘˜è¦', 'å…³è”é¡¹ç›®', 'å…³è”å‡­è¯', 'åˆ›å»ºäºº', 'åˆ›å»ºæ—¶é—´']
    worksheet.append(headers)
    
    # è®¾ç½®è¡¨å¤´æ ·å¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # æ·»åŠ æ•°æ®
    type_dict = dict(FundFlow.TYPE_CHOICES)
    for flow in fund_flows:
        row = [
            flow.flow_number,
            flow.flow_date.strftime('%Y-%m-%d') if flow.flow_date else '',
            type_dict.get(flow.flow_type, flow.flow_type),
            float(flow.amount),
            flow.account_name,
            flow.counterparty or '',
            flow.summary,
            flow.project.project_number if flow.project else '',
            flow.voucher.voucher_number if flow.voucher else '',
            flow.created_by.get_full_name() if flow.created_by else '',
            flow.created_time.strftime('%Y-%m-%d %H:%M') if flow.created_time else '',
        ]
        worksheet.append(row)
    
    # è°ƒæ•´åˆ—å®½
    column_widths = [18, 12, 10, 12, 15, 20, 30, 15, 18, 12, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('èµ„é‡‘æµæ°´_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def voucher_detail(request, voucher_id):
    """å‡­è¯è¯¦æƒ…"""
    voucher = get_object_or_404(Voucher.objects.select_related('preparer', 'reviewer', 'posted_by'), id=voucher_id)
    
    # è·å–å‡­è¯åˆ†å½•
    try:
        entries = voucher.entries.select_related('account_subject').order_by('line_number')
    except Exception:
        entries = []
    
    # æ£€æŸ¥æƒé™
    permission_codes = get_user_permission_codes(request.user)
    can_review = _permission_granted('financial_management.voucher.review', permission_codes)
    can_post = _permission_granted('financial_management.voucher.post', permission_codes)
    can_edit = _permission_granted('financial_management.voucher.manage', permission_codes) and voucher.status != 'posted'
    can_delete = _permission_granted('financial_management.voucher.manage', permission_codes) and voucher.status == 'draft'
    can_print = True  # æ‰€æœ‰ç”¨æˆ·éƒ½å¯ä»¥æ‰“å°
    
    # è®¡ç®—ç»Ÿè®¡ä¿¡æ¯
    entry_count = len(entries)
    debit_count = sum(1 for e in entries if e.debit_amount and e.debit_amount > 0)
    credit_count = sum(1 for e in entries if e.credit_amount and e.credit_amount > 0)
    
    # æ£€æŸ¥å€Ÿè´·å¹³è¡¡
    is_balanced = voucher.total_debit == voucher.total_credit
    balance_diff = abs(voucher.total_debit - voucher.total_credit)
    
    # è·å–å…³è”çš„æ€»è´¦è®°å½•æ•°
    try:
        ledger_count = Ledger.objects.filter(
            period_year=voucher.voucher_date.year if voucher.voucher_date else None,
            period_month=voucher.voucher_date.month if voucher.voucher_date else None,
        ).filter(
            account_subject__in=[e.account_subject for e in entries if e.account_subject]
        ).count() if voucher.status == 'posted' else 0
    except Exception:
        ledger_count = 0
    
    # æ„å»ºç»Ÿè®¡å¡ç‰‡
    summary_cards = []
    
    if voucher.status == 'posted':
        summary_cards.append({
            "label": "æ€»è´¦è®°å½•",
            "value": ledger_count,
            "hint": "å·²ç”Ÿæˆçš„æ€»è´¦è®°å½•æ•°"
        })
    
    context = _context(
        f"å‡­è¯è¯¦æƒ… - {voucher.voucher_number}",
        "ğŸ“",
        f"æŸ¥çœ‹è®°è´¦å‡­è¯ {voucher.voucher_number} çš„è¯¦ç»†ä¿¡æ¯å’Œåˆ†å½•",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
        'entries': entries,
        'can_review': can_review,
        'can_post': can_post,
        'can_edit': can_edit,
        'can_delete': can_delete,
        'can_print': can_print,
        'is_balanced': is_balanced,
        'balance_diff': balance_diff,
        'entry_count': entry_count,
    })
    return render(request, "financial_management/voucher_detail.html", context)


@login_required
def voucher_entry_add(request, voucher_id):
    """å¿«é€Ÿæ·»åŠ å‡­è¯åˆ†å½•ï¼ˆAJAXï¼‰"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        return JsonResponse({'success': False, 'error': 'æ‚¨æ²¡æœ‰æƒé™æ·»åŠ åˆ†å½•'}, status=403)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    
    # å·²è¿‡è´¦çš„å‡­è¯ä¸èƒ½æ·»åŠ åˆ†å½•
    if voucher.status == 'posted':
        return JsonResponse({'success': False, 'error': 'å·²è¿‡è´¦çš„å‡­è¯ä¸èƒ½æ·»åŠ åˆ†å½•'}, status=400)
    
    if request.method == 'POST':
        try:
            account_subject_id = request.POST.get('account_subject_id')
            summary = request.POST.get('summary', '')
            debit_amount_str = request.POST.get('debit_amount', '0')
            credit_amount_str = request.POST.get('credit_amount', '0')
            
            if not account_subject_id:
                return JsonResponse({'success': False, 'error': 'è¯·é€‰æ‹©ä¼šè®¡ç§‘ç›®'}, status=400)
            
            account_subject = AccountSubject.objects.get(id=account_subject_id)
            
            # è®¡ç®—ä¸‹ä¸€ä¸ªè¡Œå·
            max_line = voucher.entries.aggregate(max_line=Max('line_number'))['max_line'] or 0
            line_number = max_line + 1
            
            # åˆ›å»ºåˆ†å½•
            entry = VoucherEntry.objects.create(
                voucher=voucher,
                line_number=line_number,
                account_subject=account_subject,
                summary=summary,
                debit_amount=Decimal(debit_amount_str) if debit_amount_str else Decimal('0.00'),
                credit_amount=Decimal(credit_amount_str) if credit_amount_str else Decimal('0.00'),
            )
            
            # é‡æ–°è®¡ç®—å‡­è¯åˆè®¡
            total_debit = voucher.entries.aggregate(total=Sum('debit_amount'))['total'] or Decimal('0.00')
            total_credit = voucher.entries.aggregate(total=Sum('credit_amount'))['total'] or Decimal('0.00')
            voucher.total_debit = total_debit
            voucher.total_credit = total_credit
            voucher.save()
            
            return JsonResponse({
                'success': True,
                'entry': {
                    'id': entry.id,
                    'line_number': entry.line_number,
                    'account_subject': entry.account_subject.code + ' - ' + entry.account_subject.name,
                    'summary': entry.summary,
                    'debit_amount': str(entry.debit_amount),
                    'credit_amount': str(entry.credit_amount),
                },
                'voucher': {
                    'total_debit': str(voucher.total_debit),
                    'total_credit': str(voucher.total_credit),
                }
            })
        except AccountSubject.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'ä¼šè®¡ç§‘ç›®ä¸å­˜åœ¨'}, status=400)
        except (ValueError, InvalidOperation) as e:
            return JsonResponse({'success': False, 'error': f'é‡‘é¢æ ¼å¼é”™è¯¯ï¼š{str(e)}'}, status=400)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('æ·»åŠ åˆ†å½•å¤±è´¥: %s', str(e))
            return JsonResponse({'success': False, 'error': f'æ·»åŠ åˆ†å½•å¤±è´¥ï¼š{str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)


@login_required
def voucher_entry_update(request, voucher_id, entry_id):
    """å¿«é€Ÿç¼–è¾‘å‡­è¯åˆ†å½•ï¼ˆAJAXï¼‰"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        return JsonResponse({'success': False, 'error': 'æ‚¨æ²¡æœ‰æƒé™ç¼–è¾‘åˆ†å½•'}, status=403)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    entry = get_object_or_404(VoucherEntry, id=entry_id, voucher=voucher)
    
    # å·²è¿‡è´¦çš„å‡­è¯ä¸èƒ½ç¼–è¾‘åˆ†å½•
    if voucher.status == 'posted':
        return JsonResponse({'success': False, 'error': 'å·²è¿‡è´¦çš„å‡­è¯ä¸èƒ½ç¼–è¾‘åˆ†å½•'}, status=400)
    
    if request.method == 'POST':
        try:
            account_subject_id = request.POST.get('account_subject_id')
            summary = request.POST.get('summary', '')
            debit_amount_str = request.POST.get('debit_amount', '0')
            credit_amount_str = request.POST.get('credit_amount', '0')
            line_number_str = request.POST.get('line_number', '')
            
            if account_subject_id:
                account_subject = AccountSubject.objects.get(id=account_subject_id)
                entry.account_subject = account_subject
            
            if summary:
                entry.summary = summary
            
            if debit_amount_str:
                entry.debit_amount = Decimal(debit_amount_str) if debit_amount_str else Decimal('0.00')
            
            if credit_amount_str:
                entry.credit_amount = Decimal(credit_amount_str) if credit_amount_str else Decimal('0.00')
            
            if line_number_str:
                line_number = int(line_number_str)
                # æ£€æŸ¥è¡Œå·æ˜¯å¦å†²çª
                existing_entry = voucher.entries.filter(line_number=line_number).exclude(id=entry_id).first()
                if existing_entry:
                    return JsonResponse({'success': False, 'error': f'è¡Œå· {line_number} å·²è¢«ä½¿ç”¨'}, status=400)
                entry.line_number = line_number
            
            entry.save()
            
            # é‡æ–°è®¡ç®—å‡­è¯åˆè®¡
            total_debit = voucher.entries.aggregate(total=Sum('debit_amount'))['total'] or Decimal('0.00')
            total_credit = voucher.entries.aggregate(total=Sum('credit_amount'))['total'] or Decimal('0.00')
            voucher.total_debit = total_debit
            voucher.total_credit = total_credit
            voucher.save()
            
            return JsonResponse({
                'success': True,
                'entry': {
                    'id': entry.id,
                    'line_number': entry.line_number,
                    'account_subject': entry.account_subject.code + ' - ' + entry.account_subject.name,
                    'summary': entry.summary,
                    'debit_amount': str(entry.debit_amount),
                    'credit_amount': str(entry.credit_amount),
                },
                'voucher': {
                    'total_debit': str(voucher.total_debit),
                    'total_credit': str(voucher.total_credit),
                }
            })
        except AccountSubject.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'ä¼šè®¡ç§‘ç›®ä¸å­˜åœ¨'}, status=400)
        except (ValueError, InvalidOperation) as e:
            return JsonResponse({'success': False, 'error': f'é‡‘é¢æ ¼å¼é”™è¯¯ï¼š{str(e)}'}, status=400)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('ç¼–è¾‘åˆ†å½•å¤±è´¥: %s', str(e))
            return JsonResponse({'success': False, 'error': f'ç¼–è¾‘åˆ†å½•å¤±è´¥ï¼š{str(e)}'}, status=500)
    
    # GETè¯·æ±‚ï¼Œè¿”å›åˆ†å½•ä¿¡æ¯
    return JsonResponse({
        'success': True,
        'entry': {
            'id': entry.id,
            'line_number': entry.line_number,
            'account_subject_id': entry.account_subject.id,
            'account_subject': entry.account_subject.code + ' - ' + entry.account_subject.name,
            'summary': entry.summary,
            'debit_amount': str(entry.debit_amount),
            'credit_amount': str(entry.credit_amount),
        }
    })


@login_required
def voucher_entry_delete(request, voucher_id, entry_id):
    """å¿«é€Ÿåˆ é™¤å‡­è¯åˆ†å½•ï¼ˆAJAXï¼‰"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        return JsonResponse({'success': False, 'error': 'æ‚¨æ²¡æœ‰æƒé™åˆ é™¤åˆ†å½•'}, status=403)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    entry = get_object_or_404(VoucherEntry, id=entry_id, voucher=voucher)
    
    # å·²è¿‡è´¦çš„å‡­è¯ä¸èƒ½åˆ é™¤åˆ†å½•
    if voucher.status == 'posted':
        return JsonResponse({'success': False, 'error': 'å·²è¿‡è´¦çš„å‡­è¯ä¸èƒ½åˆ é™¤åˆ†å½•'}, status=400)
    
    if request.method == 'POST':
        try:
            entry.delete()
            
            # é‡æ–°è®¡ç®—å‡­è¯åˆè®¡
            total_debit = voucher.entries.aggregate(total=Sum('debit_amount'))['total'] or Decimal('0.00')
            total_credit = voucher.entries.aggregate(total=Sum('credit_amount'))['total'] or Decimal('0.00')
            voucher.total_debit = total_debit
            voucher.total_credit = total_credit
            voucher.save()
            
            return JsonResponse({
                'success': True,
                'voucher': {
                    'total_debit': str(voucher.total_debit),
                    'total_credit': str(voucher.total_credit),
                }
            })
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('åˆ é™¤åˆ†å½•å¤±è´¥: %s', str(e))
            return JsonResponse({'success': False, 'error': f'åˆ é™¤åˆ†å½•å¤±è´¥ï¼š{str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)


@login_required
def voucher_validate(request, voucher_id):
    """æ ¡éªŒå‡­è¯æ•°æ®ï¼ˆAJAXï¼‰"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.view', permission_codes):
        return HttpResponse('æ‚¨æ²¡æœ‰æƒé™æ ¡éªŒå‡­è¯', status=403)
    
    voucher = get_object_or_404(Voucher.objects.prefetch_related('entries'), id=voucher_id)
    
    errors = []
    warnings = []
    
    # æ£€æŸ¥æ˜¯å¦æœ‰åˆ†å½•
    if not voucher.entries.exists():
        errors.append('å‡­è¯è‡³å°‘éœ€è¦ä¸€æ¡åˆ†å½•')
    
    # æ£€æŸ¥å€Ÿè´·å¹³è¡¡
    if voucher.total_debit != voucher.total_credit:
        diff = abs(voucher.total_debit - voucher.total_credit)
        errors.append(f'å€Ÿè´·ä¸å¹³è¡¡ï¼Œå·®é¢ï¼š{diff:,.2f}')
    
    # æ£€æŸ¥åˆ†å½•ä¸­çš„ç§‘ç›®æ˜¯å¦æœ‰æ•ˆ
    for entry in voucher.entries.all():
        if not entry.account_subject.is_active:
            warnings.append(f'ç¬¬{entry.line_number}è¡Œï¼šç§‘ç›® {entry.account_subject.code} å·²åœç”¨')
        
        # æ£€æŸ¥å€Ÿæ–¹å’Œè´·æ–¹ä¸èƒ½åŒæ—¶æœ‰å€¼
        if entry.debit_amount > 0 and entry.credit_amount > 0:
            errors.append(f'ç¬¬{entry.line_number}è¡Œï¼šå€Ÿæ–¹å’Œè´·æ–¹ä¸èƒ½åŒæ—¶æœ‰é‡‘é¢')
        
        # æ£€æŸ¥å€Ÿæ–¹å’Œè´·æ–¹ä¸èƒ½åŒæ—¶ä¸º0
        if entry.debit_amount == 0 and entry.credit_amount == 0:
            warnings.append(f'ç¬¬{entry.line_number}è¡Œï¼šå€Ÿæ–¹å’Œè´·æ–¹é‡‘é¢éƒ½ä¸º0')
    
    # æ£€æŸ¥è¡Œå·æ˜¯å¦è¿ç»­
    line_numbers = sorted([e.line_number for e in voucher.entries.all()])
    if line_numbers:
        expected_lines = list(range(1, len(line_numbers) + 1))
        if line_numbers != expected_lines:
            warnings.append('åˆ†å½•è¡Œå·ä¸è¿ç»­')
    
    result = {
        'valid': len(errors) == 0,
        'errors': errors,
        'warnings': warnings,
    }
    
    return JsonResponse(result)


@login_required
def voucher_copy(request, voucher_id):
    """å¤åˆ¶å‡­è¯ï¼ˆåˆ›å»ºæ–°å‡­è¯ï¼‰"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.create', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ›å»ºå‡­è¯')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    source_voucher = get_object_or_404(Voucher.objects.prefetch_related('entries'), id=voucher_id)
    
    if request.method == 'POST':
        try:
            from django.db import transaction
            with transaction.atomic():
                # åˆ›å»ºæ–°å‡­è¯
                new_voucher = Voucher.objects.create(
                    voucher_date=request.POST.get('voucher_date', timezone.now().date()),
                    preparer=request.user,
                    status='draft',
                    attachment_count=0,
                    total_debit=source_voucher.total_debit,
                    total_credit=source_voucher.total_credit,
                    notes=f"å¤åˆ¶è‡ª {source_voucher.voucher_number}" + (f"\n{source_voucher.notes}" if source_voucher.notes else ""),
                )
                
                # è‡ªåŠ¨ç”Ÿæˆå‡­è¯å­—å·
                new_voucher.voucher_number = _generate_voucher_number(new_voucher.voucher_date)
                new_voucher.save()
                
                # å¤åˆ¶å‡­è¯åˆ†å½•
                for entry in source_voucher.entries.all():
                    VoucherEntry.objects.create(
                        voucher=new_voucher,
                        line_number=entry.line_number,
                        account_subject=entry.account_subject,
                        summary=entry.summary,
                        debit_amount=entry.debit_amount,
                        credit_amount=entry.credit_amount,
                    )
                
                messages.success(request, f'å‡­è¯å·²å¤åˆ¶ï¼Œæ–°å‡­è¯å·ï¼š{new_voucher.voucher_number}')
                return redirect('finance_pages:voucher_detail', voucher_id=new_voucher.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('å¤åˆ¶å‡­è¯å¤±è´¥: %s', str(e))
            messages.error(request, f'å¤åˆ¶å‡­è¯å¤±è´¥ï¼š{str(e)}')
            return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"å¤åˆ¶å‡­è¯ - {source_voucher.voucher_number}",
        "ğŸ“‹",
        f"å¤åˆ¶å‡­è¯ {source_voucher.voucher_number} åˆ›å»ºæ–°å‡­è¯",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'source_voucher': source_voucher,
        'entries': source_voucher.entries.select_related('account_subject').order_by('line_number'),
    })
    return render(request, "financial_management/voucher_copy.html", context)


@login_required
def voucher_print(request, voucher_id):
    """æ‰“å°å‡­è¯ï¼ˆPDFæ ¼å¼ï¼‰"""
    if not REPORTLAB_AVAILABLE:
        messages.error(request, 'PDFæ‰“å°åŠŸèƒ½éœ€è¦å®‰è£…reportlabåº“')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher.objects.select_related('preparer', 'reviewer', 'posted_by'), id=voucher_id)
    
    # è·å–å‡­è¯åˆ†å½•
    try:
        entries = voucher.entries.select_related('account_subject').order_by('line_number')
    except Exception:
        entries = []
    
    # åˆ›å»ºPDFå“åº”
    response = HttpResponse(content_type='application/pdf')
    filename = f'å‡­è¯_{voucher.voucher_number}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # åˆ›å»ºPDFæ–‡æ¡£
    doc = SimpleDocTemplate(response, pagesize=A4, 
                           rightMargin=20*mm, leftMargin=20*mm,
                           topMargin=20*mm, bottomMargin=20*mm)
    
    # æ„å»ºå†…å®¹
    story = []
    styles = getSampleStyleSheet()
    
    # æ ‡é¢˜æ ·å¼
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=12,
        alignment=1,  # å±…ä¸­
    )
    
    # æ·»åŠ æ ‡é¢˜
    story.append(Paragraph('è®°è´¦å‡­è¯', title_style))
    story.append(Spacer(1, 6*mm))
    
    # å‡­è¯åŸºæœ¬ä¿¡æ¯è¡¨æ ¼
    voucher_info_data = [
        ['å‡­è¯å­—å·', voucher.voucher_number, 'å‡­è¯æ—¥æœŸ', voucher.voucher_date.strftime('%Y-%m-%d') if voucher.voucher_date else ''],
        ['åˆ¶å•äºº', voucher.preparer.get_full_name() if voucher.preparer else '', 'é™„ä»¶æ•°', str(voucher.attachment_count)],
    ]
    
    if voucher.reviewer:
        voucher_info_data.append(['å®¡æ ¸äºº', voucher.reviewer.get_full_name(), 'å®¡æ ¸æ—¶é—´', voucher.reviewed_time.strftime('%Y-%m-%d %H:%M') if voucher.reviewed_time else ''])
    
    if voucher.posted_by:
        voucher_info_data.append(['è¿‡è´¦äºº', voucher.posted_by.get_full_name(), 'è¿‡è´¦æ—¶é—´', voucher.posted_time.strftime('%Y-%m-%d %H:%M') if voucher.posted_time else ''])
    
    voucher_info_table = Table(voucher_info_data, colWidths=[40*mm, 50*mm, 40*mm, 50*mm])
    voucher_info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(voucher_info_table)
    story.append(Spacer(1, 6*mm))
    
    # å‡­è¯åˆ†å½•è¡¨æ ¼
    entry_headers = ['è¡Œå·', 'ä¼šè®¡ç§‘ç›®', 'æ‘˜è¦', 'å€Ÿæ–¹é‡‘é¢', 'è´·æ–¹é‡‘é¢']
    entry_data = [entry_headers]
    
    for entry in entries:
        entry_data.append([
            str(entry.line_number),
            f"{entry.account_subject.code} {entry.account_subject.name}" if entry.account_subject else '',
            entry.summary,
            f"{entry.debit_amount:,.2f}" if entry.debit_amount else '0.00',
            f"{entry.credit_amount:,.2f}" if entry.credit_amount else '0.00',
        ])
    
    # æ·»åŠ åˆè®¡è¡Œ
    entry_data.append([
        '',
        '',
        'åˆè®¡',
        f"{voucher.total_debit:,.2f}",
        f"{voucher.total_credit:,.2f}",
    ])
    
    entry_table = Table(entry_data, colWidths=[15*mm, 50*mm, 60*mm, 30*mm, 30*mm])
    entry_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('FONTNAME', (0, 0), (-1, 0), 'SimSun'),
        ('BOLD', (0, 0), (-1, 0), True),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ('BOLD', (0, -1), (-1, -1), True),
    ]))
    story.append(entry_table)
    story.append(Spacer(1, 6*mm))
    
    # å¤‡æ³¨ä¿¡æ¯
    if voucher.notes:
        story.append(Paragraph(f'<b>å¤‡æ³¨ï¼š</b>{voucher.notes}', styles['Normal']))
        story.append(Spacer(1, 6*mm))
    
    # çŠ¶æ€ä¿¡æ¯
    status_dict = dict(Voucher.STATUS_CHOICES)
    status_text = status_dict.get(voucher.status, voucher.status)
    story.append(Paragraph(f'<b>çŠ¶æ€ï¼š</b>{status_text}', styles['Normal']))
    
    # ç”ŸæˆPDF
    doc.build(story)
    return response


@login_required
def voucher_submit(request, voucher_id):
    """æäº¤å‡­è¯ï¼ˆè‰ç¨¿->å·²æäº¤ï¼‰"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æäº¤å‡­è¯')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    
    # æ£€æŸ¥çŠ¶æ€
    if voucher.status != 'draft':
        messages.error(request, f'åªæœ‰è‰ç¨¿çŠ¶æ€çš„å‡­è¯æ‰èƒ½æäº¤ï¼Œå½“å‰çŠ¶æ€ï¼š{voucher.get_status_display()}')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # æ£€æŸ¥å€Ÿè´·æ˜¯å¦å¹³è¡¡
    if voucher.total_debit != voucher.total_credit:
        messages.error(request, f'å‡­è¯å€Ÿè´·ä¸å¹³è¡¡ï¼ˆå€Ÿæ–¹ï¼š{voucher.total_debit}ï¼Œè´·æ–¹ï¼š{voucher.total_credit}ï¼‰ï¼Œä¸èƒ½æäº¤')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # æ£€æŸ¥æ˜¯å¦æœ‰åˆ†å½•
    if not voucher.entries.exists():
        messages.error(request, 'å‡­è¯è‡³å°‘éœ€è¦ä¸€æ¡åˆ†å½•æ‰èƒ½æäº¤')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        voucher.status = 'submitted'
        voucher.save()
        messages.success(request, f'å‡­è¯ {voucher.voucher_number} å·²æäº¤ï¼Œç­‰å¾…å®¡æ ¸')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"æäº¤å‡­è¯ - {voucher.voucher_number}",
        "ğŸ“¤",
        f"æäº¤è®°è´¦å‡­è¯ {voucher.voucher_number} ç­‰å¾…å®¡æ ¸",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
        'entries': voucher.entries.select_related('account_subject').order_by('line_number'),
    })
    return render(request, "financial_management/voucher_submit.html", context)


@login_required
def voucher_withdraw(request, voucher_id):
    """æ’¤å›å‡­è¯ï¼ˆå·²æäº¤->è‰ç¨¿ï¼‰"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æ’¤å›å‡­è¯')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    
    # æ£€æŸ¥çŠ¶æ€
    if voucher.status != 'submitted':
        messages.error(request, f'åªæœ‰å·²æäº¤çŠ¶æ€çš„å‡­è¯æ‰èƒ½æ’¤å›ï¼Œå½“å‰çŠ¶æ€ï¼š{voucher.get_status_display()}')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # æ£€æŸ¥æƒé™ï¼šåªæœ‰åˆ¶å•äººæ‰èƒ½æ’¤å›
    if voucher.preparer != request.user:
        messages.error(request, 'åªæœ‰åˆ¶å•äººæ‰èƒ½æ’¤å›å‡­è¯')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        voucher.status = 'draft'
        voucher.save()
        messages.success(request, f'å‡­è¯ {voucher.voucher_number} å·²æ’¤å›ï¼Œå¯ä»¥ç»§ç»­ç¼–è¾‘')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"æ’¤å›å‡­è¯ - {voucher.voucher_number}",
        "â†©ï¸",
        f"æ’¤å›è®°è´¦å‡­è¯ {voucher.voucher_number} è¿”å›è‰ç¨¿çŠ¶æ€",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
    })
    return render(request, "financial_management/voucher_withdraw.html", context)


@login_required
def voucher_approve(request, voucher_id):
    """å®¡æ ¸å‡­è¯"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.review', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å®¡æ ¸å‡­è¯')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    
    # æ£€æŸ¥çŠ¶æ€
    if voucher.status not in ['submitted', 'draft']:
        messages.error(request, f'å‡­è¯çŠ¶æ€ä¸º {voucher.get_status_display()}ï¼Œä¸èƒ½å®¡æ ¸')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # æ£€æŸ¥å€Ÿè´·æ˜¯å¦å¹³è¡¡
    if voucher.total_debit != voucher.total_credit:
        messages.error(request, f'å‡­è¯å€Ÿè´·ä¸å¹³è¡¡ï¼ˆå€Ÿæ–¹ï¼š{voucher.total_debit}ï¼Œè´·æ–¹ï¼š{voucher.total_credit}ï¼‰ï¼Œä¸èƒ½å®¡æ ¸')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        action = request.POST.get('action', 'approve')
        
        if action == 'approve':
            # å®¡æ ¸é€šè¿‡
            voucher.status = 'approved'
            voucher.reviewer = request.user
            voucher.reviewed_time = timezone.now()
            voucher.save()
            messages.success(request, f'å‡­è¯ {voucher.voucher_number} å®¡æ ¸é€šè¿‡')
        elif action == 'reject':
            # å®¡æ ¸æ‹’ç»
            voucher.status = 'rejected'
            voucher.reviewer = request.user
            voucher.reviewed_time = timezone.now()
            voucher.save()
            messages.success(request, f'å‡­è¯ {voucher.voucher_number} å·²æ‹’ç»')
        
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"å®¡æ ¸å‡­è¯ - {voucher.voucher_number}",
        "âœ…",
        f"å®¡æ ¸è®°è´¦å‡­è¯ {voucher.voucher_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
    })
    return render(request, "financial_management/voucher_approve.html", context)


@login_required
def voucher_post(request, voucher_id):
    """è¿‡è´¦å‡­è¯"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.post', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™è¿‡è´¦å‡­è¯')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher.objects.prefetch_related('entries'), id=voucher_id)
    
    # æ£€æŸ¥çŠ¶æ€
    if voucher.status != 'approved':
        messages.error(request, f'åªæœ‰å·²å®¡æ ¸çš„å‡­è¯æ‰èƒ½è¿‡è´¦ï¼Œå½“å‰çŠ¶æ€ï¼š{voucher.get_status_display()}')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # æ£€æŸ¥å€Ÿè´·æ˜¯å¦å¹³è¡¡
    if voucher.total_debit != voucher.total_credit:
        messages.error(request, f'å‡­è¯å€Ÿè´·ä¸å¹³è¡¡ï¼ˆå€Ÿæ–¹ï¼š{voucher.total_debit}ï¼Œè´·æ–¹ï¼š{voucher.total_credit}ï¼‰ï¼Œä¸èƒ½è¿‡è´¦')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        from django.db import transaction
        
        try:
            with transaction.atomic():
                # æ›´æ–°å‡­è¯çŠ¶æ€
                voucher.status = 'posted'
                voucher.posted_by = request.user
                voucher.posted_time = timezone.now()
                voucher.save()
                
                # ç”Ÿæˆæ€»è´¦è®°å½•
                period_year = voucher.voucher_date.year
                period_month = voucher.voucher_date.month
                
                for entry in voucher.entries.all():
                    # è·å–æˆ–åˆ›å»ºæ€»è´¦è®°å½•
                    ledger, created = Ledger.objects.get_or_create(
                        account_subject=entry.account_subject,
                        period_year=period_year,
                        period_month=period_month,
                        period_date=voucher.voucher_date,
                        defaults={
                            'opening_balance': Decimal('0.00'),
                            'period_debit': Decimal('0.00'),
                            'period_credit': Decimal('0.00'),
                            'closing_balance': Decimal('0.00'),
                        }
                    )
                    
                    # æ›´æ–°æ€»è´¦é‡‘é¢
                    ledger.period_debit += entry.debit_amount or Decimal('0.00')
                    ledger.period_credit += entry.credit_amount or Decimal('0.00')
                    
                    # è®¡ç®—æœŸæœ«ä½™é¢ï¼ˆæ ¹æ®ç§‘ç›®ä½™é¢æ–¹å‘ï¼‰
                    if entry.account_subject.direction == 'debit':
                        # å€Ÿæ–¹ç§‘ç›®ï¼šæœŸåˆä½™é¢ + å€Ÿæ–¹ - è´·æ–¹
                        ledger.closing_balance = ledger.opening_balance + ledger.period_debit - ledger.period_credit
                    else:
                        # è´·æ–¹ç§‘ç›®ï¼šæœŸåˆä½™é¢ + è´·æ–¹ - å€Ÿæ–¹
                        ledger.closing_balance = ledger.opening_balance + ledger.period_credit - ledger.period_debit
                    
                    ledger.save()
                
                messages.success(request, f'å‡­è¯ {voucher.voucher_number} è¿‡è´¦æˆåŠŸï¼Œå·²ç”Ÿæˆæ€»è´¦è®°å½•')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('è¿‡è´¦å¤±è´¥: %s', str(e))
            messages.error(request, f'è¿‡è´¦å¤±è´¥ï¼š{str(e)}')
        
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"è¿‡è´¦å‡­è¯ - {voucher.voucher_number}",
        "ğŸ“–",
        f"å°†å‡­è¯ {voucher.voucher_number} è¿‡è´¦åˆ°æ€»è´¦",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
        'entries': voucher.entries.select_related('account_subject').order_by('line_number'),
    })
    return render(request, "financial_management/voucher_post.html", context)


@login_required
def voucher_unpost(request, voucher_id):
    """åè¿‡è´¦å‡­è¯ï¼ˆå·²è¿‡è´¦->å·²å®¡æ ¸ï¼‰"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.post', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åè¿‡è´¦å‡­è¯')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher.objects.prefetch_related('entries'), id=voucher_id)
    
    # æ£€æŸ¥çŠ¶æ€
    if voucher.status != 'posted':
        messages.error(request, f'åªæœ‰å·²è¿‡è´¦çš„å‡­è¯æ‰èƒ½åè¿‡è´¦ï¼Œå½“å‰çŠ¶æ€ï¼š{voucher.get_status_display()}')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        from django.db import transaction
        
        try:
            with transaction.atomic():
                # åˆ é™¤æ€»è´¦è®°å½•ï¼ˆå›æ»šè¿‡è´¦æ“ä½œï¼‰
                period_year = voucher.voucher_date.year
                period_month = voucher.voucher_date.month
                
                for entry in voucher.entries.all():
                    try:
                        ledger = Ledger.objects.get(
                            account_subject=entry.account_subject,
                            period_year=period_year,
                            period_month=period_month,
                            period_date=voucher.voucher_date,
                        )
                        
                        # å›æ»šæ€»è´¦é‡‘é¢
                        ledger.period_debit -= entry.debit_amount or Decimal('0.00')
                        ledger.period_credit -= entry.credit_amount or Decimal('0.00')
                        
                        # é‡æ–°è®¡ç®—æœŸæœ«ä½™é¢
                        if entry.account_subject.direction == 'debit':
                            ledger.closing_balance = ledger.opening_balance + ledger.period_debit - ledger.period_credit
                        else:
                            ledger.closing_balance = ledger.opening_balance + ledger.period_credit - ledger.period_debit
                        
                        # å¦‚æœæœ¬æœŸå€Ÿè´·éƒ½ä¸º0ï¼Œå¯ä»¥åˆ é™¤è¯¥æ€»è´¦è®°å½•
                        if ledger.period_debit == Decimal('0.00') and ledger.period_credit == Decimal('0.00'):
                            ledger.delete()
                        else:
                            ledger.save()
                    except Ledger.DoesNotExist:
                        # æ€»è´¦è®°å½•ä¸å­˜åœ¨ï¼Œè·³è¿‡
                        pass
                
                # æ›´æ–°å‡­è¯çŠ¶æ€
                voucher.status = 'approved'
                voucher.posted_by = None
                voucher.posted_time = None
                voucher.save()
                
                messages.success(request, f'å‡­è¯ {voucher.voucher_number} åè¿‡è´¦æˆåŠŸï¼Œå·²å›æ»šæ€»è´¦è®°å½•')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('åè¿‡è´¦å¤±è´¥: %s', str(e))
            messages.error(request, f'åè¿‡è´¦å¤±è´¥ï¼š{str(e)}')
        
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"åè¿‡è´¦å‡­è¯ - {voucher.voucher_number}",
        "â†©ï¸",
        f"å°†å‡­è¯ {voucher.voucher_number} åè¿‡è´¦ï¼Œå›æ»šæ€»è´¦è®°å½•",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
        'entries': voucher.entries.select_related('account_subject').order_by('line_number'),
    })
    return render(request, "financial_management/voucher_unpost.html", context)


@login_required
def voucher_delete(request, voucher_id):
    """åˆ é™¤å‡­è¯"""
    voucher = get_object_or_404(Voucher, id=voucher_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ é™¤å‡­è¯')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    # åªæœ‰è‰ç¨¿çŠ¶æ€çš„å‡­è¯å¯ä»¥åˆ é™¤
    if voucher.status != 'draft':
        messages.error(request, f'åªèƒ½åˆ é™¤è‰ç¨¿çŠ¶æ€çš„å‡­è¯ï¼Œå½“å‰çŠ¶æ€ï¼š{voucher.get_status_display()}')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    if request.method == 'POST':
        try:
            voucher_number = voucher.voucher_number
            voucher.delete()
            messages.success(request, f'å‡­è¯ {voucher_number} å·²åˆ é™¤')
            return redirect('finance_pages:voucher_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('åˆ é™¤å‡­è¯å¤±è´¥: %s', str(e))
            messages.error(request, f'åˆ é™¤å‡­è¯å¤±è´¥ï¼š{str(e)}')
            return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    context = _context(
        f"åˆ é™¤å‡­è¯ - {voucher.voucher_number}",
        "ğŸ—‘ï¸",
        f"ç¡®è®¤åˆ é™¤å‡­è¯ï¼š{voucher.voucher_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
    })
    return render(request, "financial_management/voucher_delete.html", context)


@login_required
def budget_detail(request, budget_id):
    """é¢„ç®—è¯¦æƒ…"""
    budget = get_object_or_404(Budget.objects.select_related('department', 'account_subject', 'approver', 'created_by'), id=budget_id)
    
    # è®¡ç®—ä½¿ç”¨ç‡
    usage_rate = 0
    if budget.budget_amount > 0:
        usage_rate = (budget.used_amount / budget.budget_amount) * 100
    
    # æ£€æŸ¥æƒé™
    permission_codes = get_user_permission_codes(request.user)
    can_approve = _permission_granted('financial_management.budget.approve', permission_codes)
    
    context = _context(
        f"é¢„ç®—è¯¦æƒ… - {budget.budget_number}",
        "ğŸ’°",
        f"æŸ¥çœ‹é¢„ç®— {budget.name} çš„è¯¦ç»†ä¿¡æ¯",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'budget': budget,
        'usage_rate': usage_rate,
        'can_approve': can_approve,
    })
    return render(request, "financial_management/budget_detail.html", context)


@login_required
def budget_delete(request, budget_id):
    """åˆ é™¤é¢„ç®—"""
    budget = get_object_or_404(Budget, id=budget_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.budget.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ é™¤é¢„ç®—')
        return redirect('finance_pages:budget_detail', budget_id=budget_id)
    
    # åªæœ‰è‰ç¨¿çŠ¶æ€çš„é¢„ç®—å¯ä»¥åˆ é™¤
    if budget.status != 'draft':
        messages.error(request, f'åªèƒ½åˆ é™¤è‰ç¨¿çŠ¶æ€çš„é¢„ç®—ï¼Œå½“å‰çŠ¶æ€ï¼š{budget.get_status_display()}')
        return redirect('finance_pages:budget_detail', budget_id=budget_id)
    
    if request.method == 'POST':
        try:
            budget_number = budget.budget_number
            budget.delete()
            messages.success(request, f'é¢„ç®— {budget_number} å·²åˆ é™¤')
            return redirect('finance_pages:budget_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('åˆ é™¤é¢„ç®—å¤±è´¥: %s', str(e))
            messages.error(request, f'åˆ é™¤é¢„ç®—å¤±è´¥ï¼š{str(e)}')
            return redirect('finance_pages:budget_detail', budget_id=budget_id)
    
    context = _context(
        f"åˆ é™¤é¢„ç®— - {budget.budget_number}",
        "ğŸ—‘ï¸",
        f"ç¡®è®¤åˆ é™¤é¢„ç®—ï¼š{budget.name}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'budget': budget,
    })
    return render(request, "financial_management/budget_delete.html", context)


@login_required
def budget_approve(request, budget_id):
    """å®¡æ‰¹é¢„ç®—"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.budget.approve', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å®¡æ‰¹é¢„ç®—')
        return redirect('finance_pages:budget_detail', budget_id=budget_id)
    
    budget = get_object_or_404(Budget.objects.select_related('department', 'account_subject', 'created_by'), id=budget_id)
    
    # æ£€æŸ¥çŠ¶æ€
    if budget.status != 'draft':
        messages.error(request, f'é¢„ç®—çŠ¶æ€ä¸º {budget.get_status_display()}ï¼Œä¸èƒ½å®¡æ‰¹')
        return redirect('finance_pages:budget_detail', budget_id=budget.id)
    
    if request.method == 'POST':
        action = request.POST.get('action', 'approve')
        
        if action == 'approve':
            # å®¡æ‰¹é€šè¿‡
            budget.status = 'approved'
            budget.approver = request.user
            budget.approved_time = timezone.now()
            # è‡ªåŠ¨è®¡ç®—å‰©ä½™é‡‘é¢
            budget.remaining_amount = budget.budget_amount - budget.used_amount
            budget.save()
            messages.success(request, f'é¢„ç®— {budget.budget_number} å®¡æ‰¹é€šè¿‡')
        elif action == 'reject':
            # å®¡æ‰¹æ‹’ç»ï¼ˆå–æ¶ˆï¼‰
            budget.status = 'cancelled'
            budget.approver = request.user
            budget.approved_time = timezone.now()
            budget.save()
            messages.success(request, f'é¢„ç®— {budget.budget_number} å·²å–æ¶ˆ')
        
        return redirect('finance_pages:budget_detail', budget_id=budget.id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"å®¡æ‰¹é¢„ç®— - {budget.budget_number}",
        "âœ…",
        f"å®¡æ‰¹é¢„ç®— {budget.name}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'budget': budget,
    })
    return render(request, "financial_management/budget_approve.html", context)


@login_required
def invoice_detail(request, invoice_id):
    """å‘ç¥¨è¯¦æƒ…"""
    invoice = get_object_or_404(Invoice.objects.select_related('verified_by', 'created_by'), id=invoice_id)
    
    # æ£€æŸ¥æƒé™
    permission_codes = get_user_permission_codes(request.user)
    can_verify = _permission_granted('financial_management.invoice.manage', permission_codes) and invoice.status == 'issued'
    can_edit = _permission_granted('financial_management.invoice.manage', permission_codes) and invoice.status != 'verified'
    
    context = _context(
        f"å‘ç¥¨è¯¦æƒ… - {invoice.invoice_number}",
        "ğŸ§¾",
        f"æŸ¥çœ‹å‘ç¥¨ {invoice.invoice_number} çš„è¯¦ç»†ä¿¡æ¯",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'invoice': invoice,
        'can_verify': can_verify,
        'can_edit': can_edit,
    })
    return render(request, "financial_management/invoice_detail.html", context)


@login_required
def invoice_verify(request, invoice_id):
    """è®¤è¯å‘ç¥¨"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.invoice.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™è®¤è¯å‘ç¥¨')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice_id)
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    # æ£€æŸ¥çŠ¶æ€
    if invoice.status != 'issued':
        messages.error(request, f'åªæœ‰å·²å¼€å…·çŠ¶æ€çš„å‘ç¥¨æ‰èƒ½è®¤è¯ï¼Œå½“å‰çŠ¶æ€ï¼š{invoice.get_status_display()}')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    
    if request.method == 'POST':
        invoice.status = 'verified'
        invoice.verified_by = request.user
        invoice.verified_time = timezone.now()
        invoice.save()
        messages.success(request, f'å‘ç¥¨ {invoice.invoice_number} è®¤è¯æˆåŠŸ')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"è®¤è¯å‘ç¥¨ - {invoice.invoice_number}",
        "âœ…",
        f"è®¤è¯å‘ç¥¨ {invoice.invoice_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'invoice': invoice,
    })
    return render(request, "financial_management/invoice_verify.html", context)


@login_required
def invoice_cancel(request, invoice_id):
    """ä½œåºŸå‘ç¥¨"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.invoice.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™ä½œåºŸå‘ç¥¨')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice_id)
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    # æ£€æŸ¥çŠ¶æ€ï¼šå·²è®¤è¯çš„å‘ç¥¨ä¸èƒ½ä½œåºŸ
    if invoice.status == 'verified':
        messages.error(request, 'å·²è®¤è¯çš„å‘ç¥¨ä¸èƒ½ä½œåºŸ')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    
    if request.method == 'POST':
        invoice.status = 'cancelled'
        invoice.save()
        messages.success(request, f'å‘ç¥¨ {invoice.invoice_number} å·²ä½œåºŸ')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"ä½œåºŸå‘ç¥¨ - {invoice.invoice_number}",
        "âŒ",
        f"ä½œåºŸå‘ç¥¨ {invoice.invoice_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'invoice': invoice,
    })
    return render(request, "financial_management/invoice_cancel.html", context)


@login_required
def invoice_delete(request, invoice_id):
    """åˆ é™¤å‘ç¥¨"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.invoice.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ é™¤å‘ç¥¨')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice_id)
    
    if request.method == 'POST':
        try:
            invoice_number = invoice.invoice_number
            invoice.delete()
            messages.success(request, f'å‘ç¥¨ {invoice_number} å·²åˆ é™¤')
            return redirect('finance_pages:invoice_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('åˆ é™¤å‘ç¥¨å¤±è´¥: %s', str(e))
            messages.error(request, f'åˆ é™¤å‘ç¥¨å¤±è´¥ï¼š{str(e)}')
            return redirect('finance_pages:invoice_detail', invoice_id=invoice_id)
    
    context = _context(
        f"åˆ é™¤å‘ç¥¨ - {invoice.invoice_number}",
        "ğŸ—‘ï¸",
        f"ç¡®è®¤åˆ é™¤å‘ç¥¨ï¼š{invoice.invoice_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'invoice': invoice,
    })
    return render(request, "financial_management/invoice_delete.html", context)


@login_required
def account_subject_detail(request, account_subject_id):
    """ä¼šè®¡ç§‘ç›®è¯¦æƒ…"""
    account_subject = get_object_or_404(
        AccountSubject.objects.select_related('parent', 'created_by'),
        id=account_subject_id
    )
    
    # è·å–å­ç§‘ç›®
    try:
        children = AccountSubject.objects.filter(parent=account_subject).order_by('code')
    except Exception:
        children = []
    
    # è·å–ä½¿ç”¨ç»Ÿè®¡
    try:
        voucher_entry_count = account_subject.voucher_entries.count()
        ledger_entry_count = account_subject.ledger_entries.count()
    except Exception:
        voucher_entry_count = 0
        ledger_entry_count = 0
    
    context = _context(
        f"ä¼šè®¡ç§‘ç›®è¯¦æƒ… - {account_subject.code} {account_subject.name}",
        "ğŸ“Š",
        f"æŸ¥çœ‹ä¼šè®¡ç§‘ç›® {account_subject.code} {account_subject.name} çš„è¯¦ç»†ä¿¡æ¯",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'account_subject': account_subject,
        'children': children,
        'voucher_entry_count': voucher_entry_count,
        'ledger_entry_count': ledger_entry_count,
    })
    return render(request, "financial_management/account_subject_detail.html", context)


@login_required
def ledger_detail(request, ledger_id):
    """è´¦ç°¿è¯¦æƒ…"""
    ledger = get_object_or_404(
        Ledger.objects.select_related('account_subject'),
        id=ledger_id
    )
    
    # è·å–åŒä¸€ç§‘ç›®çš„å…¶ä»–æœŸé—´è®°å½•ï¼ˆæœ€è¿‘6ä¸ªæœˆï¼‰
    try:
        related_ledgers = Ledger.objects.filter(
            account_subject=ledger.account_subject
        ).exclude(id=ledger.id).order_by('-period_date')[:6]
    except Exception:
        related_ledgers = []
    
    context = _context(
        f"è´¦ç°¿è¯¦æƒ… - {ledger.account_subject.code} {ledger.period_date}",
        "ğŸ“–",
        f"æŸ¥çœ‹ä¼šè®¡ç§‘ç›® {ledger.account_subject.code} åœ¨ {ledger.period_date} çš„è´¦åŠ¡è®°å½•",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'ledger': ledger,
        'related_ledgers': related_ledgers,
    })
    return render(request, "financial_management/ledger_detail.html", context)


@login_required
def ledger_opening_balance_setup(request):
    """è®¾ç½®æœŸåˆä½™é¢"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.ledger.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™è®¾ç½®æœŸåˆä½™é¢')
        return redirect('finance_pages:ledger_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', '1')
    
    try:
        period_year = int(period_year) if period_year else today.year
        period_month = int(period_month) if period_month else 1
    except (ValueError, TypeError):
        period_year = today.year
        period_month = 1
    
    # è·å–æ‰€æœ‰å¯ç”¨çš„ä¼šè®¡ç§‘ç›®
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    
    if request.method == 'POST':
        from django.db import transaction
        try:
            with transaction.atomic():
                success_count = 0
                for subject in account_subjects:
                    balance_key = f'balance_{subject.id}'
                    balance_str = request.POST.get(balance_key, '0')
                    
                    try:
                        balance = Decimal(balance_str) if balance_str else Decimal('0.00')
                        
                        # è·å–æˆ–åˆ›å»ºè¯¥ç§‘ç›®çš„æœŸåˆä½™é¢è®°å½•ï¼ˆä½¿ç”¨è¯¥æœŸé—´çš„ç¬¬ä¸€å¤©ï¼‰
                        period_date = timezone.datetime(period_year, period_month, 1).date()
                        
                        ledger, created = Ledger.objects.get_or_create(
                            account_subject=subject,
                            period_year=period_year,
                            period_month=period_month,
                            period_date=period_date,
                            defaults={
                                'opening_balance': balance,
                                'period_debit': Decimal('0.00'),
                                'period_credit': Decimal('0.00'),
                                'closing_balance': balance,
                            }
                        )
                        
                        if not created:
                            # æ›´æ–°æœŸåˆä½™é¢
                            ledger.opening_balance = balance
                            # é‡æ–°è®¡ç®—æœŸæœ«ä½™é¢
                            if subject.direction == 'debit':
                                ledger.closing_balance = ledger.opening_balance + ledger.period_debit - ledger.period_credit
                            else:
                                ledger.closing_balance = ledger.opening_balance + ledger.period_credit - ledger.period_debit
                            ledger.save()
                        
                        success_count += 1
                    except (ValueError, InvalidOperation):
                        continue
                
                messages.success(request, f'æˆåŠŸè®¾ç½® {success_count} ä¸ªç§‘ç›®çš„æœŸåˆä½™é¢')
                return redirect(f"{reverse('finance_pages:ledger_opening_balance_setup')}?period_year={period_year}&period_month={period_month}")
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('è®¾ç½®æœŸåˆä½™é¢å¤±è´¥: %s', str(e))
            messages.error(request, f'è®¾ç½®æœŸåˆä½™é¢å¤±è´¥ï¼š{str(e)}')
    
    # è·å–å½“å‰æœŸé—´çš„æœŸåˆä½™é¢
    period_date = timezone.datetime(period_year, period_month, 1).date()
    opening_balances = {}
    for subject in account_subjects:
        try:
            ledger = Ledger.objects.filter(
                account_subject=subject,
                period_year=period_year,
                period_month=period_month
            ).order_by('period_date').first()
            if ledger:
                opening_balances[subject.id] = ledger.opening_balance
            else:
                opening_balances[subject.id] = Decimal('0.00')
        except Exception:
            opening_balances[subject.id] = Decimal('0.00')
    
    context = _context(
        f"è®¾ç½®æœŸåˆä½™é¢ - {period_year}å¹´{period_month}æœˆ",
        "ğŸ’°",
        f"ä¸º {period_year}å¹´{period_month}æœˆ è®¾ç½®å„ç§‘ç›®çš„æœŸåˆä½™é¢",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'account_subjects': account_subjects,
        'opening_balances': opening_balances,
        'period_year': period_year,
        'period_month': period_month,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/ledger_opening_balance_setup.html", context)


@login_required
def ledger_period_closing(request):
    """æœŸæœ«ç»“è´¦"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.ledger.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™è¿›è¡ŒæœŸæœ«ç»“è´¦')
        return redirect('finance_pages:ledger_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year) if period_year else today.year
        period_month = int(period_month) if period_month else today.month
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # è®¡ç®—ä¸‹ä¸€æœŸé—´
    if period_month == 12:
        next_period_year = period_year + 1
        next_period_month = 1
    else:
        next_period_year = period_year
        next_period_month = period_month + 1
    
    # æ£€æŸ¥å½“å‰æœŸé—´æ˜¯å¦æœ‰æœªè¿‡è´¦çš„å‡­è¯
    unposted_vouchers = Voucher.objects.filter(
        voucher_date__year=period_year,
        voucher_date__month=period_month,
        status__in=['draft', 'submitted', 'approved']
    ).count()
    
    if request.method == 'POST':
        if unposted_vouchers > 0:
            messages.error(request, f'å½“å‰æœŸé—´è¿˜æœ‰ {unposted_vouchers} å¼ å‡­è¯æœªè¿‡è´¦ï¼Œä¸èƒ½ç»“è´¦')
            return redirect(f"{reverse('finance_pages:ledger_period_closing')}?period_year={period_year}&period_month={period_month}")
        
        from django.db import transaction
        try:
            with transaction.atomic():
                # è·å–å½“å‰æœŸé—´çš„æ‰€æœ‰æ€»è´¦è®°å½•
                current_ledgers = Ledger.objects.filter(
                    period_year=period_year,
                    period_month=period_month
                ).select_related('account_subject')
                
                success_count = 0
                for ledger in current_ledgers:
                    # è·å–ä¸‹ä¸€æœŸé—´çš„æœŸåˆä½™é¢è®°å½•ï¼ˆä½¿ç”¨ä¸‹ä¸€æœŸé—´çš„ç¬¬ä¸€å¤©ï¼‰
                    next_period_date = timezone.datetime(next_period_year, next_period_month, 1).date()
                    
                    next_ledger, created = Ledger.objects.get_or_create(
                        account_subject=ledger.account_subject,
                        period_year=next_period_year,
                        period_month=next_period_month,
                        period_date=next_period_date,
                        defaults={
                            'opening_balance': ledger.closing_balance,
                            'period_debit': Decimal('0.00'),
                            'period_credit': Decimal('0.00'),
                            'closing_balance': ledger.closing_balance,
                        }
                    )
                    
                    if not created:
                        # æ›´æ–°æœŸåˆä½™é¢
                        next_ledger.opening_balance = ledger.closing_balance
                        # é‡æ–°è®¡ç®—æœŸæœ«ä½™é¢
                        if ledger.account_subject.direction == 'debit':
                            next_ledger.closing_balance = next_ledger.opening_balance + next_ledger.period_debit - next_ledger.period_credit
                        else:
                            next_ledger.closing_balance = next_ledger.opening_balance + next_ledger.period_credit - next_ledger.period_debit
                        next_ledger.save()
                    
                    success_count += 1
                
                messages.success(request, f'æˆåŠŸç»“è´¦ï¼š{period_year}å¹´{period_month}æœˆ â†’ {next_period_year}å¹´{next_period_month}æœˆï¼Œå…± {success_count} ä¸ªç§‘ç›®')
                return redirect('finance_pages:ledger_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('æœŸæœ«ç»“è´¦å¤±è´¥: %s', str(e))
            messages.error(request, f'æœŸæœ«ç»“è´¦å¤±è´¥ï¼š{str(e)}')
    
    # è·å–å½“å‰æœŸé—´çš„ç§‘ç›®ç»Ÿè®¡
    current_ledgers = Ledger.objects.filter(
        period_year=period_year,
        period_month=period_month
    ).select_related('account_subject')
    
    context = _context(
        f"æœŸæœ«ç»“è´¦ - {period_year}å¹´{period_month}æœˆ",
        "ğŸ“‹",
        f"å°† {period_year}å¹´{period_month}æœˆ çš„æœŸæœ«ä½™é¢ç»“è½¬åˆ° {next_period_year}å¹´{next_period_month}æœˆ",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'period_year': period_year,
        'period_month': period_month,
        'next_period_year': next_period_year,
        'next_period_month': next_period_month,
        'unposted_vouchers': unposted_vouchers,
        'ledger_count': current_ledgers.count(),
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/ledger_period_closing.html", context)


@login_required
def subsidiary_ledger(request):
    """æ˜ç»†è´¦"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.ledger.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æŸ¥çœ‹æ˜ç»†è´¦')
        return redirect('finance_pages:ledger_management')
    
    today = timezone.now().date()
    account_subject_id = request.GET.get('account_subject_id', '')
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    try:
        period_year = int(period_year) if period_year else today.year
        period_month = int(period_month) if period_month else today.month
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # è·å–æ‰€æœ‰ä¼šè®¡ç§‘ç›®
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    
    entries = []
    account_subject = None
    opening_balance = Decimal('0.00')
    closing_balance = Decimal('0.00')
    
    if account_subject_id:
        try:
            account_subject = AccountSubject.objects.get(id=int(account_subject_id))
            
            # è·å–æœŸåˆä½™é¢ï¼ˆä¸ŠæœŸæœŸæœ«ä½™é¢ï¼‰
            if period_month > 1:
                prev_month = period_month - 1
                prev_ledger = Ledger.objects.filter(
                    account_subject=account_subject,
                    period_year=period_year,
                    period_month=prev_month
                ).order_by('-period_date').first()
                if prev_ledger:
                    opening_balance = prev_ledger.closing_balance
            else:
                prev_year = period_year - 1
                prev_ledger = Ledger.objects.filter(
                    account_subject=account_subject,
                    period_year=prev_year,
                    period_month=12
                ).order_by('-period_date').first()
                if prev_ledger:
                    opening_balance = prev_ledger.closing_balance
            
            # è·å–å‡­è¯åˆ†å½•
            voucher_entries_query = VoucherEntry.objects.filter(
                account_subject=account_subject,
                voucher__voucher_date__year=period_year,
                voucher__voucher_date__month=period_month,
                voucher__status='posted'  # åªæ˜¾ç¤ºå·²è¿‡è´¦çš„å‡­è¯
            ).select_related('voucher', 'account_subject').order_by('voucher__voucher_date', 'voucher__voucher_number', 'line_number')
            
            if date_from:
                voucher_entries_query = voucher_entries_query.filter(voucher__voucher_date__gte=date_from)
            if date_to:
                voucher_entries_query = voucher_entries_query.filter(voucher__voucher_date__lte=date_to)
            
            entries = list(voucher_entries_query)
            
            # è®¡ç®—æ¯ç¬”åˆ†å½•çš„ä½™é¢å’Œåˆè®¡
            current_balance = opening_balance
            entries_with_balance = []
            total_debit_amount = Decimal('0.00')
            total_credit_amount = Decimal('0.00')
            for entry in entries:
                if account_subject.direction == 'debit':
                    current_balance = current_balance + entry.debit_amount - entry.credit_amount
                else:
                    current_balance = current_balance + entry.credit_amount - entry.debit_amount
                entries_with_balance.append({
                    'entry': entry,
                    'balance': current_balance,
                })
                total_debit_amount += entry.debit_amount
                total_credit_amount += entry.credit_amount
            closing_balance = current_balance
            entries = entries_with_balance
            
        except (ValueError, AccountSubject.DoesNotExist):
            messages.error(request, 'æ— æ•ˆçš„ä¼šè®¡ç§‘ç›®')
    
    context = _context(
        "æ˜ç»†è´¦",
        "ğŸ“–",
        "æŸ¥çœ‹ä¼šè®¡ç§‘ç›®çš„è¯¦ç»†å‡­è¯åˆ†å½•",
        request=request,
        use_financial_nav=True
    )
    # è®¡ç®—åˆè®¡ï¼ˆå¦‚æœå·²é€‰æ‹©ç§‘ç›®ï¼‰
    total_debit_amount = Decimal('0.00')
    total_credit_amount = Decimal('0.00')
    if account_subject and entries:
        for item in entries:
            total_debit_amount += item['entry'].debit_amount
            total_credit_amount += item['entry'].credit_amount
    
    context.update({
        'account_subjects': account_subjects,
        'account_subject': account_subject,
        'entries': entries,
        'opening_balance': opening_balance,
        'closing_balance': closing_balance,
        'total_debit': total_debit_amount,
        'total_credit': total_credit_amount,
        'current_account_subject_id': account_subject_id,
        'current_period_year': period_year,
        'current_period_month': period_month,
        'current_date_from': date_from,
        'current_date_to': date_to,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/subsidiary_ledger.html", context)


@login_required
def account_balance_sheet(request):
    """ç§‘ç›®ä½™é¢è¡¨"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.ledger.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æŸ¥çœ‹ç§‘ç›®ä½™é¢è¡¨')
        return redirect('finance_pages:ledger_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year) if period_year else today.year
        period_month = int(period_month) if period_month else today.month
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # è·å–æ‰€æœ‰ä¼šè®¡ç§‘ç›®åŠå…¶ä½™é¢
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    balance_data = []
    
    total_opening_debit = Decimal('0.00')
    total_opening_credit = Decimal('0.00')
    total_period_debit = Decimal('0.00')
    total_period_credit = Decimal('0.00')
    total_closing_debit = Decimal('0.00')
    total_closing_credit = Decimal('0.00')
    
    for subject in account_subjects:
        # è·å–è¯¥ç§‘ç›®çš„æ€»è´¦è®°å½•
        ledger = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        ).order_by('-period_date').first()
        
        if ledger:
            opening_balance = ledger.opening_balance
            period_debit = ledger.period_debit
            period_credit = ledger.period_credit
            closing_balance = ledger.closing_balance
            
            # æ ¹æ®ä½™é¢æ–¹å‘è°ƒæ•´æ˜¾ç¤º
            if subject.direction == 'debit':
                opening_debit = opening_balance if opening_balance >= 0 else Decimal('0.00')
                opening_credit = -opening_balance if opening_balance < 0 else Decimal('0.00')
                closing_debit = closing_balance if closing_balance >= 0 else Decimal('0.00')
                closing_credit = -closing_balance if closing_balance < 0 else Decimal('0.00')
            else:
                opening_debit = -opening_balance if opening_balance < 0 else Decimal('0.00')
                opening_credit = opening_balance if opening_balance >= 0 else Decimal('0.00')
                closing_debit = -closing_balance if closing_balance < 0 else Decimal('0.00')
                closing_credit = closing_balance if closing_balance >= 0 else Decimal('0.00')
            
            balance_data.append({
                'subject': subject,
                'opening_debit': opening_debit,
                'opening_credit': opening_credit,
                'period_debit': period_debit,
                'period_credit': period_credit,
                'closing_debit': closing_debit,
                'closing_credit': closing_credit,
            })
            
            total_opening_debit += opening_debit
            total_opening_credit += opening_credit
            total_period_debit += period_debit
            total_period_credit += period_credit
            total_closing_debit += closing_debit
            total_closing_credit += closing_credit
    
    context = _context(
        "ç§‘ç›®ä½™é¢è¡¨",
        "ğŸ“Š",
        f"æŸ¥çœ‹ {period_year}å¹´{period_month}æœˆ æ‰€æœ‰ç§‘ç›®çš„ä½™é¢æƒ…å†µ",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'balance_data': balance_data,
        'period_year': period_year,
        'period_month': period_month,
        'total_opening_debit': total_opening_debit,
        'total_opening_credit': total_opening_credit,
        'total_period_debit': total_period_debit,
        'total_period_credit': total_period_credit,
        'total_closing_debit': total_closing_debit,
        'total_closing_credit': total_closing_credit,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/account_balance_sheet.html", context)


@login_required
def trial_balance(request):
    """è¯•ç®—å¹³è¡¡è¡¨"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.ledger.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æŸ¥çœ‹è¯•ç®—å¹³è¡¡è¡¨')
        return redirect('finance_pages:ledger_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year) if period_year else today.year
        period_month = int(period_month) if period_month else today.month
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # è·å–æ‰€æœ‰ä¼šè®¡ç§‘ç›®åŠå…¶ä½™é¢
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    trial_data = []
    
    total_opening_debit = Decimal('0.00')
    total_opening_credit = Decimal('0.00')
    total_period_debit = Decimal('0.00')
    total_period_credit = Decimal('0.00')
    total_closing_debit = Decimal('0.00')
    total_closing_credit = Decimal('0.00')
    
    for subject in account_subjects:
        # è·å–è¯¥ç§‘ç›®çš„æ€»è´¦è®°å½•
        ledger = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        ).order_by('-period_date').first()
        
        if ledger:
            opening_balance = ledger.opening_balance
            period_debit = ledger.period_debit
            period_credit = ledger.period_credit
            closing_balance = ledger.closing_balance
            
            # æ ¹æ®ä½™é¢æ–¹å‘è°ƒæ•´æ˜¾ç¤º
            if subject.direction == 'debit':
                opening_debit = opening_balance if opening_balance >= 0 else Decimal('0.00')
                opening_credit = -opening_balance if opening_balance < 0 else Decimal('0.00')
                closing_debit = closing_balance if closing_balance >= 0 else Decimal('0.00')
                closing_credit = -closing_balance if closing_balance < 0 else Decimal('0.00')
            else:
                opening_debit = -opening_balance if opening_balance < 0 else Decimal('0.00')
                opening_credit = opening_balance if opening_balance >= 0 else Decimal('0.00')
                closing_debit = -closing_balance if closing_balance < 0 else Decimal('0.00')
                closing_credit = closing_balance if closing_balance >= 0 else Decimal('0.00')
            
            trial_data.append({
                'subject': subject,
                'opening_debit': opening_debit,
                'opening_credit': opening_credit,
                'period_debit': period_debit,
                'period_credit': period_credit,
                'closing_debit': closing_debit,
                'closing_credit': closing_credit,
            })
            
            total_opening_debit += opening_debit
            total_opening_credit += opening_credit
            total_period_debit += period_debit
            total_period_credit += period_credit
            total_closing_debit += closing_debit
            total_closing_credit += closing_credit
    
    # æ£€æŸ¥æ˜¯å¦å¹³è¡¡
    opening_balanced = abs(total_opening_debit - total_opening_credit) < Decimal('0.01')
    period_balanced = abs(total_period_debit - total_period_credit) < Decimal('0.01')
    closing_balanced = abs(total_closing_debit - total_closing_credit) < Decimal('0.01')
    is_balanced = opening_balanced and period_balanced and closing_balanced
    
    context = _context(
        "è¯•ç®—å¹³è¡¡è¡¨",
        "âš–ï¸",
        f"éªŒè¯ {period_year}å¹´{period_month}æœˆ çš„å€Ÿè´·æ˜¯å¦å¹³è¡¡",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'trial_data': trial_data,
        'period_year': period_year,
        'period_month': period_month,
        'total_opening_debit': total_opening_debit,
        'total_opening_credit': total_opening_credit,
        'total_period_debit': total_period_debit,
        'total_period_credit': total_period_credit,
        'total_closing_debit': total_closing_debit,
        'total_closing_credit': total_closing_credit,
        'opening_balanced': opening_balanced,
        'period_balanced': period_balanced,
        'closing_balanced': closing_balanced,
        'is_balanced': is_balanced,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/trial_balance.html", context)


@login_required
def fund_flow_detail(request, fund_flow_id):
    """èµ„é‡‘æµæ°´è¯¦æƒ…"""
    fund_flow = get_object_or_404(
        FundFlow.objects.select_related('project', 'voucher', 'created_by'),
        id=fund_flow_id
    )
    
    context = _context(
        f"èµ„é‡‘æµæ°´è¯¦æƒ… - {fund_flow.flow_number}",
        "ğŸ’³",
        f"æŸ¥çœ‹èµ„é‡‘æµæ°´ {fund_flow.flow_number} çš„è¯¦ç»†ä¿¡æ¯",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'fund_flow': fund_flow,
    })
    return render(request, "financial_management/fund_flow_detail.html", context)


@login_required
def fund_flow_delete(request, fund_flow_id):
    """åˆ é™¤èµ„é‡‘æµæ°´"""
    fund_flow = get_object_or_404(FundFlow, id=fund_flow_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.fund_flow.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ é™¤èµ„é‡‘æµæ°´')
        return redirect('finance_pages:fund_flow_detail', fund_flow_id=fund_flow_id)
    
    if request.method == 'POST':
        try:
            flow_number = fund_flow.flow_number
            old_amount = fund_flow.amount
            
            # åˆ é™¤å‰å…ˆå›æ»šé¢„ç®—ä½¿ç”¨é‡‘é¢
            if fund_flow.flow_type == 'expense':
                # åˆ›å»ºä¸€ä¸ªä¸´æ—¶å¯¹è±¡ç”¨äºå›æ»š
                temp_flow = FundFlow(
                    flow_type=fund_flow.flow_type,
                    flow_date=fund_flow.flow_date,
                    amount=old_amount,
                    project=fund_flow.project,
                )
                # å›æ»šï¼šå‡å»æ—§é‡‘é¢
                _update_budget_from_fund_flow(temp_flow, is_create=False, old_amount=old_amount)
            
            fund_flow.delete()
            messages.success(request, f'èµ„é‡‘æµæ°´ {flow_number} å·²åˆ é™¤')
            return redirect('finance_pages:fund_flow_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('åˆ é™¤èµ„é‡‘æµæ°´å¤±è´¥: %s', str(e))
            messages.error(request, f'åˆ é™¤èµ„é‡‘æµæ°´å¤±è´¥ï¼š{str(e)}')
            return redirect('finance_pages:fund_flow_detail', fund_flow_id=fund_flow_id)
    
    context = _context(
        f"åˆ é™¤èµ„é‡‘æµæ°´ - {fund_flow.flow_number}",
        "ğŸ—‘ï¸",
        f"ç¡®è®¤åˆ é™¤èµ„é‡‘æµæ°´ï¼š{fund_flow.flow_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'fund_flow': fund_flow,
    })
    return render(request, "financial_management/fund_flow_delete.html", context)


# ==================== è´¢åŠ¡æŠ¥è¡¨ ====================

@login_required
def report_management(request):
    """è´¢åŠ¡æŠ¥è¡¨ç®¡ç†"""
    permission_codes = get_user_permission_codes(request.user)
    
    # è·å–ç­›é€‰å‚æ•°
    report_type = request.GET.get('report_type', '')
    period_year = request.GET.get('period_year', '')
    
    # è·å–æŠ¥è¡¨åˆ—è¡¨
    try:
        reports = FinancialReport.objects.select_related('generated_by').order_by('-report_date', '-generated_time')
        
        if report_type:
            reports = reports.filter(report_type=report_type)
        if period_year:
            reports = reports.filter(period_year=int(period_year))
        
        # åˆ†é¡µ
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(reports, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–æŠ¥è¡¨åˆ—è¡¨å¤±è´¥: %s', str(e))
        page_obj = None
    
    # ç»Ÿè®¡ä¿¡æ¯
    try:
        total_reports = FinancialReport.objects.count()
        balance_sheet_count = FinancialReport.objects.filter(report_type='balance_sheet').count()
        income_statement_count = FinancialReport.objects.filter(report_type='income_statement').count()
        cash_flow_count = FinancialReport.objects.filter(report_type='cash_flow').count()
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ç»Ÿè®¡ä¿¡æ¯å¤±è´¥: %s', str(e))
        summary_cards = []
    
    context = _context(
        "è´¢åŠ¡æŠ¥è¡¨ç®¡ç†",
        "ğŸ“Š",
        "ç®¡ç†è´¢åŠ¡æŠ¥è¡¨ç”Ÿæˆè®°å½•",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'reports': page_obj.object_list if page_obj else [],
        'report_type_choices': FinancialReport.REPORT_TYPE_CHOICES,
        'current_report_type': report_type,
        'current_period_year': period_year,
        'years': range(timezone.now().year - 2, timezone.now().year + 2),
    })
    return render(request, "financial_management/report_list.html", context)


@login_required
def balance_sheet_report(request):
    """èµ„äº§è´Ÿå€ºè¡¨ç”Ÿæˆ"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.report.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æŸ¥çœ‹è´¢åŠ¡æŠ¥è¡¨')
        return redirect('finance_pages:report_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year)
        period_month = int(period_month) if period_month else None
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # ç”Ÿæˆèµ„äº§è´Ÿå€ºè¡¨æ•°æ®
    report_data = {
        'assets': {},
        'liabilities': {},
        'equity': {},
    }
    
    # è·å–èµ„äº§ç±»ç§‘ç›®ä½™é¢
    asset_subjects = AccountSubject.objects.filter(
        subject_type='asset',
        is_active=True
    ).order_by('code')
    
    total_assets = Decimal('0.00')
    for subject in asset_subjects:
        # è·å–è¯¥ç§‘ç›®çš„æœŸæœ«ä½™é¢
        ledger = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        ).order_by('-period_date').first()
        
        if ledger:
            balance = ledger.closing_balance
            if subject.direction == 'credit':
                balance = -balance
            report_data['assets'][subject.code] = {
                'name': subject.name,
                'balance': balance,
            }
            total_assets += balance
    
    # è·å–è´Ÿå€ºç±»ç§‘ç›®ä½™é¢
    liability_subjects = AccountSubject.objects.filter(
        subject_type='liability',
        is_active=True
    ).order_by('code')
    
    total_liabilities = Decimal('0.00')
    for subject in liability_subjects:
        ledger = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        ).order_by('-period_date').first()
        
        if ledger:
            balance = ledger.closing_balance
            if subject.direction == 'debit':
                balance = -balance
            report_data['liabilities'][subject.code] = {
                'name': subject.name,
                'balance': balance,
            }
            total_liabilities += balance
    
    # è·å–æ‰€æœ‰è€…æƒç›Šç±»ç§‘ç›®ä½™é¢
    equity_subjects = AccountSubject.objects.filter(
        subject_type='equity',
        is_active=True
    ).order_by('code')
    
    total_equity = Decimal('0.00')
    for subject in equity_subjects:
        ledger = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        ).order_by('-period_date').first()
        
        if ledger:
            balance = ledger.closing_balance
            if subject.direction == 'debit':
                balance = -balance
            report_data['equity'][subject.code] = {
                'name': subject.name,
                'balance': balance,
            }
            total_equity += balance
    
    report_data['total_assets'] = total_assets
    report_data['total_liabilities'] = total_liabilities
    report_data['total_equity'] = total_equity
    report_data['total_liabilities_equity'] = total_liabilities + total_equity
    
    # å¦‚æœè¯·æ±‚ç”ŸæˆæŠ¥è¡¨ï¼Œä¿å­˜æŠ¥è¡¨è®°å½•
    if request.method == 'POST':
        try:
            # ç”Ÿæˆå”¯ä¸€çš„æŠ¥è¡¨ç¼–å·
            base_number = f'BS-{period_year}-{period_month:02d}' if period_month else f'BS-{period_year}'
            report_number = base_number
            counter = 1
            while FinancialReport.objects.filter(report_number=report_number).exists():
                report_number = f'{base_number}-{counter:02d}'
                counter += 1
            
            report = FinancialReport.objects.create(
                report_number=report_number,
                report_type='balance_sheet',
                period_year=period_year,
                period_month=period_month,
                report_date=today,
                report_data=report_data,
                generated_by=request.user,
            )
            messages.success(request, f'èµ„äº§è´Ÿå€ºè¡¨ç”ŸæˆæˆåŠŸï¼æŠ¥è¡¨ç¼–å·ï¼š{report_number}')
            return redirect('finance_pages:report_detail', report_id=report.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('ç”Ÿæˆèµ„äº§è´Ÿå€ºè¡¨å¤±è´¥: %s', str(e))
            messages.error(request, f'ç”ŸæˆæŠ¥è¡¨å¤±è´¥ï¼š{str(e)}')
    
    context = _context(
        "èµ„äº§è´Ÿå€ºè¡¨",
        "ğŸ“Š",
        f"{period_year}å¹´{period_month}æœˆèµ„äº§è´Ÿå€ºè¡¨",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'report_data': report_data,
        'period_year': period_year,
        'period_month': period_month,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/balance_sheet_report.html", context)


@login_required
def income_statement_report(request):
    """åˆ©æ¶¦è¡¨ç”Ÿæˆ"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.report.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æŸ¥çœ‹è´¢åŠ¡æŠ¥è¡¨')
        return redirect('finance_pages:report_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year)
        period_month = int(period_month) if period_month else None
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # ç”Ÿæˆåˆ©æ¶¦è¡¨æ•°æ®
    report_data = {
        'revenue': {},
        'expenses': {},
        'costs': {},
    }
    
    # è·å–æ”¶å…¥ç±»ç§‘ç›®å‘ç”Ÿé¢
    revenue_subjects = AccountSubject.objects.filter(
        subject_type='revenue',
        is_active=True
    ).order_by('code')
    
    total_revenue = Decimal('0.00')
    for subject in revenue_subjects:
        ledgers = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        )
        period_credit = sum(l.period_credit for l in ledgers)
        report_data['revenue'][subject.code] = {
            'name': subject.name,
            'amount': period_credit,
        }
        total_revenue += period_credit
    
    # è·å–è´¹ç”¨ç±»ç§‘ç›®å‘ç”Ÿé¢
    expense_subjects = AccountSubject.objects.filter(
        subject_type='expense',
        is_active=True
    ).order_by('code')
    
    total_expenses = Decimal('0.00')
    for subject in expense_subjects:
        ledgers = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        )
        period_debit = sum(l.period_debit for l in ledgers)
        report_data['expenses'][subject.code] = {
            'name': subject.name,
            'amount': period_debit,
        }
        total_expenses += period_debit
    
    # è·å–æˆæœ¬ç±»ç§‘ç›®å‘ç”Ÿé¢
    cost_subjects = AccountSubject.objects.filter(
        subject_type='cost',
        is_active=True
    ).order_by('code')
    
    total_costs = Decimal('0.00')
    for subject in cost_subjects:
        ledgers = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        )
        period_debit = sum(l.period_debit for l in ledgers)
        report_data['costs'][subject.code] = {
            'name': subject.name,
            'amount': period_debit,
        }
        total_costs += period_debit
    
    report_data['total_revenue'] = total_revenue
    report_data['total_costs'] = total_costs
    report_data['total_expenses'] = total_expenses
    report_data['gross_profit'] = total_revenue - total_costs
    report_data['net_profit'] = total_revenue - total_costs - total_expenses
    
    # å¦‚æœè¯·æ±‚ç”ŸæˆæŠ¥è¡¨ï¼Œä¿å­˜æŠ¥è¡¨è®°å½•
    if request.method == 'POST':
        try:
            # ç”Ÿæˆå”¯ä¸€çš„æŠ¥è¡¨ç¼–å·
            base_number = f'IS-{period_year}-{period_month:02d}' if period_month else f'IS-{period_year}'
            report_number = base_number
            counter = 1
            while FinancialReport.objects.filter(report_number=report_number).exists():
                report_number = f'{base_number}-{counter:02d}'
                counter += 1
            
            report = FinancialReport.objects.create(
                report_number=report_number,
                report_type='income_statement',
                period_year=period_year,
                period_month=period_month,
                report_date=today,
                report_data=report_data,
                generated_by=request.user,
            )
            messages.success(request, f'åˆ©æ¶¦è¡¨ç”ŸæˆæˆåŠŸï¼æŠ¥è¡¨ç¼–å·ï¼š{report_number}')
            return redirect('finance_pages:report_detail', report_id=report.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('ç”Ÿæˆåˆ©æ¶¦è¡¨å¤±è´¥: %s', str(e))
            messages.error(request, f'ç”ŸæˆæŠ¥è¡¨å¤±è´¥ï¼š{str(e)}')
    
    context = _context(
        "åˆ©æ¶¦è¡¨",
        "ğŸ“ˆ",
        f"{period_year}å¹´{period_month}æœˆåˆ©æ¶¦è¡¨",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'report_data': report_data,
        'period_year': period_year,
        'period_month': period_month,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/income_statement_report.html", context)


@login_required
def cash_flow_report(request):
    """ç°é‡‘æµé‡è¡¨ç”Ÿæˆ"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.report.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™æŸ¥çœ‹è´¢åŠ¡æŠ¥è¡¨')
        return redirect('finance_pages:report_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year)
        period_month = int(period_month) if period_month else None
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # ç”Ÿæˆç°é‡‘æµé‡è¡¨æ•°æ®ï¼ˆåŸºäºèµ„é‡‘æµæ°´ï¼‰
    if period_month:
        start_date = today.replace(year=period_year, month=period_month, day=1)
        if period_month == 12:
            end_date = today.replace(year=period_year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_date = today.replace(year=period_year, month=period_month + 1, day=1) - timedelta(days=1)
    else:
        start_date = today.replace(year=period_year, month=1, day=1)
        end_date = today.replace(year=period_year, month=12, day=31)
    
    # ç»è¥æ´»åŠ¨ç°é‡‘æµé‡
    operating_income = FundFlow.objects.filter(
        flow_date__gte=start_date,
        flow_date__lte=end_date,
        flow_type='income'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    operating_expense = FundFlow.objects.filter(
        flow_date__gte=start_date,
        flow_date__lte=end_date,
        flow_type='expense'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    operating_cash_flow = operating_income - operating_expense
    
    # æŠ•èµ„æ´»åŠ¨ç°é‡‘æµé‡ï¼ˆç®€åŒ–å¤„ç†ï¼‰
    investing_cash_flow = Decimal('0.00')
    
    # ç­¹èµ„æ´»åŠ¨ç°é‡‘æµé‡ï¼ˆç®€åŒ–å¤„ç†ï¼‰
    financing_cash_flow = Decimal('0.00')
    
    report_data = {
        'operating': {
            'income': operating_income,
            'expense': operating_expense,
            'net': operating_cash_flow,
        },
        'investing': {
            'net': investing_cash_flow,
        },
        'financing': {
            'net': financing_cash_flow,
        },
        'net_cash_flow': operating_cash_flow + investing_cash_flow + financing_cash_flow,
    }
    
    # å¦‚æœè¯·æ±‚ç”ŸæˆæŠ¥è¡¨ï¼Œä¿å­˜æŠ¥è¡¨è®°å½•
    if request.method == 'POST':
        try:
            # ç”Ÿæˆå”¯ä¸€çš„æŠ¥è¡¨ç¼–å·
            base_number = f'CF-{period_year}-{period_month:02d}' if period_month else f'CF-{period_year}'
            report_number = base_number
            counter = 1
            while FinancialReport.objects.filter(report_number=report_number).exists():
                report_number = f'{base_number}-{counter:02d}'
                counter += 1
            
            report = FinancialReport.objects.create(
                report_number=report_number,
                report_type='cash_flow',
                period_year=period_year,
                period_month=period_month,
                report_date=today,
                report_data=report_data,
                generated_by=request.user,
            )
            messages.success(request, f'ç°é‡‘æµé‡è¡¨ç”ŸæˆæˆåŠŸï¼æŠ¥è¡¨ç¼–å·ï¼š{report_number}')
            return redirect('finance_pages:report_detail', report_id=report.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('ç”Ÿæˆç°é‡‘æµé‡è¡¨å¤±è´¥: %s', str(e))
            messages.error(request, f'ç”ŸæˆæŠ¥è¡¨å¤±è´¥ï¼š{str(e)}')
    
    context = _context(
        "ç°é‡‘æµé‡è¡¨",
        "ğŸ’³",
        f"{period_year}å¹´{period_month}æœˆç°é‡‘æµé‡è¡¨",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'report_data': report_data,
        'period_year': period_year,
        'period_month': period_month,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/cash_flow_report.html", context)


@login_required
def report_export(request, report_id):
    """å¯¼å‡ºè´¢åŠ¡æŠ¥è¡¨ä¸ºExcel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.report.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å¯¼å‡ºè´¢åŠ¡æŠ¥è¡¨')
        return redirect('finance_pages:report_management')
    
    report = get_object_or_404(FinancialReport, id=report_id)
    report_data = report.report_data or {}
    
    # åˆ›å»ºExcelå·¥ä½œç°¿
    workbook = Workbook()
    worksheet = workbook.active
    
    # æ ¹æ®æŠ¥è¡¨ç±»å‹è®¾ç½®æ ‡é¢˜å’Œå†…å®¹
    report_type_dict = dict(FinancialReport.TYPE_CHOICES)
    worksheet.title = report_type_dict.get(report.report_type, 'è´¢åŠ¡æŠ¥è¡¨')
    
    # è®¾ç½®æŠ¥è¡¨æ ‡é¢˜
    title = f"{report_type_dict.get(report.report_type, 'è´¢åŠ¡æŠ¥è¡¨')} - {report.period_year}å¹´"
    if report.period_month:
        title += f"{report.period_month}æœˆ"
    worksheet.append([title])
    
    # åˆå¹¶æ ‡é¢˜å•å…ƒæ ¼
    worksheet.merge_cells(f'A1:{get_column_letter(10)}1')
    title_cell = worksheet['A1']
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet.append([])  # ç©ºè¡Œ
    
    # æ ¹æ®æŠ¥è¡¨ç±»å‹å¯¼å‡ºæ•°æ®
    if report.report_type == 'balance_sheet':
        # èµ„äº§è´Ÿå€ºè¡¨
        headers = ['é¡¹ç›®', 'æœŸåˆä½™é¢', 'æœ¬æœŸå€Ÿæ–¹', 'æœ¬æœŸè´·æ–¹', 'æœŸæœ«ä½™é¢']
        worksheet.append(headers)
        
        # è®¾ç½®è¡¨å¤´æ ·å¼
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # æ·»åŠ æ•°æ®
        for item in report_data.get('items', []):
            row = [
                item.get('name', ''),
                float(item.get('opening_balance', 0)),
                float(item.get('period_debit', 0)),
                float(item.get('period_credit', 0)),
                float(item.get('closing_balance', 0)),
            ]
            worksheet.append(row)
    
    elif report.report_type == 'income_statement':
        # åˆ©æ¶¦è¡¨
        headers = ['é¡¹ç›®', 'é‡‘é¢']
        worksheet.append(headers)
        
        # è®¾ç½®è¡¨å¤´æ ·å¼
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # æ·»åŠ æ”¶å…¥æ•°æ®
        worksheet.append(['æ”¶å…¥'])
        for code, data in report_data.get('revenue', {}).items():
            worksheet.append(['', data.get('name', ''), float(data.get('amount', 0))])
        worksheet.append(['æ”¶å…¥åˆè®¡', '', float(report_data.get('total_revenue', 0))])
        worksheet.append([])
        
        # æ·»åŠ æˆæœ¬æ•°æ®
        worksheet.append(['æˆæœ¬'])
        for code, data in report_data.get('costs', {}).items():
            worksheet.append(['', data.get('name', ''), float(data.get('amount', 0))])
        worksheet.append(['æˆæœ¬åˆè®¡', '', float(report_data.get('total_costs', 0))])
        worksheet.append([])
        
        # æ·»åŠ è´¹ç”¨æ•°æ®
        worksheet.append(['è´¹ç”¨'])
        for code, data in report_data.get('expenses', {}).items():
            worksheet.append(['', data.get('name', ''), float(data.get('amount', 0))])
        worksheet.append(['è´¹ç”¨åˆè®¡', '', float(report_data.get('total_expenses', 0))])
        worksheet.append([])
        
        # æ·»åŠ åˆ©æ¶¦æ•°æ®
        worksheet.append(['æ¯›åˆ©æ¶¦', '', float(report_data.get('gross_profit', 0))])
        worksheet.append(['å‡€åˆ©æ¶¦', '', float(report_data.get('net_profit', 0))])
    
    elif report.report_type == 'cash_flow':
        # ç°é‡‘æµé‡è¡¨
        headers = ['é¡¹ç›®', 'é‡‘é¢']
        worksheet.append(headers)
        
        # è®¾ç½®è¡¨å¤´æ ·å¼
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # ç»è¥æ´»åŠ¨ç°é‡‘æµé‡
        operating = report_data.get('operating', {})
        worksheet.append(['ç»è¥æ´»åŠ¨ç°é‡‘æµé‡'])
        worksheet.append(['', 'ç°é‡‘æµå…¥', float(operating.get('income', 0))])
        worksheet.append(['', 'ç°é‡‘æµå‡º', float(operating.get('expense', 0))])
        worksheet.append(['', 'å‡€æµé‡', float(operating.get('net', 0))])
        worksheet.append([])
        
        # æŠ•èµ„æ´»åŠ¨ç°é‡‘æµé‡
        investing = report_data.get('investing', {})
        worksheet.append(['æŠ•èµ„æ´»åŠ¨ç°é‡‘æµé‡'])
        worksheet.append(['', 'å‡€æµé‡', float(investing.get('net', 0))])
        worksheet.append([])
        
        # ç­¹èµ„æ´»åŠ¨ç°é‡‘æµé‡
        financing = report_data.get('financing', {})
        worksheet.append(['ç­¹èµ„æ´»åŠ¨ç°é‡‘æµé‡'])
        worksheet.append(['', 'å‡€æµé‡', float(financing.get('net', 0))])
        worksheet.append([])
        
        # ç°é‡‘å‡€æµé‡
        worksheet.append(['ç°é‡‘å‡€æµé‡', '', float(report_data.get('net_cash_flow', 0))])
    
    # è°ƒæ•´åˆ—å®½
    column_widths = [30, 15, 15, 15, 15]
    for i, width in enumerate(column_widths[:len(headers)], 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width
    
    # æ·»åŠ æŠ¥è¡¨ä¿¡æ¯
    worksheet.append([])
    worksheet.append(['æŠ¥è¡¨ç¼–å·', report.report_number])
    worksheet.append(['ç”Ÿæˆæ—¶é—´', report.generated_time.strftime('%Y-%m-%d %H:%M:%S') if report.generated_time else ''])
    worksheet.append(['ç”Ÿæˆäºº', report.generated_by.get_full_name() if report.generated_by else ''])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{report.report_number}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def report_detail(request, report_id):
    """æŠ¥è¡¨è¯¦æƒ…"""
    report = get_object_or_404(FinancialReport.objects.select_related('generated_by'), id=report_id)
    
    context = _context(
        f"æŠ¥è¡¨è¯¦æƒ… - {report.report_number}",
        "ğŸ“Š",
        f"æŸ¥çœ‹{report.get_report_type_display()}è¯¦æƒ…",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'report': report,
    })
    return render(request, "financial_management/report_detail.html", context)


# ==================== å¾€æ¥è´¦æ¬¾ ====================

@login_required
def receivable_management(request):
    """åº”æ”¶è´¦æ¬¾ç®¡ç†"""
    permission_codes = get_user_permission_codes(request.user)
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # è·å–åº”æ”¶è´¦æ¬¾åˆ—è¡¨
    try:
        receivables = ReceivableAccount.objects.select_related('customer', 'project', 'created_by').order_by('-receivable_date', '-account_number')
        
        if search:
            receivables = receivables.filter(
                Q(account_number__icontains=search) |
                Q(description__icontains=search)
            )
        if status:
            receivables = receivables.filter(status=status)
        if date_from:
            receivables = receivables.filter(receivable_date__gte=date_from)
        if date_to:
            receivables = receivables.filter(receivable_date__lte=date_to)
        
        # åˆ†é¡µ
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(receivables, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–åº”æ”¶è´¦æ¬¾åˆ—è¡¨å¤±è´¥: %s', str(e))
        page_obj = None
    
    # ç»Ÿè®¡ä¿¡æ¯
    try:
        total_receivables = ReceivableAccount.objects.count()
        total_amount = ReceivableAccount.objects.aggregate(total=Sum('receivable_amount'))['total'] or Decimal('0')
        received_amount = ReceivableAccount.objects.aggregate(total=Sum('received_amount'))['total'] or Decimal('0')
        remaining_amount = ReceivableAccount.objects.aggregate(total=Sum('remaining_amount'))['total'] or Decimal('0')
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ç»Ÿè®¡ä¿¡æ¯å¤±è´¥: %s', str(e))
        summary_cards = []
    
    context = _context(
        "åº”æ”¶è´¦æ¬¾ç®¡ç†",
        "ğŸ’°",
        "ç®¡ç†åº”æ”¶è´¦æ¬¾è®°å½•",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'receivables': page_obj.object_list if page_obj else [],
        'status_choices': ReceivableAccount.STATUS_CHOICES,
        'current_search': search,
        'current_status': status,
        'current_date_from': date_from,
        'current_date_to': date_to,
    })
    return render(request, "financial_management/receivable_list.html", context)


@login_required
def receivable_export(request):
    """å¯¼å‡ºåº”æ”¶è´¦æ¬¾åˆ—è¡¨ä¸ºExcel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.receivable.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å¯¼å‡ºåº”æ”¶è´¦æ¬¾')
        return redirect('finance_pages:receivable_management')
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    receivables = ReceivableAccount.objects.select_related('project', 'created_by').order_by('-receivable_date', '-account_number')
    
    if search:
        receivables = receivables.filter(
            Q(account_number__icontains=search) |
            Q(customer__icontains=search) |
            Q(description__icontains=search)
        )
    if status:
        receivables = receivables.filter(status=status)
    if date_from:
        receivables = receivables.filter(receivable_date__gte=date_from)
    if date_to:
        receivables = receivables.filter(receivable_date__lte=date_to)
    
    # åˆ›å»ºExcelå·¥ä½œç°¿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'åº”æ”¶è´¦æ¬¾'
    
    headers = ['åº”æ”¶å•å·', 'åº”æ”¶æ—¥æœŸ', 'å®¢æˆ·åç§°', 'åº”æ”¶é‡‘é¢', 'å·²æ”¶é‡‘é¢', 'æœªæ”¶é‡‘é¢', 'åˆ°æœŸæ—¥æœŸ', 'è´¦æœŸ(å¤©)', 'çŠ¶æ€', 'å…³è”é¡¹ç›®', 'å¤‡æ³¨', 'åˆ›å»ºäºº', 'åˆ›å»ºæ—¶é—´']
    worksheet.append(headers)
    
    # è®¾ç½®è¡¨å¤´æ ·å¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # æ·»åŠ æ•°æ®
    status_dict = dict(ReceivableAccount.STATUS_CHOICES)
    for receivable in receivables:
        row = [
            receivable.account_number,
            receivable.receivable_date.strftime('%Y-%m-%d') if receivable.receivable_date else '',
            receivable.customer,
            float(receivable.receivable_amount),
            float(receivable.paid_amount),
            float(receivable.remaining_amount),
            receivable.due_date.strftime('%Y-%m-%d') if receivable.due_date else '',
            receivable.payment_terms or '',
            status_dict.get(receivable.status, receivable.status),
            receivable.project.project_number if receivable.project else '',
            receivable.description or '',
            receivable.created_by.get_full_name() if receivable.created_by else '',
            receivable.created_time.strftime('%Y-%m-%d %H:%M') if receivable.created_time else '',
        ]
        worksheet.append(row)
    
    # è°ƒæ•´åˆ—å®½
    column_widths = [18, 12, 20, 12, 12, 12, 12, 10, 10, 15, 30, 12, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('åº”æ”¶è´¦æ¬¾_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def receivable_create(request):
    """æ–°å¢åº”æ”¶è´¦æ¬¾"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.receivable.create', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ›å»ºåº”æ”¶è´¦æ¬¾')
        return redirect('finance_pages:receivable_management')
    
    if request.method == 'POST':
        from .forms import ReceivableAccountForm
        form = ReceivableAccountForm(request.POST)
        if form.is_valid():
            receivable = form.save(commit=False)
            # è‡ªåŠ¨ç”Ÿæˆåº”æ”¶å•å·
            if not receivable.account_number:
                current_year = timezone.now().year
                # æŸ¥æ‰¾å½“å‰å¹´åº¦æœ€å¤§çš„åºå·
                max_receivable = ReceivableAccount.objects.filter(
                    account_number__startswith=f'AR-{current_year}-'
                ).order_by('-account_number').first()
                
                if max_receivable:
                    try:
                        # æå–åºå·éƒ¨åˆ†
                        parts = max_receivable.account_number.split('-')
                        if len(parts) >= 3:
                            seq = int(parts[-1]) + 1
                        else:
                            seq = 1
                    except (ValueError, IndexError):
                        seq = 1
                else:
                    seq = 1
                receivable.account_number = f'AR-{current_year}-{seq:04d}'
            
            # å¦‚æœè®¾ç½®äº†åº”æ”¶æ—¥æœŸå’Œè´¦æœŸï¼Œè‡ªåŠ¨è®¡ç®—åˆ°æœŸæ—¥æœŸ
            if receivable.receivable_date and receivable.payment_terms and not receivable.due_date:
                receivable.due_date = receivable.receivable_date + timedelta(days=receivable.payment_terms)
            
            receivable.created_by = request.user
            receivable.save()
            messages.success(request, f'åº”æ”¶è´¦æ¬¾ {receivable.account_number} åˆ›å»ºæˆåŠŸï¼')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable.id)
    else:
        from .forms import ReceivableAccountForm
        form = ReceivableAccountForm()
        form.fields['receivable_date'].initial = timezone.now().date()
    
    context = _context(
        "æ–°å¢åº”æ”¶è´¦æ¬¾",
        "â•",
        "åˆ›å»ºæ–°çš„åº”æ”¶è´¦æ¬¾è®°å½•",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/receivable_form.html", context)


@login_required
def receivable_update(request, receivable_id):
    """ç¼–è¾‘åº”æ”¶è´¦æ¬¾"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.receivable.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™ç¼–è¾‘åº”æ”¶è´¦æ¬¾')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    receivable = get_object_or_404(ReceivableAccount, id=receivable_id)
    
    if request.method == 'POST':
        from .forms import ReceivableAccountForm
        form = ReceivableAccountForm(request.POST, instance=receivable)
        if form.is_valid():
            receivable = form.save(commit=False)
            # å¦‚æœè®¾ç½®äº†åº”æ”¶æ—¥æœŸå’Œè´¦æœŸï¼Œè‡ªåŠ¨è®¡ç®—åˆ°æœŸæ—¥æœŸ
            if receivable.receivable_date and receivable.payment_terms and not receivable.due_date:
                receivable.due_date = receivable.receivable_date + timedelta(days=receivable.payment_terms)
            receivable.save()
            messages.success(request, f'åº”æ”¶è´¦æ¬¾ {receivable.account_number} æ›´æ–°æˆåŠŸï¼')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable.id)
    else:
        from .forms import ReceivableAccountForm
        form = ReceivableAccountForm(instance=receivable)
    
    context = _context(
        f"ç¼–è¾‘åº”æ”¶è´¦æ¬¾ - {receivable.account_number}",
        "âœï¸",
        f"ç¼–è¾‘åº”æ”¶è´¦æ¬¾ {receivable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'receivable': receivable,
        'is_create': False,
    })
    return render(request, "financial_management/receivable_form.html", context)


@login_required
def receivable_detail(request, receivable_id):
    """åº”æ”¶è´¦æ¬¾è¯¦æƒ…"""
    receivable = get_object_or_404(
        ReceivableAccount.objects.select_related('customer', 'project', 'created_by'),
        id=receivable_id
    )
    
    # è·å–æ”¶æ¬¾å†å²è®°å½•ï¼ˆé€šè¿‡èµ„é‡‘æµæ°´ï¼‰
    payment_history = FundFlow.objects.filter(
        flow_type='income',
        summary__icontains=receivable.account_number
    ).select_related('created_by', 'project').order_by('-flow_date', '-created_time')
    
    # ä¹Ÿå¯ä»¥æ ¹æ®å¯¹æ–¹å•ä½åŒ¹é…
    if receivable.customer:
        payment_history = payment_history.filter(
            Q(summary__icontains=receivable.account_number) |
            Q(counterparty__icontains=receivable.customer.name)
        )
    
    context = _context(
        f"åº”æ”¶è´¦æ¬¾è¯¦æƒ… - {receivable.account_number}",
        "ğŸ’°",
        f"æŸ¥çœ‹åº”æ”¶è´¦æ¬¾ {receivable.account_number} çš„è¯¦ç»†ä¿¡æ¯",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'receivable': receivable,
        'payment_history': payment_history[:20],  # æœ€è¿‘20æ¡è®°å½•
        'payment_history_count': payment_history.count(),
    })
    return render(request, "financial_management/receivable_detail.html", context)


@login_required
def receivable_payment(request, receivable_id):
    """è®°å½•åº”æ”¶è´¦æ¬¾æ”¶æ¬¾"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.receivable.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™è®°å½•æ”¶æ¬¾')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    receivable = get_object_or_404(ReceivableAccount, id=receivable_id)
    
    if receivable.status == 'completed':
        messages.error(request, 'è¯¥åº”æ”¶è´¦æ¬¾å·²å®Œæˆæ”¶æ¬¾ï¼Œä¸èƒ½ç»§ç»­æ”¶æ¬¾')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    if request.method == 'POST':
        payment_amount_str = request.POST.get('payment_amount', '0')
        payment_date = request.POST.get('payment_date', '')
        payment_method = request.POST.get('payment_method', '')
        payment_notes = request.POST.get('payment_notes', '')
        
        try:
            payment_amount = Decimal(payment_amount_str)
            if payment_amount <= 0:
                messages.error(request, 'æ”¶æ¬¾é‡‘é¢å¿…é¡»å¤§äº0')
                return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
            
            if payment_amount > receivable.remaining_amount:
                messages.error(request, f'æ”¶æ¬¾é‡‘é¢ä¸èƒ½è¶…è¿‡æœªæ”¶é‡‘é¢ {receivable.remaining_amount:,.2f}')
                return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
            
            # æ›´æ–°åº”æ”¶è´¦æ¬¾
            receivable.received_amount += payment_amount
            receivable.remaining_amount = receivable.receivable_amount - receivable.received_amount
            
            # è‡ªåŠ¨æ›´æ–°çŠ¶æ€
            if receivable.remaining_amount <= 0:
                receivable.status = 'completed'
            elif receivable.received_amount > 0:
                receivable.status = 'partial'
            
            receivable.save()
            
            # å¯é€‰ï¼šåˆ›å»ºèµ„é‡‘æµæ°´è®°å½•
            try:
                from django.db import transaction
                with transaction.atomic():
                    current_year = timezone.now().year
                    max_flow = FundFlow.objects.filter(
                        flow_number__startswith=f'FLOW-{current_year}-'
                    ).order_by('-flow_number').first()
                    
                    if max_flow:
                        try:
                            seq = int(max_flow.flow_number.split('-')[-1]) + 1
                        except (ValueError, IndexError):
                            seq = 1
                    else:
                        seq = 1
                    
                    flow_number = f'FLOW-{current_year}-{seq:04d}'
                    
                    FundFlow.objects.create(
                        flow_number=flow_number,
                        flow_date=payment_date or timezone.now().date(),
                        flow_type='income',
                        amount=payment_amount,
                        account_name=payment_method or 'é“¶è¡Œè´¦æˆ·',
                        counterparty=receivable.customer.name if receivable.customer else '',
                        summary=f'åº”æ”¶è´¦æ¬¾æ”¶æ¬¾ï¼š{receivable.account_number}' + (f' - {payment_notes}' if payment_notes else ''),
                        project=receivable.project,
                        created_by=request.user,
                    )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.exception('åˆ›å»ºèµ„é‡‘æµæ°´å¤±è´¥: %s', str(e))
                # ä¸å½±å“æ”¶æ¬¾è®°å½•ï¼Œåªè®°å½•æ—¥å¿—
            
            messages.success(request, f'æˆåŠŸè®°å½•æ”¶æ¬¾ {payment_amount:,.2f} å…ƒ')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable.id)
            
        except (ValueError, InvalidOperation):
            messages.error(request, 'æ”¶æ¬¾é‡‘é¢æ ¼å¼é”™è¯¯')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('è®°å½•æ”¶æ¬¾å¤±è´¥: %s', str(e))
            messages.error(request, f'è®°å½•æ”¶æ¬¾å¤±è´¥ï¼š{str(e)}')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºæ”¶æ¬¾è¡¨å•
    context = _context(
        f"è®°å½•æ”¶æ¬¾ - {receivable.account_number}",
        "ğŸ’°",
        f"è®°å½•åº”æ”¶è´¦æ¬¾ {receivable.account_number} çš„æ”¶æ¬¾",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'receivable': receivable,
    })
    return render(request, "financial_management/receivable_payment.html", context)


@login_required
def receivable_cancel(request, receivable_id):
    """å–æ¶ˆåº”æ”¶è´¦æ¬¾"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.receivable.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å–æ¶ˆåº”æ”¶è´¦æ¬¾')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    receivable = get_object_or_404(ReceivableAccount, id=receivable_id)
    
    # æ£€æŸ¥çŠ¶æ€ï¼šå·²å®Œæˆçš„åº”æ”¶è´¦æ¬¾ä¸èƒ½å–æ¶ˆ
    if receivable.status == 'completed':
        messages.error(request, 'å·²å®Œæˆçš„åº”æ”¶è´¦æ¬¾ä¸èƒ½å–æ¶ˆ')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable.id)
    
    if request.method == 'POST':
        receivable.status = 'cancelled'
        receivable.save()
        messages.success(request, f'åº”æ”¶è´¦æ¬¾ {receivable.account_number} å·²å–æ¶ˆ')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable.id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"å–æ¶ˆåº”æ”¶è´¦æ¬¾ - {receivable.account_number}",
        "âŒ",
        f"å–æ¶ˆåº”æ”¶è´¦æ¬¾ {receivable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'receivable': receivable,
    })
    return render(request, "financial_management/receivable_cancel.html", context)


@login_required
def receivable_delete(request, receivable_id):
    """åˆ é™¤åº”æ”¶è´¦æ¬¾"""
    receivable = get_object_or_404(ReceivableAccount, id=receivable_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.receivable.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ é™¤åº”æ”¶è´¦æ¬¾')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    # å¦‚æœå·²æ”¶æ¬¾ï¼Œä¸å…è®¸åˆ é™¤
    if receivable.received_amount > 0:
        messages.error(request, 'è¯¥åº”æ”¶è´¦æ¬¾å·²æœ‰æ”¶æ¬¾è®°å½•ï¼Œæ— æ³•åˆ é™¤')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    if request.method == 'POST':
        try:
            account_number = receivable.account_number
            receivable.delete()
            messages.success(request, f'åº”æ”¶è´¦æ¬¾ {account_number} å·²åˆ é™¤')
            return redirect('finance_pages:receivable_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('åˆ é™¤åº”æ”¶è´¦æ¬¾å¤±è´¥: %s', str(e))
            messages.error(request, f'åˆ é™¤åº”æ”¶è´¦æ¬¾å¤±è´¥ï¼š{str(e)}')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    context = _context(
        f"åˆ é™¤åº”æ”¶è´¦æ¬¾ - {receivable.account_number}",
        "ğŸ—‘ï¸",
        f"ç¡®è®¤åˆ é™¤åº”æ”¶è´¦æ¬¾ï¼š{receivable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'receivable': receivable,
    })
    return render(request, "financial_management/receivable_delete.html", context)


@login_required
def payable_management(request):
    """åº”ä»˜è´¦æ¬¾ç®¡ç†"""
    permission_codes = get_user_permission_codes(request.user)
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # è·å–åº”ä»˜è´¦æ¬¾åˆ—è¡¨
    try:
        payables = PayableAccount.objects.select_related('project', 'created_by').order_by('-payable_date', '-account_number')
        
        if search:
            payables = payables.filter(
                Q(account_number__icontains=search) |
                Q(supplier__icontains=search) |
                Q(description__icontains=search)
            )
        if status:
            payables = payables.filter(status=status)
        if date_from:
            payables = payables.filter(payable_date__gte=date_from)
        if date_to:
            payables = payables.filter(payable_date__lte=date_to)
        
        # åˆ†é¡µ
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(payables, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–åº”ä»˜è´¦æ¬¾åˆ—è¡¨å¤±è´¥: %s', str(e))
        page_obj = None
    
    # ç»Ÿè®¡ä¿¡æ¯
    try:
        total_payables = PayableAccount.objects.count()
        total_amount = PayableAccount.objects.aggregate(total=Sum('payable_amount'))['total'] or Decimal('0')
        paid_amount = PayableAccount.objects.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
        remaining_amount = PayableAccount.objects.aggregate(total=Sum('remaining_amount'))['total'] or Decimal('0')
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('è·å–ç»Ÿè®¡ä¿¡æ¯å¤±è´¥: %s', str(e))
        summary_cards = []
    
    context = _context(
        "åº”ä»˜è´¦æ¬¾ç®¡ç†",
        "ğŸ’¸",
        "ç®¡ç†åº”ä»˜è´¦æ¬¾è®°å½•",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'payables': page_obj.object_list if page_obj else [],
        'status_choices': PayableAccount.STATUS_CHOICES,
        'current_search': search,
        'current_status': status,
        'current_date_from': date_from,
        'current_date_to': date_to,
    })
    return render(request, "financial_management/payable_list.html", context)


@login_required
def payable_export(request):
    """å¯¼å‡ºåº”ä»˜è´¦æ¬¾åˆ—è¡¨ä¸ºExcel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.payable.view', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å¯¼å‡ºåº”ä»˜è´¦æ¬¾')
        return redirect('finance_pages:payable_management')
    
    # è·å–ç­›é€‰å‚æ•°
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    payables = PayableAccount.objects.select_related('project', 'created_by').order_by('-payable_date', '-account_number')
    
    if search:
        payables = payables.filter(
            Q(account_number__icontains=search) |
            Q(supplier__icontains=search) |
            Q(description__icontains=search)
        )
    if status:
        payables = payables.filter(status=status)
    if date_from:
        payables = payables.filter(payable_date__gte=date_from)
    if date_to:
        payables = payables.filter(payable_date__lte=date_to)
    
    # åˆ›å»ºExcelå·¥ä½œç°¿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'åº”ä»˜è´¦æ¬¾'
    
    headers = ['åº”ä»˜å•å·', 'åº”ä»˜æ—¥æœŸ', 'ä¾›åº”å•†', 'åº”ä»˜é‡‘é¢', 'å·²ä»˜é‡‘é¢', 'æœªä»˜é‡‘é¢', 'åˆ°æœŸæ—¥æœŸ', 'è´¦æœŸ(å¤©)', 'çŠ¶æ€', 'å…³è”é¡¹ç›®', 'å¤‡æ³¨', 'åˆ›å»ºäºº', 'åˆ›å»ºæ—¶é—´']
    worksheet.append(headers)
    
    # è®¾ç½®è¡¨å¤´æ ·å¼
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # æ·»åŠ æ•°æ®
    status_dict = dict(PayableAccount.STATUS_CHOICES)
    for payable in payables:
        row = [
            payable.account_number,
            payable.payable_date.strftime('%Y-%m-%d') if payable.payable_date else '',
            payable.supplier,
            float(payable.payable_amount),
            float(payable.paid_amount),
            float(payable.remaining_amount),
            payable.due_date.strftime('%Y-%m-%d') if payable.due_date else '',
            payable.payment_terms or '',
            status_dict.get(payable.status, payable.status),
            payable.project.project_number if payable.project else '',
            payable.description or '',
            payable.created_by.get_full_name() if payable.created_by else '',
            payable.created_time.strftime('%Y-%m-%d %H:%M') if payable.created_time else '',
        ]
        worksheet.append(row)
    
    # è°ƒæ•´åˆ—å®½
    column_widths = [18, 12, 20, 12, 12, 12, 12, 10, 10, 15, 30, 12, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('åº”ä»˜è´¦æ¬¾_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def payable_create(request):
    """æ–°å¢åº”ä»˜è´¦æ¬¾"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.payable.create', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ›å»ºåº”ä»˜è´¦æ¬¾')
        return redirect('finance_pages:payable_management')
    
    if request.method == 'POST':
        from .forms import PayableAccountForm
        form = PayableAccountForm(request.POST)
        if form.is_valid():
            payable = form.save(commit=False)
            # è‡ªåŠ¨ç”Ÿæˆåº”ä»˜å•å·
            if not payable.account_number:
                current_year = timezone.now().year
                # æŸ¥æ‰¾å½“å‰å¹´åº¦æœ€å¤§çš„åºå·
                max_payable = PayableAccount.objects.filter(
                    account_number__startswith=f'AP-{current_year}-'
                ).order_by('-account_number').first()
                
                if max_payable:
                    try:
                        # æå–åºå·éƒ¨åˆ†
                        parts = max_payable.account_number.split('-')
                        if len(parts) >= 3:
                            seq = int(parts[-1]) + 1
                        else:
                            seq = 1
                    except (ValueError, IndexError):
                        seq = 1
                else:
                    seq = 1
                payable.account_number = f'AP-{current_year}-{seq:04d}'
            
            # å¦‚æœè®¾ç½®äº†åº”ä»˜æ—¥æœŸå’Œè´¦æœŸï¼Œè‡ªåŠ¨è®¡ç®—åˆ°æœŸæ—¥æœŸ
            if payable.payable_date and payable.payment_terms and not payable.due_date:
                payable.due_date = payable.payable_date + timedelta(days=payable.payment_terms)
            
            payable.created_by = request.user
            payable.save()
            messages.success(request, f'åº”ä»˜è´¦æ¬¾ {payable.account_number} åˆ›å»ºæˆåŠŸï¼')
            return redirect('finance_pages:payable_detail', payable_id=payable.id)
    else:
        from .forms import PayableAccountForm
        form = PayableAccountForm()
        form.fields['payable_date'].initial = timezone.now().date()
    
    context = _context(
        "æ–°å¢åº”ä»˜è´¦æ¬¾",
        "â•",
        "åˆ›å»ºæ–°çš„åº”ä»˜è´¦æ¬¾è®°å½•",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/payable_form.html", context)


@login_required
def payable_update(request, payable_id):
    """ç¼–è¾‘åº”ä»˜è´¦æ¬¾"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.payable.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™ç¼–è¾‘åº”ä»˜è´¦æ¬¾')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    payable = get_object_or_404(PayableAccount, id=payable_id)
    
    if request.method == 'POST':
        from .forms import PayableAccountForm
        form = PayableAccountForm(request.POST, instance=payable)
        if form.is_valid():
            payable = form.save(commit=False)
            # å¦‚æœè®¾ç½®äº†åº”ä»˜æ—¥æœŸå’Œè´¦æœŸï¼Œè‡ªåŠ¨è®¡ç®—åˆ°æœŸæ—¥æœŸ
            if payable.payable_date and payable.payment_terms and not payable.due_date:
                payable.due_date = payable.payable_date + timedelta(days=payable.payment_terms)
            payable.save()
            messages.success(request, f'åº”ä»˜è´¦æ¬¾ {payable.account_number} æ›´æ–°æˆåŠŸï¼')
            return redirect('finance_pages:payable_detail', payable_id=payable.id)
    else:
        from .forms import PayableAccountForm
        form = PayableAccountForm(instance=payable)
    
    context = _context(
        f"ç¼–è¾‘åº”ä»˜è´¦æ¬¾ - {payable.account_number}",
        "âœï¸",
        f"ç¼–è¾‘åº”ä»˜è´¦æ¬¾ {payable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'payable': payable,
        'is_create': False,
    })
    return render(request, "financial_management/payable_form.html", context)


@login_required
def payable_detail(request, payable_id):
    """åº”ä»˜è´¦æ¬¾è¯¦æƒ…"""
    payable = get_object_or_404(
        PayableAccount.objects.select_related('project', 'created_by'),
        id=payable_id
    )
    
    # è·å–ä»˜æ¬¾å†å²è®°å½•ï¼ˆé€šè¿‡èµ„é‡‘æµæ°´ï¼‰
    payment_history = FundFlow.objects.filter(
        flow_type='expense',
        summary__icontains=payable.account_number
    ).select_related('created_by', 'project').order_by('-flow_date', '-created_time')
    
    # ä¹Ÿå¯ä»¥æ ¹æ®ä¾›åº”å•†åŒ¹é…
    payment_history = payment_history.filter(
        Q(summary__icontains=payable.account_number) |
        Q(counterparty__icontains=payable.supplier)
    )
    
    context = _context(
        f"åº”ä»˜è´¦æ¬¾è¯¦æƒ… - {payable.account_number}",
        "ğŸ’¸",
        f"æŸ¥çœ‹åº”ä»˜è´¦æ¬¾ {payable.account_number} çš„è¯¦ç»†ä¿¡æ¯",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'payable': payable,
        'payment_history': payment_history[:20],  # æœ€è¿‘20æ¡è®°å½•
        'payment_history_count': payment_history.count(),
    })
    return render(request, "financial_management/payable_detail.html", context)


@login_required
def payable_payment(request, payable_id):
    """è®°å½•åº”ä»˜è´¦æ¬¾ä»˜æ¬¾"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.payable.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™è®°å½•ä»˜æ¬¾')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    payable = get_object_or_404(PayableAccount, id=payable_id)
    
    if payable.status == 'completed':
        messages.error(request, 'è¯¥åº”ä»˜è´¦æ¬¾å·²å®Œæˆä»˜æ¬¾ï¼Œä¸èƒ½ç»§ç»­ä»˜æ¬¾')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    if request.method == 'POST':
        payment_amount_str = request.POST.get('payment_amount', '0')
        payment_date = request.POST.get('payment_date', '')
        payment_method = request.POST.get('payment_method', '')
        payment_notes = request.POST.get('payment_notes', '')
        
        try:
            payment_amount = Decimal(payment_amount_str)
            if payment_amount <= 0:
                messages.error(request, 'ä»˜æ¬¾é‡‘é¢å¿…é¡»å¤§äº0')
                return redirect('finance_pages:payable_detail', payable_id=payable_id)
            
            if payment_amount > payable.remaining_amount:
                messages.error(request, f'ä»˜æ¬¾é‡‘é¢ä¸èƒ½è¶…è¿‡æœªä»˜é‡‘é¢ {payable.remaining_amount:,.2f}')
                return redirect('finance_pages:payable_detail', payable_id=payable_id)
            
            # æ›´æ–°åº”ä»˜è´¦æ¬¾
            payable.paid_amount += payment_amount
            payable.remaining_amount = payable.payable_amount - payable.paid_amount
            
            # è‡ªåŠ¨æ›´æ–°çŠ¶æ€
            if payable.remaining_amount <= 0:
                payable.status = 'completed'
            elif payable.paid_amount > 0:
                payable.status = 'partial'
            
            payable.save()
            
            # å¯é€‰ï¼šåˆ›å»ºèµ„é‡‘æµæ°´è®°å½•
            try:
                from django.db import transaction
                with transaction.atomic():
                    current_year = timezone.now().year
                    max_flow = FundFlow.objects.filter(
                        flow_number__startswith=f'FLOW-{current_year}-'
                    ).order_by('-flow_number').first()
                    
                    if max_flow:
                        try:
                            seq = int(max_flow.flow_number.split('-')[-1]) + 1
                        except (ValueError, IndexError):
                            seq = 1
                    else:
                        seq = 1
                    
                    flow_number = f'FLOW-{current_year}-{seq:04d}'
                    
                    FundFlow.objects.create(
                        flow_number=flow_number,
                        flow_date=payment_date or timezone.now().date(),
                        flow_type='expense',
                        amount=payment_amount,
                        account_name=payment_method or 'é“¶è¡Œè´¦æˆ·',
                        counterparty=payable.supplier,
                        summary=f'åº”ä»˜è´¦æ¬¾ä»˜æ¬¾ï¼š{payable.account_number}' + (f' - {payment_notes}' if payment_notes else ''),
                        project=payable.project,
                        created_by=request.user,
                    )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.exception('åˆ›å»ºèµ„é‡‘æµæ°´å¤±è´¥: %s', str(e))
                # ä¸å½±å“ä»˜æ¬¾è®°å½•ï¼Œåªè®°å½•æ—¥å¿—
            
            messages.success(request, f'æˆåŠŸè®°å½•ä»˜æ¬¾ {payment_amount:,.2f} å…ƒ')
            return redirect('finance_pages:payable_detail', payable_id=payable.id)
            
        except (ValueError, InvalidOperation):
            messages.error(request, 'ä»˜æ¬¾é‡‘é¢æ ¼å¼é”™è¯¯')
            return redirect('finance_pages:payable_detail', payable_id=payable_id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('è®°å½•ä»˜æ¬¾å¤±è´¥: %s', str(e))
            messages.error(request, f'è®°å½•ä»˜æ¬¾å¤±è´¥ï¼š{str(e)}')
            return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºä»˜æ¬¾è¡¨å•
    context = _context(
        f"è®°å½•ä»˜æ¬¾ - {payable.account_number}",
        "ğŸ’¸",
        f"è®°å½•åº”ä»˜è´¦æ¬¾ {payable.account_number} çš„ä»˜æ¬¾",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'payable': payable,
    })
    return render(request, "financial_management/payable_payment.html", context)


@login_required
def payable_cancel(request, payable_id):
    """å–æ¶ˆåº”ä»˜è´¦æ¬¾"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.payable.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™å–æ¶ˆåº”ä»˜è´¦æ¬¾')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    payable = get_object_or_404(PayableAccount, id=payable_id)
    
    # æ£€æŸ¥çŠ¶æ€ï¼šå·²å®Œæˆçš„åº”ä»˜è´¦æ¬¾ä¸èƒ½å–æ¶ˆ
    if payable.status == 'completed':
        messages.error(request, 'å·²å®Œæˆçš„åº”ä»˜è´¦æ¬¾ä¸èƒ½å–æ¶ˆ')
        return redirect('finance_pages:payable_detail', payable_id=payable.id)
    
    if request.method == 'POST':
        payable.status = 'cancelled'
        payable.save()
        messages.success(request, f'åº”ä»˜è´¦æ¬¾ {payable.account_number} å·²å–æ¶ˆ')
        return redirect('finance_pages:payable_detail', payable_id=payable.id)
    
    # GETè¯·æ±‚ï¼Œæ˜¾ç¤ºç¡®è®¤é¡µé¢
    context = _context(
        f"å–æ¶ˆåº”ä»˜è´¦æ¬¾ - {payable.account_number}",
        "âŒ",
        f"å–æ¶ˆåº”ä»˜è´¦æ¬¾ {payable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'payable': payable,
    })
    return render(request, "financial_management/payable_cancel.html", context)


@login_required
def payable_delete(request, payable_id):
    """åˆ é™¤åº”ä»˜è´¦æ¬¾"""
    payable = get_object_or_404(PayableAccount, id=payable_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.payable.manage', permission_codes):
        messages.error(request, 'æ‚¨æ²¡æœ‰æƒé™åˆ é™¤åº”ä»˜è´¦æ¬¾')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    # å¦‚æœå·²ä»˜æ¬¾ï¼Œä¸å…è®¸åˆ é™¤
    if payable.paid_amount > 0:
        messages.error(request, 'è¯¥åº”ä»˜è´¦æ¬¾å·²æœ‰ä»˜æ¬¾è®°å½•ï¼Œæ— æ³•åˆ é™¤')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    if request.method == 'POST':
        try:
            account_number = payable.account_number
            payable.delete()
            messages.success(request, f'åº”ä»˜è´¦æ¬¾ {account_number} å·²åˆ é™¤')
            return redirect('finance_pages:payable_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('åˆ é™¤åº”ä»˜è´¦æ¬¾å¤±è´¥: %s', str(e))
            messages.error(request, f'åˆ é™¤åº”ä»˜è´¦æ¬¾å¤±è´¥ï¼š{str(e)}')
            return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    context = _context(
        f"åˆ é™¤åº”ä»˜è´¦æ¬¾ - {payable.account_number}",
        "ğŸ—‘ï¸",
        f"ç¡®è®¤åˆ é™¤åº”ä»˜è´¦æ¬¾ï¼š{payable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'payable': payable,
    })
    return render(request, "financial_management/payable_delete.html", context)

