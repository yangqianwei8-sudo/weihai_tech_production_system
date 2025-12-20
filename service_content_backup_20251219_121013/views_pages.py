from decimal import Decimal, InvalidOperation
import json
import csv
import io
import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Q, F
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.urls import reverse, NoReverseMatch

from backend.apps.customer_management.models import (
    ClientType,
    ClientGrade,
    Client,
    ClientContact,
    ClientProject,
    BusinessOpportunity,
    OpportunityFollowUp,
    OpportunityQuotation,
    # CustomerLead, LeadFollowUp 已删除（按《客户管理详细设计方案 v1.12》）
    CustomerRelationship,
    CustomerRelationshipUpgrade,
    BusinessExpenseApplication,
    VisitPlan,
    VisitCheckin,
    VisitReview,
    SalesActivity,
    BusinessNegotiation,
    BiddingQuotation,
    AuthorizationLetter,
    AuthorizationLetterTemplate,
    ContactEducation,
    ContactCareer,
    ContactColleague,
)

# 尝试导入沟通清单相关模型（如果存在）
try:
    from backend.apps.customer_management.models import (
        CommunicationChecklistQuestion,
        CommunicationChecklistAnswer,
        CustomerCommunicationChecklist,
    )
    HAS_COMMUNICATION_CHECKLIST_MODELS = True
except ImportError:
    HAS_COMMUNICATION_CHECKLIST_MODELS = False
# BusinessContract和BusinessPaymentPlan已迁移到production_management
from backend.apps.production_management.models import BusinessContract, BusinessPaymentPlan, DesignStage, ServiceType
from backend.apps.system_management.services import get_user_permission_codes
from backend.core.views import HOME_NAV_STRUCTURE, _permission_granted, _build_full_top_nav
from backend.apps.permission_management.utils import normalize_permission_code

logger = logging.getLogger(__name__)


# ==================== 客户管理模块左侧菜单结构（按《客户管理详细设计方案 v1.12》）====================
CUSTOMER_MANAGEMENT_MENU = [
    {
        'id': 'customer_info',
        'label': '客户信息管理',
        'icon': '👥',
        'permission': 'customer_management.client.view',
        'children': [
            {
                'id': 'customer_list',
                'label': '创建客户',
                'icon': '📋',
                'url_name': 'business_pages:customer_list',
                'permission': 'customer_management.client.view',  # 自动根据权限级别显示
            },
            {
                'id': 'customer_public_sea',
                'label': '客户公海',
                'icon': '🌊',
                'url_name': 'business_pages:customer_public_sea',
                'permission': 'customer_management.public_sea.view',
            },
        ]
    },
    {
        'id': 'customer_contact',
        'label': '人员关系管理',
        'icon': '👤',
        'permission': 'customer_management.contact.view',
        'children': [
            {
                'id': 'contact_list',
                'label': '创建联系人信息',
                'icon': '📇',
                'url_name': 'business_pages:contact_list',
                'permission': 'customer_management.contact.view',
            },
            {
                'id': 'contact_relationship_mining',
                'label': '关系挖掘',
                'icon': '🔍',
                'url_name': 'business_pages:contact_relationship_mining',
                'permission': 'customer_management.contact.view',
            },
            {
                'id': 'visit_list',
                'label': '创建联系人拜访',
                'icon': '🚪',
                'url_name': 'business_pages:customer_visit',
                'permission': 'customer_management.relationship.view',
            },
            {
                'id': 'contact_tracking_reminders',
                'label': '逾期拜访提醒',
                'icon': '🔔',
                'url_name': 'business_pages:contact_tracking_reminders',
                'permission': 'customer_management.contact.view',
            },
        ]
    },
    {
        'id': 'relationship_upgrade',
        'label': '关系升级管理',
        'icon': '📈',
        'permission': 'customer_management.relationship.view',
        'children': [
            {
                'id': 'upgrade_list',
                'label': '创建人员关系升级',
                'icon': '⬆️',
                'url_name': 'business_pages:customer_relationship_upgrade',
                'permission': 'customer_management.relationship.view',
            },
            {
                'id': 'business_expense_application',
                'label': '业务费申请',
                'icon': '💰',
                'url_name': 'business_pages:business_expense_application_list',
                'permission': 'customer_management.relationship.view',
            },
            {
                'id': 'relationship_collaboration',
                'label': '人员关系协作申请',
                'icon': '🤝',
                'url_name': 'business_pages:customer_relationship_collaboration',
                'permission': 'customer_management.relationship.view',
            },
        ]
    },
]


# ==================== 合同管理模块左侧菜单结构 =====================
CONTRACT_MANAGEMENT_MENU = [
    {
        'id': 'authorization_letter',
        'label': '业务委托书',
        'icon': '📋',
        'permission': 'customer_management.client.view',  # 使用客户管理权限（临时）
        'children': [
            {
                'id': 'authorization_letter_list',
                'label': '创建业务委托书',
                'icon': '📋',
                'url_name': 'business_pages:authorization_letter_list',
                'permission': 'customer_management.client.view',
            },
        ]
    },
    {
        'id': 'contract_signing',
        'label': '正式合同签署',
        'icon': '✍️',
        'permission': 'customer_management.client.view',  # 使用客户管理权限（临时）
        'children': [
            {
                'id': 'contract_management_list',
                'label': '创建合同草稿',
                'icon': '📄',
                'url_name': 'business_pages:contract_management_list',
                'permission': 'customer_management.client.view',
            },
            {
                'id': 'contract_negotiation_list',
                'label': '合同洽谈记录',
                'icon': '💬',
                'url_name': 'business_pages:contract_negotiation_list',
                'permission': 'customer_management.client.view',
            },
            {
                'id': 'contract_negotiation_create',
                'label': '创建合同洽谈记录',
                'icon': '➕',
                'url_name': 'business_pages:contract_negotiation_create',
                'permission': 'customer_management.client.create',
            },
            {
                'id': 'contract_finalize_list',
                'label': '合同定稿列表',
                'icon': '📋',
                'url_name': 'business_pages:contract_finalize_list',
                'permission': 'customer_management.client.view',
            },
            {
                'id': 'contract_finalize_create',
                'label': '创建合同定稿',
                'icon': '✅',
                'url_name': 'business_pages:contract_finalize_create',
                'permission': 'customer_management.client.create',
            },
        ]
    },
    {
        'id': 'contract_execution',
        'label': '合同执行',
        'icon': '📊',
        'permission': 'customer_management.client.view',  # 使用客户管理权限（临时）
        'children': [
            {
                'id': 'contract_performance',
                'label': '履约跟踪',
                'icon': '📋',
                'url_name': 'business_pages:contract_performance_track',
                'permission': 'customer_management.client.view',
            },
            {
                'id': 'contract_dispute_list',
                'label': '合同争议',
                'icon': '⚖️',
                'url_name': 'business_pages:contract_dispute_list',
                'permission': 'customer_management.client.view',
            },
        ]
    },
    {
        'id': 'contract_reminder',
        'label': '提醒与警报',
        'icon': '⚠️',
        'permission': 'customer_management.client.view',  # 使用客户管理权限（临时）
        'children': [
            {
                'id': 'contract_expiry_reminder',
                'label': '到期提醒',
                'icon': '📅',
                'url_name': 'business_pages:contract_expiry_reminder',
                'permission': 'customer_management.client.view',
            },
            {
                'id': 'contract_payment_reminder',
                'label': '付款提醒',
                'icon': '💰',
                'url_name': 'business_pages:contract_payment_reminder',
                'permission': 'customer_management.client.view',
            },
            {
                'id': 'contract_risk_warning',
                'label': '风险预警',
                'icon': '⚠️',
                'url_name': 'business_pages:contract_risk_warning',
                'permission': 'customer_management.client.view',
            },
        ]
    },
]


# ==================== 商机管理模块左侧菜单结构 =====================
OPPORTUNITY_MANAGEMENT_MENU = [
    {
        'id': 'opportunity_info',
        'label': '商机信息管理',
        'icon': '📋',
        'permission': 'customer_management.opportunity.view',
        'children': [
            {
                'id': 'opportunity_list',
                'label': '创建商机',
                'icon': '📋',
                'url_name': 'business_pages:opportunity_management',
                'permission': 'customer_management.opportunity.view',
            },
        ]
    },
    {
        'id': 'technical_support',
        'label': '技术支持',
        'icon': '🔧',
        'permission': 'customer_management.opportunity.view',
        'children': [
            {
                'id': 'evaluation_application',
                'label': '评估申请',
                'icon': '📝',
                'url_name': 'business_pages:opportunity_evaluation_application',
                'permission': 'customer_management.opportunity.manage',
            },
            {
                'id': 'drawing_evaluation',
                'label': '图纸评估',
                'icon': '📐',
                'url_name': 'business_pages:opportunity_drawing_evaluation',
                'permission': 'customer_management.opportunity.view',
            },
            {
                'id': 'tech_meeting',
                'label': '技术沟通会',
                'icon': '🤝',
                'url_name': 'business_pages:opportunity_tech_meeting',
                'permission': 'customer_management.opportunity.view',
            },
        ]
    },
    {
        'id': 'bidding_quotation',
        'label': '投标报价',
        'icon': '💰',
        'permission': 'customer_management.opportunity.view',
        'children': [
            {
                'id': 'warehouse_list',
                'label': '创建入库',
                'icon': '📥',
                'url_name': 'business_pages:opportunity_warehouse_list',
                'permission': 'customer_management.opportunity.view',
            },
            {
                'id': 'bidding_quotation_application',
                'label': '投标报价申请',
                'icon': '📋',
                'url_name': 'business_pages:opportunity_bidding_quotation_application',
                'permission': 'customer_management.opportunity.view',
            },
            {
                'id': 'bidding_quotation',
                'label': '投标报价管理',
                'icon': '📊',
                'url_name': 'business_pages:opportunity_bidding_quotation',
                'permission': 'customer_management.opportunity.view',
            },
            {
                'id': 'bidding_document_preparation',
                'label': '编制投标文件',
                'icon': '📄',
                'url_name': 'business_pages:opportunity_bidding_document_preparation',
                'permission': 'customer_management.opportunity.manage',
            },
            {
                'id': 'bidding_document_submission',
                'label': '递交投标文件',
                'icon': '📤',
                'url_name': 'business_pages:opportunity_bidding_document_submission',
                'permission': 'customer_management.opportunity.manage',
            },
        ]
    },
    {
        'id': 'opportunity_achievement',
        'label': '商机成就',
        'icon': '🎯',
        'permission': 'customer_management.opportunity.view',
        'children': [
            {
                'id': 'business_negotiation',
                'label': '商务洽谈登记',
                'icon': '💼',
                'url_name': 'business_pages:opportunity_business_negotiation',
                'permission': 'customer_management.opportunity.view',
            },
            {
                'id': 'sales_forecast',
                'label': '商机预测',
                'icon': '📈',
                'url_name': 'business_pages:opportunity_sales_forecast',
                'permission': 'customer_management.opportunity.view',
            },
            {
                'id': 'win_loss',
                'label': '赢单与输单',
                'icon': '✅',
                'url_name': 'business_pages:opportunity_win_loss',
                'permission': 'customer_management.opportunity.manage',
            },
        ]
    },
    {
        'id': 'payment_management',
        'label': '费用支付',
        'icon': '💳',
        'permission': 'customer_management.opportunity.view',
        'children': [
            {
                'id': 'bid_bond_payment',
                'label': '投标保证金支付',
                'icon': '💰',
                'url_name': 'business_pages:opportunity_bid_bond_payment',
                'permission': 'customer_management.opportunity.manage',
            },
            {
                'id': 'tender_fee_payment',
                'label': '标书费支付',
                'icon': '📄',
                'url_name': 'business_pages:opportunity_tender_fee_payment',
                'permission': 'customer_management.opportunity.manage',
            },
            {
                'id': 'tender_agent_fee_payment',
                'label': '招标代理费支付',
                'icon': '🏢',
                'url_name': 'business_pages:opportunity_tender_agent_fee_payment',
                'permission': 'customer_management.opportunity.manage',
            },
        ]
    },
]


def _build_opportunity_management_menu(permission_set, active_id=None):
    """
    生成商机管理模块左侧菜单
    
    参数:
        permission_set: 用户权限集合（set）
        active_id: 当前激活的菜单项ID
    
    返回:
        list: 菜单项列表，每个菜单项包含：
            - id: 菜单项ID
            - label: 菜单项标签
            - icon: 菜单项图标
            - url: 菜单项URL（如果有）
            - active: 是否激活
            - children: 子菜单项列表（如果有）
    """
    menu = []
    
    for menu_group in OPPORTUNITY_MANAGEMENT_MENU:
        # 检查父菜单权限
        permission = menu_group.get('permission')
        if permission and not _permission_granted(permission, permission_set):
            continue
        
        # 处理子菜单
        children = []
        for child in menu_group.get('children', []):
            # 检查子菜单权限
            child_permission = child.get('permission')
            if child_permission and not _permission_granted(child_permission, permission_set):
                continue
            
            # 获取URL
            url_name = child.get('url_name')
            url = '#'
            if url_name:
                try:
                    url = reverse(url_name)
                except NoReverseMatch:
                    url = '#'
            
            # 判断是否激活
            is_active = child.get('id') == active_id
            
            children.append({
                'id': child.get('id'),
                'label': child.get('label'),
                'icon': child.get('icon'),
                'url': url,
                'active': is_active,
            })
        
        # 如果父菜单没有可见的子菜单，跳过
        if not children:
            continue
        
        # 判断父菜单是否激活（任意子菜单激活则父菜单激活）
        group_active = any(child.get('id') == active_id for child in menu_group.get('children', []))
        
        menu.append({
            'id': menu_group.get('id'),
            'label': menu_group.get('label'),
            'icon': menu_group.get('icon'),
            'active': group_active,
            'expanded': group_active,  # 如果有激活项，默认展开（与客户管理格式一致）
            'children': children,
        })
    
    return menu


def _build_contract_management_menu(permission_set, active_id=None):
    """
    生成合同管理模块左侧菜单
    
    参数:
        permission_set: 用户权限集合（set）
        active_id: 当前激活的菜单项ID
    
    返回:
        list: 菜单项列表，每个菜单项包含：
            - id: 菜单项ID
            - label: 菜单项标签
            - icon: 菜单项图标
            - url: 菜单项URL（如果有）
            - permission: 所需权限
            - active: 是否激活
            - children: 子菜单项列表（如果有）
    """
    menu = []
    
    for menu_group in CONTRACT_MANAGEMENT_MENU:
        # 检查父菜单权限
        permission = menu_group.get('permission')
        if permission and not _check_customer_permission(permission, permission_set):
            continue
        
        # 处理子菜单
        children = []
        for child in menu_group.get('children', []):
            # 检查子菜单权限（使用_check_customer_permission以支持权限代码规范化）
            child_permission = child.get('permission')
            if child_permission and not _check_customer_permission(child_permission, permission_set):
                continue
            
            # 获取URL
            url_name = child.get('url_name')
            url = '#'
            if url_name:
                try:
                    url = reverse(url_name)
                except NoReverseMatch:
                    url = '#'
            
            # 判断是否激活
            is_active = child.get('id') == active_id
            
            children.append({
                'id': child.get('id'),
                'label': child.get('label'),
                'icon': child.get('icon'),
                'url': url,
                'active': is_active,
            })
        
        # 如果父菜单没有可见的子菜单，跳过
        if not children:
            continue
        
        # 判断父菜单是否激活（任意子菜单激活则父菜单激活）
        group_active = any(child.get('id') == active_id for child in menu_group.get('children', []))
        
        # 获取父菜单URL（如果有url_name，则使用第一个子菜单的URL作为父菜单URL）
        parent_url = '#'
        if menu_group.get('url_name'):
            try:
                parent_url = reverse(menu_group.get('url_name'))
            except NoReverseMatch:
                parent_url = '#'
        elif children:
            # 如果没有设置url_name，使用第一个子菜单的URL
            parent_url = children[0].get('url', '#')
        
        menu.append({
            'id': menu_group.get('id'),
            'label': menu_group.get('label'),
            'icon': menu_group.get('icon'),
            'url': parent_url,
            'active': group_active,
            'expanded': group_active,  # 如果有激活项，默认展开（与计划管理格式一致）
            'children': children,
        })
    
    return menu


def _check_customer_permission(permission_code, permission_set):
    """
    检查客户管理权限（支持新旧权限代码自动映射）
    
    Args:
        permission_code: 权限代码（支持旧代码和新代码）
        permission_set: 用户权限集合
    
    Returns:
        bool: 是否拥有权限
    """
    # 规范化权限代码（自动映射旧代码到新代码）
    normalized_code = normalize_permission_code(permission_code)
    return _permission_granted(normalized_code, permission_set)


def _filter_clients_by_permission(clients, user, permission_set):
    """
    根据用户权限过滤客户列表
    
    权限级别（从高到低）：
    1. view_all: 查看全部客户（总经理）
    2. view_department: 查看本部门客户（部门经理）
    3. view_assigned: 查看本人负责的客户（商务经理）
    4. view: 自动根据权限级别选择
    
    Args:
        clients: 客户查询集
        user: 用户对象
        permission_set: 用户权限集合
    
    Returns:
        过滤后的客户查询集
    """
    if not user or not getattr(user, 'is_authenticated', False):
        return clients.none()
    
    # 超级用户拥有全部权限
    if getattr(user, 'is_superuser', False):
        return clients
    
    # 检查是否有查看全部权限（最高级别）
    if _check_customer_permission('customer_management.client.view_all', permission_set):
        return clients
    
    # 检查是否有查看本部门权限
    if _check_customer_permission('customer_management.client.view_department', permission_set):
        # 获取用户所在部门
        if user.department:
            # 获取部门的所有成员
            from backend.apps.system_management.models import User
            department_users = User.objects.filter(
                department=user.department,
                is_active=True
            )
            # 过滤：负责人是本部门成员的客户 或 创建人是本部门成员的客户
            from django.db.models import Q
            return clients.filter(
                Q(responsible_user__in=department_users) | Q(created_by__in=department_users)
            )
        else:
            # 如果没有部门，降级为查看本人负责的或本人创建的
            from django.db.models import Q
            return clients.filter(
                Q(responsible_user=user) | Q(created_by=user)
            )
    
    # 检查是否有查看本人负责的权限（最低级别）
    # 包括：负责人是自己 或 创建人是自己
    if _check_customer_permission('customer_management.client.view_assigned', permission_set):
        from django.db.models import Q
        return clients.filter(
            Q(responsible_user=user) | Q(created_by=user)
        )
    
    # 检查通用view权限（向后兼容，自动选择最高可用级别）
    if _check_customer_permission('customer_management.client.view', permission_set):
        # 检查用户角色，自动判断权限级别
        # 总经理：查看全部
        if user.roles.filter(code='general_manager').exists():
            return clients
        # 部门经理：查看本部门
        if user.department and user.department.leader == user:
            from backend.apps.system_management.models import User
            from django.db.models import Q
            department_users = User.objects.filter(
                department=user.department,
                is_active=True
            )
            # 部门经理可以看到：本部门成员负责的客户 或 本部门成员创建的客户
            return clients.filter(
                Q(responsible_user__in=department_users) | Q(created_by__in=department_users)
            )
        # 商务经理或其他：查看本人负责的 或 本人创建的
        from django.db.models import Q
        return clients.filter(
            Q(responsible_user=user) | Q(created_by=user)
        )
    
    # 没有权限，返回空查询集
    return clients.none()


def _build_customer_management_menu(permission_set, active_id=None):
    """
    生成客户管理模块左侧菜单
    
    参数:
        permission_set: 用户权限集合（set）
        active_id: 当前激活的菜单项ID
    
    返回:
        list: 菜单项列表，每个菜单项包含：
            - id: 菜单项ID
            - label: 菜单项标签
            - icon: 菜单项图标
            - url: 菜单项URL（如果有）
            - permission: 所需权限
            - active: 是否激活
            - children: 子菜单项列表（如果有）
    """
    menu = []
    
    for menu_group in CUSTOMER_MANAGEMENT_MENU:
        # 检查父菜单权限
        permission = menu_group.get('permission')
        if permission and not _check_customer_permission(permission, permission_set):
            continue
        
        # 处理子菜单
        children = []
        for child in menu_group.get('children', []):
            # 检查子菜单权限（使用_check_customer_permission以支持权限代码规范化）
            child_permission = child.get('permission')
            if child_permission and not _check_customer_permission(child_permission, permission_set):
                continue
            
            # 获取URL
            url_name = child.get('url_name')
            url = '#'
            if url_name:
                try:
                    url = reverse(url_name)
                except NoReverseMatch:
                    url = '#'
            
            # 判断是否激活
            is_active = child.get('id') == active_id
            
            children.append({
                'id': child.get('id'),
                'label': child.get('label'),
                'icon': child.get('icon'),
                'url': url,
                'active': is_active,
            })
        
        # 如果父菜单没有可见的子菜单，跳过
        if not children:
            continue
        
        # 判断父菜单是否激活（任意子菜单激活则父菜单激活）
        group_active = any(child.get('id') == active_id for child in menu_group.get('children', []))
        
        # 获取父菜单URL（如果有url_name，则使用第一个子菜单的URL作为父菜单URL）
        parent_url = '#'
        if menu_group.get('url_name'):
            try:
                parent_url = reverse(menu_group.get('url_name'))
            except NoReverseMatch:
                parent_url = '#'
        elif children:
            # 如果没有设置url_name，使用第一个子菜单的URL
            parent_url = children[0].get('url', '#')
        
        menu.append({
            'id': menu_group.get('id'),
            'label': menu_group.get('label'),
            'icon': menu_group.get('icon'),
            'url': parent_url,
            'active': group_active,
            'expanded': group_active,  # 如果有激活项，默认展开（与计划管理格式一致）
            'children': children,
        })
    
    return menu


# 使用统一的顶部导航菜单生成函数（已从 backend.core.views 导入）


def _context(page_title, page_icon, description, summary_cards=None, sections=None, request=None, active_menu_id=None):
    context = {
        "page_title": page_title,
        "page_icon": page_icon,
        "description": description,
        "summary_cards": summary_cards or [],
        "sections": sections or [],
    }
    
    # 添加顶部导航菜单
    if request and request.user.is_authenticated:
        permission_set = get_user_permission_codes(request.user)
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        
        # 如果是商机管理相关页面，自动生成左侧菜单
        if request.path and '/business/opportunities' in request.path:
            # 根据路径确定激活的菜单项
            active_menu_id = None
            if '/opportunities/evaluation-application' in request.path:
                active_menu_id = 'evaluation_application'
            elif '/opportunities/drawing-evaluation' in request.path:
                active_menu_id = 'drawing_evaluation'
            elif '/opportunities/tech-meeting' in request.path:
                active_menu_id = 'tech_meeting'
            elif '/opportunities/warehouse-list' in request.path or '/opportunities/warehouse-application' in request.path:
                active_menu_id = 'warehouse_list'
            elif '/opportunities/bidding-quotation-application' in request.path:
                active_menu_id = 'bidding_quotation_application'
            elif '/opportunities/bidding-quotation' in request.path:
                active_menu_id = 'bidding_quotation'
            elif '/opportunities/bidding-document-preparation' in request.path:
                active_menu_id = 'bidding_document_preparation'
            elif '/opportunities/bidding-document-submission' in request.path:
                active_menu_id = 'bidding_document_submission'
            elif '/opportunities/business-negotiation' in request.path:
                active_menu_id = 'business_negotiation'
            elif '/opportunities/forecast' in request.path:
                active_menu_id = 'sales_forecast'
            elif '/opportunities/win-loss' in request.path:
                active_menu_id = 'win_loss'
            elif '/opportunities/bid-bond-payment' in request.path:
                active_menu_id = 'bid_bond_payment'
            elif '/opportunities/tender-fee-payment' in request.path:
                active_menu_id = 'tender_fee_payment'
            elif '/opportunities/agency-fee-payment' in request.path:
                active_menu_id = 'tender_agent_fee_payment'
            elif '/opportunities/' in request.path and '/opportunities/create' not in request.path:
                active_menu_id = 'opportunity_list'
            context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id=active_menu_id)
        # 如果是业务委托书或合同管理相关页面，生成合同管理菜单
        elif request.path and ('/business/authorization-letters' in request.path or '/business/authorization-letter-templates' in request.path or '/business/contracts' in request.path):
            # 根据路径确定激活的菜单项
            if active_menu_id is None:
                if '/business/contracts/management' in request.path:
                    active_menu_id = 'contract_management_list'
                elif '/business/contracts/dispute' in request.path:
                    active_menu_id = 'contract_dispute_list'
                elif '/business/contracts/finalize' in request.path:
                    active_menu_id = 'contract_finalize_create' if '/create' in request.path else 'contract_finalize_list'
                elif '/business/contracts/negotiation' in request.path:
                    active_menu_id = 'contract_negotiation_create' if '/create' in request.path else 'contract_negotiation_list'
                elif '/business/contracts/performance' in request.path:
                    active_menu_id = 'contract_performance'
                elif '/business/contracts/expiry-reminder' in request.path:
                    active_menu_id = 'contract_expiry_reminder'
                elif '/business/contracts/payment-reminder' in request.path:
                    active_menu_id = 'contract_payment_reminder'
                elif '/business/contracts/risk-warning' in request.path:
                    active_menu_id = 'contract_risk_warning'
                elif '/business/contracts/create' in request.path:
                    active_menu_id = 'contract_management_list'  # 创建合同页面激活合同管理菜单
                elif '/business/contracts/' in request.path and '/edit' in request.path:
                    active_menu_id = 'contract_management_list'  # 编辑合同页面激活合同管理菜单
                elif '/business/contracts/' in request.path and request.path.count('/') == 3:
                    # 合同详情页（/business/contracts/<id>/）
                    active_menu_id = 'contract_management_list'  # 合同详情页激活合同管理菜单
                elif '/business/authorization-letters' in request.path:
                    active_menu_id = 'authorization_letter_list'
            context['customer_menu'] = _build_contract_management_menu(permission_set, active_id=active_menu_id)
        # 如果是客户管理相关页面，生成客户管理菜单
        elif request.path and '/business/customers' in request.path:
            context['customer_menu'] = _build_customer_management_menu(permission_set, active_id=active_menu_id)
        # 如果是客户管理首页（/business/），生成客户管理菜单
        elif request.path == '/business/' or request.path == '/business':
            context['customer_menu'] = _build_customer_management_menu(permission_set, active_id=None)
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    
    return context


@login_required
def customer_management_home(request):
    """客户管理首页"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        permission_set = get_user_permission_codes(request.user)
        user = request.user
        
        # 检查是否是系统管理员（超级用户或staff）
        is_admin = getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False)
        
        # 统计数据（需要权限检查）
        summary_cards = []
        
        try:
            from datetime import datetime, timedelta
            today = timezone.now().date()
            yesterday = today - timedelta(days=1)
            last_week = today - timedelta(days=7)
            this_month_start = today.replace(day=1)
            last_30_days_start = today - timedelta(days=30)
            
            # 1. 客户总数（管理员或有权限的用户）
            if is_admin or _permission_granted('customer_management.client.view', permission_set):
                total_clients = Client.objects.count()
                
                # 计算昨日和上周的客户数（用于对比）
                clients_yesterday = Client.objects.filter(created_time__date__lte=yesterday).count()
                clients_last_week = Client.objects.filter(created_time__date__lte=last_week).count()
                
                # 计算变化趋势
                change_vs_yesterday = total_clients - clients_yesterday
                change_vs_last_week = total_clients - clients_last_week
                
                hint_parts = []
                if change_vs_yesterday != 0:
                    arrow = '↑' if change_vs_yesterday > 0 else '↓'
                    hint_parts.append(f'较昨日{arrow}{abs(change_vs_yesterday)}')
                if change_vs_last_week != 0:
                    arrow = '↑' if change_vs_last_week > 0 else '↓'
                    hint_parts.append(f'较上周{arrow}{abs(change_vs_last_week)}')
                hint_text = ' · '.join(hint_parts) if hint_parts else '所有客户数量'
                
                try:
                    summary_cards.append({
                        'label': '客户总数',
                        'value': total_clients,
                        'hint': hint_text,
                        'url': reverse('business_pages:customer_list'),
                        'change_vs_yesterday': change_vs_yesterday,
                        'change_vs_last_week': change_vs_last_week,
                    })
                except NoReverseMatch:
                    summary_cards.append({
                        'label': '客户总数',
                        'value': total_clients,
                        'hint': hint_text,
                        'change_vs_yesterday': change_vs_yesterday,
                        'change_vs_last_week': change_vs_last_week,
                    })
            
            # 2. 新增客户数（今日/本月）
            if is_admin or _permission_granted('customer_management.client.view', permission_set):
                new_clients_today = Client.objects.filter(created_time__date=today).count()
                new_clients_month = Client.objects.filter(created_time__gte=this_month_start).count()
                
                try:
                    summary_cards.append({
                        'label': '新增客户数',
                        'value': new_clients_month,
                        'hint': f'今日新增 {new_clients_today} 个',
                        'url': reverse('business_pages:customer_list'),
                    })
                except NoReverseMatch:
                    summary_cards.append({
                        'label': '新增客户数',
                        'value': new_clients_month,
                        'hint': f'今日新增 {new_clients_today} 个',
                    })
            
            # 3. 联系人总数（最近30天有交互或事务记录的客户数量）
            if is_admin or _permission_granted('customer_management.client.view', permission_set):
                # 获取最近30天有交互的客户（通过CustomerRelationship、VisitPlan等）
                from django.db.models import Q
                active_client_ids = set()
                
                # 通过客户关系记录
                try:
                    recent_relationships = CustomerRelationship.objects.filter(
                        created_time__gte=last_30_days_start
                    ).values_list('client_id', flat=True).distinct()
                    active_client_ids.update(recent_relationships)
                except:
                    pass
                
                # 通过拜访计划
                try:
                    recent_visits = VisitPlan.objects.filter(
                        created_time__gte=last_30_days_start
                    ).values_list('client_id', flat=True).distinct()
                    active_client_ids.update(recent_visits)
                except:
                    pass
                
                # 通过商机
                try:
                    recent_opportunities = BusinessOpportunity.objects.filter(
                        created_time__gte=last_30_days_start
                    ).values_list('client_id', flat=True).distinct()
                    active_client_ids.update(recent_opportunities)
                except:
                    pass
                
                active_clients_count = len(active_client_ids)
                total_contacts = ClientContact.objects.count()
                
                try:
                    summary_cards.append({
                        'label': '联系人总数',
                        'value': active_clients_count,
                        'hint': f'最近30天有交互记录的客户数量',
                        'url': reverse('business_pages:customer_list'),
                    })
                except NoReverseMatch:
                    summary_cards.append({
                        'label': '联系人总数',
                        'value': active_clients_count,
                        'hint': f'最近30天有交互记录的客户数量',
                    })
            
            # 4. 新增联系人数（今日/本月）
            if is_admin or _permission_granted('customer_management.client.view', permission_set):
                new_contacts_today = ClientContact.objects.filter(created_time__date=today).count()
                new_contacts_month = ClientContact.objects.filter(created_time__gte=this_month_start).count()
                
                try:
                    summary_cards.append({
                        'label': '新增联系人数',
                        'value': new_contacts_month,
                        'hint': f'今日新增 {new_contacts_today} 个',
                        'url': reverse('business_pages:contact_list'),
                    })
                except NoReverseMatch:
                    summary_cards.append({
                        'label': '新增联系人数',
                        'value': new_contacts_month,
                        'hint': f'今日新增 {new_contacts_today} 个',
                    })
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('获取统计数据失败: %s', str(e))
    
        # 功能模块区域
        sections = []
        
        # 快捷操作区域
        quick_actions = []
        
        if is_admin or _permission_granted('customer_management.client.create', permission_set):
            try:
                quick_actions.append({
                    'label': '创建新客户',
                    'icon': '➕',
                    'description': '添加新客户信息',
                    'url': reverse('business_pages:customer_create'),
                    'link_label': '创建客户 →'
                })
            except NoReverseMatch:
                pass
        
        if is_admin or _permission_granted('customer_management.client.create', permission_set):
            try:
                quick_actions.append({
                    'label': '创建联系人',
                    'icon': '👤',
                    'description': '添加客户联系人',
                    'url': reverse('business_pages:contact_create'),
                    'link_label': '创建联系人 →'
                })
            except NoReverseMatch:
                pass
        
        # 新建联系人拜访
        if is_admin or _permission_granted('customer_management.relationship.create', permission_set):
            try:
                quick_actions.append({
                    'label': '新建联系人拜访',
                    'icon': '📅',
                    'description': '创建新的拜访记录',
                    'url': reverse('business_pages:visit_plan_create'),
                    'link_label': '创建拜访 →'
                })
            except NoReverseMatch:
                pass
        
        # 新建人员关系升级
        if is_admin or _permission_granted('customer_management.relationship.upgrade', permission_set):
            try:
                quick_actions.append({
                    'label': '新建人员关系升级',
                    'icon': '⬆️',
                    'description': '记录人员关系升级',
                    'url': reverse('business_pages:customer_relationship_upgrade_create'),
                    'link_label': '创建升级 →'
                })
            except NoReverseMatch:
                pass
        
        if quick_actions:
            sections.append({
                'title': '快速操作',
                'description': '常用的快速操作入口',
                'items': quick_actions
            })
        
        # 功能模块区域
        modules = []
        
        if is_admin or _permission_granted('customer_management.client.view', permission_set):
            try:
                modules.append({
                    'label': '客户信息管理',
                    'icon': '👥',
                    'description': '管理客户基本信息，查看客户列表和详情',
                    'url': reverse('business_pages:customer_list'),
                    'link_label': '进入模块 →'
                })
            except NoReverseMatch:
                pass
        
        if is_admin or _permission_granted('customer_management.client.view', permission_set):
            try:
                modules.append({
                    'label': '人员关系管理',
                    'icon': '👤',
                    'description': '管理客户联系人信息，维护人员关系',
                    'url': reverse('business_pages:contact_list'),
                    'link_label': '进入模块 →'
                })
            except NoReverseMatch:
                pass
        
        if is_admin or _permission_granted('customer_success.opportunity.view', permission_set):
            try:
                modules.append({
                    'label': '商机管理',
                    'icon': '💼',
                    'description': '管理商机信息，跟踪商机进展',
                    'url': reverse('business_pages:opportunity_management'),
                    'link_label': '进入模块 →'
                })
            except NoReverseMatch:
                pass
        
        if is_admin or _permission_granted('customer_management.contract.view', permission_set):
            try:
                modules.append({
                    'label': '合同管理',
                    'icon': '📄',
                    'description': '管理合同信息，跟踪合同状态',
                    'url': reverse('business_pages:contract_management_list'),
                    'link_label': '进入模块 →'
                })
            except NoReverseMatch:
                pass
        
        if modules:
            sections.append({
                'title': '功能模块',
                'description': '客户管理的各个功能模块入口',
                'items': modules
            })
        
        # 最近动态/提醒
        recent_notices = []
        
        try:
            from datetime import datetime, timedelta
            today = timezone.now().date()
            
            # 逾期拜访提醒
            if is_admin or _permission_granted('customer_management.relationship.view', permission_set):
                try:
                    # VisitPlan使用plan_date字段，status字段可能有不同的值
                    overdue_visits = VisitPlan.objects.filter(
                        plan_date__date__lt=today,
                        status__in=['planned', 'in_progress']
                    ).select_related('client').order_by('plan_date')[:5]
                    
                    for visit in overdue_visits:
                        days_overdue = (today - visit.plan_date.date()).days
                        client_name = visit.client.name if visit.client else "未知客户"
                        plan_title = visit.plan_title or "拜访计划"
                        recent_notices.append({
                            'type': 'warning',
                            'icon': '⚠️',
                            'title': f'逾期拜访提醒',
                            'content': f'{client_name} - {plan_title}，已逾期 {days_overdue} 天',
                            'date': visit.plan_date.date() if hasattr(visit.plan_date, 'date') else visit.plan_date,
                        })
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f'获取逾期拜访提醒失败: {str(e)}')
            
            # 最新反馈内容摘要（通过CustomerRelationship获取）
            if is_admin or _permission_granted('customer_management.relationship.view', permission_set):
                try:
                    recent_feedbacks = CustomerRelationship.objects.filter(
                        content__isnull=False
                    ).exclude(content='').select_related('client', 'created_by', 'followup_person').order_by('-followup_time')[:5]
                    
                    for feedback in recent_feedbacks:
                        feedback_preview = feedback.content[:50] + '...' if len(feedback.content) > 50 else feedback.content
                        recent_notices.append({
                            'type': 'info',
                            'icon': '💬',
                            'title': f'最新反馈 - {feedback.client.name if feedback.client else "未知客户"}',
                            'content': feedback_preview,
                            'date': feedback.followup_time.date() if hasattr(feedback.followup_time, "date") else feedback.followup_time,
                            'author': feedback.created_by.username if feedback.created_by else '',
                        })
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f'获取最新反馈失败: {str(e)}')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('获取最近动态失败: %s', str(e))
        
        # 构建上下文
        context = _context(
            "客户管理",
            "👥",
            "客户管理首页，管理客户信息、联系人、商机等业务数据。",
            summary_cards=summary_cards,
            sections=sections,
            request=request,
        )
        
        # 添加最近动态
        context['recent_notices'] = recent_notices[:10]  # 最多显示10条
        
        return render(request, "customer_management/home.html", context)
    except Exception as e:
        logger.exception('customer_management_home 视图函数执行失败: %s', str(e))
        # 返回一个简单的错误页面，而不是让Django返回500/503错误
        messages.error(request, f'页面加载失败: {str(e)}')
        try:
            # 尝试返回一个基本的上下文
            context = _context(
                "客户管理",
                "👥",
                "客户管理首页",
                summary_cards=[],
                sections=[],
                request=request,
            )
            return render(request, "customer_management/home.html", context)
        except Exception as inner_e:
            logger.exception('渲染错误页面也失败: %s', str(inner_e))
            # 如果连错误页面都渲染不了，重定向到首页
            from django.shortcuts import redirect
            return redirect('home')


def _get_opportunities_safely(queryset, permission_set, user):
    """安全获取商机列表，处理新字段可能不存在的情况"""
    from django.db import connection, transaction
    
    # 检查新字段是否存在
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT column_name FROM information_schema.columns 
                WHERE table_name = 'business_opportunity' AND column_name = 'opportunity_type'
            """)
            has_new_fields = cursor.fetchone() is not None
    except:
        has_new_fields = False
    
    if has_new_fields:
        # 字段存在，使用正常查询
        return queryset
    else:
        # 字段不存在，使用defer排除这些字段
        try:
            return queryset.defer('opportunity_type', 'service_type')
        except:
            # 如果defer也失败，使用values()只获取需要的字段
            return queryset.values('id', 'name', 'client_id', 'client__name', 'business_manager_id')


# ==================== 客户管理模块视图函数（按《客户管理详细设计方案 v1.12》实现）====================

@login_required
def customer_list(request):
    """客户列表"""
    from django.core.paginator import Paginator
    from backend.apps.customer_management.models import Client
    
    # 获取标签页参数
    tab = request.GET.get('tab', 'all')
    
    # 获取筛选参数
    search = request.GET.get('search', '').strip()
    search_field = request.GET.get('search_field', 'name')  # 搜索字段
    client_level = request.GET.get('client_level', '')
    client_type = request.GET.get('client_type', '')
    credit_level = request.GET.get('credit_level', '')
    grade = request.GET.get('grade', '')
    industry = request.GET.get('industry', '')
    source = request.GET.get('source', '')
    legal_risk_level = request.GET.get('legal_risk_level', '')
    is_active = request.GET.get('is_active', '')
    responsible_user_id = request.GET.get('responsible_user', '')
    relationship_stage = request.GET.get('relationship_stage', '')
    department = request.GET.get('department', '')
    region = request.GET.get('region', '')
    date_range = request.GET.get('date_range', '')
    created_time_start = request.GET.get('created_time_start', '')
    created_time_end = request.GET.get('created_time_end', '')
    approval_status = request.GET.get('approval_status', '')  # 审批状态
    company_email = request.GET.get('company_email', '').strip()  # 邮箱
    legal_representative = request.GET.get('legal_representative', '').strip()  # 法定代表人
    page_size = request.GET.get('page_size', '10')
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    can_create = _check_customer_permission('customer_management.client.create', permission_set)
    
    # 检查是否有查看权限
    if not _check_customer_permission('customer_management.client.view', permission_set) and \
       not _check_customer_permission('customer_management.client.view_assigned', permission_set) and \
       not _check_customer_permission('customer_management.client.view_department', permission_set) and \
       not _check_customer_permission('customer_management.client.view_all', permission_set):
        messages.error(request, '您没有查看客户列表的权限')
        return redirect('admin:index')
    
    # 获取客户列表
    try:
        clients = Client.objects.select_related('created_by', 'responsible_user', 'responsible_user__department').prefetch_related('contacts')
        
        # 根据权限过滤客户列表（在标签页筛选之前应用）
        clients = _filter_clients_by_permission(clients, request.user, permission_set)
        
        # 根据标签页应用不同的筛选逻辑
        if tab == 'my_responsible':
            # 我负责的
            clients = clients.filter(responsible_user=request.user)
        elif tab == 'subordinate_responsible':
            # 下属负责的 - 需要获取当前用户的下属
            from backend.apps.system_management.models import User
            # 通过部门关系查找下属：如果用户是部门负责人，则部门成员是下属
            subordinates = User.objects.none()
            if request.user.department and request.user.department.leader == request.user:
                # 用户是部门负责人，获取部门所有成员（不包括自己）
                subordinates = User.objects.filter(
                    department=request.user.department,
                    is_active=True
                ).exclude(id=request.user.id)
            clients = clients.filter(responsible_user__in=subordinates)
        elif tab == 'my_collaboration':
            # 我协作的 - 需要根据协作关系筛选（这里需要根据实际模型调整）
            # 暂时使用联系人关系作为协作关系
            clients = clients.filter(contacts__user=request.user).distinct()
        elif tab == 'subordinate_collaboration':
            # 下属协作的
            from backend.apps.system_management.models import User
            # 通过部门关系查找下属：如果用户是部门负责人，则部门成员是下属
            subordinates = User.objects.none()
            if request.user.department and request.user.department.leader == request.user:
                # 用户是部门负责人，获取部门所有成员（不包括自己）
                subordinates = User.objects.filter(
                    department=request.user.department,
                    is_active=True
                ).exclude(id=request.user.id)
            clients = clients.filter(contacts__user__in=subordinates).distinct()
        elif tab == 'pending_approval':
            # 待审批的 - 需要根据审批状态筛选（这里需要根据实际审批流程调整）
            # 暂时筛选没有负责人的客户作为待审批
            clients = clients.filter(responsible_user__isnull=True)
        # tab == 'all' 时不做额外筛选
        
        # 应用搜索条件
        if search:
            if search_field == 'name':
                clients = clients.filter(name__icontains=search)
            elif search_field == 'phone':
                clients = clients.filter(
                    Q(contacts__phone__icontains=search) |
                    Q(company_phone__icontains=search)
                ).distinct()
            elif search_field == 'wechat':
                clients = clients.filter(contacts__wechat__icontains=search).distinct()
            elif search_field == 'address':
                clients = clients.filter(address__icontains=search)
            elif search_field == 'project_address1':
                clients = clients.filter(project_address1__icontains=search)
            elif search_field == 'project_address2':
                clients = clients.filter(project_address2__icontains=search)
            elif search_field == 'project_address3':
                clients = clients.filter(project_address3__icontains=search)
            elif search_field == 'approval_node':
                # 审批节点搜索（需要根据实际审批流程调整）
                clients = clients.filter(name__icontains=search)  # 临时实现
            else:
                # 默认搜索客户名称和统一信用代码
                clients = clients.filter(
                    Q(name__icontains=search) |
                    Q(unified_credit_code__icontains=search)
                )
        
        # 应用筛选条件
        if client_level:
            clients = clients.filter(client_level=client_level)
        if client_type:
            clients = clients.filter(client_type=client_type)
        if credit_level:
            clients = clients.filter(credit_level=credit_level)
        if grade:
            clients = clients.filter(grade=grade)
        if industry:
            clients = clients.filter(industry__icontains=industry)
        if source:
            clients = clients.filter(source=source)
        if legal_risk_level:
            clients = clients.filter(legal_risk_level=legal_risk_level)
        if is_active != '':
            clients = clients.filter(is_active=(is_active == '1'))
        if responsible_user_id:
            clients = clients.filter(responsible_user_id=responsible_user_id)
        if region:
            clients = clients.filter(region__icontains=region)
        if department:
            # 根据部门筛选（需要根据实际部门字段调整）
            clients = clients.filter(responsible_user__department__icontains=department)
        
        # 邮箱筛选
        if company_email:
            clients = clients.filter(company_email__icontains=company_email)
        
        # 法定代表人筛选
        if legal_representative:
            clients = clients.filter(legal_representative__icontains=legal_representative)
        
        # 审批状态筛选
        if approval_status:
            from django.contrib.contenttypes.models import ContentType
            from backend.apps.workflow_engine.models import ApprovalInstance
            client_content_type = ContentType.objects.get_for_model(Client)
            
            if approval_status == 'no_approval':
                # 无审批记录
                approval_client_ids = ApprovalInstance.objects.filter(
                    content_type=client_content_type
                ).values_list('object_id', flat=True).distinct()
                clients = clients.exclude(id__in=approval_client_ids)
            else:
                # 有审批记录且状态匹配
                approval_client_ids = ApprovalInstance.objects.filter(
                    content_type=client_content_type,
                    status=approval_status
                ).values_list('object_id', flat=True).distinct()
                clients = clients.filter(id__in=approval_client_ids)
        
        # 日期范围筛选
        if date_range:
            from datetime import datetime, timedelta
            today = timezone.now().date()
            
            if date_range == 'today':
                clients = clients.filter(created_time__date=today)
            elif date_range == 'yesterday':
                yesterday = today - timedelta(days=1)
                clients = clients.filter(created_time__date=yesterday)
            elif date_range == 'this_week':
                week_start = today - timedelta(days=today.weekday())
                clients = clients.filter(created_time__date__gte=week_start)
            elif date_range == 'last_week':
                week_start = today - timedelta(days=today.weekday() + 7)
                week_end = today - timedelta(days=today.weekday() + 1)
                clients = clients.filter(created_time__date__gte=week_start, created_time__date__lte=week_end)
            elif date_range == 'this_month':
                month_start = today.replace(day=1)
                clients = clients.filter(created_time__date__gte=month_start)
            elif date_range == 'last_month':
                first_day_this_month = today.replace(day=1)
                last_day_last_month = first_day_this_month - timedelta(days=1)
                first_day_last_month = last_day_last_month.replace(day=1)
                clients = clients.filter(created_time__date__gte=first_day_last_month, created_time__date__lte=last_day_last_month)
            elif date_range == 'custom':
                if created_time_start:
                    clients = clients.filter(created_time__date__gte=created_time_start)
                if created_time_end:
                    clients = clients.filter(created_time__date__lte=created_time_end)
        
        # 按创建时间倒序排列
        clients = clients.order_by('-created_time')
        
        # 分页
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        
        paginator = Paginator(clients, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        # 统计信息
        total_clients = Client.objects.count()
        active_clients = Client.objects.filter(is_active=True).count()
        vip_clients = Client.objects.filter(client_level='vip').count()
        public_sea_clients = Client.objects.filter(responsible_user__isnull=True).count()
        total_contract_amount = Client.objects.aggregate(
            total=Sum('total_contract_amount')
        )['total'] or Decimal('0')
        
        # 重点客户（按合同金额排序，取前5个）
        key_clients = Client.objects.filter(
            total_contract_amount__gt=0
        ).order_by('-total_contract_amount')[:5]
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取客户列表失败: %s', str(e))
        messages.error(request, f'获取客户列表失败：{str(e)}')
        page_obj = None
        total_clients = 0
        active_clients = 0
        vip_clients = 0
        public_sea_clients = 0
        total_contract_amount = Decimal('0')
        key_clients = []
    
    # 统计卡片
    summary_cards = []
    
    context = _context(
        "客户列表",
        "👥",
        "管理所有客户信息，查看客户详情和统计数据。",
        summary_cards=summary_cards,
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='customer_list'
    )
    
    # 获取用户列表（用于高级筛选）
    from backend.apps.system_management.models import User, Department
    users = User.objects.filter(is_active=True).order_by('username')
    
    # 获取部门列表（用于筛选）
    departments = User.objects.filter(
        is_active=True,
        department__isnull=False
    ).values_list('department__name', flat=True).distinct().order_by('department__name')
    
    # 获取审批状态选项
    from backend.apps.workflow_engine.models import ApprovalInstance
    approval_status_choices = [
        ('', '全部'),
        ('no_approval', '无审批'),
        ('draft', '草稿'),
        ('pending', '审批中'),
        ('approved', '已通过'),
        ('rejected', '已驳回'),
        ('withdrawn', '已撤回'),
        ('cancelled', '已取消'),
    ]
    
    context.update({
        'page_obj': page_obj,
        'tab': tab,
        'search': search,
        'search_field': search_field,
        'client_level': client_level,
        'client_type': client_type,
        'credit_level': credit_level,
        'grade': grade,
        'industry': industry,
        'source': source,
        'legal_risk_level': legal_risk_level,
        'is_active': is_active,
        'responsible_user_id': responsible_user_id,
        'relationship_stage': relationship_stage,
        'department': department,
        'region': region,
        'date_range': date_range,
        'created_time_start': created_time_start,
        'created_time_end': created_time_end,
        'approval_status': approval_status,
        'company_email': company_email,
        'legal_representative': legal_representative,
        'key_clients': key_clients,
        'can_create': can_create,
        'users': users,
        'departments': departments,
        'client_level_choices': Client.CLIENT_LEVELS,
        'client_type_choices': [(ct.id, ct.name) for ct in ClientType.objects.filter(is_active=True).order_by('display_order', 'name')],
        'credit_level_choices': Client.CREDIT_LEVELS,
        'source_choices': Client.SOURCE_CHOICES,
        'grade_choices': [(cg.id, cg.name) for cg in ClientGrade.objects.filter(is_active=True).order_by('display_order', 'name')],
        'approval_status_choices': approval_status_choices,
    })
    return render(request, "customer_management/customer_list.html", context)


@login_required
def customer_create(request):
    """创建客户"""
    from backend.apps.customer_management.models import Client, ClientType
    from backend.apps.customer_management.forms import CustomerForm
    
    permission_set = get_user_permission_codes(request.user)
    if not _check_customer_permission('customer_management.client.edit', permission_set):
        messages.error(request, '您没有权限创建客户')
        return redirect('business_pages:customer_list')
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, user=request.user)
        if form.is_valid():
            client = form.save(commit=False)
            client.created_by = request.user
            # 如果表单中没有设置负责人，则默认设置为创建人
            if not client.responsible_user:
                client.responsible_user = request.user
            
            # 确保 client_type 有值（强制检查，避免数据库错误）
            # 检查 client_type 和 client_type_id
            if not hasattr(client, 'client_type') or client.client_type is None or (hasattr(client, 'client_type_id') and client.client_type_id is None):
                # 尝试获取默认的客户类型
                default_client_type = ClientType.objects.filter(is_active=True).order_by('display_order', 'id').first()
                if default_client_type:
                    client.client_type = default_client_type
                else:
                    # 如果没有可用的客户类型，返回错误
                    messages.error(request, '创建失败：请选择客户类型，或联系管理员配置客户类型')
                    context = _context(
                        "创建客户",
                        "➕",
                        "创建新客户信息",
                        request=request,
                    )
                    permission_set = get_user_permission_codes(request.user)
                    context['customer_menu'] = _build_customer_management_menu(
                        permission_set, 
                        active_id='customer_create'
                    )
                    context.update({
                        'form': form,
                        'client_type_choices': [(ct.id, ct.name) for ct in ClientType.objects.filter(is_active=True).order_by('display_order', 'name')],
                        'source_choices': Client.SOURCE_CHOICES,
                    })
                    return render(request, "customer_management/customer_form.html", context)
            
            # 最终检查：确保 client_type 有值（强制设置，避免数据库错误）
            if not client.client_type or client.client_type_id is None:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'客户创建时 client_type 为空，尝试设置默认值。表单数据: {form.cleaned_data}')
                
                # 强制获取一个默认的客户类型
                default_client_type = ClientType.objects.filter(is_active=True).order_by('display_order', 'id').first()
                if default_client_type:
                    client.client_type = default_client_type
                    logger.info(f'已设置默认客户类型: {default_client_type.id} - {default_client_type.name}')
                else:
                    logger.error('没有可用的客户类型，无法创建客户')
                    messages.error(request, '创建失败：客户类型不能为空，请联系管理员配置客户类型')
                    context = _context(
                        "创建客户",
                        "➕",
                        "创建新客户信息",
                        request=request,
                    )
                    permission_set = get_user_permission_codes(request.user)
                    context['customer_menu'] = _build_customer_management_menu(
                        permission_set, 
                        active_id='customer_create'
                    )
                    context.update({
                        'form': form,
                        'client_type_choices': [(ct.id, ct.name) for ct in ClientType.objects.filter(is_active=True).order_by('display_order', 'name')],
                        'source_choices': Client.SOURCE_CHOICES,
                    })
                    return render(request, "customer_management/customer_form.html", context)
            
            # 最后一次检查：确保 client_type_id 不为 None
            # 如果仍然为None，尝试从POST数据中直接获取
            if client.client_type_id is None:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'保存前 client_type_id 仍为 None，尝试从POST数据获取。POST数据: {request.POST.get("client_type")}')
                
                # 尝试从POST数据中获取client_type_id
                post_client_type_id = request.POST.get('client_type')
                if post_client_type_id:
                    try:
                        post_client_type_id = int(post_client_type_id)
                        # 尝试获取该ID的ClientType
                        post_client_type = ClientType.objects.filter(id=post_client_type_id).first()
                        if post_client_type:
                            client.client_type = post_client_type
                            logger.info(f'从POST数据设置 client_type: {post_client_type.id} - {post_client_type.name}')
                        else:
                            logger.warning(f'POST数据中的client_type ID {post_client_type_id} 不存在')
                    except (ValueError, TypeError):
                        logger.error(f'POST数据中的client_type格式错误: {post_client_type_id}')
                
                # 如果仍然为None，使用默认值
                if client.client_type_id is None:
                    logger.error('保存前 client_type_id 仍为 None，强制设置默认值')
                    default_client_type = ClientType.objects.filter(is_active=True).order_by('display_order', 'id').first()
                    if default_client_type:
                        client.client_type_id = default_client_type.id
                        logger.info(f'已设置默认客户类型: {default_client_type.id} - {default_client_type.name}')
                    else:
                        messages.error(request, '创建失败：无法设置客户类型，请联系管理员')
                        return redirect('business_pages:customer_list')
            
            # 确保外键字段被正确设置
            # 如果 client_type 对象存在但 client_type_id 为 None，强制设置
            if client.client_type and client.client_type_id is None:
                client.client_type_id = client.client_type.id
            # 如果 client_type_id 存在但 client_type 对象不存在，确保对象存在
            elif client.client_type_id and not client.client_type:
                try:
                    client.client_type = ClientType.objects.get(id=client.client_type_id)
                except ClientType.DoesNotExist:
                    # 如果 ID 不存在，重置为 None，后续会设置默认值
                    client.client_type_id = None
                    client.client_type = None
            
            # 最终验证：确保 client_type_id 不为 None
            # 如果仍然为 None，强制设置默认值
            if client.client_type_id is None or client.client_type is None:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning('保存前 client_type_id 为 None，强制设置默认值')
                
                # 强制获取并设置默认客户类型
                default_client_type = ClientType.objects.filter(is_active=True).order_by('display_order', 'id').first()
                if default_client_type:
                    client.client_type_id = default_client_type.id
                    client.client_type = default_client_type
                    logger.info(f'已强制设置默认客户类型: {default_client_type.id} - {default_client_type.name}')
                else:
                    logger.critical('没有可用的客户类型，无法创建客户')
                    messages.error(request, '创建失败：客户类型不能为空，请联系管理员配置客户类型')
                    context = _context(
                        "创建客户",
                        "➕",
                        "创建新客户信息",
                        request=request,
                    )
                    permission_set = get_user_permission_codes(request.user)
                    context['customer_menu'] = _build_customer_management_menu(
                        permission_set, 
                        active_id='customer_create'
                    )
                    context.update({
                        'form': form,
                        'client_type_choices': [(ct.id, ct.name) for ct in ClientType.objects.filter(is_active=True).order_by('display_order', 'name')],
                        'source_choices': Client.SOURCE_CHOICES,
                    })
                    return render(request, "customer_management/customer_form.html", context)
            
            # 最后一次验证：确保 client_type_id 不为 None（防止意外情况）
            if client.client_type_id is None:
                import logging
                logger = logging.getLogger(__name__)
                logger.critical('保存前最终验证失败：client_type_id 仍为 None，阻止保存操作')
                messages.error(request, '创建失败：客户类型设置失败，请联系管理员')
                return redirect('business_pages:customer_list')
            
            # 确保 client_type_id 和 client_type 对象一致
            if client.client_type.id != client.client_type_id:
                client.client_type_id = client.client_type.id
            
            client.save()
            
            # 自动启动审批流程
            try:
                from django.contrib.contenttypes.models import ContentType
                from backend.apps.workflow_engine.models import WorkflowTemplate, ApprovalInstance
                from backend.apps.workflow_engine.services import ApprovalEngine
                
                # 检查是否已有正在进行的审批
                content_type = ContentType.objects.get_for_model(Client)
                existing_instance = ApprovalInstance.objects.filter(
                    content_type=content_type,
                    object_id=client.id,
                    status='pending'
                ).first()
                
                if not existing_instance:
                    # 获取客户管理审批流程
                    try:
                        workflow = WorkflowTemplate.objects.get(
                            code='customer_management_approval',
                            status='active'
                        )
                        
                        # 启动审批流程
                        comment = f'申请创建客户：{client.name}（统一信用代码：{client.unified_credit_code or "未填写"}）'
                        instance = ApprovalEngine.start_approval(
                            workflow=workflow,
                            content_object=client,
                            applicant=request.user,
                            comment=comment
                        )
                        
                        messages.success(
                            request, 
                            f'客户创建成功，已自动提交审批（审批编号：{instance.instance_number}）'
                        )
                    except WorkflowTemplate.DoesNotExist:
                        messages.warning(
                            request, 
                            '客户创建成功，但审批流程未配置，请联系管理员配置审批流程'
                        )
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.exception('启动客户审批流程失败: %s', str(e))
                        messages.warning(
                            request, 
                            f'客户创建成功，但启动审批流程失败：{str(e)}，请联系管理员'
                        )
                else:
                    messages.success(
                        request, 
                        f'客户创建成功，该客户已有正在进行的审批流程（审批编号：{existing_instance.instance_number}）'
                    )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.exception('客户创建后处理审批流程时出错: %s', str(e))
                messages.success(request, '客户创建成功')
            
            return redirect('business_pages:customer_detail', client_id=client.id)
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = CustomerForm(user=request.user)
    
    context = _context(
        "创建客户",
        "➕",
        "创建新客户信息",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='customer_create'
    )
    
    context.update({
        'form': form,
        'client_type_choices': [(ct.id, ct.name) for ct in ClientType.objects.filter(is_active=True).order_by('display_order', 'name')],
        'source_choices': Client.SOURCE_CHOICES,
    })
    return render(request, "customer_management/customer_form.html", context)


@login_required
def customer_detail(request, client_id):
    """客户详情"""
    from backend.apps.customer_management.models import Client, ClientContact
    from backend.apps.production_management.models import Project
    
    client = get_object_or_404(Client, id=client_id)
    permission_set = get_user_permission_codes(request.user)
    
    # 获取关联数据
    # 关联项目列表
    try:
        related_projects = Project.objects.filter(client=client).select_related('project_manager', 'service_type').order_by('-created_time')[:10]
    except Exception:
        related_projects = []
    
    # 关联商机列表
    try:
        related_opportunities = BusinessOpportunity.objects.filter(client=client).select_related('business_manager').order_by('-created_time')[:10]
    except Exception:
        related_opportunities = []
    
    # 关联合同列表
    try:
        related_contracts = BusinessContract.objects.filter(client=client).select_related('project').order_by('-created_time')[:10]
    except Exception:
        related_contracts = []
    
    # 关联联系人列表
    try:
        related_contacts = ClientContact.objects.filter(client=client).order_by('-is_primary', '-created_time')[:10]
    except Exception:
        related_contacts = []
    
    # 获取被执行记录
    try:
        from backend.apps.customer_management.models import ExecutionRecord
        execution_records = ExecutionRecord.objects.filter(client=client).order_by('-filing_date', '-created_time')
        execution_count = execution_records.count()
    except Exception:
        execution_records = []
        execution_count = 0
    
    # 获取审批信息
    approval_instance = None
    approval_records = []
    approval_path_nodes = []
    can_submit_approval = False
    try:
        from django.contrib.contenttypes.models import ContentType
        from backend.apps.workflow_engine.models import ApprovalInstance, ApprovalRecord
        from collections import defaultdict
        
        content_type = ContentType.objects.get_for_model(Client)
        approval_instance = ApprovalInstance.objects.filter(
            content_type=content_type,
            object_id=client.id
        ).select_related('workflow', 'applicant', 'current_node').prefetch_related(
            'workflow__nodes'
        ).order_by('-created_time').first()
        
        if approval_instance:
            approval_records = ApprovalRecord.objects.filter(
                instance=approval_instance
            ).select_related('node', 'approver', 'transferred_to').order_by('node__sequence', 'approval_time')
            
            # 准备审批路径数据
            workflow_nodes = approval_instance.workflow.nodes.all().order_by('sequence')
            records_by_node = defaultdict(list)
            for record in approval_records:
                records_by_node[record.node_id].append(record)
            
            # 构建审批路径节点列表
            for node in workflow_nodes:
                node_records = records_by_node.get(node.id, [])
                node_status = 'pending'  # 默认待审批
                
                # 判断节点状态
                if any(r.result == 'approved' for r in node_records):
                    node_status = 'approved'
                elif any(r.result == 'rejected' for r in node_records):
                    node_status = 'rejected'
                elif any(r.result == 'pending' for r in node_records):
                    node_status = 'pending'
                elif node.node_type in ['start', 'end']:
                    node_status = 'completed'
                else:
                    # 检查是否是当前节点
                    if approval_instance.current_node and approval_instance.current_node.id == node.id:
                        node_status = 'current'
                    else:
                        node_status = 'waiting'
                
                approval_path_nodes.append({
                    'node': node,
                    'records': node_records,
                    'status': node_status,
                    'is_current': approval_instance.current_node and approval_instance.current_node.id == node.id,
                })
        
        # 检查是否可以提交审批（有权限且没有正在进行的审批）
        can_submit_approval = _check_customer_permission('customer_management.client.approve', permission_set) and not approval_instance
    except Exception:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取审批信息失败')
        pass
    
    context = _context(
        f"客户详情 - {client.name}",
        "👤",
        f"{client.name}",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='customer_list'
    )
    
    can_manage = _check_customer_permission('customer_management.client.edit', permission_set)
    context.update({
        'client': client,
        'can_edit': can_manage,
        'can_manage': can_manage,
        'projects': related_projects,
        'opportunities': related_opportunities,
        'contracts': related_contracts,
        'contacts': related_contacts,
        'execution_records': execution_records,
        'execution_count': execution_count,
        'total_execution_amount': client.total_execution_amount or 0,
        'approval_instance': approval_instance,
        'approval_records': approval_records,
        'approval_path_nodes': approval_path_nodes,
        'can_submit_approval': can_submit_approval,
    })
    return render(request, "customer_management/customer_detail.html", context)


@login_required
def execution_records_export(request, client_id):
    """导出被执行记录"""
    from backend.apps.customer_management.models import Client, ExecutionRecord
    from django.http import HttpResponse
    import csv
    from django.utils import timezone
    
    client = get_object_or_404(Client, id=client_id)
    permission_set = get_user_permission_codes(request.user)
    
    if not _check_customer_permission('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限导出被执行记录')
        return redirect('business_pages:customer_detail', client_id=client_id)
    
    try:
        records = ExecutionRecord.objects.filter(client=client).order_by('-filing_date', '-created_time')
        export_format = request.GET.get('format', 'xlsx')
        
        # 导出为CSV
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = f'attachment; filename="execution_records_{client.name}_{timezone.now().strftime("%Y%m%d")}.csv"'
            
            writer = csv.writer(response)
            # 写入表头
            writer.writerow(['案号', '执行状态', '执行法院', '立案日期', '执行金额', '数据来源', '创建时间'])
            
            # 写入数据
            for record in records:
                writer.writerow([
                    record.case_number or '',
                    record.get_execution_status_display(),
                    record.execution_court or '',
                    record.filing_date.strftime('%Y-%m-%d') if record.filing_date else '',
                    str(record.execution_amount) if record.execution_amount else '0',
                    record.get_source_display(),
                    record.created_time.strftime('%Y-%m-%d %H:%M:%S') if record.created_time else ''
                ])
            
            return response
        
        # 导出为Excel（需要openpyxl库）
        elif export_format == 'xlsx':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = '被执行记录'
                
                # 设置表头
                headers = ['案号', '执行状态', '执行法院', '立案日期', '执行金额', '数据来源', '创建时间']
                ws.append(headers)
                
                # 设置表头样式
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 写入数据
                for record in records:
                    ws.append([
                        record.case_number or '',
                        record.get_execution_status_display(),
                        record.execution_court or '',
                        record.filing_date.strftime('%Y-%m-%d') if record.filing_date else '',
                        str(record.execution_amount) if record.execution_amount else '0',
                        record.get_source_display(),
                        record.created_time.strftime('%Y-%m-%d %H:%M:%S') if record.created_time else ''
                    ])
                
                # 调整列宽
                column_widths = [25, 15, 25, 12, 15, 15, 20]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[chr(64 + i)].width = width
                
                # 设置文本对齐
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f'execution_records_{client.name}_{timezone.now().strftime("%Y%m%d")}.xlsx'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                wb.save(response)
                return response
                
            except ImportError:
                # 如果没有openpyxl，返回CSV格式
                messages.warning(request, 'Excel导出功能需要安装openpyxl库，已改为CSV格式导出')
                return execution_records_export(request, client_id)
        
        else:
            messages.error(request, '不支持的导出格式')
            return redirect('business_pages:customer_detail', client_id=client_id)
            
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('导出被执行记录失败: %s', str(e))
        messages.error(request, f'导出失败：{str(e)}')
        return redirect('business_pages:customer_detail', client_id=client_id)


@login_required
def customer_edit(request, client_id):
    """编辑客户"""
    from backend.apps.customer_management.models import Client, ExecutionRecord
    from backend.apps.customer_management.forms import CustomerForm
    
    client = get_object_or_404(Client, id=client_id)
    permission_set = get_user_permission_codes(request.user)
    
    if not _check_customer_permission('customer_management.client.edit', permission_set):
        messages.error(request, '您没有权限编辑客户')
        return redirect('business_pages:customer_detail', client_id=client_id)
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=client, user=request.user)
        if form.is_valid():
            updated_client = form.save(commit=False)
            
            # 确保 client_type 有值（表单验证应该已经处理，但为了保险起见再次检查）
            # ClientType 已在文件顶部导入，不需要再次导入
            if not updated_client.client_type or updated_client.client_type_id is None:
                default_client_type = ClientType.objects.filter(is_active=True).order_by('display_order', 'id').first()
                if default_client_type:
                    updated_client.client_type = default_client_type
                else:
                    messages.error(request, '更新失败：客户类型不能为空，请联系管理员配置客户类型')
                    return redirect('business_pages:customer_detail', client_id=client_id)
            
            updated_client.save()
            messages.success(request, '客户信息已更新')
            return redirect('business_pages:customer_detail', client_id=client.id)
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = CustomerForm(instance=client, user=request.user)
    
    # 获取被执行记录
    from backend.apps.customer_management.models import ExecutionRecord
    execution_records = ExecutionRecord.objects.filter(
        client=client
    ).order_by('-filing_date', '-created_time')
    
    context = _context(
        f"编辑客户 - {client.name}",
        "✏️",
        f"{client.name}",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='customer_list'
    )
    
    context.update({
        'client': client,
        'form': form,
        'grade_choices': [(cg.id, cg.name) for cg in ClientGrade.objects.filter(is_active=True).order_by('display_order', 'name')],
        'client_type_choices': [(ct.id, ct.name) for ct in ClientType.objects.filter(is_active=True).order_by('display_order', 'name')],
        'source_choices': Client.SOURCE_CHOICES,
        'execution_records': execution_records,
        'execution_count': execution_records.count(),
        'total_execution_amount': client.total_execution_amount or 0,
    })
    return render(request, "customer_management/customer_form.html", context)


@login_required
def customer_delete(request, client_id):
    """删除客户"""
    from backend.apps.customer_management.models import Client, ClientContact
    
    client = get_object_or_404(Client, id=client_id)
    permission_set = get_user_permission_codes(request.user)
    
    if not _check_customer_permission('customer_management.client.edit', permission_set):
        messages.error(request, '您没有权限删除客户')
        return redirect('business_pages:customer_detail', client_id=client_id)

    # 检查关联关系
    related_projects_count = 0
    related_opportunities_count = 0
    related_contracts_count = 0
    related_contacts_count = 0
    
    try:
        from backend.apps.production_management.models import Project
        related_projects_count = Project.objects.filter(client=client).count()
    except Exception:
        pass
    
    try:
        related_opportunities_count = BusinessOpportunity.objects.filter(client=client).count()
    except Exception:
        pass
    
    try:
        related_contracts_count = BusinessContract.objects.filter(client=client).count()
    except Exception:
        pass
    
    try:
        related_contacts_count = ClientContact.objects.filter(client=client).count()
    except Exception:
        pass
    
    has_relations = (related_projects_count > 0 or related_opportunities_count > 0 or 
                     related_contracts_count > 0 or related_contacts_count > 0)
    
    if request.method == 'POST':
        if has_relations:
            messages.error(request, f'无法删除客户：该客户关联了 {related_projects_count} 个项目、{related_opportunities_count} 个商机、{related_contracts_count} 个合同、{related_contacts_count} 个联系人。请先解除关联关系。')
            return redirect('business_pages:customer_detail', client_id=client_id)
        
        try:
            client.delete()
            messages.success(request, '客户已删除')
            return redirect('business_pages:customer_list')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除客户失败: %s', str(e))
            messages.error(request, f'删除客户失败：{str(e)}')
            return redirect('business_pages:customer_detail', client_id=client_id)
    
    # GET 请求，显示确认删除页面
    context = _context(
        "删除客户",
        "🗑️",
        f"确认删除客户：{client.name}",
        request=request,
    )
    context.update({
        'client': client,
        'related_projects_count': related_projects_count,
        'related_opportunities_count': related_opportunities_count,
        'related_contracts_count': related_contracts_count,
        'related_contacts_count': related_contacts_count,
        'has_relations': has_relations,
    })
    return render(request, "customer_management/customer_delete.html", context)
@login_required
def customer_batch_delete(request):
    """批量删除客户"""
    from backend.apps.customer_management.models import Client
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '仅支持POST请求'}, status=405)
    
    permission_set = get_user_permission_codes(request.user)
    if not _check_customer_permission('customer_management.client.edit', permission_set):
        return JsonResponse({'success': False, 'message': '您没有权限删除客户'}, status=403)
    
    try:
        # 获取参数
        client_ids_str = request.POST.get('client_ids', '')
        
        if not client_ids_str:
            return JsonResponse({'success': False, 'message': '请选择要删除的客户'}, status=400)
        
        # 解析客户ID列表
        client_ids = [int(id.strip()) for id in client_ids_str.split(',') if id.strip()]
        
        if not client_ids:
            return JsonResponse({'success': False, 'message': '无效的客户ID列表'}, status=400)
        
        # 批量删除（检查关联关系）
        clients = Client.objects.filter(id__in=client_ids)
        deleted_count = 0
        failed_clients = []
        
        for client in clients:
            # 检查关联关系
            has_relations = False
            try:
                from backend.apps.production_management.models import Project
                if Project.objects.filter(client=client).exists():
                    has_relations = True
            except Exception:
                pass
            
            try:
                if BusinessOpportunity.objects.filter(client=client).exists():
                    has_relations = True
            except Exception:
                pass
            
            try:
                from backend.apps.production_management.models import BusinessContract
                if BusinessContract.objects.filter(client=client).exists():
                    has_relations = True
            except Exception:
                pass
            
            try:
                from backend.apps.customer_management.models import ClientContact
                if ClientContact.objects.filter(client=client).exists():
                    has_relations = True
            except Exception:
                pass
            
            if has_relations:
                failed_clients.append(client.name)
                continue
            
            try:
                client.delete()
                deleted_count += 1
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.exception('删除客户失败: %s', str(e))
                failed_clients.append(client.name)
        
        message = f'成功删除 {deleted_count} 个客户'
        if failed_clients:
            message += f'，{len(failed_clients)} 个客户删除失败（存在关联关系）'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'failed_count': len(failed_clients),
            'failed_clients': failed_clients
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('批量删除客户失败: %s', str(e))
        return JsonResponse({'success': False, 'message': f'批量删除失败：{str(e)}'}, status=500)


@login_required
def customer_export(request):
    """导出客户数据"""
    from backend.apps.customer_management.models import Client
    from django.http import HttpResponse
    import csv
    import json
    
    permission_set = get_user_permission_codes(request.user)
    if not _check_customer_permission('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限导出客户数据')
        return redirect('business_pages:customer_list')
    
    try:
        # 获取筛选参数
        client_ids_str = request.GET.get('ids', '')
        export_format = request.GET.get('format', 'xlsx')
        
        # 获取客户列表
        if client_ids_str:
            client_ids = [int(id.strip()) for id in client_ids_str.split(',') if id.strip()]
            clients = Client.objects.filter(id__in=client_ids).select_related('responsible_user', 'created_by')
        else:
            # 如果没有指定ID，使用当前筛选条件
            clients = Client.objects.select_related('responsible_user', 'created_by').all()
            
            # 应用筛选条件（简化版，可以根据需要扩展）
            search = request.GET.get('search', '').strip()
            if search:
                clients = clients.filter(name__icontains=search)
        
        # 导出为CSV
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = 'attachment; filename="customers.csv"'
            
            writer = csv.writer(response)
            # 写入表头
            writer.writerow(['客户名称', '统一信用代码', '客户等级', '客户类型', '信用等级', '负责人', '创建时间'])
            
            # 写入数据
            for client in clients:
                writer.writerow([
                    client.name,
                    client.unified_credit_code or '',
                    client.get_client_level_display(),
                    client.get_client_type_display() if client.client_type else '',
                    client.get_credit_level_display(),
                    client.responsible_user.get_full_name() if client.responsible_user else '公海',
                    client.created_time.strftime('%Y-%m-%d %H:%M:%S') if client.created_time else ''
                ])
            
            return response
        
        # 导出为Excel（需要openpyxl库）
        elif export_format == 'xlsx':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = '客户列表'
                
                # 设置表头
                headers = ['客户名称', '统一信用代码', '客户等级', '客户类型', '信用等级', '负责人', '创建时间']
                ws.append(headers)
                
                # 设置表头样式
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # 写入数据
                for client in clients:
                    ws.append([
                        client.name,
                        client.unified_credit_code or '',
                        client.get_client_level_display(),
                        client.get_client_type_display() if client.client_type else '',
                        client.get_credit_level_display(),
                        client.responsible_user.get_full_name() if client.responsible_user else '公海',
                        client.created_time.strftime('%Y-%m-%d %H:%M:%S') if client.created_time else ''
                    ])
                
                # 调整列宽
                column_widths = [20, 20, 15, 15, 15, 15, 20]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[chr(64 + i)].width = width
                
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'
                wb.save(response)
                return response
                
            except ImportError:
                # 如果没有openpyxl，返回CSV格式
                messages.warning(request, 'Excel导出功能需要安装openpyxl库，已改为CSV格式导出')
                return customer_export(request)
        
        else:
            messages.error(request, '不支持的导出格式')
            return redirect('business_pages:customer_list')
            
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('导出客户数据失败: %s', str(e))
        messages.error(request, f'导出失败：{str(e)}')
        return redirect('business_pages:customer_list')


@login_required
def customer_submit_approval(request, client_id):
    """提交客户审批"""
    from backend.apps.customer_management.models import Client
    
    client = get_object_or_404(Client, id=client_id)
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _check_customer_permission('customer_management.client.edit', permission_set):
        messages.error(request, '您没有权限提交客户审批')
        return redirect('business_pages:customer_detail', client_id=client_id)
    
    if request.method == 'POST':
        try:
            from django.contrib.contenttypes.models import ContentType
            from backend.apps.workflow_engine.models import WorkflowTemplate, ApprovalInstance
            from backend.apps.workflow_engine.services import ApprovalEngine
            
            # 检查是否已有正在进行的审批
            content_type = ContentType.objects.get_for_model(Client)
            existing_instance = ApprovalInstance.objects.filter(
                content_type=content_type,
                object_id=client.id,
                status__in=['pending', 'in_progress']
            ).first()
            
            if existing_instance:
                messages.warning(request, f'该客户已有正在进行的审批流程（审批编号：{existing_instance.instance_number}）')
                return redirect('business_pages:customer_detail', client_id=client_id)
            
            # 获取客户管理审批流程
            try:
                workflow = WorkflowTemplate.objects.get(
                    code='customer_management_approval',
                    status='active'
                )
            except WorkflowTemplate.DoesNotExist:
                messages.error(request, '客户管理审批流程未配置，请联系管理员')
                return redirect('business_pages:customer_detail', client_id=client_id)
            
            # 启动审批流程
            comment = request.POST.get('comment', f'申请审批客户：{client.name}')
            instance = ApprovalEngine.start_approval(
                workflow=workflow,
                content_object=client,
                applicant=request.user,
                comment=comment
            )
            
            messages.success(request, f'客户审批已提交（审批编号：{instance.instance_number}）')
            return redirect('business_pages:customer_detail', client_id=client_id)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('提交客户审批失败: %s', str(e))
            messages.error(request, f'提交客户审批失败：{str(e)}')
            return redirect('business_pages:customer_detail', client_id=client_id)
    
    # GET 请求，显示提交审批确认页面
    from django.contrib.contenttypes.models import ContentType
    from backend.apps.workflow_engine.models import ApprovalInstance
    
    # 检查是否已有正在进行的审批
    content_type = ContentType.objects.get_for_model(Client)
    existing_instance = ApprovalInstance.objects.filter(
        content_type=content_type,
        object_id=client.id,
        status__in=['pending', 'in_progress']
    ).first()
    
    context = _context(
        "提交客户审批",
        "📋",
        f"提交客户 {client.name} 进行审批",
        request=request,
    )
    context.update({
        'client': client,
        'existing_instance': existing_instance,
    })
    return render(request, "customer_management/customer_submit_approval.html", context)


@login_required
def customer_public_sea(request):
    """客户公海"""
    from django.core.paginator import Paginator
    from backend.apps.customer_management.models import Client
    
    # 获取搜索参数（保留搜索功能）
    search = request.GET.get('search', '').strip()
    # 筛选参数将通过新的筛选模块处理，这里不再单独获取
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    can_claim = _check_customer_permission('customer_management.public_sea.claim', permission_set)
    
    # 获取公海客户列表（responsible_user为空）
    try:
        clients = Client.objects.filter(responsible_user__isnull=True).select_related('created_by')
        
        # 应用搜索条件
        if search:
            clients = clients.filter(
                Q(name__icontains=search) |
                Q(unified_credit_code__icontains=search)
            )
        
        # 筛选条件将通过新的筛选模块处理，这里使用通用方式获取所有GET参数
        # 支持通过GET参数进行筛选（由前端筛选模块提交）
        filter_params = {}
        for key, value in request.GET.items():
            if key not in ['search', 'page', 'page_size', 'tab'] and value:
                filter_params[key] = value
        
        # 应用筛选条件
        if filter_params.get('public_sea_reason'):
            clients = clients.filter(public_sea_reason=filter_params['public_sea_reason'])
        if filter_params.get('client_level'):
            clients = clients.filter(client_level=filter_params['client_level'])
        if filter_params.get('industry'):
            clients = clients.filter(industry__icontains=filter_params['industry'])
        if filter_params.get('region'):
            clients = clients.filter(region__icontains=filter_params['region'])
        
        # 按进入公海时间倒序排列
        clients = clients.order_by('-public_sea_entry_time', '-created_time')
        
        # 分页
        paginator = Paginator(clients, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        # 统计信息
        total_public_sea = Client.objects.filter(responsible_user__isnull=True).count()
        unassigned_count = Client.objects.filter(responsible_user__isnull=True, public_sea_reason='unassigned').count()
        released_count = Client.objects.filter(responsible_user__isnull=True, public_sea_reason='released').count()
        auto_entry_count = Client.objects.filter(responsible_user__isnull=True, public_sea_reason='auto_entry').count()
        
        # 今日认领数
        from datetime import date, timedelta
        today = date.today()
        today_claimed = Client.objects.filter(
            responsible_user__isnull=False,
            public_sea_entry_time__isnull=True,
            updated_time__date=today
        ).count()
        
        # 本周认领数
        week_start = today - timedelta(days=today.weekday())
        week_claimed = Client.objects.filter(
            responsible_user__isnull=False,
            public_sea_entry_time__isnull=True,
            updated_time__date__gte=week_start
        ).count()
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取公海客户列表失败: %s', str(e))
        messages.error(request, f'获取公海客户列表失败：{str(e)}')
        page_obj = None
        total_public_sea = 0
        unassigned_count = 0
        released_count = 0
        auto_entry_count = 0
        today_claimed = 0
        week_claimed = 0
    
    # 统计卡片
    summary_cards = []
    
    context = _context(
        "客户公海",
        "🌊",
        "查看和认领公海客户",
        summary_cards=summary_cards,
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='customer_public_sea'
    )
    
    # 获取筛选参数（用于前端显示当前筛选状态）
    public_sea_reason = request.GET.get('public_sea_reason', '')
    client_level = request.GET.get('client_level', '')
    industry = request.GET.get('industry', '')
    region = request.GET.get('region', '')
    
    context.update({
        'page_obj': page_obj,
        'search': search,
        'public_sea_reason': public_sea_reason,
        'client_level': client_level,
        'industry': industry,
        'region': region,
        'can_claim': can_claim,
        'public_sea_reason_choices': Client.PUBLIC_SEA_REASON_CHOICES,
        'client_level_choices': Client.CLIENT_LEVELS,
    })
    return render(request, "customer_management/customer_public_sea.html", context)


@login_required
def customer_public_sea_claim(request, client_id):
    """认领公海客户"""
    from backend.apps.customer_management.models import Client
    
    client = get_object_or_404(Client, id=client_id)
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _check_customer_permission('customer_management.public_sea.view', permission_set):
        messages.error(request, '您没有权限认领公海客户')
        return redirect('business_pages:customer_public_sea')
    
    # 检查是否在公海
    if client.responsible_user is not None:
        messages.warning(request, '该客户不在公海，无法认领')
        return redirect('business_pages:customer_public_sea')
    
    if request.method == 'POST':
        try:
            # 认领客户
            client.claim_from_public_sea(request.user)
            messages.success(request, f'成功认领客户：{client.name}')
            return redirect('business_pages:customer_detail', client_id=client.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('认领公海客户失败: %s', str(e))
            messages.error(request, f'认领公海客户失败：{str(e)}')
            return redirect('business_pages:customer_public_sea')
    
    # GET 请求，显示认领确认页面
    context = _context(
        "认领公海客户",
        "✅",
        f"确认认领客户：{client.name}",
        request=request,
    )
    context.update({
        'client': client,
    })
    return render(request, "customer_management/customer_public_sea_claim.html", context)


# ==================== 人员关系管理视图函数 =====================

@login_required
def contact_list(request):
    """创建联系人信息"""
    from django.core.paginator import Paginator
    from backend.apps.customer_management.models import ClientContact, Client
    
    # 获取搜索参数（保留搜索功能）
    search = request.GET.get('search', '').strip()
    # 筛选参数将通过新的筛选模块处理，这里使用通用方式获取所有GET参数
    filter_params = {}
    for key, value in request.GET.items():
        if key not in ['search', 'page', 'page_size', 'tab'] and value:
            filter_params[key] = value
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    can_create = _check_customer_permission('customer_management.contact.create', permission_set)
    
    # 获取联系人列表
    try:
        contacts = ClientContact.objects.select_related('client', 'created_by')
        
        # 应用搜索条件
        if search:
            contacts = contacts.filter(
                Q(name__icontains=search) |
                Q(phone__icontains=search) |
                Q(email__icontains=search) |
                Q(client__name__icontains=search)
            )
        
        # 应用筛选条件
        if filter_params.get('client'):
            contacts = contacts.filter(client_id=filter_params['client'])
        if filter_params.get('role'):
            contacts = contacts.filter(role=filter_params['role'])
        if filter_params.get('relationship_level'):
            contacts = contacts.filter(relationship_level=filter_params['relationship_level'])
        
        # 按创建时间倒序排列
        contacts = contacts.order_by('-is_primary', '-created_time')
        
        # 分页
        paginator = Paginator(contacts, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取联系人列表失败: %s', str(e))
        messages.error(request, f'获取联系人列表失败：{str(e)}')
        page_obj = None
    
    # 获取客户列表（用于筛选）
    clients = Client.objects.all().order_by('name')
    
    context = _context(
        "创建联系人信息",
        "📇",
        "管理所有客户联系人信息",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='contact_list'
    )
    
    context.update({
        'page_obj': page_obj,
        'search': search,
        'client_id': filter_params.get('client', ''),
        'role': filter_params.get('role', ''),
        'relationship_level': filter_params.get('relationship_level', ''),
        'clients': clients,
        'can_create': can_create,
        'role_choices': ClientContact.ROLE_CHOICES,
        'relationship_level_choices': ClientContact.RELATIONSHIP_LEVEL_CHOICES,
    })
    return render(request, "customer_management/contact_list.html", context)


@login_required
def contact_create(request):
    """创建联系人信息"""
    from backend.apps.customer_management.models import ClientContact, Client, ContactCareer, ContactColleague
    from backend.apps.customer_management.forms import ContactForm, ContactCareerFormSet, ContactEducationFormSet
    
    permission_set = get_user_permission_codes(request.user)
    if not _check_customer_permission('customer_management.contact.edit', permission_set):
        messages.error(request, '您没有权限创建联系人')
        return redirect('business_pages:contact_list')
    
    if request.method == 'POST':
        # 判断是保存草稿还是提交
        action = request.POST.get('action', 'submit')
        is_draft = (action == 'save_draft')
        
        form = ContactForm(request.POST, request.FILES, is_draft=is_draft)
        career_formset = ContactCareerFormSet(request.POST, prefix='careers')
        education_formset = ContactEducationFormSet(request.POST, prefix='educations')
        
        # 如果是保存草稿，修改表单集的验证规则
        if is_draft:
            career_formset.min_num = 0
            career_formset.validate_min = False
            education_formset.min_num = 0
            education_formset.validate_min = False
            # 为表单集中的每个表单设置is_draft
            for form_item in career_formset.forms:
                form_item.is_draft = True
            for form_item in education_formset.forms:
                form_item.is_draft = True
        
        # 保存草稿时，直接保存，不进行验证
        # 提交时，必须通过验证才能保存
        if is_draft:
            # 保存草稿：直接保存，不检查验证结果
            # 但需要确保数据能正确保存，所以先尝试清理数据
            form.full_clean()
            # 清除所有验证错误，允许保存
            form.errors.clear()
            # 表单集也清除错误
            for form_item in career_formset.forms:
                form_item.errors.clear()
            for form_item in education_formset.forms:
                form_item.errors.clear()
            # 标记表单和表单集为有效，以便保存
            form._errors = {}
            career_formset._errors = {}
            education_formset._errors = {}
            can_save = True
        else:
            # 提交：必须通过验证
            can_save = form.is_valid() and career_formset.is_valid() and education_formset.is_valid()
        
        if can_save:
            contact = form.save(commit=False)
            contact.created_by = request.user
            if is_draft:
                contact.approval_status = 'draft'  # 保存草稿状态
            else:
                contact.approval_status = 'pending'  # 设置为待审批状态
            
            # 处理多选字段
            contact.preferred_contact_methods = form.cleaned_data.get('preferred_contact_methods', [])
            contact.tags = form.cleaned_data.get('tags', [])
            
            # 处理简历文件上传时间
            if form.cleaned_data.get('resume_file'):
                from django.utils import timezone
                contact.resume_upload_time = timezone.now()
            
            contact.save()
            form.save_m2m()  # 保存多对多关系
            
            # 保存职业信息
            careers = career_formset.save(commit=False)
            saved_careers = []
            for idx, career in enumerate(careers):
                career.contact = contact
                career.save()
                saved_careers.append((idx, career))
            
            # 删除标记为删除的职业记录
            for career in career_formset.deleted_objects:
                # 删除职业信息时，同时删除关联的同事关系人员
                career.colleagues.all().delete()
                career.delete()
            
            # 保存同事关系人员
            # 遍历所有保存的职业信息，查找对应的同事关系人员数据
            for form_idx, career in saved_careers:
                # 查找该职业信息对应的同事关系人员数据
                # 数据格式：careers-{form_idx}-colleagues-{j}-{field}
                colleague_prefix = f'careers-{form_idx}-colleagues'
                
                # 获取所有以该前缀开头的字段
                colleague_data = {}
                for key, value in request.POST.items():
                    if key.startswith(colleague_prefix + '-'):
                        # 解析字段名：careers-0-colleagues-0-name -> (0, name)
                        remaining = key[len(colleague_prefix) + 1:]  # 去掉前缀和连字符
                        parts = remaining.split('-')
                        if len(parts) >= 2:
                            try:
                                colleague_index = int(parts[0])
                                field_name = '-'.join(parts[1:])
                                
                                if colleague_index not in colleague_data:
                                    colleague_data[colleague_index] = {}
                                colleague_data[colleague_index][field_name] = value
                            except (ValueError, IndexError):
                                continue
                
                # 删除该职业信息的所有现有同事关系人员
                career.colleagues.all().delete()
                
                # 保存新的同事关系人员
                for colleague_index, colleague_fields in colleague_data.items():
                    # 检查是否有DELETE标记
                    delete_key = f'{colleague_prefix}-{colleague_index}-DELETE'
                    if request.POST.get(delete_key) == 'on':
                        continue
                    
                    # 检查必填字段
                    if not colleague_fields.get('name'):
                        continue
                    
                    ContactColleague.objects.create(
                        career=career,
                        department=colleague_fields.get('department', ''),
                        name=colleague_fields.get('name', ''),
                        position=colleague_fields.get('position', ''),
                        phone=colleague_fields.get('phone', '')
                    )
            
            # 只有提交时才启动审批流程，保存草稿时不启动
            if not is_draft:
                try:
                    from backend.apps.workflow_engine.models import WorkflowTemplate
                    from backend.apps.workflow_engine.services import ApprovalEngine
                    
                    workflow = WorkflowTemplate.objects.get(
                        code='contact_approval',
                        status='active'
                    )
                    
                    ApprovalEngine.start_approval(
                        workflow=workflow,
                        content_object=contact,
                        applicant=request.user,
                        comment=f'申请创建联系人：{contact.name}（客户：{contact.client.name}）'
                    )
                    
                    messages.success(request, f'联系人创建成功，已提交审批。联系人：{contact.name}')
                except WorkflowTemplate.DoesNotExist:
                    messages.warning(request, '联系人创建成功，但审批流程未配置，请联系管理员')
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('启动审批流程失败: %s', str(e))
                    messages.warning(request, f'联系人创建成功，但启动审批流程失败：{str(e)}')
            else:
                messages.success(request, f'联系人草稿保存成功。联系人：{contact.name}')
            
            return redirect('business_pages:contact_list')
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = ContactForm()
        career_formset = ContactCareerFormSet(prefix='careers')
        education_formset = ContactEducationFormSet(prefix='educations')
    
    # 获取客户列表（用于下拉选择）
    clients = Client.objects.all().order_by('name')
    
    context = _context(
        "创建联系人信息",
        "➕",
        "创建新联系人信息",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='contact_create'
    )
    
    context.update({
        'form': form,
        'career_formset': career_formset,
        'education_formset': education_formset,
        'clients': clients,
        'role_choices': ClientContact.ROLE_CHOICES,
        'relationship_level_choices': ClientContact.RELATIONSHIP_LEVEL_CHOICES,
        'gender_choices': ClientContact.GENDER_CHOICES,
        'decision_influence_choices': ClientContact.DECISION_INFLUENCE_CHOICES,
    })
    return render(request, "customer_management/contact_form.html", context)


@login_required
def contact_detail(request, contact_id):
    """联系人详情"""
    from backend.apps.customer_management.models import (
        ClientContact, ContactCareer, ContactEducation, ContactWorkExperience,
        ContactJobChange, ContactCooperation, ContactTracking
    )
    
    contact = get_object_or_404(ClientContact, id=contact_id)
    permission_set = get_user_permission_codes(request.user)
    can_edit = _check_customer_permission('customer_management.contact.edit', permission_set)
    
    # 获取关联数据
    careers = ContactCareer.objects.filter(contact=contact).order_by('-join_date')
    educations = ContactEducation.objects.filter(contact=contact).order_by('-start_date')
    work_experiences = ContactWorkExperience.objects.filter(contact=contact).order_by('-start_date')
    job_changes = ContactJobChange.objects.filter(contact=contact).order_by('-change_date')
    cooperations = ContactCooperation.objects.filter(contact=contact).order_by('-cooperation_date')
    trackings = ContactTracking.objects.filter(contact=contact).order_by('-tracking_date')
    
    context = _context(
        f"联系人详情 - {contact.name}",
        "📇",
        f"查看联系人 {contact.name} 的详细信息",
        request=request,
    )
    context.update({
        'contact': contact,
        'careers': careers,
        'educations': educations,
        'work_experiences': work_experiences,
        'job_changes': job_changes,
        'cooperations': cooperations,
        'trackings': trackings,
        'can_edit': can_edit,
    })
    return render(request, "customer_management/contact_detail.html", context)


@login_required
def contact_edit(request, contact_id):
    """编辑联系人"""
    from backend.apps.customer_management.models import ClientContact, Client, ContactCareer, ContactColleague
    from backend.apps.customer_management.forms import ContactForm, ContactCareerFormSet, ContactEducationFormSet
    
    contact = get_object_or_404(ClientContact, id=contact_id)
    permission_set = get_user_permission_codes(request.user)
    
    if not _check_customer_permission('customer_management.contact.edit', permission_set):
        messages.error(request, '您没有权限编辑联系人')
        return redirect('business_pages:contact_list')
    
    if request.method == 'POST':
        # 判断是保存草稿还是提交
        action = request.POST.get('action', 'submit')
        is_draft = (action == 'save_draft')
        
        form = ContactForm(request.POST, request.FILES, instance=contact, is_draft=is_draft)
        career_formset = ContactCareerFormSet(request.POST, instance=contact, prefix='careers')
        education_formset = ContactEducationFormSet(request.POST, instance=contact, prefix='educations')
        
        # 如果是保存草稿，修改表单集的验证规则
        if is_draft:
            career_formset.min_num = 0
            career_formset.validate_min = False
            education_formset.min_num = 0
            education_formset.validate_min = False
            # 为表单集中的每个表单设置is_draft
            for form_item in career_formset.forms:
                form_item.is_draft = True
            for form_item in education_formset.forms:
                form_item.is_draft = True
        
        # 保存草稿时，直接保存，不进行验证
        # 提交时，必须通过验证才能保存
        if is_draft:
            # 保存草稿：直接保存，不检查验证结果
            # 但需要确保数据能正确保存，所以先尝试清理数据
            form.full_clean()
            # 清除所有验证错误，允许保存
            form.errors.clear()
            # 表单集也清除错误
            for form_item in career_formset.forms:
                form_item.errors.clear()
            for form_item in education_formset.forms:
                form_item.errors.clear()
            # 标记表单和表单集为有效，以便保存
            form._errors = {}
            career_formset._errors = {}
            education_formset._errors = {}
            can_save = True
        else:
            # 提交：必须通过验证
            can_save = form.is_valid() and career_formset.is_valid() and education_formset.is_valid()
        
        if can_save:
            # 保存表单
            contact = form.save(commit=False)
            if is_draft:
                contact.approval_status = 'draft'  # 保存草稿状态
            else:
                contact.approval_status = 'pending'  # 设置为待审批状态
            
            # 处理多选字段
            contact.preferred_contact_methods = form.cleaned_data.get('preferred_contact_methods', [])
            contact.tags = form.cleaned_data.get('tags', [])
            
            # 处理简历文件上传时间
            if form.cleaned_data.get('resume_file') and not contact.resume_upload_time:
                from django.utils import timezone
                contact.resume_upload_time = timezone.now()
            
            contact.save()
            form.save_m2m()
            
            # 保存职业信息
            careers = career_formset.save(commit=False)
            saved_careers = []
            for idx, career in enumerate(careers):
                career.contact = contact
                career.save()
                saved_careers.append((idx, career))
            
            # 删除标记为删除的职业记录
            for career in career_formset.deleted_objects:
                # 删除职业信息时，同时删除关联的同事关系人员
                career.colleagues.all().delete()
                career.delete()
            
            # 保存同事关系人员
            # 遍历所有保存的职业信息，查找对应的同事关系人员数据
            for form_idx, career in saved_careers:
                # 查找该职业信息对应的同事关系人员数据
                # 数据格式：careers-{form_idx}-colleagues-{j}-{field}
                colleague_prefix = f'careers-{form_idx}-colleagues'
                
                # 获取所有以该前缀开头的字段
                colleague_data = {}
                for key, value in request.POST.items():
                    if key.startswith(colleague_prefix + '-'):
                        # 解析字段名：careers-0-colleagues-0-name -> (0, name)
                        remaining = key[len(colleague_prefix) + 1:]  # 去掉前缀和连字符
                        parts = remaining.split('-')
                        if len(parts) >= 2:
                            try:
                                colleague_index = int(parts[0])
                                field_name = '-'.join(parts[1:])
                                
                                if colleague_index not in colleague_data:
                                    colleague_data[colleague_index] = {}
                                colleague_data[colleague_index][field_name] = value
                            except (ValueError, IndexError):
                                continue
                
                # 删除该职业信息的所有现有同事关系人员
                career.colleagues.all().delete()
                
                # 保存新的同事关系人员
                for colleague_index, colleague_fields in colleague_data.items():
                    # 检查是否有DELETE标记
                    delete_key = f'{colleague_prefix}-{colleague_index}-DELETE'
                    if request.POST.get(delete_key) == 'on':
                        continue
                    
                    # 检查必填字段
                    if not colleague_fields.get('name'):
                        continue
                    
                    ContactColleague.objects.create(
                        career=career,
                        department=colleague_fields.get('department', ''),
                        name=colleague_fields.get('name', ''),
                        position=colleague_fields.get('position', ''),
                        phone=colleague_fields.get('phone', '')
                    )
            
            # 保存教育信息
            educations = education_formset.save(commit=False)
            for education in educations:
                education.contact = contact
                education.save()
            
            # 删除标记为删除的教育记录
            for education in education_formset.deleted_objects:
                education.delete()
            
            # 只有提交时才启动审批流程，保存草稿时不启动
            if not is_draft:
                try:
                    from backend.apps.workflow_engine.models import WorkflowTemplate
                    from backend.apps.workflow_engine.services import ApprovalEngine
                    
                    workflow = WorkflowTemplate.objects.get(
                        code='contact_approval',
                        status='active'
                    )
                    
                    ApprovalEngine.start_approval(
                        workflow=workflow,
                        content_object=contact,
                        applicant=request.user,
                        comment=f'申请修改联系人：{contact.name}（客户：{contact.client.name}）'
                    )
                    
                    messages.success(request, f'联系人信息已更新，已提交审批。联系人：{contact.name}')
                except WorkflowTemplate.DoesNotExist:
                    messages.warning(request, '联系人信息已更新，但审批流程未配置，请联系管理员')
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('启动审批流程失败: %s', str(e))
                    messages.warning(request, f'联系人信息已更新，但启动审批流程失败：{str(e)}')
            else:
                messages.success(request, f'联系人草稿保存成功。联系人：{contact.name}')
            
            return redirect('business_pages:contact_list')
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = ContactForm(instance=contact)
        career_formset = ContactCareerFormSet(instance=contact, prefix='careers')
        education_formset = ContactEducationFormSet(instance=contact, prefix='educations')
    
    # 获取客户列表（用于下拉选择）
    clients = Client.objects.all().order_by('name')
    
    context = _context(
        f"编辑联系人 - {contact.name}",
        "✏️",
        f"{contact.name}",
        request=request,
    )
    context.update({
        'contact': contact,
        'form': form,
        'career_formset': career_formset,
        'education_formset': education_formset,
        'clients': clients,
        'role_choices': ClientContact.ROLE_CHOICES,
        'relationship_level_choices': ClientContact.RELATIONSHIP_LEVEL_CHOICES,
        'gender_choices': ClientContact.GENDER_CHOICES,
        'decision_influence_choices': ClientContact.DECISION_INFLUENCE_CHOICES,
    })
    return render(request, "customer_management/contact_form.html", context)


@login_required
def contact_delete(request, contact_id):
    """删除联系人"""
    from backend.apps.customer_management.models import ClientContact
    
    contact = get_object_or_404(ClientContact, id=contact_id)
    permission_set = get_user_permission_codes(request.user)
    
    if not _check_customer_permission('customer_management.contact.edit', permission_set):
        messages.error(request, '您没有权限删除联系人')
        return redirect('business_pages:contact_list')
    
    if request.method == 'POST':
        try:
            contact_name = contact.name
            contact.delete()
            messages.success(request, f'联系人 {contact_name} 已删除')
            return redirect('business_pages:contact_list')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除联系人失败: %s', str(e))
            messages.error(request, f'删除联系人失败：{str(e)}')
            return redirect('business_pages:contact_list')
    
    # GET 请求，显示确认删除页面
    context = _context(
        "删除联系人",
        "🗑️",
        f"确认删除联系人：{contact.name}",
        request=request,
    )
    context.update({
        'contact': contact,
    })
    return render(request, "customer_management/contact_delete.html", context)


@login_required
def contact_relationship_mining(request):
    """关系挖掘：通过电话或客户公司查找相关人员的关系网络"""
    from django.db.models import Q
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    can_view = _check_customer_permission('customer_management.contact.view', permission_set)
    
    if not can_view:
        messages.error(request, '您没有权限访问此功能')
        return redirect('business_pages:contact_list')
    
    phone = request.GET.get('phone', '').strip()
    client_id = request.GET.get('client_id', '').strip()
    search_type = request.GET.get('search_type', 'phone')  # 'phone' 或 'client'
    
    target_contact = None
    target_client = None
    client_contacts = []  # 目标客户公司的所有联系人
    same_education_contacts = []
    same_birthplace_contacts = []
    same_career_contacts = []
    client_company_contacts = []  # 目标客户公司内的人员关系网
    
    try:
        if search_type == 'phone' and phone:
            # 方式1：通过电话查找目标人员
            target_contact = ClientContact.objects.filter(
                Q(phone__icontains=phone) | Q(telephone__icontains=phone)
            ).select_related('client', 'created_by').first()
            
            if target_contact:
                # 挖掘目标人物的关系网
                same_education_contacts, same_birthplace_contacts, same_career_contacts = _mine_contact_relationships(target_contact)
        
        elif search_type == 'client' and client_id:
            # 方式2：通过客户公司查找
            try:
                target_client = Client.objects.get(id=int(client_id))
                # 获取该客户公司的所有联系人
                client_contacts = ClientContact.objects.filter(
                    client=target_client
                ).select_related('client', 'created_by').order_by('-relationship_score', 'name')
                
                # 挖掘客户公司内所有联系人的关系网络
                if client_contacts.exists():
                    client_company_contacts = _mine_client_company_relationships(target_client, client_contacts)
            
            except (Client.DoesNotExist, ValueError):
                messages.error(request, '客户不存在')
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('关系挖掘查询失败: %s', str(e))
        messages.error(request, f'查询失败：{str(e)}')
    
    context = _context(
        "关系挖掘",
        "🔍",
        "通过电话或客户公司查找相关人员的关系网络",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='contact_relationship_mining'
    )
    
    # 获取客户列表（用于下拉选择）
    clients = Client.objects.filter(is_active=True).order_by('name')
    
    context.update({
        'phone': phone,
        'client_id': client_id,
        'search_type': search_type,
        'target_contact': target_contact,
        'target_client': target_client,
        'client_contacts': client_contacts,
        'same_education_contacts': same_education_contacts,
        'same_birthplace_contacts': same_birthplace_contacts,
        'same_career_contacts': same_career_contacts,
        'client_company_contacts': client_company_contacts,
        'clients': clients,
    })
    
    return render(request, "customer_management/contact_relationship_mining.html", context)


def _mine_contact_relationships(target_contact):
    """挖掘单个联系人的关系网络"""
    from django.db.models import Q
    
    same_education_contacts = []
    same_birthplace_contacts = []
    same_career_contacts = []
    
    # 1. 查找相同教育背景的人员
    target_schools = set()
    target_school_names = set()
    
    for education in target_contact.educations.all():
        if education.school:
            target_schools.add(education.school.id)
        if education.school_name:
            target_school_names.add(education.school_name.strip())
    
    if target_schools or target_school_names:
        same_education_query = Q()
        if target_schools:
            same_education_query |= Q(educations__school_id__in=target_schools)
        if target_school_names:
            for school_name in target_school_names:
                same_education_query |= Q(educations__school_name__icontains=school_name)
        
        same_education_contacts = ClientContact.objects.filter(
            same_education_query
        ).exclude(id=target_contact.id).distinct().select_related('client', 'created_by')[:50]
    
    # 2. 查找相同籍贯的人员
    if target_contact.birthplace:
        birthplace_keywords = target_contact.birthplace.split()
        if birthplace_keywords:
            birthplace_query = Q()
            for keyword in birthplace_keywords:
                if len(keyword) > 1:
                    birthplace_query |= Q(birthplace__icontains=keyword)
            
            same_birthplace_contacts = ClientContact.objects.filter(
                birthplace_query
            ).exclude(id=target_contact.id).exclude(birthplace='').distinct().select_related('client', 'created_by')[:50]
    
    # 3. 查找相同职业信息的人员
    target_companies = set()
    target_departments = set()
    target_positions = set()
    
    for career in target_contact.careers.all():
        if career.company:
            target_companies.add(career.company.strip())
        if career.department:
            target_departments.add(career.department.strip())
        if career.position:
            target_positions.add(career.position.strip())
    
    if target_companies or target_departments or target_positions:
        career_query = Q()
        if target_companies:
            for company in target_companies:
                if len(company) > 1:
                    career_query |= Q(careers__company__icontains=company)
        if target_departments:
            for dept in target_departments:
                if len(dept) > 1:
                    career_query |= Q(careers__department__icontains=dept)
        if target_positions:
            for pos in target_positions:
                if len(pos) > 1:
                    career_query |= Q(careers__position__icontains=pos)
        
        same_career_contacts = ClientContact.objects.filter(
            career_query
        ).exclude(id=target_contact.id).distinct().select_related('client', 'created_by')[:50]
    
    return same_education_contacts, same_birthplace_contacts, same_career_contacts


def _mine_client_company_relationships(target_client, client_contacts):
    """挖掘客户公司内所有联系人的关系网络"""
    from django.db.models import Q
    from collections import defaultdict
    
    # 收集所有联系人的关系数据
    all_schools = set()
    all_school_names = set()
    all_birthplaces = set()
    all_companies = set()
    all_departments = set()
    all_positions = set()
    
    contact_ids = [c.id for c in client_contacts]
    
    # 收集教育背景
    educations = ContactEducation.objects.filter(contact_id__in=contact_ids)
    for edu in educations:
        if edu.school:
            all_schools.add(edu.school.id)
        if edu.school_name:
            all_school_names.add(edu.school_name.strip())
    
    # 收集籍贯
    for contact in client_contacts:
        if contact.birthplace:
            all_birthplaces.add(contact.birthplace.strip())
    
    # 收集职业信息
    careers = ContactCareer.objects.filter(contact_id__in=contact_ids)
    for career in careers:
        if career.company:
            all_companies.add(career.company.strip())
        if career.department:
            all_departments.add(career.department.strip())
        if career.position:
            all_positions.add(career.position.strip())
    
    # 查找与客户公司内人员有关系的其他联系人
    related_contacts = []
    
    # 1. 相同教育背景
    if all_schools or all_school_names:
        education_query = Q()
        if all_schools:
            education_query |= Q(educations__school_id__in=all_schools)
        if all_school_names:
            for school_name in all_school_names:
                if len(school_name) > 1:
                    education_query |= Q(educations__school_name__icontains=school_name)
        
        education_contacts = ClientContact.objects.filter(
            education_query
        ).exclude(id__in=contact_ids).distinct().select_related('client', 'created_by')[:100]
        
        for contact in education_contacts:
            related_contacts.append({
                'contact': contact,
                'relation_type': 'education',
                'relation_desc': '相同教育背景'
            })
    
    # 2. 相同籍贯
    if all_birthplaces:
        birthplace_query = Q()
        for birthplace in all_birthplaces:
            if len(birthplace) > 1:
                birthplace_query |= Q(birthplace__icontains=birthplace)
        
        birthplace_contacts = ClientContact.objects.filter(
            birthplace_query
        ).exclude(id__in=contact_ids).exclude(birthplace='').distinct().select_related('client', 'created_by')[:100]
        
        for contact in birthplace_contacts:
            # 检查是否已添加（避免重复）
            if not any(c['contact'].id == contact.id for c in related_contacts):
                related_contacts.append({
                    'contact': contact,
                    'relation_type': 'birthplace',
                    'relation_desc': '相同籍贯'
                })
    
    # 3. 相同职业信息（同一公司、部门或职位）
    if all_companies or all_departments or all_positions:
        career_query = Q()
        if all_companies:
            for company in all_companies:
                if len(company) > 1:
                    career_query |= Q(careers__company__icontains=company)
        if all_departments:
            for dept in all_departments:
                if len(dept) > 1:
                    career_query |= Q(careers__department__icontains=dept)
        if all_positions:
            for pos in all_positions:
                if len(pos) > 1:
                    career_query |= Q(careers__position__icontains=pos)
        
        career_contacts = ClientContact.objects.filter(
            career_query
        ).exclude(id__in=contact_ids).distinct().select_related('client', 'created_by')[:100]
        
        for contact in career_contacts:
            # 检查是否已添加（避免重复）
            if not any(c['contact'].id == contact.id for c in related_contacts):
                related_contacts.append({
                    'contact': contact,
                    'relation_type': 'career',
                    'relation_desc': '相同职业信息'
                })
    
    # 4. 同事关系（通过ContactColleague）
    colleague_careers = ContactCareer.objects.filter(contact_id__in=contact_ids)
    colleague_contact_ids = set()
    
    for career in colleague_careers:
        colleagues = ContactColleague.objects.filter(career=career)
        for colleague in colleagues:
            # 尝试通过姓名和电话匹配联系人
            if colleague.name and colleague.phone:
                matched_contacts = ClientContact.objects.filter(
                    Q(name__icontains=colleague.name) | Q(phone__icontains=colleague.phone)
                ).exclude(id__in=contact_ids)
                for matched in matched_contacts:
                    colleague_contact_ids.add(matched.id)
    
    if colleague_contact_ids:
        colleague_contacts = ClientContact.objects.filter(
            id__in=colleague_contact_ids
        ).select_related('client', 'created_by')[:50]
        
        for contact in colleague_contacts:
            if not any(c['contact'].id == contact.id for c in related_contacts):
                related_contacts.append({
                    'contact': contact,
                    'relation_type': 'colleague',
                    'relation_desc': '同事关系'
                })
    
    # 按关系评分排序
    related_contacts.sort(key=lambda x: x['contact'].relationship_score, reverse=True)
    
    return related_contacts


def contact_tracking_reminders(request):
    """逾期拜访提醒列表"""
    from django.db.models import Q
    from datetime import timedelta
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    can_view = _check_customer_permission('customer_management.contact.view', permission_set)
    
    if not can_view:
        messages.error(request, '您没有权限访问此功能')
        return redirect('business_pages:contact_list')
    
    # 获取查询参数
    days_ahead = int(request.GET.get('days_ahead', 7))  # 提前提醒天数
    filter_type = request.GET.get('filter_type', 'all')  # all, overdue, upcoming
    
    # 获取当前用户创建的联系人（或根据权限获取）
    contacts = ClientContact.objects.select_related('client', 'created_by').all()
    
    # 权限过滤：如果用户不是管理员，只显示自己创建的联系人
    if not request.user.is_superuser:
        # 检查是否有查看所有联系人的权限
        if not _check_customer_permission('customer_management.contact.view_all', permission_set):
            contacts = contacts.filter(created_by=request.user)
    
    # 计算提醒信息
    reminders = []
    today = timezone.now().date()
    
    for contact in contacts:
        next_date = contact.get_next_tracking_date()
        days_until = (next_date - today).days
        is_overdue = days_until < 0
        
        # 根据筛选条件过滤
        if filter_type == 'overdue' and not is_overdue:
            continue
        if filter_type == 'upcoming' and (is_overdue or days_until > days_ahead):
            continue
        if filter_type == 'all' and not is_overdue and days_until > days_ahead:
            continue
        
        # 确定优先级
        if contact.role == 'decision_maker':
            priority = 'high'
        elif contact.role == 'promoter':
            priority = 'medium'
        else:
            priority = 'normal'
        
        reminders.append({
            'contact': contact,
            'next_date': next_date,
            'days_until': days_until,
            'is_overdue': is_overdue,
            'overdue_days': abs(days_until) if is_overdue else 0,
            'priority': priority,
            'tracking_cycle': contact.calculate_tracking_cycle(),
        })
    
    # 排序：超期 > 优先级 > 日期
    reminders.sort(key=lambda x: (
        not x['is_overdue'],  # 超期的在前
        x['priority'] != 'high',  # 高优先级在前
        x['days_until']  # 日期越近越前
    ))
    
    # 统计信息
    stats = {
        'total': len(reminders),
        'overdue': sum(1 for r in reminders if r['is_overdue']),
        'upcoming': sum(1 for r in reminders if not r['is_overdue']),
        'high_priority': sum(1 for r in reminders if r['priority'] == 'high'),
    }
    
    context = _context(
        "逾期拜访提醒",
        "🔔",
        "客户人员逾期拜访提醒列表",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='contact_tracking_reminders'
    )
    
    context.update({
        'reminders': reminders,
        'stats': stats,
        'days_ahead': days_ahead,
        'filter_type': filter_type,
    })
    
    return render(request, "customer_management/contact_tracking_reminders.html", context)


@login_required
def contact_info_change_create(request):
    """创建联系人信息变更申请"""
    # 检查权限
    permission_set = get_user_permission_codes(request.user)
    if not _check_customer_permission('customer_management.contact.edit', permission_set):
        messages.error(request, '您没有权限创建联系人信息变更申请')
        return redirect('business_pages:contact_list')
    
    # 检查 ContactInfoChange 模型是否存在
    try:
        from backend.apps.customer_management.models import ContactInfoChange
        from backend.apps.customer_management.forms import ContactInfoChangeForm
    except (ImportError, AttributeError):
        # 模型不存在，重定向到联系人列表
        messages.warning(request, '联系人信息变更功能暂时不可用，请直接编辑联系人信息')
        return redirect('business_pages:contact_list')
    
    # 如果模型为 None（在 forms.py 中可能被设置为 None）
    if ContactInfoChange is None:
        messages.warning(request, '联系人信息变更功能暂时不可用，请直接编辑联系人信息')
        return redirect('business_pages:contact_list')
    
    # 处理表单提交
    if request.method == 'POST':
        form = ContactInfoChangeForm(request.POST, user=request.user)
        if form.is_valid():
            info_change = form.save(commit=False)
            info_change.created_by = request.user
            info_change.approval_status = 'draft'  # 默认为草稿状态
            info_change.save()
            messages.success(request, '联系人信息变更申请已创建')
            return redirect('business_pages:contact_list')
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = ContactInfoChangeForm(user=request.user)
    
    # 获取联系人列表（用于下拉选择）
    contacts = ClientContact.objects.all().order_by('name')
    
    context = _context(
        "创建联系人信息变更申请",
        "📝",
        "创建新的联系人信息变更申请",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='contact_list'
    )
    
    context.update({
        'form': form,
        'contacts': contacts,
    })
    
    return render(request, "customer_management/contact_info_change_create.html", context)


# ==================== 跟进与拜访管理视图函数 =====================

@login_required
def customer_visit(request):
    """创建联系人拜访"""
    from django.core.paginator import Paginator
    from backend.apps.customer_management.models import CustomerRelationship
    
    # 获取筛选参数（使用通用方式支持新筛选模块）
    search = request.GET.get('search', '').strip()
    
    # 获取通用筛选参数
    filter_params = {}
    for key, value in request.GET.items():
        if key not in ['search', 'page', 'page_size'] and value:
            filter_params[key] = value
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    can_create = _check_customer_permission('customer_management.relationship.create', permission_set)
    
    # 获取拜访记录列表（record_type='visit'）
    try:
        relationships = CustomerRelationship.objects.filter(
            record_type='visit'
        ).select_related('client', 'followup_person', 'created_by').prefetch_related('related_contacts')
        
        # 应用搜索条件
        if search:
            relationships = relationships.filter(
                Q(client__name__icontains=search) |
                Q(content__icontains=search)
            )
        
        # 应用通用筛选条件
        if filter_params.get('client'):
            relationships = relationships.filter(client_id=filter_params['client'])
        if filter_params.get('visit_type'):
            relationships = relationships.filter(visit_type=filter_params['visit_type'])
        
        # 按跟进时间倒序排列
        relationships = relationships.order_by('-followup_time')
        
        # 分页
        paginator = Paginator(relationships, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取拜访记录列表失败: %s', str(e))
        messages.error(request, f'获取拜访记录列表失败：{str(e)}')
        page_obj = None
    
    # 获取客户列表（用于筛选）
    from backend.apps.customer_management.models import Client
    clients = Client.objects.all().order_by('name')
    
    context = _context(
        "创建联系人拜访",
        "🚪",
        "查看和管理客户拜访记录",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='visit_list'
    )
    
    context.update({
        'page_obj': page_obj,
        'search': search,
        'client_id': filter_params.get('client', ''),
        'visit_type': filter_params.get('visit_type', ''),
        'clients': clients,
        'can_create': can_create,
        'visit_type_choices': CustomerRelationship.VISIT_TYPE_CHOICES,
    })
    return render(request, "customer_management/customer_visit.html", context)


# ==================== 关系升级管理视图函数 =====================

@login_required
def customer_relationship_upgrade(request):
    """创建人员关系升级"""
    from django.core.paginator import Paginator
    from backend.apps.customer_management.models import CustomerRelationshipUpgrade
    
    # 获取筛选参数
    search = request.GET.get('search', '').strip()
    client_id = request.GET.get('client', '')
    approval_status = request.GET.get('approval_status', '')
    to_level = request.GET.get('to_level', '')
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    can_create = _check_customer_permission('customer_management.relationship.create', permission_set)
    
    # 获取关系升级记录列表
    try:
        upgrades = CustomerRelationshipUpgrade.objects.select_related(
            'client', 'created_by', 'approval_instance'
        ).prefetch_related('related_contacts')
        
        # 应用搜索条件
        if search:
            upgrades = upgrades.filter(
                Q(client__name__icontains=search) |
                Q(upgrade_reason__icontains=search)
            )
        
        # 应用筛选条件
        if client_id:
            upgrades = upgrades.filter(client_id=client_id)
        if approval_status:
            upgrades = upgrades.filter(approval_status=approval_status)
        if to_level:
            upgrades = upgrades.filter(to_level=to_level)
        
        # 按创建时间倒序排列
        upgrades = upgrades.order_by('-created_time')
        
        # 分页
        paginator = Paginator(upgrades, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取关系升级记录列表失败: %s', str(e))
        messages.error(request, f'获取关系升级记录列表失败：{str(e)}')
        page_obj = None
    
    # 获取客户列表（用于筛选）
    from backend.apps.customer_management.models import Client
    clients = Client.objects.all().order_by('name')
    
    context = _context(
        "创建人员关系升级",
        "⬆️",
        "创建和管理客户关系升级记录",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='upgrade_list'
    )
    
    context.update({
        'page_obj': page_obj,
        'search': search,
        'client_id': client_id,
        'approval_status': approval_status,
        'to_level': to_level,
        'clients': clients,
        'can_create': can_create,
        'approval_status_choices': CustomerRelationshipUpgrade.APPROVAL_STATUS_CHOICES,
        'relationship_level_choices': CustomerRelationshipUpgrade.RELATIONSHIP_LEVEL_CHOICES,
    })
    return render(request, "customer_management/customer_relationship_upgrade.html", context)


@login_required
def customer_relationship_upgrade_create(request):
    """创建关系升级申请"""
    from backend.apps.customer_management.models import CustomerRelationshipUpgrade, Client, ClientContact
    from backend.apps.customer_management.forms import RelationshipUpgradeForm
    
    permission_set = get_user_permission_codes(request.user)
    if not _check_customer_permission('customer_management.relationship.edit', permission_set):
        messages.error(request, '您没有权限创建关系升级申请')
        return redirect('business_pages:customer_relationship_upgrade')
    
    if request.method == 'POST':
        form = RelationshipUpgradeForm(request.POST)
        if form.is_valid():
            upgrade = form.save(commit=False)
            upgrade.created_by = request.user
            
            # 判断是否需要审批
            if upgrade.requires_approval():
                # 需要审批，设置状态为待审批
                upgrade.approval_status = 'pending'
                
                # 启动审批流程
                try:
                    from django.contrib.contenttypes.models import ContentType
                    from backend.apps.workflow_engine.models import WorkflowTemplate
                    from backend.apps.workflow_engine.services import ApprovalEngine
                    
                    workflow = WorkflowTemplate.objects.get(
                        code='customer_relationship_upgrade_approval',
                        status='active'
                    )
                    
                    instance = ApprovalEngine.start_approval(
                        workflow=workflow,
                        content_object=upgrade,
                        applicant=request.user,
                        comment=f'申请将客户 {upgrade.client.name} 的关系等级从{upgrade.get_from_level_display()}升级为{upgrade.get_to_level_display()}'
                    )
                    
                    upgrade.approval_instance = instance
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('启动关系升级审批流程失败: %s', str(e))
                    messages.error(request, '关系等级升级审批流程未配置或已停用，请联系管理员')
                    return redirect('business_pages:customer_relationship_upgrade')
            else:
                # 无需审批，直接生效
                upgrade.approval_status = 'approved'
                upgrade.approved_time = timezone.now()
                
                # 更新关联联系人的关系等级
                for contact in form.cleaned_data.get('related_contacts', []):
                    contact.relationship_level = upgrade.to_level
                    contact.save(update_fields=['relationship_level'])
            
            upgrade.save()
            form.save_m2m()  # 保存多对多关系（related_contacts）
            
            messages.success(request, '关系升级申请已提交' if upgrade.requires_approval() else '关系等级已更新')
            return redirect('business_pages:customer_relationship_upgrade')
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = RelationshipUpgradeForm()
    
    # 获取客户列表和联系人列表（用于下拉选择）
    clients = Client.objects.all().order_by('name')
    contacts = ClientContact.objects.all().select_related('client').order_by('client__name', 'name')
    
    context = _context(
        "创建关系升级申请",
        "➕",
        "创建新的客户关系升级申请",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='upgrade_list'
    )
    
    context.update({
        'form': form,
        'clients': clients,
        'contacts': contacts,
        'relationship_level_choices': CustomerRelationshipUpgrade.RELATIONSHIP_LEVEL_CHOICES,
    })
    return render(request, "customer_management/customer_relationship_upgrade_form.html", context)


@login_required
def business_expense_application_list(request):
    """业务费申请列表"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    from backend.apps.customer_management.models import BusinessExpenseApplication, Client
    
    # 获取筛选参数
    search = request.GET.get('search', '').strip()
    client_id = request.GET.get('client', '')
    approval_status = request.GET.get('approval_status', '')
    expense_type = request.GET.get('expense_type', '')
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    can_create = _check_customer_permission('customer_management.relationship.create', permission_set)
    
    # 获取业务费申请列表
    try:
        expenses = BusinessExpenseApplication.objects.select_related(
            'client', 'created_by', 'approval_instance'
        ).prefetch_related('related_contacts')
        
        # 应用搜索条件
        if search:
            expenses = expenses.filter(
                Q(application_number__icontains=search) |
                Q(client__name__icontains=search) |
                Q(description__icontains=search)
            )
        
        # 应用筛选条件
        if client_id:
            expenses = expenses.filter(client_id=client_id)
        if approval_status:
            expenses = expenses.filter(approval_status=approval_status)
        if expense_type:
            expenses = expenses.filter(expense_type=expense_type)
        
        # 按创建时间倒序排列
        expenses = expenses.order_by('-created_time')
        
        # 分页
        paginator = Paginator(expenses, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取业务费申请列表失败: %s', str(e))
        messages.error(request, f'获取业务费申请列表失败：{str(e)}')
        page_obj = None
    
    # 获取客户列表（用于筛选）
    clients = Client.objects.all().order_by('name')
    
    context = _context(
        "业务费申请列表",
        "💰",
        "查看和管理业务费申请记录",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='business_expense_application'
    )
    
    context.update({
        'page_obj': page_obj,
        'search': search,
        'client_id': client_id,
        'approval_status': approval_status,
        'expense_type': expense_type,
        'clients': clients,
        'can_create': can_create,
        'approval_status_choices': BusinessExpenseApplication.APPROVAL_STATUS_CHOICES,
        'expense_type_choices': BusinessExpenseApplication.EXPENSE_TYPE_CHOICES,
    })
    return render(request, "customer_management/business_expense_application_list.html", context)


@login_required
def business_expense_application_create(request):
    """创建业务费申请"""
    from backend.apps.customer_management.models import BusinessExpenseApplication, Client, ClientContact
    from backend.apps.customer_management.forms import BusinessExpenseApplicationForm
    
    permission_set = get_user_permission_codes(request.user)
    if not _check_customer_permission('customer_management.relationship.create', permission_set):
        messages.error(request, '您没有权限创建业务费申请')
        return redirect('business_pages:business_expense_application_list')
    
    if request.method == 'POST':
        action = request.POST.get('action', 'submit')
        is_draft = (action == 'save_draft')
        
        form = BusinessExpenseApplicationForm(request.POST, request.FILES, is_draft=is_draft)
        
        if is_draft or form.is_valid():
            expense = form.save(commit=False)
            expense.created_by = request.user
            expense.approval_status = 'draft' if is_draft else 'pending'
            
            # 如果需要审批且不是草稿，启动审批流程
            if not is_draft:
                try:
                    from django.contrib.contenttypes.models import ContentType
                    from backend.apps.workflow_engine.models import WorkflowTemplate
                    from backend.apps.workflow_engine.services import ApprovalEngine
                    
                    workflow = WorkflowTemplate.objects.get(
                        code='business_expense_approval',
                        status='active'
                    )
                    
                    instance = ApprovalEngine.start_approval(
                        workflow=workflow,
                        content_object=expense,
                        applicant=request.user,
                        comment=f'申请业务费：{expense.get_expense_type_display()} - ¥{expense.amount}（客户：{expense.client.name}）'
                    )
                    
                    expense.approval_instance = instance
                    expense.approval_status = 'pending'
                except WorkflowTemplate.DoesNotExist:
                    messages.warning(request, '业务费审批流程未配置，申请已保存为草稿，请联系管理员')
                    expense.approval_status = 'draft'
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('启动业务费审批流程失败: %s', str(e))
                    messages.warning(request, f'启动审批流程失败：{str(e)}，申请已保存为草稿')
                    expense.approval_status = 'draft'
            
            expense.save()
            form.save_m2m()  # 保存多对多关系（related_contacts）
            
            if is_draft:
                messages.success(request, f'业务费申请草稿保存成功。申请单号：{expense.application_number}')
            else:
                messages.success(request, f'业务费申请已提交。申请单号：{expense.application_number}')
            
            return redirect('business_pages:business_expense_application_list')
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = BusinessExpenseApplicationForm()
    
    # 获取客户列表和联系人列表（用于下拉选择）
    clients = Client.objects.all().order_by('name')
    contacts = ClientContact.objects.all().select_related('client').order_by('client__name', 'name')
    
    context = _context(
        "创建业务费申请",
        "💰",
        "创建新的业务费申请",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='business_expense_application'
    )
    
    context.update({
        'form': form,
        'clients': clients,
        'contacts': contacts,
        'expense_type_choices': BusinessExpenseApplication.EXPENSE_TYPE_CHOICES,
    })
    return render(request, "customer_management/business_expense_application_form.html", context)


@login_required
def customer_relationship_collaboration(request):
    """人员关系协作申请列表"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    from backend.apps.customer_management.models import CustomerRelationshipCollaboration, Client
    
    # 获取筛选参数
    search = request.GET.get('search', '').strip()
    client_id = request.GET.get('client', '')
    task_type = request.GET.get('task_type', '')
    status = request.GET.get('status', '')
    priority = request.GET.get('priority', '')
    responsible_user_id = request.GET.get('responsible_user', '')
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    can_create = _check_customer_permission('customer_management.relationship.create', permission_set)
    
    # 获取人员关系协作申请列表
    try:
        collaborations = CustomerRelationshipCollaboration.objects.select_related(
            'client', 'created_by'
        ).prefetch_related('responsible_users', 'collaborators')
        
        # 应用搜索条件
        if search:
            collaborations = collaborations.filter(
                Q(description__icontains=search) |
                Q(client__name__icontains=search)
            )
        
        # 应用筛选条件
        if client_id:
            collaborations = collaborations.filter(client_id=client_id)
        if task_type:
            collaborations = collaborations.filter(task_type=task_type)
        if status:
            collaborations = collaborations.filter(status=status)
        if priority:
            collaborations = collaborations.filter(priority=priority)
        if responsible_user_id:
            collaborations = collaborations.filter(responsible_users__id=responsible_user_id)
        
        # 按创建时间倒序排列
        collaborations = collaborations.order_by('-created_time')
        
        # 分页
        paginator = Paginator(collaborations, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取人员关系协作申请列表失败: %s', str(e))
        messages.error(request, f'获取人员关系协作申请列表失败：{str(e)}')
        page_obj = None
    
    # 获取客户列表（用于筛选）
    clients = Client.objects.filter(is_active=True).order_by('name')
    
    # 获取用户列表（用于筛选负责人）
    from backend.apps.system_management.models import User
    users = User.objects.filter(is_active=True).order_by('username')
    
    context = _context(
        "人员关系协作申请",
        "🤝",
        "创建和管理人员关系协作申请",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='relationship_collaboration'
    )
    
    context.update({
        'page_obj': page_obj,
        'search': search,
        'client_id': client_id,
        'task_type': task_type,
        'status': status,
        'priority': priority,
        'responsible_user_id': responsible_user_id,
        'clients': clients,
        'users': users,
        'can_create': can_create,
        'task_type_choices': CustomerRelationshipCollaboration.TASK_TYPE_CHOICES,
        'status_choices': CustomerRelationshipCollaboration.STATUS_CHOICES,
        'priority_choices': CustomerRelationshipCollaboration.PRIORITY_CHOICES,
    })
    
    return render(request, "customer_management/customer_relationship_collaboration.html", context)


@login_required
def customer_relationship_collaboration_create(request):
    """创建人员关系协作申请"""
    from backend.apps.customer_management.models import CustomerRelationshipCollaboration, Client, ClientContact
    from backend.apps.customer_management.forms import CollaborationTaskForm
    
    permission_set = get_user_permission_codes(request.user)
    if not _check_customer_permission('customer_management.relationship.edit', permission_set):
        messages.error(request, '您没有权限创建人员关系协作申请')
        return redirect('business_pages:customer_relationship_collaboration')
    
    if request.method == 'POST':
        form = CollaborationTaskForm(request.POST)
        if form.is_valid():
            collaboration = form.save(commit=False)
            collaboration.created_by = request.user
            collaboration.status = 'pending'
            collaboration.progress = 0
            collaboration.priority = 'medium'  # 默认优先级为中
            # 如果没有提供title，使用描述的前50个字符作为title
            if not collaboration.title and collaboration.description:
                collaboration.title = collaboration.description[:50]
            collaboration.save()
            form.save_m2m()  # 保存多对多关系
            
            # 默认将创建人设置为负责人
            collaboration.responsible_users.add(request.user)
            
            # 创建执行记录
            from backend.apps.customer_management.models import CustomerRelationshipCollaborationExecution
            CustomerRelationshipCollaborationExecution.objects.create(
                collaboration=collaboration,
                user=request.user,
                action='created',
                content='创建了人员关系协作申请'
            )
            
            messages.success(request, '人员关系协作申请创建成功')
            return redirect('business_pages:customer_relationship_collaboration')
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = CollaborationTaskForm()
    
    # 获取客户列表和联系人列表（用于下拉选择）
    clients = Client.objects.filter(is_active=True).order_by('name')
    contacts = ClientContact.objects.all().select_related('client').order_by('client__name', 'name')
    
    context = _context(
        "创建人员关系协作申请",
        "➕",
        "创建新的人员关系协作申请",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='relationship_collaboration'
    )
    
    context.update({
        'form': form,
        'clients': clients,
        'contacts': contacts,
        'task_type_choices': CustomerRelationshipCollaboration.TASK_TYPE_CHOICES,
        'priority_choices': CustomerRelationshipCollaboration.PRIORITY_CHOICES,
    })
    
    return render(request, "customer_management/customer_relationship_collaboration_form.html", context)


@login_required
def customer_relationship_collaboration_detail(request, collaboration_id):
    """人员关系协作申请详情"""
    from backend.apps.customer_management.models import (
        CustomerRelationshipCollaboration,
        CustomerRelationshipCollaborationComment,
        CustomerRelationshipCollaborationAttachment,
        CustomerRelationshipCollaborationExecution
    )
    
    collaboration = get_object_or_404(
        CustomerRelationshipCollaboration.objects.select_related(
            'client', 'created_by'
        ).prefetch_related(
            'responsible_users', 'collaborators', 'related_contacts', 'related_relationships'
        ),
        id=collaboration_id
    )
    
    permission_set = get_user_permission_codes(request.user)
    can_edit = _check_customer_permission('customer_management.relationship.edit', permission_set)
    can_comment = _check_customer_permission('customer_management.relationship.view', permission_set)
    
    # 检查用户是否有权限查看（负责人、协作者或创建人）
    is_responsible = collaboration.responsible_users.filter(id=request.user.id).exists()
    is_collaborator = collaboration.collaborators.filter(id=request.user.id).exists()
    is_creator = collaboration.created_by == request.user
    
    if not (is_responsible or is_collaborator or is_creator or can_edit):
        messages.error(request, '您没有权限查看此人员关系协作申请')
        return redirect('business_pages:customer_relationship_collaboration')
    
    # 处理POST请求（更新状态、进度、添加评论等）
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_status':
            # 更新状态
            if not (is_responsible or is_collaborator or is_creator or can_edit):
                messages.error(request, '您没有权限更新申请状态')
            else:
                new_status = request.POST.get('status')
                old_status = collaboration.status
                if new_status in dict(CustomerRelationshipCollaboration.STATUS_CHOICES):
                    collaboration.status = new_status
                    
                    # 如果状态变为进行中，记录开始时间
                    if new_status == 'in_progress' and not collaboration.start_time:
                        collaboration.start_time = timezone.now()
                    
                    # 如果状态变为已完成，记录完成时间和进度
                    if new_status == 'completed':
                        collaboration.completed_time = timezone.now()
                        collaboration.progress = 100
                    
                    collaboration.save()
                    
                    # 创建执行记录
                    CustomerRelationshipCollaborationExecution.objects.create(
                        collaboration=collaboration,
                        user=request.user,
                        action='status_changed',
                        content=f'状态从 {dict(CustomerRelationshipCollaboration.STATUS_CHOICES).get(old_status, old_status)} 变更为 {dict(CustomerRelationshipCollaboration.STATUS_CHOICES).get(new_status, new_status)}'
                    )
                    
                    messages.success(request, '申请状态已更新')
                else:
                    messages.error(request, '无效的状态值')
        
        elif action == 'update_progress':
            # 更新进度
            if not (is_responsible or is_collaborator or is_creator or can_edit):
                messages.error(request, '您没有权限更新申请进度')
            else:
                try:
                    new_progress = int(request.POST.get('progress', 0))
                    if 0 <= new_progress <= 100:
                        old_progress = collaboration.progress
                        collaboration.progress = new_progress
                        
                        # 如果进度达到100%，自动更新状态为已完成
                        if new_progress == 100 and collaboration.status != 'completed':
                            collaboration.status = 'completed'
                            collaboration.completed_time = timezone.now()
                        
                        collaboration.save()
                        
                        # 创建执行记录
                        CustomerRelationshipCollaborationExecution.objects.create(
                            collaboration=collaboration,
                            user=request.user,
                            action='progress_updated',
                            content=f'进度从 {old_progress}% 更新为 {new_progress}%'
                        )
                        
                        messages.success(request, '申请进度已更新')
                    else:
                        messages.error(request, '进度值必须在0-100之间')
                except ValueError:
                    messages.error(request, '无效的进度值')
        
        elif action == 'add_comment':
            # 添加评论
            if not can_comment:
                messages.error(request, '您没有权限添加评论')
            else:
                comment_content = request.POST.get('content', '').strip()
                if comment_content:
                    CustomerRelationshipCollaborationComment.objects.create(
                        collaboration=collaboration,
                        user=request.user,
                        content=comment_content
                    )
                    
                    # 创建执行记录
                    CustomerRelationshipCollaborationExecution.objects.create(
                        collaboration=collaboration,
                        user=request.user,
                        action='commented',
                        content=f'添加了评论：{comment_content[:50]}...'
                    )
                    
                    messages.success(request, '评论已添加')
                else:
                    messages.error(request, '评论内容不能为空')
        
        elif action == 'upload_attachment':
            # 上传附件
            if not (is_responsible or is_collaborator or is_creator or can_edit):
                messages.error(request, '您没有权限上传附件')
            else:
                from .forms import CollaborationAttachmentForm
                form = CollaborationAttachmentForm(request.POST, request.FILES)
                if form.is_valid():
                    attachment = form.save(commit=False)
                    attachment.collaboration = collaboration
                    attachment.uploaded_by = request.user
                    
                    # 如果文件名称未填写，使用上传文件名
                    if not attachment.file_name:
                        attachment.file_name = request.FILES['file'].name
                    
                    # 自动计算文件大小
                    if request.FILES.get('file'):
                        attachment.file_size = request.FILES['file'].size
                    
                    attachment.save()
                    
                    # 创建执行记录
                    CustomerRelationshipCollaborationExecution.objects.create(
                        collaboration=collaboration,
                        user=request.user,
                        action='attachment_added',
                        content=f'上传了附件：{attachment.file_name}'
                    )
                    
                    messages.success(request, f'附件 {attachment.file_name} 上传成功')
                else:
                    messages.error(request, '附件上传失败，请检查输入')
        
        elif action == 'delete_attachment':
            # 删除附件
            attachment_id = request.POST.get('attachment_id')
            if not (is_responsible or is_collaborator or is_creator or can_edit):
                messages.error(request, '您没有权限删除附件')
            elif attachment_id:
                try:
                    attachment = CustomerRelationshipCollaborationAttachment.objects.get(
                        id=attachment_id,
                        collaboration=collaboration
                    )
                    file_name = attachment.file_name
                    attachment.delete()
                    
                    # 创建执行记录
                    CustomerRelationshipCollaborationExecution.objects.create(
                        collaboration=collaboration,
                        user=request.user,
                        action='attachment_deleted',
                        content=f'删除了附件：{file_name}'
                    )
                    
                    messages.success(request, f'附件 {file_name} 已删除')
                except CustomerRelationshipCollaborationAttachment.DoesNotExist:
                    messages.error(request, '附件不存在')
        
        return redirect('business_pages:customer_relationship_collaboration_detail', collaboration_id=collaboration_id)
    
    # 获取评论列表
    comments = CustomerRelationshipCollaborationComment.objects.filter(
        collaboration=collaboration
    ).select_related('user').order_by('-created_time')
    
    # 获取附件列表
    attachments = CustomerRelationshipCollaborationAttachment.objects.filter(
        collaboration=collaboration
    ).select_related('uploaded_by').order_by('-uploaded_time')
    
    # 获取执行记录
    executions = CustomerRelationshipCollaborationExecution.objects.filter(
        collaboration=collaboration
    ).select_related('user').order_by('-execution_time')
    
    context = _context(
        "人员关系协作申请详情",
        "🤝",
        "查看人员关系协作申请详细信息",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='relationship_collaboration'
    )
    
    context.update({
        'collaboration': collaboration,
        'comments': comments,
        'attachments': attachments,
        'executions': executions,
        'can_edit': can_edit,
        'can_comment': can_comment,
        'is_responsible': is_responsible,
        'is_collaborator': is_collaborator,
        'is_creator': is_creator,
        'status_choices': CustomerRelationshipCollaboration.STATUS_CHOICES,
    })
    
    return render(request, "customer_management/customer_relationship_collaboration_detail.html", context)


# ==================== 合同管理公共函数 =====================

def _apply_contract_filters(queryset, filters):
    """
    应用合同筛选条件（公共函数）
    
    Args:
        queryset: 合同查询集
        filters: 筛选条件字典，包含：
            - search: 搜索关键词
            - status: 状态筛选
            - contract_type: 合同类型筛选
            - client_id: 客户ID筛选
            - project_id: 项目ID筛选
            - date_from: 开始日期筛选
            - date_to: 结束日期筛选
    
    Returns:
        QuerySet: 应用筛选条件后的查询集
    """
    from django.db.models import Q
    
    if filters.get('search'):
        search = filters['search']
        queryset = queryset.filter(
            Q(project_number__icontains=search) |
            Q(contract_name__icontains=search) |
            Q(client__name__icontains=search) |
            Q(project__project_number__icontains=search) |
            Q(project__name__icontains=search)
        )
    
    if filters.get('status'):
        queryset = queryset.filter(status=filters['status'])
    
    if filters.get('contract_type'):
        queryset = queryset.filter(contract_type=filters['contract_type'])
    
    if filters.get('client_id'):
        queryset = queryset.filter(client_id=filters['client_id'])
    
    if filters.get('project_id'):
        queryset = queryset.filter(project_id=filters['project_id'])
    
    if filters.get('date_from'):
        queryset = queryset.filter(contract_date__gte=filters['date_from'])
    
    if filters.get('date_to'):
        queryset = queryset.filter(contract_date__lte=filters['date_to'])
    
    return queryset

@login_required
def contract_management_list(request):
    """
    合同管理列表页面（显示所有状态的合同）
    
    功能：
    - 显示所有状态的合同列表
    - 支持多维度筛选（状态、类型、客户、日期范围）
    - 支持分页显示
    """
    import logging
    from django.core.paginator import Paginator
    
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.contract.view', permission_set):
        messages.error(request, '您没有权限访问合同管理')
        return redirect('business_pages:customer_management_home')
    
    # 获取筛选参数
    filters = {
        'search': request.GET.get('search', ''),
        'status': request.GET.get('status', ''),
        'contract_type': request.GET.get('contract_type', ''),
        'client_id': request.GET.get('client_id', ''),
        'project_id': request.GET.get('project_id', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }
    
    # 获取合同列表
    try:
        contracts = BusinessContract.objects.select_related(
            'client', 'project', 'created_by'
        ).order_by('-created_time')
        
        # 应用筛选条件
        contracts = _apply_contract_filters(contracts, filters)
        
        # 分页
        paginator = Paginator(contracts, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        logger.exception('获取合同列表失败: %s', str(e))
        messages.error(request, f'获取合同列表失败：{str(e)}')
        page_obj = None
    
    # 统计卡片已删除，设置为空列表
    summary_cards = []
    
    # 获取筛选选项
    try:
        clients = Client.objects.filter(is_active=True).order_by('name')[:100]
    except Exception as e:
        logger.exception('获取客户列表失败: %s', str(e))
        clients = []
    
    # 获取项目列表（仅获取有合同的项目）
    projects = []
    try:
        from backend.apps.production_management.models import Project
        contract_project_ids = BusinessContract.objects.filter(
            project__isnull=False
        ).values_list('project_id', flat=True).distinct()[:50]
        
        if contract_project_ids:
            projects = Project.objects.filter(
                id__in=contract_project_ids
            ).order_by('-created_time')[:50]
    except Exception as e:
        logger.exception('获取项目列表失败: %s', str(e))
        projects = []
    
    # 检查创建权限
    can_create = _permission_granted('customer_management.contract.create', permission_set)
    
    context = _context(
        "合同管理",
        "📄",
        "管理所有业务合同",
        summary_cards=summary_cards,
        request=request,
        active_menu_id='contract_management_list',
    )
    
    # 为每个合同对象添加权限属性
    if page_obj:
        for contract in page_obj:
            # 判断是否可以编辑（创建人或具有编辑权限，且状态为草稿）
            contract.can_edit = (
                contract.status == 'draft' and (
                    contract.created_by == request.user or 
                    _permission_granted('customer_management.contract.manage', permission_set)
                )
            )
            # 判断是否可以删除（创建人或具有删除权限，且状态为草稿）
            contract.can_delete = (
                contract.status == 'draft' and (
                    contract.created_by == request.user or 
                    _permission_granted('customer_management.contract.manage', permission_set)
                )
            )
    
    context.update({
        'page_obj': page_obj,
        'search': filters['search'],
        'status': filters['status'],
        'contract_type': filters['contract_type'],
        'client_id': filters['client_id'],
        'project_id': filters['project_id'],
        'date_from': filters['date_from'],
        'date_to': filters['date_to'],
        'clients': clients,
        'projects': projects,
        'status_choices': BusinessContract.CONTRACT_STATUS_CHOICES,
        'type_choices': BusinessContract.CONTRACT_TYPE_CHOICES,
        'can_create': can_create,
    })
    
    return render(request, "customer_management/contract_management_list.html", context)


@login_required
def contract_detail(request, contract_id):
    """
    合同详情页面
    
    功能：
    - 显示合同完整信息
    - 显示关联数据（回款计划、文件、变更记录、子合同等）
    - 支持状态流转操作
    - 支持文件上传和管理
    - 支持创建变更记录
    - 显示审批流程和记录
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.contract.view', permission_set):
        messages.error(request, '您没有权限查看合同详情')
        return redirect('business_pages:contract_management_list')
    
    contract = get_object_or_404(
        BusinessContract.objects.select_related(
            'client', 'project', 'parent_contract', 'created_by', 'opportunity', 'opportunity__business_manager', 'opportunity__client'
        ), 
        id=contract_id
    )
    
    # 获取关联数据
    payment_plans = contract.payment_plans.all().order_by('planned_date')
    
    # 获取回款记录（通过回款计划关联）
    payment_records = []
    try:
        from backend.apps.settlement_management.models import PaymentRecord
        # 获取该合同所有回款计划的ID
        payment_plan_ids = list(payment_plans.values_list('id', flat=True))
        if payment_plan_ids:
            payment_records = PaymentRecord.objects.filter(
                payment_plan_type='business',
                payment_plan_id__in=payment_plan_ids
            ).select_related('created_by', 'confirmed_by').order_by('-payment_date', '-created_time')
    except Exception as e:
        logger.warning(f"获取回款记录失败: {str(e)}")
        payment_records = []
    
    files = contract.files.all().order_by('-uploaded_time')
    approvals = contract.approvals.all().order_by('approval_level', '-created_time')
    changes = contract.changes.all().order_by('-created_time')
    sub_contracts = contract.sub_contracts.all().order_by('-created_time')
    status_logs = contract.status_logs.all().order_by('-created_time')
    
    # 获取可流转的状态列表（包含状态代码和标签）
    valid_transition_codes = BusinessContract.get_valid_transitions(contract.status)
    status_choices_dict = dict(BusinessContract.CONTRACT_STATUS_CHOICES)
    valid_transitions = [
        {'code': code, 'label': status_choices_dict.get(code, code)}
        for code in valid_transition_codes
    ]
    
    # 为状态日志添加状态标签
    status_logs_list = []
    for log in status_logs:
        log_dict = {
            'id': log.id,
            'from_status': log.from_status,
            'from_status_label': status_choices_dict.get(log.from_status, log.from_status) if log.from_status else '初始状态',
            'to_status': log.to_status,
            'to_status_label': status_choices_dict.get(log.to_status, log.to_status),
            'actor': log.actor,
            'comment': log.comment,
            'created_time': log.created_time,
        }
        status_logs_list.append(log_dict)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    can_manage = _check_customer_permission('customer_management.client.edit', permission_set)
    can_edit = can_manage and contract.status == 'draft'  # 只有草稿状态才能编辑
    
    # 获取审批信息
    approval_instance = None
    approval_records = []
    can_submit_approval = False
    try:
        from django.contrib.contenttypes.models import ContentType
        from backend.apps.workflow_engine.models import ApprovalInstance, ApprovalRecord
        
        content_type = ContentType.objects.get_for_model(BusinessContract)
        approval_instance = ApprovalInstance.objects.filter(
            content_type=content_type,
            object_id=contract.id
        ).select_related('workflow', 'applicant', 'current_node').order_by('-created_time').first()
        
        if approval_instance:
            approval_records = ApprovalRecord.objects.filter(
                instance=approval_instance
            ).select_related('node', 'approver', 'transferred_to').order_by('-approval_time')
        
        # 检查是否可以提交审批（有权限且合同状态为草稿、争议或定稿，且没有正在进行的审批）
        can_submit_approval = (
            can_manage and 
            contract.status in ['draft', 'dispute', 'finalized'] and
            (not approval_instance or approval_instance.status not in ['pending', 'in_progress'])
        )
    except Exception:
        pass
    
    # 使用统一的上下文构建函数
    base_context = _context(
        f'合同详情 - {contract.project_number or contract.contract_name or "未命名"}',
        '📃',
        '查看合同详细信息和关联数据',
        request=request,
        active_menu_id='contract_management_list',
    )
    
    # 添加合同详情相关数据
    base_context.update({
        'contract': contract,
        'payment_plans': payment_plans,
        'payment_records': payment_records,
        'files': files,
        'approvals': approvals,
        'changes': changes,
        'sub_contracts': sub_contracts,
        'status_logs': status_logs_list,
        'valid_transitions': valid_transitions,
        'status_choices': status_choices_dict,
        'can_manage': can_manage,
        'can_edit': can_edit,
        'approval_instance': approval_instance,
        'approval_records': approval_records,
        'can_submit_approval': can_submit_approval,
    })
    
    # 调试：确保opportunity被加载
    if hasattr(contract, 'opportunity'):
        logger.info(f"合同 {contract.id} 关联商机: {contract.opportunity}")
    else:
        logger.info(f"合同 {contract.id} 未关联商机")
    
    return render(request, "customer_management/contract_detail.html", base_context)


@login_required
def contract_create(request):
    """
    新建合同页面
    
    功能：
    - 创建新合同
    - 支持从业务委托书转换创建
    - 自动生成合同编号
    - 表单验证和错误处理
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.contract.create', permission_set):
        messages.error(request, '您没有权限创建合同')
        return redirect('business_pages:contract_management_list')
    
    # 检查是否从业务委托书转换而来
    authorization_letter_id = request.GET.get('authorization_letter')
    authorization_letter = None
    if authorization_letter_id:
        try:
            authorization_letter = AuthorizationLetter.objects.get(id=authorization_letter_id)
            if not authorization_letter.can_convert_to_contract():
                messages.warning(request, '只有已确认状态的委托书可以转换为合同')
                authorization_letter = None
        except AuthorizationLetter.DoesNotExist:
            pass
    
    if request.method == 'POST':
        # 处理表单提交
        try:
            from django.db import transaction
            from .forms import ContractForm
            form = ContractForm(request.POST, user=request.user, permission_set=permission_set)
            if form.is_valid():
                with transaction.atomic():
                    contract = form.save(commit=False)
                    contract.created_by = request.user
                    # 合同状态由系统自动判断，默认为合同草稿
                    if not contract.status:
                        contract.status = 'draft'
                    
                    # 如果是从委托书转换而来，继承项目编号
                    if authorization_letter_id:
                        try:
                            letter = AuthorizationLetter.objects.get(id=authorization_letter_id)
                            # 继承业务委托书的项目编号
                            if letter.project_number:
                                contract.project_number = letter.project_number
                            contract.save()
                            messages.success(request, f'合同创建成功（从委托书转换）。')
                        except AuthorizationLetter.DoesNotExist:
                            contract.save()
                            messages.success(request, f'合同创建成功。')
                    else:
                        contract.save()
                        messages.success(request, f'合同创建成功。')
                    
                    # 处理服务内容项
                    from backend.apps.production_management.models import ContractServiceContent, ServiceType, DesignStage, BusinessType, ServiceProfession
                    # 删除旧的服务内容项
                    ContractServiceContent.objects.filter(contract=contract).delete()
                    # 保存新的服务内容项
                    service_contents_data = {}
                    service_professions_data = {}  # 存储每个服务内容项的专业ID列表
                    
                    for key, value in request.POST.items():
                        if key.startswith('service_contents['):
                            # 解析 service_contents[0][service_type] 格式
                            import re
                            match = re.match(r'service_contents\[(\d+)\]\[(\w+)\]', key)
                            if match:
                                index = int(match.group(1))
                                field = match.group(2)
                                if index not in service_contents_data:
                                    service_contents_data[index] = {}
                                service_contents_data[index][field] = value
                            # 解析服务专业复选框 service_contents[0][service_professions]
                            match_profession = re.match(r'service_contents\[(\d+)\]\[service_professions\]', key)
                            if match_profession:
                                index = int(match_profession.group(1))
                                if index not in service_professions_data:
                                    service_professions_data[index] = []
                                if value:  # 复选框被选中
                                    try:
                                        service_professions_data[index].append(int(value))
                                    except ValueError:
                                        pass
                    
                    # 保存服务内容项
                    for index, content_data in service_contents_data.items():
                        # 至少需要服务类型才保存
                        if content_data.get('service_type'):
                            try:
                                service_type_id = int(content_data.get('service_type', 0)) or None
                                design_stage_id = int(content_data.get('design_stage', 0)) or None if content_data.get('design_stage') else None
                                business_type_id = int(content_data.get('business_type', 0)) or None if content_data.get('business_type') else None
                                
                                service_content = ContractServiceContent.objects.create(
                                    contract=contract,
                                    service_type_id=service_type_id,
                                    design_stage_id=design_stage_id,
                                    business_type_id=business_type_id,
                                    description=content_data.get('description', ''),
                                    order=index,
                                )
                                
                                # 保存服务专业（多对多关系）
                                if index in service_professions_data and service_professions_data[index]:
                                    profession_ids = service_professions_data[index]
                                    professions = ServiceProfession.objects.filter(id__in=profession_ids)
                                    service_content.service_professions.set(professions)
                            except (ValueError, TypeError) as e:
                                logger.warning(f'保存服务内容项失败: {str(e)}')
                                continue
                    
                    try:
                        from decimal import Decimal
                        import re
                        
                        # 先删除所有旧的结算方案（重新创建）
                        for key, value in request.POST.items():
                            pass
                    except Exception as e:
                        # 如果保存结算方案失败，记录错误但不影响合同创建
                        logger.warning(f'保存结算方案失败: {str(e)}')
                
                return redirect('business_pages:contract_detail', contract_id=contract.id)
            else:
                messages.error(request, '表单验证失败，请检查输入。')
        except Exception as e:
            logger.exception('创建合同失败: %s', str(e))
            messages.error(request, f'创建合同失败：{str(e)}')
    else:
        from .forms import ContractForm
        # 传递user和permission_set给表单，以便应用权限过滤
        form = ContractForm(user=request.user, permission_set=permission_set)
        
        # 设置责任部门和责任人员（系统自动填充，不可修改）
        if request.user.is_authenticated:
            # 责任部门：当前登录账号对应的部门
            if hasattr(request.user, 'department') and request.user.department:
                form.initial['responsible_department'] = request.user.department.name
            else:
                form.initial['responsible_department'] = '未设置部门'
            # 责任人员：当前登录账号对应的人员姓名
            form.initial['responsible_person'] = request.user.get_full_name() or request.user.username
        
        # 如果是从委托书转换而来，预填充表单
        if authorization_letter:
            # 预填充合同信息
            if authorization_letter.project:
                form.fields['project'].initial = authorization_letter.project
            if authorization_letter.opportunity and authorization_letter.opportunity.client:
                # 尝试找到对应的客户
                try:
                    client = Client.objects.get(name=authorization_letter.client_name)
                    form.fields['client'].initial = client
                except Client.DoesNotExist:
                    pass
            
            # 预填充合同名称
            if not form.initial.get('contract_name'):
                form.initial['contract_name'] = f"{authorization_letter.project_name} - 服务合同"
            
            # 预填充金额
            if authorization_letter.provisional_price:
                form.initial['contract_amount'] = authorization_letter.provisional_price
            
            # 预填充日期
            if authorization_letter.letter_date:
                form.initial['contract_date'] = authorization_letter.letter_date
                form.initial['effective_date'] = authorization_letter.letter_date
                if authorization_letter.start_date:
                    form.initial['start_date'] = authorization_letter.start_date
                if authorization_letter.end_date:
                    form.initial['end_date'] = authorization_letter.end_date
            
            # 预填充签约主体信息
            form.initial['party_a_name'] = authorization_letter.client_name
            form.initial['party_b_name'] = authorization_letter.trustee_name
            
            # 预填充项目编号（继承业务委托书的项目编号）
            if authorization_letter.project_number:
                form.initial['project_number'] = authorization_letter.project_number
    
    # 使用统一的上下文构建函数
    base_context = _context(
        '创建合同草稿',
        '➕',
        '创建新的业务合同',
        request=request,
        active_menu_id='contract_management_list',
    )
    
    from datetime import datetime
    import json
    # 从数据库获取我方单位列表
    from backend.apps.system_management.models import OurCompany
    our_units_list = list(OurCompany.objects.filter(is_active=True).order_by('order', 'id').values_list('company_name', flat=True))
    # 如果没有配置，使用默认值
    if not our_units_list:
        our_units_list = [
            '四川维海科技有限公司',
            '重庆维海科技有限公司',
            '云南维海科技有限公司',
            '西安维海科技有限公司',
            '禾间成都建筑设计咨询有限公司',
            '成都宏天升荣科技有限公司',
        ]
    # 转换为JSON字符串供JavaScript使用
    our_units = json.dumps(our_units_list, ensure_ascii=False)
    # 从后台引入服务内容相关选项
    from backend.apps.production_management.models import BusinessType, ServiceType, DesignStage, ServiceProfession, SettlementNodeType, AfterSalesNodeType
    business_types = BusinessType.objects.filter(is_active=True).order_by('order', 'id')
    service_types = ServiceType.objects.all().order_by('order', 'id')
    design_stages = DesignStage.objects.filter(is_active=True).order_by('order', 'id')
    service_professions = ServiceProfession.objects.all().order_by('service_type__order', 'order', 'id')
    settlement_node_types = SettlementNodeType.objects.filter(is_active=True).order_by('order', 'id')
    after_sales_node_types = AfterSalesNodeType.objects.filter(is_active=True).order_by('order', 'id')
    
    # 获取成果文件类型（用于服务内容的成果清单）
    from backend.apps.production_management.models import ResultFileType
    result_file_types = ResultFileType.objects.filter(is_active=True).order_by('service_category', 'order', 'id')
    
    # 获取结算方式（用于价款信息）
    from backend.apps.settlement_center.models import SettlementMethod
    settlement_methods = SettlementMethod.objects.filter(is_active=True).order_by('sort_order', 'name')
    
    # 获取已有的服务内容项（创建时为空）
    from backend.apps.production_management.models import ContractServiceContent
    existing_service_contents = ContractServiceContent.objects.none()
    
    # 约定管辖选项
    GOVERNING_LAW_CHOICES = [
        ('party_a_location', '甲方所在地'),
        ('party_b_location', '乙方所在地'),
        ('project_location', '项目所在地'),
        ('not_specified', '未约定'),
        ('legal_default', '法定管辖'),
    ]
        # 获取客户数据（用于自动填充客户方信息）
    from backend.apps.customer_management.models import Client, ClientContact
    clients = Client.objects.filter(is_active=True).select_related().prefetch_related('contacts').order_by('name')
    
    # 获取我方签约主体、项目负责人、商务负责人数据
    from backend.apps.system_management.models import User
    # 我方签约主体（从配置中获取，已在our_units中）
    # 项目负责人（所有活跃用户）
    project_managers = User.objects.filter(is_active=True).order_by('username')
    # 商务负责人（默认当前用户）
    business_managers = User.objects.filter(is_active=True).order_by('username')
    

    base_context.update({
        'form': form,
        'contract': None,  # 创建合同时contract为None
        'authorization_letter': authorization_letter,
        'current_year': datetime.now().year,
        'our_units': our_units,
        'business_types': business_types,
        'service_types': service_types,
        'design_stages': design_stages,
        'service_professions': service_professions,
        'settlement_node_types': settlement_node_types,
        'after_sales_node_types': after_sales_node_types,
        'result_file_types': result_file_types,
        'settlement_methods': settlement_methods,
        'existing_service_contents': existing_service_contents,
                'clients': clients,
                'project_managers': project_managers,
        'business_managers': business_managers,
        'governing_law_choices': GOVERNING_LAW_CHOICES,
    })
    
    return render(request, "customer_management/contract_form.html", base_context)


@login_required
def contract_edit(request, contract_id):
    """
    编辑合同页面
    
    功能：
    - 编辑合同信息
    - 仅允许编辑草稿状态的合同
    - 权限检查（创建人或具有编辑权限）
    """
    import logging
    logger = logging.getLogger(__name__)
    
    contract = get_object_or_404(BusinessContract, id=contract_id)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    can_edit = (
        contract.status == 'draft' and (
            contract.created_by == request.user or 
            _permission_granted('customer_management.contract.manage', permission_set)
        )
    )
    
    if not can_edit:
        messages.error(request, '您没有权限编辑此合同，或合同状态不允许编辑（仅草稿状态可编辑）')
        return redirect('business_pages:contract_detail', contract_id=contract.id)
    
    if request.method == 'POST':
        # 处理表单提交
        try:
            from django.db import transaction
            from .forms import ContractForm
            form = ContractForm(request.POST, instance=contract, user=request.user, permission_set=permission_set)
            if form.is_valid():
                with transaction.atomic():
                    contract = form.save(commit=False)
                contract.save()
                
                # 处理服务内容项
                from backend.apps.production_management.models import ContractServiceContent, ServiceType, DesignStage, BusinessType, ServiceProfession
                # 删除旧的服务内容项
                ContractServiceContent.objects.filter(contract=contract).delete()
                # 保存新的服务内容项
                service_contents_data = {}
                service_professions_data = {}  # 存储每个服务内容项的专业ID列表
                
                for key, value in request.POST.items():
                    if key.startswith('service_contents['):
                        # 解析 service_contents[0][service_type] 格式
                        import re
                        match = re.match(r'service_contents\[(\d+)\]\[(\w+)\]', key)
                        if match:
                            index = int(match.group(1))
                            field = match.group(2)
                            if index not in service_contents_data:
                                service_contents_data[index] = {}
                            service_contents_data[index][field] = value
                        # 解析服务专业复选框 service_contents[0][service_professions]
                        match_profession = re.match(r'service_contents\[(\d+)\]\[service_professions\]', key)
                        if match_profession:
                            index = int(match_profession.group(1))
                            if index not in service_professions_data:
                                service_professions_data[index] = []
                            if value:  # 复选框被选中
                                try:
                                    service_professions_data[index].append(int(value))
                                except ValueError:
                                    pass
                
                # 保存服务内容项
                for index, content_data in service_contents_data.items():
                    # 至少需要服务类型才保存
                    if content_data.get('service_type'):
                        try:
                            service_type_id = int(content_data.get('service_type', 0)) or None
                            design_stage_id = int(content_data.get('design_stage', 0)) or None if content_data.get('design_stage') else None
                            business_type_id = int(content_data.get('business_type', 0)) or None if content_data.get('business_type') else None
                            
                            service_content = ContractServiceContent.objects.create(
                                contract=contract,
                                service_type_id=service_type_id,
                                design_stage_id=design_stage_id,
                                business_type_id=business_type_id,
                                description=content_data.get('description', ''),
                                order=index,
                            )
                            
                            # 保存服务专业（多对多关系）
                            if index in service_professions_data and service_professions_data[index]:
                                profession_ids = service_professions_data[index]
                                professions = ServiceProfession.objects.filter(id__in=profession_ids)
                                service_content.service_professions.set(professions)
                        except (ValueError, TypeError) as e:
                            logger.warning(f'保存服务内容项失败: {str(e)}')
                            continue
                
                try:
                    from decimal import Decimal
                    import re
                    
                    # 先删除所有旧的结算方案（重新创建）
                    for key, value in request.POST.items():
                        pass
                except Exception as e:
                    # 如果保存结算方案失败，记录错误但不影响合同更新
                    logger.warning(f'保存结算方案失败: {str(e)}')
                
                messages.success(request, f'合同 {contract.contract_number} 更新成功。')
                return redirect('business_pages:contract_detail', contract_id=contract.id)
            else:
                messages.error(request, '表单验证失败，请检查输入。')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('更新合同失败: %s', str(e))
            messages.error(request, f'更新合同失败：{str(e)}')
    else:
        from .forms import ContractForm
        form = ContractForm(instance=contract, user=request.user, permission_set=permission_set)
        
        # 设置责任部门和责任人员（系统自动填充，不可修改）
        if request.user.is_authenticated:
            # 责任部门：当前登录账号对应的部门
            if hasattr(request.user, 'department') and request.user.department:
                form.initial['responsible_department'] = request.user.department.name
            else:
                form.initial['responsible_department'] = '未设置部门'
            # 责任人员：当前登录账号对应的人员姓名
            form.initial['responsible_person'] = request.user.get_full_name() or request.user.username
    
    # 使用统一的上下文构建函数
    base_context = _context(
        f'编辑合同 - {contract.contract_number}',
        '✏️',
        '编辑合同信息',
        request=request,
        active_menu_id='contract_management_list',
    )
    
    from datetime import datetime
    import json
    # 从数据库获取我方单位列表
    from backend.apps.system_management.models import OurCompany
    our_units_list = list(OurCompany.objects.filter(is_active=True).order_by('order', 'id').values_list('company_name', flat=True))
    # 如果没有配置，使用默认值
    if not our_units_list:
        our_units_list = [
            '四川维海科技有限公司',
            '重庆维海科技有限公司',
            '云南维海科技有限公司',
            '西安维海科技有限公司',
            '禾间成都建筑设计咨询有限公司',
            '成都宏天升荣科技有限公司',
        ]
    # 转换为JSON字符串供JavaScript使用
    our_units = json.dumps(our_units_list, ensure_ascii=False)
    # 从后台引入服务内容相关选项
    from backend.apps.production_management.models import BusinessType, ServiceType, DesignStage, ServiceProfession, SettlementNodeType, AfterSalesNodeType
    business_types = BusinessType.objects.filter(is_active=True).order_by('order', 'id')
    service_types = ServiceType.objects.all().order_by('order', 'id')
    design_stages = DesignStage.objects.filter(is_active=True).order_by('order', 'id')
    service_professions = ServiceProfession.objects.all().order_by('service_type__order', 'order', 'id')
    settlement_node_types = SettlementNodeType.objects.filter(is_active=True).order_by('order', 'id')
    after_sales_node_types = AfterSalesNodeType.objects.filter(is_active=True).order_by('order', 'id')
    
    # 获取成果文件类型（用于生产阶段的节点）
    # 使用交付信息中的文件类型映射（从服务类型获取）
    delivery_file_types = []
    
    # 获取已有的服务内容项
    from backend.apps.production_management.models import ContractServiceContent
    existing_service_contents = ContractServiceContent.objects.filter(
        contract=contract
    ).select_related('service_type', 'design_stage', 'business_type').prefetch_related('service_professions').order_by('order', 'id')
    
    
    base_context.update({
        'form': form,
        'contract': contract,
        'current_year': datetime.now().year,
        'our_units': our_units,
        'business_types': business_types,
        'service_types': service_types,
        'design_stages': design_stages,
        'service_professions': service_professions,
        'settlement_node_types': settlement_node_types,
        'after_sales_node_types': after_sales_node_types,
        'delivery_file_types': delivery_file_types,
        'existing_service_contents': existing_service_contents,
    })
    
    return render(request, "customer_management/contract_form.html", base_context)


@login_required
def contract_delete(request, contract_id):
    """
    删除合同
    
    功能：
    - 仅允许删除草稿状态的合同
    - 检查关联数据，存在关联数据时不允许删除
    - 删除后重定向到合同管理列表
    """
    import logging
    logger = logging.getLogger(__name__)
    
    contract = get_object_or_404(BusinessContract, id=contract_id)
    
    # 权限检查：需要有合同管理权限
    permission_set = get_user_permission_codes(request.user)
    can_delete = (
        contract.status == 'draft' and (
            contract.created_by == request.user or 
            _permission_granted('customer_management.contract.manage', permission_set)
        )
    )
    
    if not can_delete:
        messages.error(request, '您没有权限删除此合同，或合同状态不允许删除（仅草稿状态可删除）')
        return redirect('business_pages:contract_detail', contract_id=contract.id)
    
    if request.method == 'POST':
        try:
            # 检查关联关系
            has_sub_contracts = contract.sub_contracts.exists()
            has_payment_plans = contract.payment_plans.exists()
            
            if has_sub_contracts or has_payment_plans:
                error_msg = '无法删除合同，存在以下关联数据：'
                if has_sub_contracts:
                    error_msg += '子合同、'
                if has_payment_plans:
                    error_msg += '回款计划、'
                error_msg = error_msg.rstrip('、')
                messages.error(request, error_msg)
                return redirect('business_pages:contract_detail', contract_id=contract.id)
            
            contract_number = contract.contract_number
            contract.delete()
            messages.success(request, f'合同 {contract_number} 已删除')
            return redirect('business_pages:contract_management_list')
        except Exception as e:
            logger.exception('删除合同失败: %s', str(e))
            messages.error(request, f'删除合同失败：{str(e)}')
    
    return redirect('business_pages:contract_detail', contract_id=contract.id)


@login_required
def contract_submit_approval(request, contract_id):
    """提交合同审批"""
    contract = get_object_or_404(BusinessContract, id=contract_id)
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _check_customer_permission('customer_management.client.edit', permission_set):
        messages.error(request, '您没有权限提交合同审批')
        return redirect('business_pages:contract_detail', contract_id=contract_id)
    
    # 状态检查：只有草稿或待审核状态的合同才能提交审批
    if contract.status not in ['draft', 'pending_review']:
        messages.error(request, f'合同状态为{contract.get_status_display()}，无法提交审批')
        return redirect('business_pages:contract_detail', contract_id=contract_id)
    
    if request.method == 'POST':
        try:
            from django.contrib.contenttypes.models import ContentType
            from backend.apps.workflow_engine.models import WorkflowTemplate, ApprovalInstance
            from backend.apps.workflow_engine.services import ApprovalEngine
            
            # 检查是否已有正在进行的审批
            content_type = ContentType.objects.get_for_model(BusinessContract)
            existing_instance = ApprovalInstance.objects.filter(
                content_type=content_type,
                object_id=contract.id,
                status__in=['pending', 'in_progress']
            ).first()
            
            if existing_instance:
                messages.warning(request, f'该合同已有正在进行的审批（审批编号：{existing_instance.instance_number}）')
                return redirect('business_pages:contract_detail', contract_id=contract_id)
            
            # 获取审批流程模板
            try:
                workflow = WorkflowTemplate.objects.get(
                    code='contract_approval',
                    status='active'
                )
            except WorkflowTemplate.DoesNotExist:
                # 如果合同审批流程不存在，尝试使用客户管理审批流程
                try:
                    workflow = WorkflowTemplate.objects.get(
                        code='customer_management_approval',
                        status='active'
                    )
                except WorkflowTemplate.DoesNotExist:
                    messages.error(request, '合同审批流程未配置，请联系管理员')
                    return redirect('business_pages:contract_detail', contract_id=contract_id)
            
            # 启动审批流程
            comment = request.POST.get('comment', f'申请审批合同：{contract.contract_number} - {contract.contract_name}')
            instance = ApprovalEngine.start_approval(
                workflow=workflow,
                content_object=contract,
                applicant=request.user,
                comment=comment
            )
            
            # 更新合同状态为待审核
            if contract.status == 'draft':
                contract.status = 'pending_review'
                contract.save()
            
            messages.success(request, f'合同审批已提交（审批编号：{instance.instance_number}）')
            return redirect('business_pages:contract_detail', contract_id=contract_id)
            
        except Exception as e:
            logger.exception('提交合同审批失败: %s', str(e))
            messages.error(request, f'提交合同审批失败：{str(e)}')
            return redirect('business_pages:contract_detail', contract_id=contract_id)
    
    # GET 请求，显示提交审批确认页面
    from django.contrib.contenttypes.models import ContentType
    from backend.apps.workflow_engine.models import ApprovalInstance
    
    # 检查是否已有正在进行的审批
    content_type = ContentType.objects.get_for_model(BusinessContract)
    existing_instance = ApprovalInstance.objects.filter(
        content_type=content_type,
        object_id=contract.id,
        status__in=['pending', 'in_progress']
    ).first()
    
    # 使用统一的上下文构建函数
    base_context = _context(
        f'提交审批 - {contract.contract_number}',
        '📋',
        '提交合同审批流程',
        request=request,
        active_menu_id='contract_management_list',
    )
    
    base_context.update({
        'contract': contract,
        'existing_instance': existing_instance,
    })
    
    return render(request, "customer_management/contract_submit_approval.html", base_context)


@login_required
def contract_dispute_list(request):
    """
    合同争议列表页面
    
    功能：
    - 显示合同争议状态的合同（状态为dispute）
    - 支持筛选和搜索
    - 支持分页显示
    """
    import logging
    from django.core.paginator import Paginator
    
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限访问合同争议')
        return redirect('business_pages:contract_management_list')
    
    # 获取筛选参数
    filters = {
        'search': request.GET.get('search', ''),
        'contract_type': request.GET.get('contract_type', ''),
        'client_id': request.GET.get('client_id', ''),
        'project_id': request.GET.get('project_id', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }
    
    # 获取合同争议状态的合同列表
    try:
        contracts = BusinessContract.objects.filter(
            status='dispute'
        ).select_related('client', 'project', 'created_by').order_by('-created_time')
        
        # 应用筛选条件
        contracts = _apply_contract_filters(contracts, filters)
        
        # 分页
        paginator = Paginator(contracts, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        logger.exception('获取合同争议列表失败: %s', str(e))
        messages.error(request, f'获取合同争议列表失败：{str(e)}')
        page_obj = None
    
    # 统计信息
    try:
        total_count = BusinessContract.objects.filter(status='dispute').count()
        summary_cards = []
    except Exception as e:
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    # 检查创建权限
    can_create = _permission_granted('customer_management.client.create', permission_set)
    
    # 获取筛选选项
    try:
        clients = Client.objects.filter(is_active=True).order_by('name')[:100]
    except Exception as e:
        logger.exception('获取客户列表失败: %s', str(e))
        clients = []
    
    # 获取项目选项
    try:
        contract_project_ids = BusinessContract.objects.filter(
            status='dispute',
            project__isnull=False
        ).values_list('project_id', flat=True).distinct()[:50]
        projects = Project.objects.filter(id__in=contract_project_ids).order_by('name')[:50]
    except Exception as e:
        logger.exception('获取项目列表失败: %s', str(e))
        projects = []
    
    # 获取类型选项
    try:
        type_choices = BusinessContract.CONTRACT_TYPE_CHOICES
    except AttributeError as e:
        logger.exception('获取合同类型选项失败: %s', str(e))
        type_choices = []
    
    context = _context(
        "合同争议",
        "⚖️",
        "管理处于争议状态的合同",
        summary_cards=summary_cards,
        request=request,
        active_menu_id='contract_dispute_list',
    )
    
    context.update({
        'page_obj': page_obj,
        'clients': clients,
        'projects': projects,
        'type_choices': type_choices,
        'search': filters['search'],
        'selected_type': filters['contract_type'],
        'selected_client_id': filters['client_id'],
        'selected_project_id': filters['project_id'],
        'date_from': filters['date_from'],
        'date_to': filters['date_to'],
        'can_create': can_create,
    })
    
    return render(request, "customer_management/contract_list.html", context)


@login_required
def contract_finalize_list(request):
    """
    合同定稿列表页面
    
    功能：
    - 显示合同定稿状态的合同（状态为finalized）
    - 支持筛选和搜索
    - 支持分页显示
    """
    import logging
    from django.core.paginator import Paginator
    
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限访问合同定稿')
        return redirect('business_pages:contract_management_list')
    
    # 获取筛选参数
    filters = {
        'search': request.GET.get('search', ''),
        'contract_type': request.GET.get('contract_type', ''),
        'client_id': request.GET.get('client_id', ''),
        'project_id': request.GET.get('project_id', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }
    
    # 获取合同定稿状态的合同列表
    try:
        contracts = BusinessContract.objects.filter(
            status='finalized'
        ).select_related('client', 'project', 'created_by').order_by('-created_time')
        
        # 应用筛选条件
        contracts = _apply_contract_filters(contracts, filters)
        
        # 分页
        paginator = Paginator(contracts, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        logger.exception('获取合同定稿列表失败: %s', str(e))
        messages.error(request, f'获取合同定稿列表失败：{str(e)}')
        page_obj = None
    
    # 统计信息
    try:
        total_count = BusinessContract.objects.filter(status='finalized').count()
        summary_cards = []
    except Exception as e:
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    # 检查创建权限
    can_create = _permission_granted('customer_management.client.create', permission_set)
    
    # 获取筛选选项
    try:
        clients = Client.objects.filter(is_active=True).order_by('name')[:100]
    except Exception as e:
        logger.exception('获取客户列表失败: %s', str(e))
        clients = []
    
    # 获取项目选项
    try:
        contract_project_ids = BusinessContract.objects.filter(
            status='finalized',
            project__isnull=False
        ).values_list('project_id', flat=True).distinct()[:50]
        projects = Project.objects.filter(id__in=contract_project_ids).order_by('name')[:50]
    except Exception as e:
        logger.exception('获取项目列表失败: %s', str(e))
        projects = []
    
    # 获取类型选项
    try:
        type_choices = BusinessContract.CONTRACT_TYPE_CHOICES
    except AttributeError as e:
        logger.exception('获取合同类型选项失败: %s', str(e))
        type_choices = []
    
    context = _context(
        "合同定稿",
        "📝",
        "管理已定稿的合同（创建流程第三步）",
        summary_cards=summary_cards,
        request=request,
        active_menu_id='contract_finalize_list',
    )
    
    context.update({
        'page_obj': page_obj,
        'clients': clients,
        'projects': projects,
        'type_choices': type_choices,
        'search': filters['search'],
        'selected_type': filters['contract_type'],
        'selected_client_id': filters['client_id'],
        'selected_project_id': filters['project_id'],
        'date_from': filters['date_from'],
        'date_to': filters['date_to'],
        'can_create': can_create,
    })
    
    return render(request, "customer_management/contract_list.html", context)


@login_required
def contract_negotiation_create(request):
    """
    创建合同洽谈记录页面
    
    功能：
    - 创建新的合同洽谈记录
    - 记录洽谈内容、参与人员、时间等信息
    - 关联到具体合同
    """
    import logging
    from .models import ContractNegotiation
    from .forms import ContractNegotiationForm
    
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.create', permission_set):
        messages.error(request, '您没有权限创建合同洽谈记录')
        return redirect('business_pages:contract_management_list')
    
    # 获取关联合同ID（如果从合同详情页跳转）
    contract_id = request.GET.get('contract_id')
    contract = None
    if contract_id:
        try:
            contract = BusinessContract.objects.get(id=contract_id)
        except BusinessContract.DoesNotExist:
            messages.warning(request, '关联的合同不存在')
    
    if request.method == 'POST':
        # 处理表单提交
        try:
            form = ContractNegotiationForm(request.POST, user=request.user)
            if form.is_valid():
                negotiation = form.save(commit=False)
                negotiation.created_by = request.user
                
                # 如果从合同详情页跳转，自动关联合同
                if contract and not negotiation.contract:
                    negotiation.contract = contract
                
                # 如果关联了合同，自动填充客户
                if negotiation.contract and negotiation.contract.client:
                    negotiation.client = negotiation.contract.client
                
                negotiation.save()
                form.save_m2m()  # 保存多对多关系（参与人员）
                
                messages.success(request, '合同洽谈记录创建成功')
                
                # 根据来源决定跳转页面
                if contract:
                    return redirect('business_pages:contract_detail', contract_id=contract.id)
                else:
                    return redirect('business_pages:contract_management_list')
            else:
                messages.error(request, '表单验证失败，请检查输入。')
        except Exception as e:
            logger.exception('创建合同洽谈记录失败: %s', str(e))
            messages.error(request, f'创建合同洽谈记录失败：{str(e)}')
    else:
        # GET请求，显示创建页面
        form = ContractNegotiationForm(user=request.user)
        
        # 如果从合同详情页跳转，预填充合同信息
        if contract:
            form.fields['contract'].initial = contract
            if contract.client:
                form.fields['client'].initial = contract.client
            if contract.project:
                form.fields['project'].initial = contract.project
        
        # 默认参与人员包含当前用户
        form.fields['participants'].initial = [request.user.id]
    
    context = _context(
        '创建合同洽谈记录',
        '💬',
        '记录合同洽谈过程中的关键信息',
        request=request,
        active_menu_id='contract_negotiation_create',
    )
    
    context.update({
        'form': form,
        'contract': contract,
    })
    
    return render(request, "customer_management/contract_negotiation_form.html", context)


@login_required
def contract_negotiation_list(request):
    """
    合同洽谈记录列表页面
    
    功能：
    - 显示所有合同洽谈记录
    - 支持筛选和搜索
    - 支持分页显示
    """
    import logging
    from django.core.paginator import Paginator
    from .models import ContractNegotiation
    from django.db.models import Q
    
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限访问合同洽谈记录')
        return redirect('business_pages:contract_management_list')
    
    # 获取筛选参数
    filters = {
        'search': request.GET.get('search', ''),
        'negotiation_type': request.GET.get('negotiation_type', ''),
        'status': request.GET.get('status', ''),
        'client_id': request.GET.get('client_id', ''),
        'contract_id': request.GET.get('contract_id', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }
    
    # 获取洽谈记录列表
    try:
        negotiations = ContractNegotiation.objects.select_related(
            'contract', 'client', 'project', 'created_by'
        ).prefetch_related('participants').order_by('-negotiation_date', '-created_time')
        
        # 应用筛选条件
        if filters['search']:
            search = filters['search']
            negotiations = negotiations.filter(
                Q(title__icontains=search) |
                Q(content__icontains=search) |
                Q(negotiation_number__icontains=search) |
                Q(client__name__icontains=search) |
                Q(contract__contract_number__icontains=search)
            )
        
        if filters['negotiation_type']:
            negotiations = negotiations.filter(negotiation_type=filters['negotiation_type'])
        
        if filters['status']:
            negotiations = negotiations.filter(status=filters['status'])
        
        if filters['client_id']:
            negotiations = negotiations.filter(client_id=filters['client_id'])
        
        if filters['contract_id']:
            negotiations = negotiations.filter(contract_id=filters['contract_id'])
        
        if filters['date_from']:
            negotiations = negotiations.filter(negotiation_date__gte=filters['date_from'])
        
        if filters['date_to']:
            negotiations = negotiations.filter(negotiation_date__lte=filters['date_to'])
        
        # 分页
        paginator = Paginator(negotiations, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        logger.exception('获取合同洽谈记录列表失败: %s', str(e))
        messages.error(request, f'获取合同洽谈记录列表失败：{str(e)}')
        page_obj = None
    
    # 统计信息
    try:
        total_count = ContractNegotiation.objects.count()
        ongoing_count = ContractNegotiation.objects.filter(status='ongoing').count()
        completed_count = ContractNegotiation.objects.filter(status='completed').count()
        summary_cards = []
    except Exception as e:
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    # 检查创建权限
    can_create = _permission_granted('customer_management.client.create', permission_set)
    
    # 获取筛选选项
    try:
        clients = Client.objects.filter(is_active=True).order_by('name')[:100]
    except Exception as e:
        logger.exception('获取客户列表失败: %s', str(e))
        clients = []
    
    # 获取合同选项
    contracts = []
    try:
        contract_ids = ContractNegotiation.objects.filter(
            contract__isnull=False
        ).values_list('contract_id', flat=True).distinct()[:50]
        if contract_ids:
            contracts = BusinessContract.objects.filter(
                id__in=contract_ids
            ).order_by('-created_time')[:50]
    except Exception as e:
        logger.exception('获取合同列表失败: %s', str(e))
        contracts = []
    
    # 获取类型选项
    type_choices = ContractNegotiation.NEGOTIATION_TYPE_CHOICES
    status_choices = ContractNegotiation.STATUS_CHOICES
    
    context = _context(
        "合同洽谈记录",
        "💬",
        "管理所有合同洽谈记录",
        summary_cards=summary_cards,
        request=request,
        active_menu_id='contract_negotiation_create',
    )
    
    context.update({
        'page_obj': page_obj,
        'clients': clients,
        'contracts': contracts,
        'type_choices': type_choices,
        'status_choices': status_choices,
        'search': filters['search'],
        'selected_type': filters['negotiation_type'],
        'selected_status': filters['status'],
        'selected_client_id': filters['client_id'],
        'selected_contract_id': filters['contract_id'],
        'date_from': filters['date_from'],
        'date_to': filters['date_to'],
        'can_create': can_create,
    })
    
    return render(request, "customer_management/contract_negotiation_list.html", context)


@login_required
def contract_negotiation_detail(request, negotiation_id):
    """
    合同洽谈记录详情页面
    """
    import logging
    from .models import ContractNegotiation
    
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限查看合同洽谈记录')
        return redirect('business_pages:contract_negotiation_list')
    
    negotiation = get_object_or_404(
        ContractNegotiation.objects.select_related(
            'contract', 'client', 'project', 'created_by'
        ).prefetch_related('participants'),
        id=negotiation_id
    )
    
    # 检查编辑权限
    can_edit = (
        negotiation.created_by == request.user or
        _permission_granted('customer_management.client.edit', permission_set)
    )
    
    context = _context(
        f'合同洽谈记录详情 - {negotiation.title}',
        '💬',
        '查看合同洽谈记录的详细信息',
        request=request,
        active_menu_id='contract_negotiation_create',
    )
    
    context.update({
        'negotiation': negotiation,
        'can_edit': can_edit,
    })
    
    return render(request, "customer_management/contract_negotiation_detail.html", context)


@login_required
def contract_finalize_create(request):
    """
    创建合同定稿页面
    
    功能：
    - 创建新合同并直接设置为定稿状态
    - 或者从现有合同创建定稿版本
    - 支持从业务委托书转换创建
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.create', permission_set):
        messages.error(request, '您没有权限创建合同定稿')
        return redirect('business_pages:contract_finalize_list')
    
    # 检查是否从业务委托书转换而来
    authorization_letter_id = request.GET.get('authorization_letter')
    authorization_letter = None
    if authorization_letter_id:
        try:
            authorization_letter = AuthorizationLetter.objects.get(id=authorization_letter_id)
            if not authorization_letter.can_convert_to_contract():
                messages.warning(request, '只有已确认状态的委托书可以转换为合同')
                authorization_letter = None
        except AuthorizationLetter.DoesNotExist:
            pass
    
    # 检查是否从现有合同创建定稿
    contract_id = request.GET.get('contract_id')
    source_contract = None
    if contract_id:
        try:
            source_contract = BusinessContract.objects.get(id=contract_id)
        except BusinessContract.DoesNotExist:
            messages.warning(request, '源合同不存在')
    
    if request.method == 'POST':
        # 处理表单提交
        try:
            from django.db import transaction
            from .forms import ContractForm
            form = ContractForm(request.POST, user=request.user, permission_set=permission_set)
            if form.is_valid():
                with transaction.atomic():
                    contract = form.save(commit=False)
                    contract.created_by = request.user
                    
                    # 合同定稿流程：直接设置为定稿状态
                    contract.status = 'finalized'
                
                # 如果是从委托书转换而来，继承项目编号
                if authorization_letter_id:
                    try:
                        letter = AuthorizationLetter.objects.get(id=authorization_letter_id)
                        if letter.project_number:
                            contract.project_number = letter.project_number
                        contract.save()
                        messages.success(request, f'合同定稿创建成功（从委托书转换），已进入定稿状态。')
                    except AuthorizationLetter.DoesNotExist:
                        contract.save()
                        messages.success(request, f'合同定稿创建成功，已进入定稿状态。')
                elif source_contract:
                    # 从现有合同创建定稿版本
                    contract.save()
                    messages.success(request, f'合同定稿创建成功，已进入定稿状态。')
                else:
                    contract.save()
                    messages.success(request, f'合同定稿创建成功，已进入定稿状态。')
                
                # 处理服务内容项（与contract_create保持一致）
                from backend.apps.production_management.models import ContractServiceContent, ServiceType, DesignStage, BusinessType, ServiceProfession
                # 删除旧的服务内容项
                ContractServiceContent.objects.filter(contract=contract).delete()
                # 保存新的服务内容项
                service_contents_data = {}
                service_professions_data = {}  # 存储每个服务内容项的专业ID列表
                
                for key, value in request.POST.items():
                    if key.startswith('service_contents['):
                        # 解析 service_contents[0][service_type] 格式
                        import re
                        match = re.match(r'service_contents\[(\d+)\]\[(\w+)\]', key)
                        if match:
                            index = int(match.group(1))
                            field = match.group(2)
                            if index not in service_contents_data:
                                service_contents_data[index] = {}
                            service_contents_data[index][field] = value
                        # 解析服务专业复选框 service_contents[0][service_professions]
                        match_profession = re.match(r'service_contents\[(\d+)\]\[service_professions\]', key)
                        if match_profession:
                            index = int(match_profession.group(1))
                            if index not in service_professions_data:
                                service_professions_data[index] = []
                            if value:  # 复选框被选中
                                try:
                                    service_professions_data[index].append(int(value))
                                except ValueError:
                                    pass
                
                # 保存服务内容项
                for index, content_data in service_contents_data.items():
                    # 至少需要服务类型才保存
                    if content_data.get('service_type'):
                        try:
                            service_type_id = int(content_data.get('service_type', 0)) or None
                            design_stage_id = int(content_data.get('design_stage', 0)) or None if content_data.get('design_stage') else None
                            business_type_id = int(content_data.get('business_type', 0)) or None if content_data.get('business_type') else None
                            
                            service_content = ContractServiceContent.objects.create(
                                contract=contract,
                                service_type_id=service_type_id,
                                design_stage_id=design_stage_id,
                                business_type_id=business_type_id,
                                description=content_data.get('description', ''),
                                order=index,
                            )
                            
                            # 保存服务专业（多对多关系）
                            if index in service_professions_data and service_professions_data[index]:
                                profession_ids = service_professions_data[index]
                                professions = ServiceProfession.objects.filter(id__in=profession_ids)
                                service_content.service_professions.set(professions)
                        except (ValueError, TypeError) as e:
                            logger.warning(f'保存服务内容项失败: {str(e)}')
                            continue
                
                try:
                    from decimal import Decimal
                    import re
                    
                    # 先删除所有旧的结算方案（重新创建）
                    for key, value in request.POST.items():
                        pass
                except Exception as e:
                    # 如果保存结算方案失败，记录错误但不影响合同创建
                    logger.warning(f'保存结算方案失败: {str(e)}')
                
                # 创建成功后跳转到合同定稿列表页面
                return redirect('business_pages:contract_finalize_list')
            else:
                messages.error(request, '表单验证失败，请检查输入。')
        except Exception as e:
            logger.exception('创建合同定稿失败: %s', str(e))
            messages.error(request, f'创建合同定稿失败：{str(e)}')
    else:
        from .forms import ContractForm
        form = ContractForm(user=request.user, permission_set=permission_set)
        
        # 设置责任部门和责任人员（系统自动填充，不可修改）
        if request.user.is_authenticated:
            # 责任部门：当前登录账号对应的部门
            if hasattr(request.user, 'department') and request.user.department:
                form.initial['responsible_department'] = request.user.department.name
            else:
                form.initial['responsible_department'] = '未设置部门'
            # 责任人员：当前登录账号对应的人员姓名
            form.initial['responsible_person'] = request.user.get_full_name() or request.user.username
        
        # 合同定稿流程：默认状态为"合同定稿"
        form.initial['status'] = 'finalized'
        
        # 如果是从委托书转换而来，预填充表单
        if authorization_letter:
            if authorization_letter.project:
                form.fields['project'].initial = authorization_letter.project
            if authorization_letter.opportunity and authorization_letter.opportunity.client:
                try:
                    client = Client.objects.get(name=authorization_letter.client_name)
                    form.fields['client'].initial = client
                except Client.DoesNotExist:
                    pass
            
            if not form.initial.get('contract_name'):
                form.initial['contract_name'] = f"{authorization_letter.project_name} - 服务合同"
            
            if authorization_letter.provisional_price:
                form.initial['contract_amount'] = authorization_letter.provisional_price
            
            if authorization_letter.letter_date:
                form.initial['contract_date'] = authorization_letter.letter_date
                form.initial['effective_date'] = authorization_letter.letter_date
                if authorization_letter.start_date:
                    form.initial['start_date'] = authorization_letter.start_date
                if authorization_letter.end_date:
                    form.initial['end_date'] = authorization_letter.end_date
            
            form.initial['party_a_name'] = authorization_letter.client_name
            form.initial['party_b_name'] = authorization_letter.trustee_name
            
            if authorization_letter.project_number:
                form.initial['project_number'] = authorization_letter.project_number
        
        # 如果是从现有合同创建定稿，预填充表单
        if source_contract:
            form.initial['client'] = source_contract.client
            form.initial['project'] = source_contract.project
            form.initial['contract_name'] = source_contract.contract_name
            form.initial['contract_amount'] = source_contract.contract_amount
            form.initial['contract_date'] = source_contract.contract_date
            form.initial['effective_date'] = source_contract.effective_date
            form.initial['start_date'] = source_contract.start_date
            form.initial['end_date'] = source_contract.end_date
            form.initial['project_number'] = source_contract.project_number
    
    base_context = _context(
        '创建合同定稿',
        '✅',
        '创建新的合同定稿，直接进入定稿状态',
        request=request,
        active_menu_id='contract_finalize_create',
    )
    
    from datetime import datetime
    from backend.apps.system_management.models import OurCompany
    our_units = list(OurCompany.objects.filter(is_active=True).order_by('order', 'id').values_list('company_name', flat=True))
    if not our_units:
        our_units = [
            '四川维海科技有限公司',
            '重庆维海科技有限公司',
            '云南维海科技有限公司',
            '西安维海科技有限公司',
            '禾间成都建筑设计咨询有限公司',
            '成都宏天升荣科技有限公司',
        ]
    
    base_context.update({
        'form': form,
        'authorization_letter': authorization_letter,
        'source_contract': source_contract,
        'is_finalize_create': True,  # 标记这是合同定稿创建页面
        'current_year': datetime.now().year,
        'our_units': our_units,
    })
    
    return render(request, "customer_management/contract_form.html", base_context)


@login_required
def contract_performance_track(request):
    """
    履约跟踪页面
    
    功能：
    - 显示执行中的合同列表
    - 跟踪合同履约情况
    - 显示履约进度和关键指标
    """
    import logging
    from django.core.paginator import Paginator
    from django.db.models import Sum, Q
    from django.utils import timezone
    from datetime import timedelta
    
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限访问履约跟踪')
        return redirect('business_pages:contract_management_list')
    
    # 获取筛选参数
    filters = {
        'search': request.GET.get('search', ''),
        'contract_type': request.GET.get('contract_type', ''),
        'client_id': request.GET.get('client_id', ''),
    }
    
    # 获取执行中的合同列表
    try:
        contracts = BusinessContract.objects.filter(
            status__in=['executing', 'effective']
        ).select_related('client', 'project', 'created_by').order_by('-start_date', '-created_time')
        
        # 应用筛选条件
        contracts = _apply_contract_filters(contracts, filters)
        
        # 分页
        paginator = Paginator(contracts, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        logger.exception('获取履约跟踪列表失败: %s', str(e))
        messages.error(request, f'获取履约跟踪列表失败：{str(e)}')
        page_obj = None
    
    # 统计信息
    try:
        executing_count = BusinessContract.objects.filter(status='executing').count()
        effective_count = BusinessContract.objects.filter(status='effective').count()
        total_count = executing_count + effective_count
        
        # 计算履约率（已回款/合同金额）
        total_amount = BusinessContract.objects.filter(
            status__in=['executing', 'effective']
        ).aggregate(total=Sum('contract_amount'))['total'] or 0
        total_payment = BusinessContract.objects.filter(
            status__in=['executing', 'effective']
        ).aggregate(total=Sum('payment_amount'))['total'] or 0
        performance_rate = (total_payment / total_amount * 100) if total_amount > 0 else 0
        
        summary_cards = []
    except Exception as e:
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    # 检查创建权限
    can_create = _permission_granted('customer_management.client.create', permission_set)
    
    # 获取筛选选项
    try:
        clients = Client.objects.filter(is_active=True).order_by('name')[:100]
    except Exception as e:
        logger.exception('获取客户列表失败: %s', str(e))
        clients = []
    
    # 获取类型选项
    try:
        type_choices = BusinessContract.CONTRACT_TYPE_CHOICES
    except AttributeError as e:
        logger.exception('获取合同类型选项失败: %s', str(e))
        type_choices = []
    
    context = _context(
        "履约跟踪",
        "📋",
        "跟踪合同履约情况和执行进度",
        summary_cards=summary_cards,
        request=request,
        active_menu_id='contract_performance',
    )
    
    context.update({
        'page_obj': page_obj,
        'clients': clients,
        'type_choices': type_choices,
        'search': filters['search'],
        'selected_type': filters['contract_type'],
        'selected_client_id': filters['client_id'],
        'can_create': can_create,
    })
    
    return render(request, "customer_management/contract_list.html", context)


@login_required
def contract_expiry_reminder(request):
    """
    到期提醒页面
    
    功能：
    - 显示即将到期的合同
    - 支持设置提醒天数
    - 显示到期时间倒计时
    """
    import logging
    from django.core.paginator import Paginator
    from django.utils import timezone
    from datetime import timedelta
    
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限访问到期提醒')
        return redirect('business_pages:contract_management_list')
    
    # 获取提醒天数（默认30天）
    days_ahead = int(request.GET.get('days', 30))
    
    # 计算到期日期范围
    today = timezone.now().date()
    expiry_date = today + timedelta(days=days_ahead)
    
    # 获取筛选参数
    filters = {
        'search': request.GET.get('search', ''),
        'contract_type': request.GET.get('contract_type', ''),
    }
    
    # 获取即将到期的合同列表
    try:
        contracts = BusinessContract.objects.filter(
            status__in=['executing', 'effective'],
            end_date__isnull=False,
            end_date__lte=expiry_date,
            end_date__gte=today
        ).select_related('client', 'project', 'created_by').order_by('end_date')
        
        # 应用筛选条件
        contracts = _apply_contract_filters(contracts, filters)
        
        # 分页
        paginator = Paginator(contracts, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        logger.exception('获取到期提醒列表失败: %s', str(e))
        messages.error(request, f'获取到期提醒列表失败：{str(e)}')
        page_obj = None
    
    # 统计信息
    try:
        total_count = BusinessContract.objects.filter(
            status__in=['executing', 'effective'],
            end_date__isnull=False,
            end_date__lte=expiry_date,
            end_date__gte=today
        ).count()
        
        # 按到期时间分组统计
        expired_soon = BusinessContract.objects.filter(
            status__in=['executing', 'effective'],
            end_date__isnull=False,
            end_date__lte=today + timedelta(days=7),
            end_date__gte=today
        ).count()
        
        expired_this_month = BusinessContract.objects.filter(
            status__in=['executing', 'effective'],
            end_date__isnull=False,
            end_date__lte=today + timedelta(days=30),
            end_date__gte=today + timedelta(days=7)
        ).count()
        
        summary_cards = []
    except Exception as e:
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    # 获取类型选项
    try:
        type_choices = BusinessContract.CONTRACT_TYPE_CHOICES
    except AttributeError as e:
        logger.exception('获取合同类型选项失败: %s', str(e))
        type_choices = []
    
    context = _context(
        "到期提醒",
        "📅",
        f"提醒未来{days_ahead}天内到期的合同",
        summary_cards=summary_cards,
        request=request,
        active_menu_id='contract_expiry_reminder',
    )
    
    context.update({
        'page_obj': page_obj,
        'type_choices': type_choices,
        'search': filters['search'],
        'selected_type': filters['contract_type'],
        'days_ahead': days_ahead,
    })
    
    return render(request, "customer_management/contract_list.html", context)


@login_required
def contract_payment_reminder(request):
    """
    付款提醒页面
    
    功能：
    - 显示需要付款的合同
    - 跟踪回款进度
    - 显示逾期未回款合同
    """
    import logging
    from django.core.paginator import Paginator
    from django.db.models import Q, F
    from django.utils import timezone
    from datetime import timedelta
    
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限访问付款提醒')
        return redirect('business_pages:contract_management_list')
    
    # 获取筛选参数
    filters = {
        'search': request.GET.get('search', ''),
        'contract_type': request.GET.get('contract_type', ''),
        'overdue_only': request.GET.get('overdue_only', ''),
    }
    
    # 获取需要付款的合同列表（有未回款金额的合同）
    try:
        contracts = BusinessContract.objects.filter(
            status__in=['executing', 'effective', 'signed'],
            contract_amount__gt=0
        ).select_related('client', 'project', 'created_by').order_by('-contract_date')
        
        # 计算未回款金额
        contracts = contracts.annotate(
            unpaid=F('contract_amount') - F('payment_amount')
        ).filter(unpaid__gt=0)
        
        # 应用筛选条件
        contracts = _apply_contract_filters(contracts, filters)
        
        # 如果只显示逾期合同
        if filters['overdue_only']:
            # 获取有回款计划的合同，检查是否有逾期
            from backend.apps.production_management.models import BusinessPaymentPlan
            overdue_contract_ids = BusinessPaymentPlan.objects.filter(
                planned_date__lt=timezone.now().date(),
                actual_payment_date__isnull=True
            ).values_list('contract_id', flat=True).distinct()
            contracts = contracts.filter(id__in=overdue_contract_ids)
        
        # 分页
        paginator = Paginator(contracts, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        logger.exception('获取付款提醒列表失败: %s', str(e))
        messages.error(request, f'获取付款提醒列表失败：{str(e)}')
        page_obj = None
    
    # 统计信息
    try:
        from django.db.models import Sum, ExpressionWrapper, DecimalField
        # 计算待回款总额
        contracts_with_unpaid = BusinessContract.objects.filter(
            status__in=['executing', 'effective', 'signed'],
            contract_amount__gt=0
        ).annotate(
            unpaid=ExpressionWrapper(F('contract_amount') - F('payment_amount'), output_field=DecimalField())
        ).filter(unpaid__gt=0)
        
        total_unpaid = contracts_with_unpaid.aggregate(total=Sum('unpaid'))['total'] or 0
        unpaid_count = contracts_with_unpaid.count()
        
        # 计算逾期合同数量
        from backend.apps.production_management.models import BusinessPaymentPlan
        overdue_count = BusinessPaymentPlan.objects.filter(
            planned_date__lt=timezone.now().date(),
            actual_payment_date__isnull=True
        ).values('contract').distinct().count()
        
        summary_cards = []
    except Exception as e:
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    # 获取类型选项
    try:
        type_choices = BusinessContract.CONTRACT_TYPE_CHOICES
    except AttributeError as e:
        logger.exception('获取合同类型选项失败: %s', str(e))
        type_choices = []
    
    context = _context(
        "付款提醒",
        "💰",
        "跟踪合同回款情况和付款提醒",
        summary_cards=summary_cards,
        request=request,
        active_menu_id='contract_payment_reminder',
    )
    
    context.update({
        'page_obj': page_obj,
        'type_choices': type_choices,
        'search': filters['search'],
        'selected_type': filters['contract_type'],
        'overdue_only': filters['overdue_only'],
    })
    
    return render(request, "customer_management/contract_list.html", context)


@login_required
def contract_risk_warning(request):
    """
    风险预警页面
    
    功能：
    - 显示有风险的合同
    - 识别各种风险类型（逾期、金额异常、状态异常等）
    - 提供风险等级评估
    """
    import logging
    from django.core.paginator import Paginator
    from django.db.models import Q, F
    from django.utils import timezone
    from datetime import timedelta
    
    logger = logging.getLogger(__name__)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限访问风险预警')
        return redirect('business_pages:contract_management_list')
    
    # 获取筛选参数
    filters = {
        'search': request.GET.get('search', ''),
        'risk_type': request.GET.get('risk_type', ''),
    }
    
    # 识别有风险的合同
    try:
        today = timezone.now().date()
        
        # 高风险：已到期但未完成
        high_risk = BusinessContract.objects.filter(
            Q(status__in=['executing', 'effective']) &
            Q(end_date__lt=today)
        )
        
        # 中风险：即将到期（30天内）
        medium_risk = BusinessContract.objects.filter(
            Q(status__in=['executing', 'effective']) &
            Q(end_date__gte=today) &
            Q(end_date__lte=today + timedelta(days=30))
        )
        
        # 低风险：回款异常（未回款金额超过合同金额的50%）
        from django.db.models import ExpressionWrapper, DecimalField, Case, When, Value
        low_risk = BusinessContract.objects.filter(
            Q(status__in=['executing', 'effective', 'signed']) &
            Q(contract_amount__gt=0)
        ).annotate(
            unpaid=ExpressionWrapper(F('contract_amount') - F('payment_amount'), output_field=DecimalField()),
            payment_rate=Case(
                When(contract_amount__gt=0, then=ExpressionWrapper(F('payment_amount') * 100 / F('contract_amount'), output_field=DecimalField())),
                default=Value(0),
                output_field=DecimalField()
            )
        ).filter(
            Q(payment_rate__lt=50) | Q(unpaid__gt=F('contract_amount') * 0.5)
        )
        
        # 合并所有风险合同
        risk_contract_ids = set()
        risk_contract_ids.update(high_risk.values_list('id', flat=True))
        risk_contract_ids.update(medium_risk.values_list('id', flat=True))
        risk_contract_ids.update(low_risk.values_list('id', flat=True))
        
        contracts = BusinessContract.objects.filter(
            id__in=risk_contract_ids
        ).select_related('client', 'project', 'created_by').order_by('-end_date', '-created_time')
        
        # 应用筛选条件
        contracts = _apply_contract_filters(contracts, filters)
        
        # 分页
        paginator = Paginator(contracts, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        logger.exception('获取风险预警列表失败: %s', str(e))
        messages.error(request, f'获取风险预警列表失败：{str(e)}')
        page_obj = None
    
    # 统计信息
    try:
        today = timezone.now().date()
        high_risk_count = BusinessContract.objects.filter(
            Q(status__in=['executing', 'effective']) &
            Q(end_date__lt=today)
        ).count()
        
        medium_risk_count = BusinessContract.objects.filter(
            Q(status__in=['executing', 'effective']) &
            Q(end_date__gte=today) &
            Q(end_date__lte=today + timedelta(days=30))
        ).count()
        
        from django.db.models import ExpressionWrapper, DecimalField, Case, When, Value
        low_risk_count = BusinessContract.objects.filter(
            Q(status__in=['executing', 'effective', 'signed']) &
            Q(contract_amount__gt=0)
        ).annotate(
            payment_rate=Case(
                When(contract_amount__gt=0, then=ExpressionWrapper(F('payment_amount') * 100 / F('contract_amount'), output_field=DecimalField())),
                default=Value(0),
                output_field=DecimalField()
            )
        ).filter(payment_rate__lt=50).count()
        
        summary_cards = []
    except Exception as e:
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    context = _context(
        "风险预警",
        "⚠️",
        "识别和预警合同风险",
        summary_cards=summary_cards,
        request=request,
        active_menu_id='contract_risk_warning',
    )
    
    context.update({
        'page_obj': page_obj,
        'search': filters['search'],
        'selected_risk_type': filters['risk_type'],
    })
    
    return render(request, "customer_management/contract_list.html", context)


@login_required
def project_settlement(request):
    settlements = BusinessPaymentPlan.objects.select_related("contract__project")
    status_counts = settlements.values("status").annotate(total=Count("id"))
    status_map = {row["status"]: row["total"] for row in status_counts}
    summary_cards = []
    latest_settlements = settlements.order_by("-planned_date")[:6]
    section_items = []
    for plan in latest_settlements:
        project = plan.contract.project if plan.contract and plan.contract.project_id else None
        section_items.append({
            'label': f"{project.project_number if project else '未关联'} · {plan.phase_name}",
            'description': f"计划金额 ¥{plan.planned_amount:,.0f} · 状态 {plan.get_status_display()}",
            'url': '#',
            'icon': '💰',
        })
    context = _context(
        "项目结算",
        "🧾",
        "统筹项目回款计划、结算单以及内部核算任务。",
        summary_cards=summary_cards,
        sections=[
            {
                "title": "结算进度",
                "description": "按项目维度查看结算节点和状态。",
                "items": section_items or [
                    {
                        "label": "暂无结算数据",
                        "description": "尚未创建结算计划。",
                        "url": "#",
                        "icon": "ℹ️",
                    }
                ],
            }
        ],
        request=request,
    )
    return render(request, "shared/center_dashboard.html", context)


@login_required
def output_analysis(request):
    from decimal import Decimal  # 在函数开头导入，避免作用域问题
    contracts = BusinessContract.objects.select_related('project')
    payments = BusinessPaymentPlan.objects.all()
    total_contract = contracts.aggregate(total=Sum('contract_amount'))['total'] or Decimal('0')
    total_payment = payments.aggregate(total=Sum('actual_amount'))['total'] or Decimal('0')
    summary_cards = []
    context = _context(
        "产值分析",
        "📊",
        "汇总商务合同与回款数据，为经营分析提供支持。",
        summary_cards=summary_cards,
        sections=[
            {
                "title": "常用报表",
                "description": "产值分析所需的核心报表与数据视图。",
                "items": [
                    {"label": "合同执行情况", "description": "查看合同签订、变更与执行情况。", "url": "#", "icon": "📑"},
                    {"label": "回款趋势分析", "description": "跟踪月度回款走势与贡献度。", "url": "#", "icon": "📈"},
                    {"label": "客户贡献榜", "description": "识别合同金额贡献度较高的客户。", "url": "#", "icon": "🏆"},
                ],
            }
        ],
        request=request,
    )
    return render(request, "shared/center_dashboard.html", context)


@login_required
def payment_tracking(request):
    from decimal import Decimal  # 在函数开头导入，避免作用域问题
    plans = BusinessPaymentPlan.objects.select_related("contract__project").order_by("planned_date")[:8]
    outstanding = sum(
        max((plan.planned_amount or Decimal("0")) - (plan.actual_amount or Decimal("0")), Decimal("0"))
        for plan in plans
        if plan.status in {"pending", "partial", "overdue"}
    )
    summary_cards = []
    section_items = []
    for plan in plans:
        project = plan.contract.project if plan.contract and plan.contract.project_id else None
        section_items.append({
            'label': f"{project.project_number if project else '未关联'} · {plan.phase_name}",
            'description': f"计划金额 ¥{plan.planned_amount:,.0f} · 状态 {plan.get_status_display()}",
            'url': '#',
            'icon': '⏰',
        })
    context = _context(
        "收款跟踪",
        "💵",
        "统一跟踪项目回款节点、提醒通知与实际到账情况。",
        summary_cards=summary_cards,
        sections=[
            {
                "title": "回款计划",
                "description": "重点关注即将到期的回款与提醒。",
                "items": section_items or [
                    {
                        "label": "暂无回款计划",
                        "description": "请在项目中配置回款计划。",
                        "url": "#",
                        "icon": "ℹ️",
                    }
                ],
            }
        ],
        request=request,
    )
    return render(request, "shared/center_dashboard.html", context)


def _calc_progress(summary):
    expected = summary.get("planned_total") or Decimal("0")
    actual = summary.get("actual_total") or Decimal("0")
    if expected == 0:
        return "--"
    return f"{(actual / expected * 100):.0f}%"


def _calc_ratio(value, base):
    if not base:
        return "--"
    return f"{(value / base * 100):.1f}%"


# ==================== 商机管理视图 ====================

@login_required
def opportunity_management(request):
    """商机管理列表页面（根据商机管理专项设计方案）"""
    from django.core.paginator import Paginator
    from datetime import datetime
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    client_id = request.GET.get('client_id', '')
    business_manager_id = request.GET.get('business_manager_id', '')
    urgency = request.GET.get('urgency', '')
    expected_sign_date_from = request.GET.get('expected_sign_date_from', '')
    expected_sign_date_to = request.GET.get('expected_sign_date_to', '')
    tab = request.GET.get('tab', 'all')
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    
    # 获取商机列表
    try:
        opportunities = BusinessOpportunity.objects.select_related(
            'client', 'business_manager', 'created_by'
        ).prefetch_related('followups').order_by('-created_time')
        
        # 权限过滤：普通商务经理只能看自己负责的商机
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            opportunities = opportunities.filter(business_manager=request.user)
        
        # 标签页过滤
        if tab == 'my':
            opportunities = opportunities.filter(business_manager=request.user)
        elif tab == 'subordinate':
            # 下属负责的（需要根据实际业务逻辑实现）
            pass
        
        # 应用筛选条件
        if search:
            opportunities = opportunities.filter(
                Q(opportunity_number__icontains=search) |
                Q(name__icontains=search) |
                Q(project_name__icontains=search) |
                Q(client__name__icontains=search)
            )
        if status:
            opportunities = opportunities.filter(status=status)
        if client_id:
            opportunities = opportunities.filter(client_id=client_id)
        if business_manager_id:
            opportunities = opportunities.filter(business_manager_id=business_manager_id)
        if urgency:
            opportunities = opportunities.filter(urgency=urgency)
        if expected_sign_date_from:
            opportunities = opportunities.filter(expected_sign_date__gte=expected_sign_date_from)
        if expected_sign_date_to:
            opportunities = opportunities.filter(expected_sign_date__lte=expected_sign_date_to)
        
        # 分页
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(opportunities, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取商机列表失败: %s', str(e))
        messages.error(request, f'获取商机列表失败：{str(e)}')
        page_obj = None
    
    # 统计信息
    try:
        # 基础查询集（考虑权限）
        base_queryset = BusinessOpportunity.objects.all()
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            base_queryset = base_queryset.filter(business_manager=request.user)
        
        total_opportunities = base_queryset.count()
        
        # 活跃商机（排除已结束状态）
        active_queryset = base_queryset.exclude(status__in=['won', 'lost', 'cancelled'])
        active_opportunities = active_queryset.count()
        
        # 预计金额总和
        total_estimated = active_queryset.aggregate(total=Sum('estimated_amount'))['total'] or Decimal('0')
        
        # 加权金额总和
        total_weighted_amount = active_queryset.aggregate(total=Sum('weighted_amount'))['total'] or Decimal('0')
        
        # 本月新增（当前月份创建的商机）
        now = timezone.now()
        monthly_new = base_queryset.filter(
            created_time__year=now.year,
            created_time__month=now.month
        ).count()
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    # 获取筛选选项
    clients = Client.objects.filter(is_active=True).order_by('name')
    try:
        business_managers = request.user.__class__.objects.filter(
            roles__code='business_manager'
        ).distinct().order_by('username')
    except:
        business_managers = request.user.__class__.objects.all().order_by('username')[:50]
    
    context = _context(
        "商机管理",
        "💼",
        "从潜在客户到签约项目的全流程数字化管理，实现销售漏斗可视化和过程标准化。",
        summary_cards=summary_cards,
        request=request,
    )
    # 使用完整的顶部菜单
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        # 生成左侧菜单（商机列表页面，激活商机列表项）
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='opportunity_list')
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    context.update({
        'page_obj': page_obj,
        'search': search,
        'status': status,
        'client_id': client_id,
        'business_manager_id': business_manager_id,
        'urgency': urgency,
        'expected_sign_date_from': expected_sign_date_from,
        'expected_sign_date_to': expected_sign_date_to,
        'tab': tab,
        'clients': clients,
        'business_managers': business_managers,
        'status_choices': BusinessOpportunity.STATUS_CHOICES,
        'urgency_choices': BusinessOpportunity.URGENCY_CHOICES,
        'can_create': _permission_granted('customer_management.opportunity.create', permission_set),
    })
    return render(request, "customer_management/opportunity_list.html", context)


@login_required
def opportunity_detail(request, opportunity_id):
    """商机详情页面（根据商机管理专项设计方案）"""
    opportunity = get_object_or_404(
        BusinessOpportunity.objects.select_related('client', 'business_manager', 'created_by', 'approver'),
        id=opportunity_id
    )
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        if opportunity.business_manager != request.user:
            messages.error(request, '您没有权限查看此商机')
            return redirect('business_pages:opportunity_management')
    
    # 获取关联数据
    followups = opportunity.followups.select_related('created_by').order_by('-follow_date', '-created_time')
    quotations = opportunity.quotations.select_related('created_by').order_by('-version_number')[:10]
    
    # 获取审批信息
    approval_instance = None
    approval_records = []
    can_submit_approval = False
    try:
        from django.contrib.contenttypes.models import ContentType
        from backend.apps.workflow_engine.models import ApprovalInstance, ApprovalRecord
        
        content_type = ContentType.objects.get_for_model(BusinessOpportunity)
        approval_instance = ApprovalInstance.objects.filter(
            content_type=content_type,
            object_id=opportunity.id
        ).select_related('workflow', 'applicant', 'current_node').order_by('-created_time').first()
        
        if approval_instance:
            approval_records = ApprovalRecord.objects.filter(
                instance=approval_instance
            ).select_related('node', 'approver', 'transferred_to').order_by('-approval_time')
        
        # 检查是否可以提交审批（有权限且没有正在进行的审批）
        can_submit_approval = _permission_granted('customer_management.opportunity.edit', permission_set) and not approval_instance
    except Exception:
        pass
    
    # 计算健康度评分（如果未计算或需要更新）
    if not opportunity.health_score or opportunity.health_score == 0:
        try:
            # 调用模型的save方法更新健康度
            opportunity.save()
            # 重新获取以获取更新后的健康度
            opportunity.refresh_from_db()
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f'更新商机健康度失败: {str(e)}')
    
    context = _context(
        f"商机详情 - {opportunity.name}",
        "💼",
        f"商机编号：{opportunity.opportunity_number or '未编号'}",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        # 生成左侧菜单（商机详情页面，无激活项）
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id=None)
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    context.update({
        'opportunity': opportunity,
        'followups': followups,
        'quotations': quotations,
        'approval_instance': approval_instance,
        'approval_records': approval_records,
        'can_submit_approval': can_submit_approval,
        'status_choices': BusinessOpportunity.STATUS_CHOICES,
        'urgency_choices': BusinessOpportunity.URGENCY_CHOICES,
        'can_edit': _permission_granted('customer_management.opportunity.edit', permission_set) or opportunity.business_manager == request.user,
        'user': request.user,
    })
    return render(request, "customer_management/opportunity_detail.html", context)


@login_required
def opportunity_create(request):
    """创建商机（根据商机管理专项设计方案）"""
    try:
        permission_set = get_user_permission_codes(request.user)
        if not _permission_granted('customer_management.opportunity.create', permission_set):
            messages.error(request, '您没有权限创建商机')
            return redirect('business_pages:opportunity_management')
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('权限检查失败: %s', str(e))
        messages.error(request, f'权限检查失败：{str(e)}')
        return redirect('business_pages:opportunity_management')
    
    if request.method == 'POST':
        try:
            # 获取并验证必填字段
            client_id = request.POST.get('client_id')
            
            if not client_id:
                messages.error(request, '请选择关联客户')
                return redirect('business_pages:opportunity_create')
            
            # 获取客户信息
            client = Client.objects.get(id=client_id)
            
            # 获取项目名称，用于生成默认商机名称
            project_name = request.POST.get('project_name', '').strip()
            
            # 自动生成商机名称：客户名称 + 项目名称（如果有）
            if project_name:
                name = f"{client.name} - {project_name}"
            else:
                name = client.name
            
            # 获取数值字段
            estimated_amount = Decimal(request.POST.get('estimated_amount', '0') or '0')
            success_probability = int(request.POST.get('success_probability', 10))
            building_area = request.POST.get('building_area')
            
            # 获取服务类型ID
            service_type_id = request.POST.get('service_type_id') or None
            
            # 获取图纸阶段ID
            drawing_stage_id = request.POST.get('drawing_stage') or None
            drawing_stage_obj = None
            if drawing_stage_id:
                try:
                    drawing_stage_obj = DesignStage.objects.filter(id=drawing_stage_id, is_active=True).first()
                except (ValueError, TypeError):
                    pass
            
            opportunity = BusinessOpportunity.objects.create(
                name=name,
                client_id=client_id,
                business_manager=request.user,  # 表单由谁填写，商务就是谁
                status='potential',  # 新建商机默认状态为潜在客户
                opportunity_type=request.POST.get('opportunity_type') or None,
                service_type_id=service_type_id,
                urgency=request.POST.get('urgency', 'normal'),
                project_name=request.POST.get('project_name', '').strip(),
                project_address=request.POST.get('project_address', '').strip(),
                project_type=request.POST.get('project_type', '').strip(),
                building_area=Decimal(building_area) if building_area else None,
                drawing_stage=drawing_stage_obj,
                estimated_amount=estimated_amount,
                success_probability=success_probability,
                expected_sign_date=request.POST.get('expected_sign_date') or None,
                description=request.POST.get('description', '').strip(),
                created_by=request.user,
            )
            # 计算加权金额
            opportunity.weighted_amount = estimated_amount * Decimal(success_probability) / Decimal('100')
            opportunity.save()
            messages.success(request, f'商机 "{opportunity.name}" 创建成功')
            return redirect('business_pages:opportunity_detail', opportunity_id=opportunity.id)
        except ValueError as e:
            messages.error(request, f'数据格式错误：{str(e)}')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('创建商机失败: %s', str(e))
            messages.error(request, f'创建商机失败：{str(e)}')
    
    # GET请求，显示表单
    try:
        from backend.apps.production_management.models import ServiceType, Project
        from django.db.models import Max
        from datetime import datetime
        
        clients = Client.objects.filter(is_active=True).order_by('name')
        service_types = ServiceType.objects.all().order_by('order', 'name')
        design_stages = DesignStage.objects.filter(is_active=True).order_by('order', 'id')
        
        # 生成商机编号预览
        current_date = datetime.now().strftime('%Y%m%d')
        date_prefix = f'SJ-{current_date}-'
        max_opp = BusinessOpportunity.objects.filter(
            opportunity_number__startswith=date_prefix
        ).aggregate(max_num=Max('opportunity_number'))['max_num']
        
        if max_opp:
            try:
                seq = int(max_opp.split('-')[-1]) + 1
            except (ValueError, IndexError):
                seq = 1
        else:
            seq = 1
        
        preview_opportunity_number = f'{date_prefix}{seq:04d}'
        
        context = _context(
            "创建商机",
            "➕",
            "填写以下信息创建新商机",
            request=request,
        )
        if request and request.user.is_authenticated:
            context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
            # 生成左侧菜单（商机创建页面，激活"商机创建"菜单项）
            context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='opportunity_create')
        else:
            context['full_top_nav'] = []
            context['customer_menu'] = []
        context.update({
            'clients': clients,
            'service_types': service_types,
            'design_stages': design_stages,
            'urgency_choices': BusinessOpportunity.URGENCY_CHOICES,
            'business_types': Project.BUSINESS_TYPES,
            'preview_opportunity_number': preview_opportunity_number,
        })
        return render(request, "customer_management/opportunity_form.html", context)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('创建商机表单加载失败: %s', str(e))
        messages.error(request, f'加载创建商机表单失败：{str(e)}')
        return redirect('business_pages:opportunity_management')


@login_required
def opportunity_edit(request, opportunity_id):
    """编辑商机（根据商机管理专项设计方案）"""
    opportunity = get_object_or_404(BusinessOpportunity, id=opportunity_id)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.opportunity.edit', permission_set):
        if opportunity.business_manager != request.user:
            messages.error(request, '您没有权限编辑此商机')
            return redirect('business_pages:opportunity_detail', opportunity_id=opportunity.id)
    
    if request.method == 'POST':
        try:
            # 获取并验证必填字段
            name = request.POST.get('name', '').strip()
            client_id = request.POST.get('client_id')
            
            if not client_id:
                messages.error(request, '请选择关联客户')
                return redirect('business_pages:opportunity_edit', opportunity_id=opportunity.id)
            
            # 获取客户信息
            client = Client.objects.get(id=client_id)
            
            # 获取项目名称，用于生成默认商机名称
            project_name = request.POST.get('project_name', '').strip()
            
            # 如果表单中没有提供商机名称，则自动生成
            if not name:
                # 自动生成商机名称：客户名称 + 项目名称（如果有）
                if project_name:
                    name = f"{client.name} - {project_name}"
                else:
                    name = client.name
            
            # 获取数值字段
            estimated_amount = Decimal(request.POST.get('estimated_amount', '0') or '0')
            success_probability = int(request.POST.get('success_probability', 10))
            building_area = request.POST.get('building_area')
            
            opportunity.name = name
            opportunity.client_id = client_id
            # 负责商务和商机状态不可在编辑时修改
            # business_manager 保持不变（由创建人决定）
            # status 保持不变（通过状态流转功能修改）
            opportunity.opportunity_type = request.POST.get('opportunity_type') or None
            opportunity.service_type_id = request.POST.get('service_type_id') or None
            opportunity.urgency = request.POST.get('urgency')
            opportunity.project_name = request.POST.get('project_name', '').strip()
            opportunity.project_address = request.POST.get('project_address', '').strip()
            opportunity.project_type = request.POST.get('project_type', '').strip()
            opportunity.building_area = Decimal(building_area) if building_area else None
            
            # 获取图纸阶段ID
            drawing_stage_id = request.POST.get('drawing_stage') or None
            drawing_stage_obj = None
            if drawing_stage_id:
                try:
                    drawing_stage_obj = DesignStage.objects.filter(id=drawing_stage_id, is_active=True).first()
                except (ValueError, TypeError):
                    pass
            opportunity.drawing_stage = drawing_stage_obj
            opportunity.estimated_amount = estimated_amount
            opportunity.success_probability = success_probability
            opportunity.expected_sign_date = request.POST.get('expected_sign_date') or None
            opportunity.description = request.POST.get('description', '').strip()
            # 计算加权金额
            opportunity.weighted_amount = estimated_amount * Decimal(success_probability) / Decimal('100')
            opportunity.save()
            messages.success(request, f'商机 "{opportunity.name}" 更新成功')
            return redirect('business_pages:opportunity_detail', opportunity_id=opportunity.id)
        except ValueError as e:
            messages.error(request, f'数据格式错误：{str(e)}')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('更新商机失败: %s', str(e))
            messages.error(request, f'更新商机失败：{str(e)}')
    
    # GET请求，显示表单
    from backend.apps.production_management.models import ServiceType, Project
    
    clients = Client.objects.filter(is_active=True).select_related('responsible_user').order_by('name')
    service_types = ServiceType.objects.all().order_by('order', 'name')
    design_stages = DesignStage.objects.filter(is_active=True).order_by('order', 'id')
    
    context = _context(
        f"编辑商机 - {opportunity.name}",
        "✏️",
        f"商机编号：{opportunity.opportunity_number or '未编号'}",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context.update({
        'opportunity': opportunity,
        'clients': clients,
        'service_types': service_types,
        'design_stages': design_stages,
        'urgency_choices': BusinessOpportunity.URGENCY_CHOICES,
        'business_types': Project.BUSINESS_TYPES,
    })
    return render(request, "customer_management/opportunity_form.html", context)


@login_required
def opportunity_delete(request, opportunity_id):
    """删除商机（根据商机管理专项设计方案）"""
    opportunity = get_object_or_404(BusinessOpportunity, id=opportunity_id)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.opportunity.delete', permission_set):
        if opportunity.business_manager != request.user:
            messages.error(request, '您没有权限删除此商机')
            return redirect('business_pages:opportunity_detail', opportunity_id=opportunity.id)
    
    if request.method == 'POST':
        try:
            opportunity_name = opportunity.name
            opportunity.delete()
            messages.success(request, f'商机 "{opportunity_name}" 已删除')
            return redirect('business_pages:opportunity_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除商机失败: %s', str(e))
            messages.error(request, f'删除商机失败：{str(e)}')
    
    # GET请求，显示确认页面
    context = _context(
        "删除商机",
        "🗑️",
        f"确认删除商机：{opportunity.name}",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context.update({
        'opportunity': opportunity,
    })
    return render(request, "customer_management/opportunity_delete.html", context)


@login_required
def opportunity_status_transition(request, opportunity_id):
    """商机状态流转页面（根据总体设计方案）"""
    from .models import OpportunityStatusLog
    
    opportunity = get_object_or_404(
        BusinessOpportunity.objects.select_related('client', 'business_manager'),
        id=opportunity_id
    )
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.opportunity.edit', permission_set):
        if opportunity.business_manager != request.user:
            messages.error(request, '您没有权限修改此商机状态')
            return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
    
    # 获取可流转的状态
    valid_transitions = BusinessOpportunity.get_valid_transitions(opportunity.status)
    transition_choices = [(status, dict(BusinessOpportunity.STATUS_CHOICES).get(status, status)) 
                          for status in valid_transitions]
    
    # 获取状态流转历史
    status_logs = opportunity.status_logs.select_related('actor').order_by('-created_time')[:20]
    
    if request.method == 'POST':
        target_status = request.POST.get('target_status')
        comment = request.POST.get('comment', '').strip()
        
        if not target_status:
            messages.error(request, '请选择目标状态')
        elif target_status not in valid_transitions:
            messages.error(request, '无效的状态流转')
        else:
            try:
                opportunity.transition_to(target_status, actor=request.user, comment=comment)
                messages.success(request, f'商机状态已从 {opportunity.get_status_display()} 流转到 {dict(BusinessOpportunity.STATUS_CHOICES).get(target_status, target_status)}')
                return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.exception('状态流转失败: %s', str(e))
                messages.error(request, f'状态流转失败：{str(e)}')
    
    context = _context(
        f"状态流转 - {opportunity.name}",
        "🔄",
        f"商机编号：{opportunity.opportunity_number or '未编号'}",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context.update({
        'opportunity': opportunity,
        'transition_choices': transition_choices,
        'status_logs': status_logs,
        'status_choices': BusinessOpportunity.STATUS_CHOICES,
    })
    return render(request, "customer_management/opportunity_status_transition.html", context)


@login_required
def opportunity_followup_create(request, opportunity_id):
    """创建商机跟进记录（根据总体设计方案）"""
    from .models import OpportunityFollowUp
    from datetime import date
    
    opportunity = get_object_or_404(BusinessOpportunity, id=opportunity_id)
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        if opportunity.business_manager != request.user:
            messages.error(request, '您没有权限为此商机创建跟进记录')
            return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
    
    if request.method == 'POST':
        try:
            follow_date = request.POST.get('follow_date')
            follow_type = request.POST.get('follow_type', 'phone')
            participants = request.POST.get('participants', '').strip()
            content = request.POST.get('content', '').strip()
            customer_feedback = request.POST.get('customer_feedback', '').strip()
            next_plan = request.POST.get('next_plan', '').strip()
            next_follow_date = request.POST.get('next_follow_date') or None
            
            # 验证必填字段
            if not follow_date:
                messages.error(request, '跟进日期不能为空')
            elif not content:
                messages.error(request, '跟进内容不能为空')
            else:
                followup = OpportunityFollowUp.objects.create(
                    opportunity=opportunity,
                    follow_date=follow_date,
                    follow_type=follow_type,
                    participants=participants,
                    content=content,
                    customer_feedback=customer_feedback,
                    next_plan=next_plan,
                    next_follow_date=next_follow_date,
                    created_by=request.user,
                )
                messages.success(request, '跟进记录创建成功')
                return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('创建跟进记录失败: %s', str(e))
            messages.error(request, f'创建跟进记录失败：{str(e)}')
    
    context = _context(
        f"创建跟进记录 - {opportunity.name}",
        "📝",
        f"商机编号：{opportunity.opportunity_number or '未编号'}",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context.update({
        'opportunity': opportunity,
        'follow_type_choices': OpportunityFollowUp.FOLLOW_TYPE_CHOICES,
        'default_follow_date': date.today().isoformat(),
    })
    return render(request, "customer_management/opportunity_followup_form.html", context)


@login_required
def opportunity_followup_edit(request, opportunity_id, followup_id):
    """编辑商机跟进记录（根据总体设计方案）"""
    from .models import OpportunityFollowUp
    from datetime import date
    
    opportunity = get_object_or_404(BusinessOpportunity, id=opportunity_id)
    followup = get_object_or_404(OpportunityFollowUp, id=followup_id, opportunity=opportunity)
    
    # 权限检查：仅创建人或管理员可编辑
    permission_set = get_user_permission_codes(request.user)
    if followup.created_by != request.user and not _permission_granted('customer_management.opportunity.edit', permission_set):
        messages.error(request, '您没有权限编辑此跟进记录')
        return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
    
    if request.method == 'POST':
        try:
            follow_date = request.POST.get('follow_date')
            follow_type = request.POST.get('follow_type', 'phone')
            participants = request.POST.get('participants', '').strip()
            content = request.POST.get('content', '').strip()
            customer_feedback = request.POST.get('customer_feedback', '').strip()
            next_plan = request.POST.get('next_plan', '').strip()
            next_follow_date = request.POST.get('next_follow_date') or None
            
            # 验证必填字段
            if not follow_date:
                messages.error(request, '跟进日期不能为空')
            elif not content:
                messages.error(request, '跟进内容不能为空')
            else:
                followup.follow_date = follow_date
                followup.follow_type = follow_type
                followup.participants = participants
                followup.content = content
                followup.customer_feedback = customer_feedback
                followup.next_plan = next_plan
                followup.next_follow_date = next_follow_date
                followup.save()
                messages.success(request, '跟进记录已更新')
                return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('更新跟进记录失败: %s', str(e))
            messages.error(request, f'更新跟进记录失败：{str(e)}')
    
    context = _context(
        f"编辑跟进记录 - {opportunity.name}",
        "✏️",
        f"商机编号：{opportunity.opportunity_number or '未编号'}",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context.update({
        'opportunity': opportunity,
        'followup': followup,
        'follow_type_choices': OpportunityFollowUp.FOLLOW_TYPE_CHOICES,
    })
    return render(request, "customer_management/opportunity_followup_form.html", context)


@login_required
def opportunity_followup_delete(request, opportunity_id, followup_id):
    """删除商机跟进记录（根据总体设计方案）"""
    from .models import OpportunityFollowUp
    
    opportunity = get_object_or_404(BusinessOpportunity, id=opportunity_id)
    followup = get_object_or_404(OpportunityFollowUp, id=followup_id, opportunity=opportunity)
    
    # 权限检查：仅创建人或管理员可删除
    permission_set = get_user_permission_codes(request.user)
    if followup.created_by != request.user and not _permission_granted('customer_management.opportunity.delete', permission_set):
        messages.error(request, '您没有权限删除此跟进记录')
        return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
    
    if request.method == 'POST':
        try:
            followup.delete()
            messages.success(request, '跟进记录已删除')
            return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除跟进记录失败: %s', str(e))
            messages.error(request, f'删除跟进记录失败：{str(e)}')
    
    context = _context(
        f"删除跟进记录 - {opportunity.name}",
        "🗑️",
        f"确认删除跟进记录：{followup.follow_date}",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context.update({
        'opportunity': opportunity,
        'followup': followup,
    })
    return render(request, "customer_management/opportunity_followup_delete.html", context)


@login_required
def opportunity_evaluation_application(request):
    """评估申请页面（根据总体设计方案）"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问评估申请功能')
        return redirect('business_pages:opportunity_management')
    
    # 获取商机列表（用于表单下拉框）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    
    if request.method == 'POST':
        # TODO: 处理表单提交
        messages.success(request, '评估申请已提交')
        return redirect('business_pages:opportunity_evaluation_application')
    
    context = _context(
        "评估申请",
        "📋",
        "提交图纸评估申请",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context['opportunities'] = opportunities[:100]  # 限制显示数量
    return render(request, "customer_management/opportunity_evaluation_application.html", context)


@login_required
def opportunity_warehouse_application(request):
    """入库申请页面"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问入库申请功能')
        return redirect('business_pages:opportunity_management')
    
    # 获取商机列表（用于表单下拉框）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    opportunities = _get_opportunities_safely(opportunities, permission_set, request.user)
    
    if request.method == 'POST':
        # TODO: 处理表单提交
        messages.success(request, '入库申请已提交')
        return redirect('business_pages:opportunity_warehouse_application')
    
    context = _context(
        "入库申请",
        "📦",
        "提交入库申请",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context['opportunities'] = opportunities[:100]  # 限制显示数量
    return render(request, "customer_management/opportunity_warehouse_application.html", context)


@login_required
def opportunity_warehouse_list(request):
    """入库列表页面"""
    from django.core.paginator import Paginator
    
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问入库列表')
        return redirect('business_pages:opportunity_management')
    
    # 获取筛选参数
    search = request.GET.get('search', '').strip()
    opportunity_id = request.GET.get('opportunity_id', '')
    status = request.GET.get('status', '')
    page_size = request.GET.get('page_size', '20')
    
    # 获取入库申请列表（这里暂时使用商机列表作为占位，实际应该查询入库申请记录）
    try:
        # TODO: 如果有入库申请模型，应该查询入库申请记录
        # warehouse_applications = WarehouseApplication.objects.select_related('opportunity', 'created_by').order_by('-created_time')
        
        # 暂时使用空列表，实际应该从数据库查询
        warehouse_applications = []
        
        # 应用搜索条件
        if search:
            # TODO: 如果有模型，应该应用搜索条件
            pass
        
        # 应用筛选条件
        if opportunity_id:
            # TODO: 如果有模型，应该应用筛选条件
            pass
        if status:
            # TODO: 如果有模型，应该应用筛选条件
            pass
        
        # 分页
        paginator = Paginator(warehouse_applications, int(page_size))
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取入库列表失败: %s', str(e))
        messages.error(request, f'获取入库列表失败：{str(e)}')
        page_obj = None
    
    # 获取商机列表（用于筛选）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    opportunities = _get_opportunities_safely(opportunities, permission_set, request.user)
    
    # 检查创建权限
    can_create = _permission_granted('customer_management.opportunity.manage', permission_set)
    
    context = _context(
        "入库列表",
        "📥",
        "管理所有入库申请记录",
        request=request,
    )
    
    # 生成左侧菜单
    context['customer_menu'] = _build_opportunity_management_menu(
        permission_set, 
        active_id='warehouse_list'
    )
    
    context.update({
        'page_obj': page_obj,
        'search': search,
        'opportunity_id': opportunity_id,
        'status': status,
        'opportunities': opportunities[:100],  # 限制显示数量
        'can_create': can_create,
    })
    return render(request, "customer_management/opportunity_warehouse_list.html", context)


@login_required
def opportunity_bid_bond_payment(request):
    """投标保证金支付申请页面"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问投标保证金支付申请功能')
        return redirect('business_pages:opportunity_management')
    
    # 获取商机列表（用于表单下拉框）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    opportunities = _get_opportunities_safely(opportunities, permission_set, request.user)
    
    if request.method == 'POST':
        # TODO: 处理表单提交
        messages.success(request, '投标保证金支付申请已提交')
        return redirect('business_pages:opportunity_bid_bond_payment')
    
    context = _context(
        "投标保证金支付申请",
        "💳",
        "提交投标保证金支付申请",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context['opportunities'] = opportunities[:100]  # 限制显示数量
    return render(request, "customer_management/opportunity_bid_bond_payment.html", context)


@login_required
def opportunity_tender_fee_payment(request):
    """标书费支付申请页面"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问标书费支付申请功能')
        return redirect('business_pages:opportunity_management')
    
    # 获取商机列表（用于表单下拉框）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    opportunities = _get_opportunities_safely(opportunities, permission_set, request.user)
    
    if request.method == 'POST':
        # TODO: 处理表单提交
        messages.success(request, '标书费支付申请已提交')
        return redirect('business_pages:opportunity_tender_fee_payment')
    
    context = _context(
        "标书费支付申请",
        "💵",
        "提交标书费支付申请",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context['opportunities'] = opportunities[:100]  # 限制显示数量
    return render(request, "customer_management/opportunity_tender_fee_payment.html", context)


@login_required
def opportunity_agency_fee_payment(request):
    """招标代理费支付申请页面"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问招标代理费支付申请功能')
        return redirect('business_pages:opportunity_management')
    
    # 获取商机列表（用于表单下拉框）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    opportunities = _get_opportunities_safely(opportunities, permission_set, request.user)
    
    if request.method == 'POST':
        # TODO: 处理表单提交
        messages.success(request, '招标代理费支付申请已提交')
        return redirect('business_pages:opportunity_agency_fee_payment')
    
    context = _context(
        "招标代理费支付申请",
        "💴",
        "提交招标代理费支付申请",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context['opportunities'] = opportunities[:100]  # 限制显示数量
    return render(request, "customer_management/opportunity_agency_fee_payment.html", context)


@login_required
def opportunity_drawing_evaluation(request):
    """图纸评估页面（根据总体设计方案）"""
    from backend.apps.production_management.models import ServiceProfession
    
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问图纸评估功能')
        return redirect('business_pages:opportunity_management')
    
    # 获取商机列表（用于表单下拉框）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    
    # 获取所有服务专业（用于成本节省评估）
    service_professions = ServiceProfession.objects.select_related('service_type').order_by('service_type__order', 'order', 'name')
    
    if request.method == 'POST':
        # TODO: 处理表单提交
        messages.success(request, '图纸评估记录已保存')
        return redirect('business_pages:opportunity_drawing_evaluation')
    
    context = _context(
        "图纸评估",
        "📐",
        "商机图纸评估功能",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        # 生成左侧菜单（图纸评估页面，激活"图纸评估"菜单项）
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='drawing_evaluation')
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    context.update({
        'opportunities': opportunities[:100],  # 限制显示数量
        'service_professions': service_professions,
    })
    return render(request, "customer_management/opportunity_drawing_evaluation.html", context)


@login_required
def opportunity_bidding_quotation(request):
    """投标报价页面（根据总体设计方案，整合资源管理信息）"""
    permission_set = get_user_permission_codes(request.user)
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    opportunity_id = request.GET.get('opportunity_id', '')
    status = request.GET.get('status', '')
    
    # 获取投标报价列表
    try:
        from backend.apps.customer_management.models import BiddingQuotation
        bidding_quotations = BiddingQuotation.objects.select_related(
            'opportunity', 'opportunity__client', 'opportunity__business_manager', 'created_by'
        ).order_by('-bidding_date', '-created_time')
        
        # 权限过滤：只能查看自己创建的或关联商机是自己负责的投标报价
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            bidding_quotations = bidding_quotations.filter(
                Q(created_by=request.user) |
                Q(opportunity__business_manager=request.user)
            )
        
        # 应用筛选条件
        if search:
            bidding_quotations = bidding_quotations.filter(
                Q(bidding_number__icontains=search) |
                Q(opportunity__name__icontains=search) |
                Q(opportunity__opportunity_number__icontains=search) |
                Q(opportunity__client__name__icontains=search)
            )
        if opportunity_id:
            bidding_quotations = bidding_quotations.filter(opportunity_id=opportunity_id)
        if status:
            bidding_quotations = bidding_quotations.filter(status=status)
        
        # 分页
        from django.core.paginator import Paginator
        paginator = Paginator(bidding_quotations, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取投标报价列表失败: %s', str(e))
        messages.error(request, f'获取投标报价列表失败：{str(e)}')
        page_obj = None
    
    # 获取资源管理信息（用于投标报价）
    try:
        # 获取已完成项目（类似业绩）
        from backend.apps.production_management.models import Project
        completed_projects = Project.objects.filter(
            status__in=['completed', 'delivered']
        ).select_related('client').order_by('-end_date')[:50]
        
        # 获取员工档案（用于人员证书）
        from backend.apps.personnel_management.models import Employee, EmployeeArchive
        employees = Employee.objects.filter(status='active').select_related('department')[:100]
        employee_certificates = EmployeeArchive.objects.filter(
            category__in=['certificate', 'qualification', 'license', 'education']
        ).select_related('employee')[:100]
        
        # 获取技术方案（从资源标准模块）
        from backend.apps.resource_standard.models import TechnicalSolution
        technical_solutions = TechnicalSolution.objects.all()[:50]
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取资源管理信息失败: %s', str(e))
        completed_projects = []
        employees = []
        employee_certificates = []
        technical_solutions = []
    
    context = _context(
        "投标报价",
        "💰",
        "商机投标报价管理（整合资源管理信息）",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        # 生成左侧菜单（投标报价申请页面，激活"投标报价申请"菜单项）
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='bidding_quotation_application')
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    # 获取商机列表（用于筛选下拉框）
    try:
        opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            opportunities = opportunities.filter(business_manager=request.user)
        opportunities = opportunities[:100]  # 限制显示数量
    except Exception as e:
        opportunities = []
    
    # 获取状态选项
    from backend.apps.customer_management.models import BiddingQuotation
    from django.utils import timezone
    status_choices = BiddingQuotation.STATUS_CHOICES
    
    context.update({
        'page_obj': page_obj,
        'search': search,
        'opportunity_id': opportunity_id,
        'status': status,
        'opportunities': opportunities,
        'status_choices': status_choices,
        'today': timezone.now().date(),
        'completed_projects': completed_projects,
        'employees': employees,
        'employee_certificates': employee_certificates,
        'technical_solutions': technical_solutions,
    })
    return render(request, "customer_management/opportunity_bidding_quotation.html", context)


@login_required
def opportunity_bidding_quotation_application(request):
    """投标报价申请页面（第一步）"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问投标报价申请功能')
        return redirect('business_pages:opportunity_management')
    
    # 获取商机列表（用于表单下拉框）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    opportunities = _get_opportunities_safely(opportunities, permission_set, request.user)
    
    if request.method == 'POST':
        # TODO: 处理表单提交
        messages.success(request, '投标报价申请已提交')
        return redirect('business_pages:opportunity_bidding_quotation_application')
    
    context = _context(
        "投标报价申请",
        "📝",
        "提交投标报价申请",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context['opportunities'] = opportunities[:100]  # 限制显示数量
    return render(request, "customer_management/opportunity_bidding_quotation_application.html", context)


@login_required
def opportunity_bidding_document_preparation(request):
    """编制投标文件页面（第二步）"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问编制投标文件功能')
        return redirect('business_pages:opportunity_management')
    
    # 获取商机列表（用于表单下拉框）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    opportunities = _get_opportunities_safely(opportunities, permission_set, request.user)
    
    if request.method == 'POST':
        # TODO: 处理表单提交
        messages.success(request, '投标文件编制信息已保存')
        return redirect('business_pages:opportunity_bidding_document_preparation')
    
    context = _context(
        "编制投标文件",
        "📄",
        "编制投标文件信息管理",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context['opportunities'] = opportunities[:100]  # 限制显示数量
    return render(request, "customer_management/opportunity_bidding_document_preparation.html", context)


@login_required
def opportunity_bidding_document_submission(request):
    """递交投标文件页面（第三步）"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问递交投标文件功能')
        return redirect('business_pages:opportunity_management')
    
    # 获取商机列表（用于表单下拉框）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    opportunities = _get_opportunities_safely(opportunities, permission_set, request.user)
    
    if request.method == 'POST':
        # TODO: 处理表单提交
        messages.success(request, '投标文件递交信息已保存')
        return redirect('business_pages:opportunity_bidding_document_submission')
    
    context = _context(
        "递交投标文件",
        "📤",
        "递交投标文件信息管理",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context['opportunities'] = opportunities[:100]  # 限制显示数量
    return render(request, "customer_management/opportunity_bidding_document_submission.html", context)


@login_required
def bidding_quotation_create(request):
    """创建投标报价页面"""
    permission_set = get_user_permission_codes(request.user)
    
    if request.method == 'POST':
        try:
            # 获取并验证必填字段
            opportunity_id = request.POST.get('opportunity_id')
            bidding_date = request.POST.get('bidding_date')
            submission_deadline = request.POST.get('submission_deadline')
            
            if not opportunity_id:
                messages.error(request, '请选择关联商机')
                return redirect('business_pages:bidding_quotation_create')
            if not bidding_date:
                messages.error(request, '投标日期不能为空')
                return redirect('business_pages:bidding_quotation_create')
            if not submission_deadline:
                messages.error(request, '提交截止日期不能为空')
                return redirect('business_pages:bidding_quotation_create')
            
            # 获取商机
            opportunity = get_object_or_404(BusinessOpportunity, id=opportunity_id)
            
            # 创建投标报价记录
            bidding_quotation = BiddingQuotation.objects.create(
                opportunity=opportunity,
                bidding_number=request.POST.get('bidding_number', '').strip(),
                bidding_date=bidding_date,
                submission_deadline=submission_deadline,
                status=request.POST.get('status', 'draft'),
                tender_requirements=request.POST.get('tender_requirements', '').strip(),
                notes=request.POST.get('notes', '').strip(),
                created_by=request.user,
            )
            
            messages.success(request, f'投标报价 "{bidding_quotation.bidding_number or "新建"}" 创建成功')
            return redirect('business_pages:opportunity_bidding_quotation')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('创建投标报价失败: %s', str(e))
            messages.error(request, f'创建投标报价失败：{str(e)}')
    
    # GET请求，显示表单
    # 获取可用的商机列表
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    
    # 权限过滤
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    
    context = _context(
        "创建投标报价",
        "➕",
        "填写以下信息创建新的投标报价",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        # 生成左侧菜单（投标报价页面，激活"投标报价"菜单项）
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='bidding_quotation')
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    context.update({
        'opportunities': opportunities[:100],  # 限制显示数量
        'status_choices': BiddingQuotation.STATUS_CHOICES,
    })
    return render(request, "customer_management/bidding_quotation_form.html", context)


@login_required
def bidding_quotation_detail(request, bidding_id):
    """投标报价详情页面"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限查看投标报价详情')
        return redirect('business_pages:opportunity_bidding_quotation')
    
    try:
        from django.shortcuts import get_object_or_404
        
        bidding_quotation = get_object_or_404(
            BiddingQuotation.objects.select_related(
                'opportunity', 'opportunity__client', 'opportunity__business_manager', 'created_by'
            ),
            id=bidding_id
        )
        
        # 权限过滤：只能查看自己创建的或关联商机是自己负责的投标报价
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            if bidding_quotation.created_by != request.user and bidding_quotation.opportunity.business_manager != request.user:
                messages.error(request, '您没有权限查看此投标报价')
                return redirect('business_pages:opportunity_bidding_quotation')
        
        # 获取关联的类似业绩
        similar_projects = bidding_quotation.similar_projects.select_related('client')[:20]
        
        context = _context(
            f"投标报价详情 - {bidding_quotation.bidding_number or '未编号'}",
            "📋",
            "查看投标报价详细信息",
            request=request,
        )
        if request and request.user.is_authenticated:
            context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        else:
            context['full_top_nav'] = []
        from django.utils import timezone
        context.update({
            'bidding_quotation': bidding_quotation,
            'similar_projects': similar_projects,
            'today': timezone.now().date(),
        })
        return render(request, "customer_management/bidding_quotation_detail.html", context)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('查看投标报价详情失败: %s', str(e))
        messages.error(request, f'查看投标报价详情失败：{str(e)}')
        return redirect('business_pages:opportunity_bidding_quotation')


@login_required
def bidding_quotation_edit(request, bidding_id):
    """投标报价编辑页面"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限编辑投标报价')
        return redirect('business_pages:opportunity_bidding_quotation')
    
    try:
        from django.shortcuts import get_object_or_404
        
        bidding_quotation = get_object_or_404(
            BiddingQuotation.objects.select_related('opportunity', 'opportunity__client'),
            id=bidding_id
        )
        
        # 权限过滤：只能编辑自己创建的或关联商机是自己负责的投标报价
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            if bidding_quotation.created_by != request.user and bidding_quotation.opportunity.business_manager != request.user:
                messages.error(request, '您没有权限编辑此投标报价')
                return redirect('business_pages:opportunity_bidding_quotation')
        
        if request.method == 'POST':
            # 处理表单提交
            bidding_quotation.bidding_number = request.POST.get('bidding_number', '').strip() or bidding_quotation.bidding_number
            bidding_quotation.bidding_date = request.POST.get('bidding_date') or bidding_quotation.bidding_date
            bidding_quotation.submission_deadline = request.POST.get('submission_deadline') or bidding_quotation.submission_deadline
            bidding_quotation.status = request.POST.get('status', bidding_quotation.status)
            bidding_quotation.tender_requirements = request.POST.get('tender_requirements', '').strip()
            bidding_quotation.notes = request.POST.get('notes', '').strip()
            
            # 处理技术标信息（JSON格式）
            technical_proposal = {}
            technical_proposal['technical_solution'] = request.POST.get('technical_solution', '').strip()
            technical_proposal['technical_capability'] = request.POST.get('technical_capability', '').strip()
            technical_proposal['technical_team'] = request.POST.get('technical_team', '').strip()
            technical_proposal['implementation_plan'] = request.POST.get('implementation_plan', '').strip()
            bidding_quotation.technical_proposal = technical_proposal
            
            # 处理商务标信息（JSON格式）
            commercial_proposal = {}
            commercial_proposal['quotation_mode'] = request.POST.get('quotation_mode', 'rate')
            commercial_proposal['saved_amount'] = float(request.POST.get('saved_amount', 0) or 0)
            commercial_proposal['mode_params'] = {}
            
            # 根据报价模式处理参数
            if commercial_proposal['quotation_mode'] == 'rate':
                commercial_proposal['mode_params']['rate'] = float(request.POST.get('rate', 0) or 0) / 100
            elif commercial_proposal['quotation_mode'] == 'base_fee_rate':
                commercial_proposal['mode_params']['base_fee'] = float(request.POST.get('base_fee', 0) or 0)
                commercial_proposal['mode_params']['rate'] = float(request.POST.get('rate', 0) or 0) / 100
            elif commercial_proposal['quotation_mode'] == 'fixed':
                commercial_proposal['mode_params']['fixed_amount'] = float(request.POST.get('fixed_amount', 0) or 0)
            
            commercial_proposal['cap_fee'] = float(request.POST.get('cap_fee', 0) or 0) if request.POST.get('cap_fee') else None
            commercial_proposal['service_fee'] = float(request.POST.get('service_fee', 0) or 0)
            commercial_proposal['payment_method'] = request.POST.get('payment_method', '').strip()
            commercial_proposal['service_commitment'] = request.POST.get('service_commitment', '').strip()
            bidding_quotation.commercial_proposal = commercial_proposal
            
            # 处理类似业绩（多对多关系）
            similar_project_ids = request.POST.getlist('similar_projects')
            bidding_quotation.save()
            if similar_project_ids:
                from backend.apps.production_management.models import Project
                similar_projects = Project.objects.filter(id__in=similar_project_ids)
                bidding_quotation.similar_projects.set(similar_projects)
            
            messages.success(request, f'投标报价 "{bidding_quotation.bidding_number or "未编号"}" 更新成功')
            return redirect('business_pages:bidding_quotation_detail', bidding_id=bidding_quotation.id)
        
        # GET请求，显示编辑表单
        # 获取可用的商机列表
        opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            opportunities = opportunities.filter(business_manager=request.user)
        
        # 获取已完成项目（类似业绩）
        from backend.apps.production_management.models import Project
        completed_projects = Project.objects.filter(
            status__in=['completed', 'delivered']
        ).select_related('client').order_by('-end_date')[:50]
        
        # 获取报价模式选项
        from backend.apps.customer_management.models import OpportunityQuotation
        quotation_mode_choices = OpportunityQuotation._meta.get_field('quotation_mode').choices
        
        context = _context(
            f"编辑投标报价 - {bidding_quotation.bidding_number or '未编号'}",
            "✏️",
            "编辑投标报价信息",
            request=request,
        )
        if request and request.user.is_authenticated:
            context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        else:
            context['full_top_nav'] = []
        context.update({
            'bidding_quotation': bidding_quotation,
            'opportunities': opportunities[:100],
            'completed_projects': completed_projects,
            'status_choices': BiddingQuotation.STATUS_CHOICES,
            'quotation_mode_choices': quotation_mode_choices,
        })
        return render(request, "customer_management/bidding_quotation_edit.html", context)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('编辑投标报价失败: %s', str(e))
        messages.error(request, f'编辑投标报价失败：{str(e)}')
        return redirect('business_pages:opportunity_bidding_quotation')


@login_required
def opportunity_tech_meeting(request):
    """技术沟通会页面（根据总体设计方案）"""
    permission_set = get_user_permission_codes(request.user)
    
    # 权限检查
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限访问技术沟通会功能')
        return redirect('business_pages:opportunity_management')
    
    # 获取商机列表（用于表单下拉框）
    opportunities = BusinessOpportunity.objects.select_related('client', 'business_manager').order_by('-created_time')
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        opportunities = opportunities.filter(business_manager=request.user)
    
    if request.method == 'POST':
        # TODO: 处理表单提交
        messages.success(request, '技术沟通会记录已保存')
        return redirect('business_pages:opportunity_tech_meeting')
    
    context = _context(
        "技术沟通会",
        "🤝",
        "商机技术沟通会功能",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        # 生成左侧菜单（技术沟通会页面，激活"技术沟通会"菜单项）
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='tech_meeting')
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    context['opportunities'] = opportunities[:100]  # 限制显示数量
    return render(request, "customer_management/opportunity_tech_meeting.html", context)


@login_required
def opportunity_followup_list(request):
    """跟进记录列表页面（根据总体设计方案）"""
    permission_set = get_user_permission_codes(request.user)
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    opportunity_id = request.GET.get('opportunity_id', '')
    follow_type = request.GET.get('follow_type', '')
    
    # 获取跟进记录
    try:
        followups = OpportunityFollowUp.objects.select_related(
            'opportunity', 'created_by', 'opportunity__client'
        ).order_by('-follow_date', '-created_time')
        
        # 权限过滤：普通商务经理只能看自己负责的商机的跟进记录
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            followups = followups.filter(opportunity__business_manager=request.user)
        
        # 应用筛选条件
        if search:
            followups = followups.filter(
                Q(content__icontains=search) |
                Q(opportunity__name__icontains=search) |
                Q(opportunity__opportunity_number__icontains=search)
            )
        if opportunity_id:
            followups = followups.filter(opportunity_id=opportunity_id)
        if follow_type:
            followups = followups.filter(follow_type=follow_type)
        
        # 分页
        from django.core.paginator import Paginator
        paginator = Paginator(followups, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取跟进记录列表失败: %s', str(e))
        messages.error(request, f'获取跟进记录列表失败：{str(e)}')
        page_obj = None
    
    context = _context(
        "跟进记录",
        "📝",
        "商机跟进记录管理",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context.update({
        'page_obj': page_obj,
        'search': search,
        'opportunity_id': opportunity_id,
        'follow_type': follow_type,
        'follow_type_choices': OpportunityFollowUp.FOLLOW_TYPE_CHOICES,
        'opportunities': BusinessOpportunity.objects.filter(
            business_manager=request.user
        ).order_by('-created_time')[:50] if not _permission_granted('customer_management.opportunity.view_all', permission_set) 
        else BusinessOpportunity.objects.all().order_by('-created_time')[:100],
    })
    return render(request, "customer_management/opportunity_followup_list.html", context)


@login_required
def opportunity_sales_forecast(request):
    """商机预测页面（根据总体设计方案，API已实现）"""
    from datetime import datetime
    from calendar import monthrange
    from django.db.models import Sum
    from django.utils import timezone
    
    permission_set = get_user_permission_codes(request.user)
    
    # 获取预测月份
    forecast_month = request.GET.get('month', '')
    if not forecast_month:
        today = timezone.now().date()
        forecast_month = f"{today.year}-{today.month:02d}"
    
    try:
        year, month = map(int, forecast_month.split('-'))
        start_date = datetime(year, month, 1).date()
        days_in_month = monthrange(year, month)[1]
        end_date = datetime(year, month, days_in_month).date()
    except (ValueError, IndexError):
        today = timezone.now().date()
        start_date = datetime(today.year, today.month, 1).date()
        days_in_month = monthrange(today.year, today.month)[1]
        end_date = datetime(today.year, today.month, days_in_month).date()
        forecast_month = f"{today.year}-{today.month:02d}"
    
    # 获取活跃商机
    active_opportunities = BusinessOpportunity.objects.exclude(
        status__in=['won', 'lost', 'cancelled']
    )
    
    # 权限过滤
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        active_opportunities = active_opportunities.filter(business_manager=request.user)
    
    # 计算本月预计签约的商机
    month_opportunities = active_opportunities.filter(
        expected_sign_date__gte=start_date,
        expected_sign_date__lte=end_date
    )
    
    # 统计基础数据
    total_active = active_opportunities.count()
    total_weighted_amount = float(active_opportunities.aggregate(
        total=Sum('weighted_amount')
    )['total'] or 0)
    month_weighted_amount = float(month_opportunities.aggregate(
        total=Sum('weighted_amount')
    )['total'] or 0)
    
    # 计算历史转化率
    historical_queryset = BusinessOpportunity.objects.filter(
        status__in=['initial_contact', 'requirement_confirmed', 'quotation', 'negotiation', 'won']
    )
    if not _permission_granted('customer_management.opportunity.view_all', permission_set):
        historical_queryset = historical_queryset.filter(business_manager=request.user)
    
    historical_initial = historical_queryset.count()
    historical_won = historical_queryset.filter(status='won').count()
    
    historical_conversion_rate = 35.0  # 默认值
    if historical_initial > 0:
        historical_conversion_rate = (historical_won / historical_initial) * 100
    
    # 计算预测值（转换为万元）
    optimistic_forecast = (month_weighted_amount * (historical_conversion_rate / 100) * 1.2) / 10000
    neutral_forecast = (month_weighted_amount * (historical_conversion_rate / 100)) / 10000
    conservative_forecast = (month_weighted_amount * (historical_conversion_rate / 100) * 0.8) / 10000
    
    # 目标差距分析
    monthly_target = (total_weighted_amount * 0.6) / 10000
    target_gap = monthly_target - neutral_forecast
    
    # 生成建议
    suggestions = []
    if target_gap > 0:
        suggestions.append('预测金额低于月度目标，建议加大商机开拓力度')
        suggestions.append('建议提升在途商机的转化率')
        suggestions.append('建议重点关注高价值商机，加快推进速度')
    else:
        suggestions.append('预测金额达到月度目标，继续保持')
        suggestions.append('建议持续跟进在途商机，确保按时签约')
    
    forecast_data = {
        'month': forecast_month,
        'active_opportunities': total_active,
        'weighted_amount': total_weighted_amount / 10000,  # 转换为万元
        'historical_conversion_rate': historical_conversion_rate,
        'optimistic': optimistic_forecast,
        'neutral': neutral_forecast,
        'conservative': conservative_forecast,
        'target_gap': {
            'monthly_target': monthly_target,
            'gap': target_gap,
            'suggestions': '\n'.join(suggestions)
        }
    }
    
    context = _context(
        "商机预测",
        "📈",
        "销售预测分析",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        # 生成左侧菜单（商机预测页面，激活"商机预测"菜单项）
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='sales_forecast')
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    context['forecast_data'] = forecast_data
    
    return render(request, "customer_management/opportunity_sales_forecast.html", context)


@login_required
def opportunity_win_loss(request):
    """赢单与输单管理页面（根据商机管理专项设计方案）"""
    from django.core.paginator import Paginator
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')  # 'won' 或 'lost'
    client_id = request.GET.get('client_id', '')
    business_manager_id = request.GET.get('business_manager_id', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限查看赢单与输单信息')
        return redirect('business_pages:opportunity_management')
    
    # 获取赢单和输单商机列表
    try:
        opportunities = BusinessOpportunity.objects.select_related(
            'client', 'business_manager', 'created_by'
        ).filter(status__in=['won', 'lost']).order_by('-updated_time')
        
        # 权限过滤：普通商务经理只能看自己负责的商机
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            opportunities = opportunities.filter(business_manager=request.user)
        
        # 应用筛选条件
        if search:
            opportunities = opportunities.filter(
                Q(opportunity_number__icontains=search) |
                Q(name__icontains=search) |
                Q(project_name__icontains=search) |
                Q(client__name__icontains=search)
            )
        if status_filter in ['won', 'lost']:
            opportunities = opportunities.filter(status=status_filter)
        if client_id:
            opportunities = opportunities.filter(client_id=client_id)
        if business_manager_id:
            opportunities = opportunities.filter(business_manager_id=business_manager_id)
        if date_from:
            opportunities = opportunities.filter(updated_time__gte=date_from)
        if date_to:
            opportunities = opportunities.filter(updated_time__lte=date_to)
        
        # 分页
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(opportunities, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取赢单与输单列表失败: %s', str(e))
        messages.error(request, f'获取赢单与输单列表失败：{str(e)}')
        page_obj = None
    
    # 统计信息
    try:
        # 基础查询集（考虑权限）
        base_queryset = BusinessOpportunity.objects.filter(status__in=['won', 'lost'])
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            base_queryset = base_queryset.filter(business_manager=request.user)
        
        total_count = base_queryset.count()
        won_count = base_queryset.filter(status='won').count()
        lost_count = base_queryset.filter(status='lost').count()
        
        # 赢单金额统计
        won_amount = base_queryset.filter(status='won').aggregate(
            total=Sum('actual_amount')
        )['total'] or Decimal('0')
        
        # 输单金额统计（预计金额）
        lost_amount = base_queryset.filter(status='lost').aggregate(
            total=Sum('estimated_amount')
        )['total'] or Decimal('0')
        
        # 赢单率
        win_rate = 0.0
        if total_count > 0:
            win_rate = (won_count / total_count) * 100
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    # 获取筛选选项
    clients = Client.objects.filter(is_active=True).order_by('name')
    try:
        business_managers = request.user.__class__.objects.filter(
            roles__code='business_manager'
        ).distinct().order_by('username')
    except:
        business_managers = request.user.__class__.objects.all().order_by('username')[:50]
    
    context = _context(
        "赢单与输单",
        "✅",
        "商机赢单与输单管理，记录商机最终结果和原因分析",
        summary_cards=summary_cards,
        request=request,
    )
    # 使用完整的顶部菜单
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        # 生成左侧菜单（赢单与输单页面，激活"赢单与输单"菜单项）
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='win_loss')
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    context.update({
        'page_obj': page_obj,
        'search': search,
        'status': status_filter,
        'client_id': client_id,
        'business_manager_id': business_manager_id,
        'date_from': date_from,
        'date_to': date_to,
        'clients': clients,
        'business_managers': business_managers,
        'status_choices': [('won', '赢单'), ('lost', '输单')],
        'won_amount': won_amount,
        'lost_amount': lost_amount,
    })
    return render(request, "customer_management/opportunity_win_loss.html", context)


@login_required
def opportunity_win_loss_select(request):
    """选择商机并标记为赢单/输单页面"""
    from django.core.paginator import Paginator
    
    # 获取目标状态（won 或 lost）
    target_status = request.GET.get('target_status', '')
    if target_status not in ['won', 'lost']:
        messages.error(request, '无效的目标状态')
        return redirect('business_pages:opportunity_win_loss')
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    client_id = request.GET.get('client_id', '')
    business_manager_id = request.GET.get('business_manager_id', '')
    
    # 获取权限
    permission_set = get_user_permission_codes(request.user)
    
    if not _permission_granted('customer_management.opportunity.edit', permission_set):
        messages.error(request, '您没有权限标记商机为赢单/输单')
        return redirect('business_pages:opportunity_win_loss')
    
    # 获取可以转换为赢单/输单的商机
    # 包括：1) 状态为"商务谈判"的商机 2) 有商务洽谈记录的商机（无论状态）
    try:
        # 获取有商务洽谈记录的商机ID列表
        negotiation_opportunity_ids = BusinessNegotiation.objects.values_list('opportunity_id', flat=True).distinct()
        
        # 获取可以转换的商机：状态为"商务谈判"或有商务洽谈记录
        opportunities = BusinessOpportunity.objects.select_related(
            'client', 'business_manager', 'created_by'
        ).filter(
            Q(status='negotiation') | Q(id__in=negotiation_opportunity_ids)
        ).exclude(
            status__in=['won', 'lost', 'cancelled']  # 排除已结束的商机
        ).order_by('-updated_time')
        
        # 权限过滤：普通商务经理只能看自己负责的商机
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            opportunities = opportunities.filter(business_manager=request.user)
        
        # 应用筛选条件
        if search:
            opportunities = opportunities.filter(
                Q(opportunity_number__icontains=search) |
                Q(name__icontains=search) |
                Q(project_name__icontains=search) |
                Q(client__name__icontains=search)
            )
        if client_id:
            opportunities = opportunities.filter(client_id=client_id)
        if business_manager_id:
            opportunities = opportunities.filter(business_manager_id=business_manager_id)
        
        # 分页
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(opportunities, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取商机列表失败: %s', str(e))
        messages.error(request, f'获取商机列表失败：{str(e)}')
        page_obj = None
    
    # 获取筛选选项
    clients = Client.objects.filter(is_active=True).order_by('name')
    try:
        business_managers = request.user.__class__.objects.filter(
            roles__code='business_manager'
        ).distinct().order_by('username')
    except:
        business_managers = request.user.__class__.objects.all().order_by('username')[:50]
    
    status_label = '赢单' if target_status == 'won' else '输单'
    
    context = _context(
        f"选择商机 - 标记为{status_label}",
        "✅" if target_status == 'won' else "❌",
        f"选择要标记为{status_label}的商机",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='win_loss')
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    context.update({
        'page_obj': page_obj,
        'search': search,
        'client_id': client_id,
        'business_manager_id': business_manager_id,
        'clients': clients,
        'business_managers': business_managers,
        'target_status': target_status,
        'status_label': status_label,
    })
    return render(request, "customer_management/opportunity_win_loss_select.html", context)


@login_required
def opportunity_mark_win_loss(request, opportunity_id):
    """快速标记商机为赢单或输单"""
    opportunity = get_object_or_404(
        BusinessOpportunity.objects.select_related('client', 'business_manager'),
        id=opportunity_id
    )
    
    # 权限检查
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.opportunity.edit', permission_set):
        if opportunity.business_manager != request.user:
            messages.error(request, '您没有权限修改此商机状态')
            return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
    
    # 获取目标状态
    target_status = request.GET.get('target_status', '')
    if target_status not in ['won', 'lost']:
        messages.error(request, '无效的目标状态')
        return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
    
    # 检查是否可以转换
    # 允许转换的情况：1) 状态转换规则允许 2) 有商务洽谈记录（说明已进入商务阶段）
    can_transition = opportunity.can_transition_to(target_status)
    has_negotiation = BusinessNegotiation.objects.filter(opportunity=opportunity).exists()
    
    if not can_transition and not has_negotiation:
        messages.error(request, f'当前商机状态为"{opportunity.get_status_display()}"，无法直接标记为{"赢单" if target_status == "won" else "输单"}。请先将商机状态转换为"商务谈判"，或创建商务洽谈记录。')
        return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
    
    # 如果有商务洽谈记录但状态不允许直接转换，先更新状态为"商务谈判"
    if not can_transition and has_negotiation and opportunity.status != 'negotiation':
        # 如果当前状态可以转换为"商务谈判"，先转换状态
        if opportunity.can_transition_to('negotiation'):
            try:
                opportunity.transition_to('negotiation', actor=request.user, comment='自动转换为商务谈判状态（因为有商务洽谈记录）')
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'自动转换状态失败: {str(e)}')
                # 继续执行，允许直接标记
    
    if request.method == 'POST':
        comment = request.POST.get('comment', '').strip()
        try:
            # 先更新额外信息字段（在状态转换之前）
            if target_status == 'won':
                actual_amount = request.POST.get('actual_amount', '').strip()
                contract_number = request.POST.get('contract_number', '').strip()
                actual_sign_date = request.POST.get('actual_sign_date', '').strip()
                win_reason = request.POST.get('win_reason', '').strip()
                
                if actual_amount:
                    try:
                        opportunity.actual_amount = Decimal(actual_amount)
                    except (ValueError, InvalidOperation):
                        pass
                if contract_number:
                    opportunity.contract_number = contract_number
                if actual_sign_date:
                    try:
                        from datetime import datetime
                        opportunity.actual_sign_date = datetime.strptime(actual_sign_date, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                if win_reason:
                    opportunity.win_reason = win_reason
            elif target_status == 'lost':
                loss_reason = request.POST.get('loss_reason', '').strip()
                if loss_reason:
                    opportunity.loss_reason = loss_reason
            
            # 执行状态流转（这会保存所有字段，包括状态）
            opportunity.transition_to(target_status, actor=request.user, comment=comment)
            
            # 从数据库重新加载对象以确保状态已更新
            opportunity.refresh_from_db()
            
            # 验证状态是否已更新
            if opportunity.status != target_status:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'状态更新失败：期望状态={target_status}，实际状态={opportunity.status}')
                messages.error(request, '状态更新失败，请重试')
                return redirect('business_pages:opportunity_detail', opportunity_id=opportunity_id)
            
            status_label = '赢单' if target_status == 'won' else '输单'
            messages.success(request, f'商机已成功标记为{status_label}')
            return redirect('business_pages:opportunity_win_loss')
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('标记商机失败: %s', str(e))
            messages.error(request, f'标记商机失败：{str(e)}')
    
    # GET 请求，显示确认表单
    status_label = '赢单' if target_status == 'won' else '输单'
    context = _context(
        f"标记为{status_label} - {opportunity.name}",
        "✅" if target_status == 'won' else "❌",
        f"商机编号：{opportunity.opportunity_number or '未编号'}",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='win_loss')
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    context.update({
        'opportunity': opportunity,
        'target_status': target_status,
        'status_label': status_label,
    })
    return render(request, "customer_management/opportunity_mark_win_loss.html", context)


@login_required
def opportunity_business_negotiation(request):
    """商务洽谈页面（根据总体设计方案）"""
    permission_set = get_user_permission_codes(request.user)
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    opportunity_id = request.GET.get('opportunity_id', '')
    
    # 获取商务洽谈记录列表
    try:
        negotiations = BusinessNegotiation.objects.select_related(
            'opportunity', 'opportunity__client', 'opportunity__business_manager', 'created_by'
        ).order_by('-negotiation_date', '-created_time')
        
        # 权限过滤：普通商务经理只能看自己负责的商机的洽谈记录
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            negotiations = negotiations.filter(opportunity__business_manager=request.user)
        
        # 应用筛选条件
        if search:
            negotiations = negotiations.filter(
                Q(opportunity__name__icontains=search) |
                Q(opportunity__opportunity_number__icontains=search) |
                Q(opportunity__client__name__icontains=search) |
                Q(content__icontains=search)
            )
        if opportunity_id:
            negotiations = negotiations.filter(opportunity_id=opportunity_id)
        
        # 分页
        from django.core.paginator import Paginator
        paginator = Paginator(negotiations, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取商务洽谈记录失败: %s', str(e))
        messages.error(request, f'获取商务洽谈记录失败：{str(e)}')
        page_obj = None
    
    # 获取商机列表（用于筛选下拉框）
    try:
        opportunities_for_filter = BusinessOpportunity.objects.select_related(
            'client', 'business_manager'
        ).order_by('-created_time')
        
        # 权限过滤
        if not _permission_granted('customer_management.opportunity.view_all', permission_set):
            opportunities_for_filter = opportunities_for_filter.filter(business_manager=request.user)
        
        opportunities_for_filter = opportunities_for_filter[:100]  # 限制数量
    except Exception as e:
        opportunities_for_filter = []
    
    context = _context(
        "商务洽谈登记",
        "💬",
        "商机商务洽谈登记管理",
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
        # 生成左侧菜单（商务洽谈登记页面，激活"商务洽谈登记"菜单项）
        context['customer_menu'] = _build_opportunity_management_menu(permission_set, active_id='business_negotiation')
    else:
        context['full_top_nav'] = []
        context['customer_menu'] = []
    context.update({
        'page_obj': page_obj,
        'search': search,
        'opportunity_id': opportunity_id,
        'opportunities': opportunities_for_filter,
    })
    return render(request, "customer_management/opportunity_business_negotiation.html", context)


@login_required
def opportunity_business_negotiation_form(request, opportunity_id=None):
    """商务洽谈表单页面（创建/编辑）"""
    permission_set = get_user_permission_codes(request.user)
    
    if opportunity_id:
        opportunity = get_object_or_404(BusinessOpportunity, id=opportunity_id)
        # 权限检查
        if not _permission_granted('customer_management.opportunity.view', permission_set):
            if opportunity.business_manager != request.user:
                messages.error(request, '您没有权限查看此商机')
                return redirect('business_pages:opportunity_business_negotiation')
    else:
        opportunity = None
    
    if request.method == 'POST':
        try:
            # 获取表单数据
            opportunity_id = request.POST.get('opportunity_id')
            if not opportunity_id:
                messages.error(request, '请选择关联商机')
                return redirect('business_pages:opportunity_business_negotiation_form', opportunity_id=opportunity_id) if opportunity_id else redirect('business_pages:opportunity_business_negotiation_form')
            
            opp = get_object_or_404(BusinessOpportunity, id=opportunity_id)
            
            # 权限检查
            if not _permission_granted('customer_management.opportunity.view', permission_set):
                if opp.business_manager != request.user:
                    messages.error(request, '您没有权限为此商机创建洽谈登记')
                    return redirect('business_pages:opportunity_business_negotiation')
            
            # 创建商务洽谈记录
            negotiation = BusinessNegotiation.objects.create(
                opportunity=opp,
                negotiation_date=request.POST.get('negotiation_date'),
                negotiation_type=request.POST.get('negotiation_type'),
                participants=request.POST.get('participants', ''),
                content=request.POST.get('content'),
                client_feedback=request.POST.get('client_feedback', ''),
                next_plan=request.POST.get('next_plan', ''),
                discussed_amount=request.POST.get('discussed_amount') or None,
                payment_terms=request.POST.get('payment_terms', ''),
                contract_terms=request.POST.get('contract_terms', ''),
                created_by=request.user
            )
            
            messages.success(request, '商务洽谈登记已保存')
            return redirect('business_pages:opportunity_business_negotiation')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('保存商务洽谈记录失败: %s', str(e))
            messages.error(request, f'保存失败：{str(e)}')
    
    description = f"商机：{opportunity.name}" if opportunity else "创建新的商务洽谈登记"
    context = _context(
        f"{'编辑' if opportunity_id else '创建'}商务洽谈登记",
        "💬",
        description,
        request=request,
    )
    if request and request.user.is_authenticated:
        context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
    else:
        context['full_top_nav'] = []
    context.update({
        'opportunity': opportunity,
        'opportunities': BusinessOpportunity.objects.filter(
            business_manager=request.user
        ).order_by('-created_time')[:50] if not _permission_granted('customer_management.opportunity.view_all', permission_set) 
        else BusinessOpportunity.objects.all().order_by('-created_time')[:100],
    })
    return render(request, "customer_management/opportunity_business_negotiation_form.html", context)




# ==================== 拜访四步流程视图函数 =====================

@login_required
def visit_plan_flow(request, plan_id=None):
    """统一的客户拜访流程视图（四步流程在同一页面完成）"""
    from backend.apps.customer_management.forms import VisitPlanForm, VisitChecklistForm, VisitCheckinForm, VisitReviewForm
    
    permission_set = get_user_permission_codes(request.user)
    if not _check_customer_permission('customer_management.relationship.edit', permission_set):
        messages.error(request, '您没有权限创建客户拜访')
        return redirect('business_pages:customer_visit')
    
    visit_plan = None
    current_step = 1
    
    # 如果有plan_id，获取已有的拜访计划
    if plan_id:
        visit_plan = get_object_or_404(VisitPlan, id=plan_id)
        # 检查权限：只有创建人或管理员可以编辑
        if visit_plan.created_by != request.user and not _permission_granted('customer_management.manage', permission_set):
            messages.error(request, '您没有权限编辑此拜访计划')
            return redirect('business_pages:customer_visit')
        # 获取当前步骤
        current_step = visit_plan.get_current_step()
    
    # 处理POST请求
    if request.method == 'POST':
        step = int(request.POST.get('step', 1))
        
        # 如果有plan_id，先获取visit_plan
        post_plan_id = request.POST.get('plan_id')
        if post_plan_id and not visit_plan:
            visit_plan = get_object_or_404(VisitPlan, id=post_plan_id)
            # 检查权限
            if visit_plan.created_by != request.user and not _permission_granted('customer_management.manage', permission_set):
                messages.error(request, '您没有权限编辑此拜访计划')
                return redirect('business_pages:customer_visit')
        
        if step == 1:
            # 第一步：创建拜访计划
            if visit_plan:
                permission_set = get_user_permission_codes(request.user)
                form = VisitPlanForm(request.POST, instance=visit_plan, user=request.user, permission_set=permission_set)
            else:
                permission_set = get_user_permission_codes(request.user)
                form = VisitPlanForm(request.POST, user=request.user, permission_set=permission_set)
            
            if form.is_valid():
                visit_plan = form.save(commit=False)
                visit_plan.created_by = request.user
                visit_plan.status = 'planned'
                visit_plan.save()
                current_step = 2  # 自动进入第二步
                messages.success(request, '拜访计划创建成功，请继续准备沟通清单')
            else:
                messages.error(request, '表单验证失败，请检查输入')
        
        elif step == 2:
            # 第二步：沟通清单准备
            if not visit_plan:
                messages.error(request, '请先创建拜访计划')
                return redirect('business_pages:visit_plan_flow')
            
            form = VisitChecklistForm(request.POST, instance=visit_plan)
            if form.is_valid():
                visit_plan = form.save(commit=False)
                visit_plan.checklist_prepared = True
                visit_plan.checklist_prepared_time = timezone.now()
                visit_plan.save()
                
                # 保存沟通清单问题的答案（如果模型存在）
                if HAS_COMMUNICATION_CHECKLIST_MODELS:
                    questions = CommunicationChecklistQuestion.objects.filter(is_active=True).order_by('part', 'order')
                    questions_by_part = {}
                    for question in questions:
                        if question.part not in questions_by_part:
                            questions_by_part[question.part] = []
                        questions_by_part[question.part].append(question)
                    
                    if questions_by_part:
                        checklist, created = CustomerCommunicationChecklist.objects.get_or_create(
                            client=visit_plan.client,
                            communication_date=visit_plan.plan_date,
                            defaults={
                                'title': f'{visit_plan.plan_title} - 沟通清单',
                                'location': visit_plan.location or '',
                                'status': 'before',
                                'created_by': request.user,
                                'opportunity': visit_plan.related_opportunity,
                            }
                        )
                        
                        for question in questions:
                            answer_value = request.POST.get(f'question_{question.id}', 'unknown')
                            note_before = request.POST.get(f'note_before_{question.id}', '').strip()
                            
                            answer, answer_created = CommunicationChecklistAnswer.objects.get_or_create(
                                checklist=checklist,
                                question=question,
                                defaults={
                                    'answer': answer_value,
                                    'note_before': note_before,
                                }
                            )
                            if not answer_created:
                                answer.answer = answer_value
                                answer.note_before = note_before
                                answer.save()
                
                current_step = 3  # 自动进入第三步
                messages.success(request, '沟通清单准备完成，可以进行拜访定位打卡')
            else:
                messages.error(request, '表单验证失败，请检查输入')
        
        elif step == 3:
            # 第三步：拜访定位打卡
            if not visit_plan:
                messages.error(request, '请先完成前面的步骤')
                return redirect('business_pages:visit_plan_flow')
            
            checkin = visit_plan.checkins.first()
            if checkin:
                form = VisitCheckinForm(request.POST, instance=checkin)
            else:
                form = VisitCheckinForm(request.POST)
            
            if form.is_valid():
                checkin = form.save(commit=False)
                checkin.visit_plan = visit_plan
                checkin.client = visit_plan.client
                checkin.created_by = request.user
                if not checkin.checkin_time:
                    checkin.checkin_time = timezone.now()
                checkin.save()
                
                # 更新拜访计划状态
                visit_plan.status = 'in_progress'
                visit_plan.save()
                
                current_step = 4  # 自动进入第四步
                messages.success(request, '拜访打卡成功，请进行拜访结果复盘')
            else:
                messages.error(request, '表单验证失败，请检查输入')
        
        elif step == 4:
            # 第四步：拜访结果复盘
            if not visit_plan:
                messages.error(request, '请先完成前面的步骤')
                return redirect('business_pages:visit_plan_flow')
            
            checkin = visit_plan.checkins.first()
            if not checkin:
                messages.warning(request, '请先完成拜访定位打卡')
                current_step = 3
            else:
                review, created = VisitReview.objects.get_or_create(
                    visit_plan=visit_plan,
                    defaults={'created_by': request.user, 'visit_checkin': checkin}
                )
                
                form = VisitReviewForm(request.POST, instance=review)
                if form.is_valid():
                    review = form.save(commit=False)
                    review.visit_checkin = checkin
                    if not review.created_by:
                        review.created_by = request.user
                    review.save()
                    
                    # 更新拜访计划状态为已完成
                    visit_plan.status = 'completed'
                    visit_plan.save()
                    
                    messages.success(request, '拜访结果复盘完成，拜访流程已全部完成')
                    return redirect('business_pages:visit_plan_detail', plan_id=visit_plan.id)
                else:
                    messages.error(request, '表单验证失败，请检查输入')
        
        # POST处理完成后，如果有visit_plan，需要重新获取以确保数据最新
        if visit_plan and visit_plan.id:
            visit_plan = get_object_or_404(VisitPlan, id=visit_plan.id)
            current_step = visit_plan.get_current_step()
    
    # 准备表单和上下文
    forms = {}
    questions_by_part = {}
    existing_answers = {}
    
    # 第一步表单
    if visit_plan:
        permission_set = get_user_permission_codes(request.user)
        forms['step1'] = VisitPlanForm(instance=visit_plan, user=request.user, permission_set=permission_set)
    else:
        permission_set = get_user_permission_codes(request.user)
        forms['step1'] = VisitPlanForm(user=request.user, permission_set=permission_set)
    
    # 第二步表单和沟通清单问题
    if visit_plan:
        forms['step2'] = VisitChecklistForm(instance=visit_plan)
        
        if HAS_COMMUNICATION_CHECKLIST_MODELS:
            questions = CommunicationChecklistQuestion.objects.filter(is_active=True).order_by('part', 'order')
            
            # 先获取现有答案
            if visit_plan and visit_plan.client:
                try:
                    existing_checklist = CustomerCommunicationChecklist.objects.filter(
                        client=visit_plan.client,
                        communication_date=visit_plan.plan_date
                    ).first()
                    if existing_checklist:
                        for answer in existing_checklist.answers.all():
                            existing_answers[answer.question_id] = {
                                'answer': answer.answer,
                                'note_before': answer.note_before,
                            }
                except Exception:
                    pass
            
            # 为每个问题准备答案数据，方便模板访问
            for question in questions:
                if question.part not in questions_by_part:
                    questions_by_part[question.part] = []
                questions_by_part[question.part].append(question)
                # 设置答案数据
                if question.id in existing_answers:
                    question.answer_data = existing_answers[question.id]
                else:
                    question.answer_data = {'answer': 'unknown', 'note_before': ''}
    else:
        forms['step2'] = VisitChecklistForm()
    
    # 第三步表单
    if visit_plan:
        checkin = visit_plan.checkins.first()
        if checkin:
            forms['step3'] = VisitCheckinForm(instance=checkin)
        else:
            forms['step3'] = VisitCheckinForm(initial={
                'checkin_time': timezone.now(),
                'checkin_location': visit_plan.location or '',
            })
    else:
        forms['step3'] = VisitCheckinForm()
    
    # 第四步表单
    if visit_plan:
        checkin = visit_plan.checkins.first()
        if checkin:
            review, created = VisitReview.objects.get_or_create(
                visit_plan=visit_plan,
                defaults={'created_by': request.user, 'visit_checkin': checkin}
            )
            forms['step4'] = VisitReviewForm(instance=review)
        else:
            forms['step4'] = VisitReviewForm()
    else:
        forms['step4'] = VisitReviewForm()
    
    context = _context(
        "客户拜访流程" if not visit_plan else f"客户拜访流程 - {visit_plan.plan_title}",
        "📅",
        "完成客户拜访的四步流程",
        request=request,
    )
    
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='visit_list'
    )
    
    context.update({
        'visit_plan': visit_plan,
        'current_step': current_step,
        'forms': forms,
        'questions_by_part': questions_by_part,
        'existing_answers': existing_answers,
    })
    
    return render(request, "customer_management/visit_plan_flow.html", context)


@login_required
def visit_plan_create(request):
    """第一步：创建拜访计划"""
    from backend.apps.customer_management.forms import VisitPlanForm
    
    permission_set = get_user_permission_codes(request.user)
    if not _check_customer_permission('customer_management.relationship.edit', permission_set):
        messages.error(request, '您没有权限创建拜访计划')
        return redirect('business_pages:customer_visit')
    
    if request.method == 'POST':
        form = VisitPlanForm(request.POST, user=request.user, permission_set=permission_set)
        if form.is_valid():
            visit_plan = form.save(commit=False)
            visit_plan.created_by = request.user
            visit_plan.status = 'planned'
            visit_plan.save()
            
            messages.success(request, '拜访计划创建成功，请继续准备沟通清单')
            return redirect('business_pages:visit_plan_checklist', plan_id=visit_plan.id)
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = VisitPlanForm(user=request.user, permission_set=permission_set)
    
    context = _context(
        "创建拜访计划",
        "📅",
        "第一步：创建拜访计划",
        request=request,
    )
    
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='visit_list'
    )
    
    # 获取所有客户的地址信息，用于前端自动填充
    # 只获取已审批通过的、该用户负责的客户
    clients_with_address = {}
    clients_opportunities = {}
    
    if request.user:
        from django.contrib.contenttypes.models import ContentType
        from backend.apps.workflow_engine.models import ApprovalInstance
        
        # 获取已审批通过的客户（通过 approval_status 或 ApprovalInstance）
        client_content_type = ContentType.objects.get_for_model(Client)
        approved_instance_ids = ApprovalInstance.objects.filter(
            content_type=client_content_type,
            status='approved'
        ).values_list('object_id', flat=True)
        
        # 只显示该用户作为负责人的、已审批通过的客户
        if approved_instance_ids:
            approved_clients = Client.objects.filter(
                is_active=True,
                responsible_user=request.user,
                id__in=approved_instance_ids
            ).distinct()
        else:
            approved_clients = Client.objects.none()
        
        # 获取客户地址信息
        for client in approved_clients.values('id', 'company_address'):
            clients_with_address[str(client['id'])] = client['company_address'] or ''
        
        # 获取客户及其对应的商机，用于前端动态过滤
        opportunities = BusinessOpportunity.objects.filter(
            client__in=approved_clients,
            status__in=['potential', 'initial_contact', 'requirement_confirmed', 'quotation', 'negotiation']
        ).select_related('client').order_by('-created_time')
        
        for opp in opportunities:
            client_id = str(opp.client.id) if opp.client else ''
            if client_id and client_id not in clients_opportunities:
                clients_opportunities[client_id] = []
            if client_id:
                clients_opportunities[client_id].append({
                    'id': opp.id,
                    'name': opp.name,
                    'client_name': opp.client.name if opp.client else ''
                })
    else:
        clients_queryset = Client.objects.filter(is_active=True).values('id', 'company_address')
        for client in clients_queryset:
            clients_with_address[str(client['id'])] = client['company_address'] or ''
    
    context.update({
        'form': form,
        'step': 1,
        'step_title': '创建计划',
        'clients_with_address_json': json.dumps(clients_with_address),
        'clients_opportunities_json': json.dumps(clients_opportunities),
    })
    return render(request, "customer_management/visit_plan_step_form.html", context)


@login_required
def visit_plan_checklist(request, plan_id):
    """第二步：沟通清单准备"""
    from backend.apps.customer_management.forms import VisitChecklistForm
    
    permission_set = get_user_permission_codes(request.user)
    visit_plan = get_object_or_404(VisitPlan, id=plan_id)
    
    # 检查权限：只有创建人或管理员可以编辑
    if visit_plan.created_by != request.user and not _permission_granted('customer_management.manage', permission_set):
        messages.error(request, '您没有权限编辑此拜访计划')
        return redirect('business_pages:visit_plan_detail', plan_id=plan_id)
    
    # 获取启用的沟通清单问题，按部分和排序分组（如果模型存在）
    questions_by_part = {}
    existing_answers = {}
    if HAS_COMMUNICATION_CHECKLIST_MODELS:
        questions = CommunicationChecklistQuestion.objects.filter(is_active=True).order_by('part', 'order')
        for question in questions:
            if question.part not in questions_by_part:
                questions_by_part[question.part] = []
            questions_by_part[question.part].append(question)
        
        # 获取已有的答案（如果有）
        if visit_plan.client:
            try:
                existing_checklist = CustomerCommunicationChecklist.objects.filter(
                    client=visit_plan.client,
                    communication_date=visit_plan.plan_date
                ).first()
                if existing_checklist:
                    for answer in existing_checklist.answers.all():
                        existing_answers[answer.question_id] = {
                            'answer': answer.answer,
                            'note_before': answer.note_before,
                        }
            except Exception:
                pass
    
    if request.method == 'POST':
        form = VisitChecklistForm(request.POST, instance=visit_plan)
        if form.is_valid():
            visit_plan = form.save(commit=False)
            visit_plan.checklist_prepared = True
            visit_plan.checklist_prepared_time = timezone.now()
            visit_plan.save()
            
            # 保存沟通清单问题的答案（如果模型存在）
            if HAS_COMMUNICATION_CHECKLIST_MODELS and questions_by_part:
                # 获取或创建沟通清单记录
                checklist, created = CustomerCommunicationChecklist.objects.get_or_create(
                    client=visit_plan.client,
                    communication_date=visit_plan.plan_date,
                    defaults={
                        'title': f'{visit_plan.plan_title} - 沟通清单',
                        'location': visit_plan.location or '',
                        'status': 'before',
                        'created_by': request.user,
                        'opportunity': visit_plan.related_opportunity,
                    }
                )
                
                # 保存每个问题的答案
                questions = CommunicationChecklistQuestion.objects.filter(is_active=True).order_by('part', 'order')
                for question in questions:
                    answer_value = request.POST.get(f'question_{question.id}', 'unknown')
                    note_before = request.POST.get(f'note_before_{question.id}', '').strip()
                    
                    answer, answer_created = CommunicationChecklistAnswer.objects.get_or_create(
                        checklist=checklist,
                        question=question,
                        defaults={
                            'answer': answer_value,
                            'note_before': note_before,
                        }
                    )
                    if not answer_created:
                        answer.answer = answer_value
                        answer.note_before = note_before
                        answer.save()
            
            messages.success(request, '沟通清单准备完成，可以进行拜访定位打卡')
            return redirect('business_pages:visit_plan_checkin', plan_id=visit_plan.id)
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = VisitChecklistForm(instance=visit_plan)
    
    context = _context(
        f"沟通清单准备 - {visit_plan.plan_title}",
        "📋",
        "第二步：沟通清单准备",
        request=request,
    )
    
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='visit_list'
    )
    
    context.update({
        'form': form,
        'visit_plan': visit_plan,
        'step': 2,
        'step_title': '沟通清单准备',
        'questions_by_part': questions_by_part,
        'existing_answers': existing_answers,
    })
    return render(request, "customer_management/visit_plan_step_form.html", context)


@login_required
def visit_plan_checkin(request, plan_id):
    """第三步：拜访定位打卡"""
    from backend.apps.customer_management.forms import VisitCheckinForm
    
    permission_set = get_user_permission_codes(request.user)
    visit_plan = get_object_or_404(VisitPlan, id=plan_id)
    
    # 检查权限：只有创建人或管理员可以打卡
    if visit_plan.created_by != request.user and not _permission_granted('customer_management.manage', permission_set):
        messages.error(request, '您没有权限进行拜访打卡')
        return redirect('business_pages:visit_plan_detail', plan_id=plan_id)
    
    if request.method == 'POST':
        form = VisitCheckinForm(request.POST)
        if form.is_valid():
            checkin = form.save(commit=False)
            checkin.visit_plan = visit_plan
            checkin.client = visit_plan.client
            checkin.created_by = request.user
            if not checkin.checkin_time:
                checkin.checkin_time = timezone.now()
            checkin.save()
            
            # 更新拜访计划状态
            visit_plan.status = 'in_progress'
            visit_plan.save()
            
            messages.success(request, '拜访打卡成功，请进行拜访结果复盘')
            return redirect('business_pages:visit_plan_review', plan_id=visit_plan.id)
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        # 初始化表单，设置默认值
        form = VisitCheckinForm(initial={
            'checkin_time': timezone.now(),
            'checkin_location': visit_plan.location or '',
        })
    
    context = _context(
        f"拜访定位打卡 - {visit_plan.plan_title}",
        "📍",
        "第三步：拜访定位打卡",
        request=request,
    )
    
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='visit_list'
    )
    
    context.update({
        'form': form,
        'visit_plan': visit_plan,
        'step': 3,
        'step_title': '拜访定位打卡',
    })
    return render(request, "customer_management/visit_plan_step_form.html", context)


@login_required
def visit_plan_review(request, plan_id):
    """第四步：拜访结果复盘"""
    from backend.apps.customer_management.forms import VisitReviewForm
    
    permission_set = get_user_permission_codes(request.user)
    visit_plan = get_object_or_404(VisitPlan, id=plan_id)
    
    # 检查权限：只有创建人或管理员可以复盘
    if visit_plan.created_by != request.user and not _permission_granted('customer_management.manage', permission_set):
        messages.error(request, '您没有权限进行拜访复盘')
        return redirect('business_pages:visit_plan_detail', plan_id=plan_id)
    
    # 检查是否已打卡
    checkin = visit_plan.checkins.first()
    if not checkin:
        messages.warning(request, '请先完成拜访定位打卡')
        return redirect('business_pages:visit_plan_checkin', plan_id=plan_id)
    
    # 获取或创建复盘记录
    review, created = VisitReview.objects.get_or_create(
        visit_plan=visit_plan,
        defaults={'created_by': request.user, 'visit_checkin': checkin}
    )
    
    if request.method == 'POST':
        form = VisitReviewForm(request.POST, instance=review)
        if form.is_valid():
            review = form.save(commit=False)
            review.visit_checkin = checkin
            if not review.created_by:
                review.created_by = request.user
            review.save()
            
            # 更新拜访计划状态为已完成
            visit_plan.status = 'completed'
            visit_plan.save()
            
            messages.success(request, '拜访结果复盘完成，拜访流程已全部完成')
            return redirect('business_pages:visit_plan_detail', plan_id=visit_plan.id)
        else:
            messages.error(request, '表单验证失败，请检查输入')
    else:
        form = VisitReviewForm(instance=review)
    
    context = _context(
        f"拜访结果复盘 - {visit_plan.plan_title}",
        "📊",
        "第四步：拜访结果复盘",
        request=request,
    )
    
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='visit_list'
    )
    
    context.update({
        'form': form,
        'visit_plan': visit_plan,
        'checkin': checkin,
        'review': review,
        'step': 4,
        'step_title': '拜访结果复盘',
    })
    return render(request, "customer_management/visit_plan_step_form.html", context)


@login_required
def visit_plan_detail(request, plan_id):
    """拜访计划详情（显示四步流程）"""
    permission_set = get_user_permission_codes(request.user)
    visit_plan = get_object_or_404(VisitPlan, id=plan_id)
    
    # 检查权限
    if visit_plan.created_by != request.user and not _permission_granted('customer_management.relationship.view', permission_set):
        messages.error(request, '您没有权限查看此拜访计划')
        return redirect('business_pages:customer_visit')
    
    # 获取当前步骤
    current_step = visit_plan.get_current_step()
    
    # 获取打卡记录
    checkin = visit_plan.checkins.first()
    
    # 获取复盘记录
    try:
        review = visit_plan.review
    except VisitReview.DoesNotExist:
        review = None
    
    context = _context(
        f"拜访计划详情 - {visit_plan.plan_title}",
        "📋",
        f"客户：{visit_plan.client.name}",
        request=request,
    )
    
    context['customer_menu'] = _build_customer_management_menu(
        permission_set, 
        active_id='visit_list'
    )
    
    context.update({
        'visit_plan': visit_plan,
        'checkin': checkin,
        'review': review,
        'current_step': current_step,
        'can_edit': visit_plan.created_by == request.user or _permission_granted('customer_management.relationship.manage', permission_set),
    })
    return render(request, "customer_management/visit_plan_detail.html", context)


# ==================== 业务委托书管理模块 ====================

@login_required
def authorization_letter_list(request):
    """业务委托书列表页面"""
    from django.core.paginator import Paginator
    from .forms import AuthorizationLetterForm
    
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限访问业务委托书列表')
        return redirect('business_pages:customer_list')
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    client_name = request.GET.get('client_name', '')
    opportunity_id = request.GET.get('opportunity_id', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # 获取委托书列表
    try:
        letters = AuthorizationLetter.objects.select_related('opportunity', 'project', 'created_by').order_by('-created_time')
        
        # 应用筛选条件
        if search:
            letters = letters.filter(
                Q(letter_number__icontains=search) |
                Q(project_name__icontains=search) |
                Q(client_name__icontains=search) |
                Q(trustee_name__icontains=search)
            )
        if status:
            letters = letters.filter(status=status)
        if client_name:
            letters = letters.filter(client_name__icontains=client_name)
        if opportunity_id:
            letters = letters.filter(opportunity_id=opportunity_id)
        if date_from:
            letters = letters.filter(created_time__date__gte=date_from)
        if date_to:
            letters = letters.filter(created_time__date__lte=date_to)
        
        # 分页
        paginator = Paginator(letters, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        from django.db import OperationalError, ProgrammingError
        logger = logging.getLogger(__name__)
        logger.exception('获取委托书列表失败: %s', str(e))
        
        # 检查是否是表不存在的错误
        error_msg = str(e)
        if 'does not exist' in error_msg.lower() or 'relation' in error_msg.lower():
            messages.error(
                request, 
                '数据库表不存在，请运行迁移或联系系统管理员。错误详情：表 business_authorization_letter 不存在。'
            )
        else:
            messages.error(request, f'获取委托书列表失败：{error_msg}')
        page_obj = None
    
    # 统计信息（应用当前筛选条件）
    try:
        base_queryset = AuthorizationLetter.objects.all()
        
        # 应用相同的筛选条件到统计查询
        if search:
            base_queryset = base_queryset.filter(
                Q(letter_number__icontains=search) |
                Q(project_name__icontains=search) |
                Q(client_name__icontains=search) |
                Q(trustee_name__icontains=search)
            )
        if status:
            base_queryset = base_queryset.filter(status=status)
        if client_name:
            base_queryset = base_queryset.filter(client_name__icontains=client_name)
        if opportunity_id:
            base_queryset = base_queryset.filter(opportunity_id=opportunity_id)
        if date_from:
            base_queryset = base_queryset.filter(created_time__date__gte=date_from)
        if date_to:
            base_queryset = base_queryset.filter(created_time__date__lte=date_to)
        
        total_count = base_queryset.count()
        confirmed_count = base_queryset.filter(status='confirmed').count()
        submitted_count = base_queryset.filter(status='submitted').count()
        draft_count = base_queryset.filter(status='draft').count()
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    # 获取筛选选项
    clients = Client.objects.filter(is_active=True).order_by('name')[:100]  # 限制数量
    opportunities = BusinessOpportunity.objects.filter(
        status__in=['potential', 'initial_contact', 'requirement_confirmed', 'quotation', 'negotiation']
    ).order_by('-created_time')[:100]
    
    # 检查创建权限
    can_create = _permission_granted('customer_management.client.create', permission_set)
    
    context = _context(
        "创建业务委托书",
        "📋",
        "管理业务委托书",
        request=request,
        active_menu_id='authorization_letter_list',
    )
    
    # 为每个委托书对象添加权限属性
    if page_obj:
        for letter in page_obj:
            # 判断是否可以编辑（创建人或具有编辑权限）
            letter.can_edit = (
                letter.created_by == request.user or 
                _permission_granted('customer_management.client.edit', permission_set)
            )
            # 判断是否可以删除（创建人或具有删除权限）
            letter.can_delete = (
                letter.created_by == request.user or 
                _permission_granted('customer_management.client.delete', permission_set)
            )
    
    context.update({
        'page_obj': page_obj,
        'summary_cards': summary_cards,
        'search': search,
        'status': status,
        'client_name': client_name,
        'opportunity_id': opportunity_id,
        'date_from': date_from,
        'date_to': date_to,
        'clients': clients,
        'opportunities': opportunities,
        'status_choices': AuthorizationLetter.STATUS_CHOICES,
        'can_create': can_create,
    })
    
    return render(request, "customer_management/authorization_letter_list.html", context)


@login_required
def authorization_letter_create(request):
    """创建业务委托书"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        from .forms import AuthorizationLetterForm
        
        permission_set = get_user_permission_codes(request.user)
        if not _permission_granted('customer_management.client.create', permission_set):
            messages.error(request, '您没有权限创建业务委托书')
            return redirect('business_pages:authorization_letter_list')
        
        if request.method == 'POST':
            form = AuthorizationLetterForm(request.POST)
            if form.is_valid():
                letter = form.save(commit=False)
                letter.created_by = request.user
                letter.save()
                messages.success(request, f'业务委托书 "{letter.project_name}" 创建成功')
                return redirect('business_pages:authorization_letter_list')
        else:
            form = AuthorizationLetterForm()
        
        context = _context(
            "创建业务委托书",
            "➕",
            "填写业务委托书信息",
            request=request,
            active_menu_id='authorization_letter_create',
        )
        
        context.update({
            'form': form,
            'is_create': True,
        })
        
        return render(request, "customer_management/authorization_letter_form.html", context)
    except Exception as e:
        logger.exception('创建业务委托书页面加载失败: %s', str(e))
        messages.error(request, f'页面加载失败：{str(e)}')
        return redirect('business_pages:authorization_letter_list')


@login_required
def authorization_letter_detail(request, letter_id):
    """业务委托书详情"""
    permission_set = get_user_permission_codes(request.user)
    letter = get_object_or_404(AuthorizationLetter, id=letter_id)
    
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限查看此业务委托书')
        return redirect('business_pages:authorization_letter_list')
    
    context = _context(
        f"业务委托书详情 - {letter.project_name}",
        "📋",
        f"委托书编号：{letter.letter_number}",
        request=request,
        active_menu_id='authorization_letter_list',
    )
    
    context.update({
        'letter': letter,
        'can_edit': letter.can_edit() and _permission_granted('customer_management.client.edit', permission_set),
        'can_delete': letter.can_delete() and _permission_granted('customer_management.client.delete', permission_set),
        'can_convert': letter.can_convert_to_contract() and _permission_granted('customer_management.client.create', permission_set),
    })
    
    return render(request, "customer_management/authorization_letter_detail.html", context)


@login_required
def authorization_letter_edit(request, letter_id):
    """编辑业务委托书"""
    from .forms import AuthorizationLetterForm
    
    permission_set = get_user_permission_codes(request.user)
    letter = get_object_or_404(AuthorizationLetter, id=letter_id)
    
    if not letter.can_edit():
        messages.error(request, '只有草稿状态的委托书可以编辑')
        return redirect('business_pages:authorization_letter_detail', letter_id=letter_id)
    
    if not _permission_granted('customer_management.client.edit', permission_set):
        messages.error(request, '您没有权限编辑此业务委托书')
        return redirect('business_pages:authorization_letter_detail', letter_id=letter_id)
    
    if request.method == 'POST':
        form = AuthorizationLetterForm(request.POST, instance=letter)
        if form.is_valid():
            letter = form.save()
            messages.success(request, f'业务委托书 "{letter.project_name}" 更新成功')
            return redirect('business_pages:authorization_letter_detail', letter_id=letter_id)
    else:
        form = AuthorizationLetterForm(instance=letter)
    
    context = _context(
        f"编辑业务委托书 - {letter.project_name}",
        "✏️",
        f"委托书编号：{letter.letter_number}",
        request=request,
        active_menu_id='authorization_letter_list',
    )
    
    context.update({
        'form': form,
        'letter': letter,
        'is_create': False,
    })
    
    return render(request, "customer_management/authorization_letter_form.html", context)


@login_required
def authorization_letter_delete(request, letter_id):
    """删除业务委托书"""
    permission_set = get_user_permission_codes(request.user)
    letter = get_object_or_404(AuthorizationLetter, id=letter_id)
    
    if not letter.can_delete():
        messages.error(request, '只有草稿状态的委托书可以删除')
        return redirect('business_pages:authorization_letter_detail', letter_id=letter_id)
    
    if not _permission_granted('customer_management.client.delete', permission_set):
        messages.error(request, '您没有权限删除此业务委托书')
        return redirect('business_pages:authorization_letter_detail', letter_id=letter_id)
    
    if request.method == 'POST':
        letter_name = letter.project_name
        letter.delete()
        messages.success(request, f'业务委托书 "{letter_name}" 已删除')
        return redirect('business_pages:authorization_letter_list')
    
    context = _context(
        f"删除业务委托书 - {letter.project_name}",
        "🗑️",
        f"确认删除委托书编号：{letter.letter_number}",
        request=request,
        active_menu_id='authorization_letter_list',
    )
    
    context.update({
        'letter': letter,
    })
    
    return render(request, "customer_management/authorization_letter_delete.html", context)


@login_required
def authorization_letter_status_transition(request, letter_id):
    """业务委托书状态流转"""
    permission_set = get_user_permission_codes(request.user)
    letter = get_object_or_404(AuthorizationLetter, id=letter_id)
    
    if not _permission_granted('customer_management.client.edit', permission_set):
        messages.error(request, '您没有权限操作此业务委托书')
        return redirect('business_pages:authorization_letter_detail', letter_id=letter_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'submit':
            if letter.status == 'draft':
                letter.status = 'submitted'
                letter.save()
                messages.success(request, '委托书已提交')
            else:
                messages.error(request, '只能提交草稿状态的委托书')
        elif action == 'confirm':
            if letter.status == 'submitted':
                letter.status = 'confirmed'
                letter.save()
                messages.success(request, '委托书已确认')
            else:
                messages.error(request, '只能确认已提交状态的委托书')
        elif action == 'cancel':
            if letter.status in ['draft', 'submitted']:
                letter.status = 'cancelled'
                letter.save()
                messages.success(request, '委托书已作废')
            else:
                messages.error(request, '只能作废草稿或已提交状态的委托书')
        else:
            messages.error(request, '无效的操作')
    
    return redirect('business_pages:authorization_letter_detail', letter_id=letter_id)


# ==================== 业务委托书模板管理 ====================

@login_required
def authorization_letter_template_list(request):
    """业务委托书模板列表页面"""
    from django.core.paginator import Paginator
    from .forms import AuthorizationLetterTemplateForm
    from .models import AuthorizationLetterTemplate
    
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限访问业务委托书模板列表')
        return redirect('business_pages:authorization_letter_list')
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    template_type = request.GET.get('template_type', '')
    status = request.GET.get('status', '')
    
    # 获取模板列表
    try:
        templates = AuthorizationLetterTemplate.objects.select_related('created_by', 'updated_by').order_by('-created_time')
        
        # 应用筛选条件
        if search:
            templates = templates.filter(
                Q(template_name__icontains=search) |
                Q(category__icontains=search) |
                Q(description__icontains=search)
            )
        if template_type:
            templates = templates.filter(template_type=template_type)
        if status:
            templates = templates.filter(status=status)
        
        # 分页
        paginator = Paginator(templates, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取模板列表失败: %s', str(e))
        messages.error(request, f'获取模板列表失败：{str(e)}')
        page_obj = None
    
    # 统计信息
    try:
        total_count = AuthorizationLetterTemplate.objects.count()
        active_count = AuthorizationLetterTemplate.objects.filter(status='active').count()
        draft_count = AuthorizationLetterTemplate.objects.filter(status='draft').count()
        total_usage = AuthorizationLetterTemplate.objects.aggregate(total=Sum('usage_count'))['total'] or 0
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    context = _context(
        "业务委托书模板列表",
        "📄",
        "管理业务委托书模板，快速创建委托书",
        summary_cards=summary_cards,
        request=request,
        active_menu_id='authorization_letter_template',
    )
    
    context.update({
        'page_obj': page_obj,
        'search': search,
        'template_type': template_type,
        'status': status,
        'template_type_choices': AuthorizationLetterTemplate.TEMPLATE_TYPE_CHOICES,
        'status_choices': AuthorizationLetterTemplate.STATUS_CHOICES,
    })
    
    return render(request, "customer_management/authorization_letter_template_list.html", context)


@login_required
def authorization_letter_template_create(request):
    """创建业务委托书模板"""
    from .forms import AuthorizationLetterTemplateForm
    from .models import AuthorizationLetterTemplate
    
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.create', permission_set):
        messages.error(request, '您没有权限创建业务委托书模板')
        return redirect('business_pages:authorization_letter_template_list')
    
    if request.method == 'POST':
        import json
        form = AuthorizationLetterTemplateForm(request.POST, request.FILES)
        
        # 处理JSON字段
        if 'template_content' in request.POST:
            try:
                template_content = json.loads(request.POST.get('template_content', '{}'))
                form.data = form.data.copy()
                form.data['template_content'] = template_content
            except json.JSONDecodeError:
                messages.error(request, '模板内容格式错误')
        
        if 'variables' in request.POST:
            try:
                variables = json.loads(request.POST.get('variables', '[]'))
                form.data = form.data.copy()
                form.data['variables'] = variables
            except json.JSONDecodeError:
                messages.error(request, '变量列表格式错误')
        
        if form.is_valid():
            template = form.save(commit=False)
            template.created_by = request.user
            template.save()
            messages.success(request, f'业务委托书模板 "{template.template_name}" 创建成功')
            return redirect('business_pages:authorization_letter_template_list')
    else:
        form = AuthorizationLetterTemplateForm()
    
    context = _context(
        "创建业务委托书模板",
        "➕",
        "填写模板信息，支持变量占位符",
        request=request,
        active_menu_id='authorization_letter_template',
    )
    
    context.update({
        'form': form,
        'is_create': True,
    })
    
    return render(request, "customer_management/authorization_letter_template_form.html", context)


@login_required
def authorization_letter_template_edit(request, template_id):
    """编辑业务委托书模板"""
    from .forms import AuthorizationLetterTemplateForm
    from .models import AuthorizationLetterTemplate
    
    permission_set = get_user_permission_codes(request.user)
    template = get_object_or_404(AuthorizationLetterTemplate, id=template_id)
    
    if not _permission_granted('customer_management.client.edit', permission_set):
        messages.error(request, '您没有权限编辑此业务委托书模板')
        return redirect('business_pages:authorization_letter_template_list')
    
    if request.method == 'POST':
        import json
        form = AuthorizationLetterTemplateForm(request.POST, request.FILES, instance=template)
        
        # 处理JSON字段
        if 'template_content' in request.POST:
            try:
                template_content = json.loads(request.POST.get('template_content', '{}'))
                form.data = form.data.copy()
                form.data['template_content'] = template_content
            except json.JSONDecodeError:
                messages.error(request, '模板内容格式错误')
        
        if 'variables' in request.POST:
            try:
                variables = json.loads(request.POST.get('variables', '[]'))
                form.data = form.data.copy()
                form.data['variables'] = variables
            except json.JSONDecodeError:
                messages.error(request, '变量列表格式错误')
        
        if form.is_valid():
            template = form.save(commit=False)
            template.updated_by = request.user
            template.save()
            messages.success(request, f'业务委托书模板 "{template.template_name}" 更新成功')
            return redirect('business_pages:authorization_letter_template_list')
    else:
        form = AuthorizationLetterTemplateForm(instance=template)
    
    context = _context(
        f"编辑业务委托书模板 - {template.template_name}",
        "✏️",
        f"模板类型：{template.get_template_type_display()}",
        request=request,
        active_menu_id='authorization_letter_template',
    )
    
    context.update({
        'form': form,
        'template': template,
        'is_create': False,
    })
    
    return render(request, "customer_management/authorization_letter_template_form.html", context)


@login_required
def authorization_letter_template_delete(request, template_id):
    """删除业务委托书模板"""
    from .models import AuthorizationLetterTemplate
    
    permission_set = get_user_permission_codes(request.user)
    template = get_object_or_404(AuthorizationLetterTemplate, id=template_id)
    
    if not _permission_granted('customer_management.client.delete', permission_set):
        messages.error(request, '您没有权限删除此业务委托书模板')
        return redirect('business_pages:authorization_letter_template_list')
    
    if request.method == 'POST':
        template_name = template.template_name
        template.delete()
        messages.success(request, f'业务委托书模板 "{template_name}" 已删除')
        return redirect('business_pages:authorization_letter_template_list')
    
    context = _context(
        f"删除业务委托书模板 - {template.template_name}",
        "🗑️",
        f"确认删除模板：{template.template_name}",
        request=request,
        active_menu_id='authorization_letter_template',
    )
    
    context.update({
        'template': template,
    })
    
    return render(request, "customer_management/authorization_letter_template_delete.html", context)


@login_required
def authorization_letter_create_from_template(request, template_id):
    """从模板创建业务委托书"""
    from .forms import AuthorizationLetterForm
    from .models import AuthorizationLetterTemplate, AuthorizationLetter
    
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.create', permission_set):
        messages.error(request, '您没有权限创建业务委托书')
        return redirect('business_pages:authorization_letter_list')
    
    template = get_object_or_404(AuthorizationLetterTemplate, id=template_id)
    
    if request.method == 'POST':
        form = AuthorizationLetterForm(request.POST)
        if form.is_valid():
            letter = form.save(commit=False)
            letter.created_by = request.user
            letter.save()
            
            # 增加模板使用次数
            template.increment_usage()
            
            messages.success(request, f'业务委托书 "{letter.project_name}" 创建成功（来自模板：{template.template_name}）')
            return redirect('business_pages:authorization_letter_detail', letter_id=letter.id)
    else:
        # 从模板填充表单初始值
        form = AuthorizationLetterForm()
        template_content = template.template_content or {}
        
        # 填充表单字段
        for field_name, field_value in template_content.items():
            if hasattr(form, 'fields') and field_name in form.fields:
                form.initial[field_name] = field_value
    
    context = _context(
        f"从模板创建业务委托书 - {template.template_name}",
        "📄",
        f"模板类型：{template.get_template_type_display()}",
        request=request,
        active_menu_id='authorization_letter_create',
    )
    
    context.update({
        'form': form,
        'template': template,
        'is_create': True,
        'from_template': True,
    })
    
    return render(request, "customer_management/authorization_letter_form.html", context)


@login_required
def authorization_letter_template_file_preview(request, template_id):
    """预览业务委托书模板文件"""
    from django.http import FileResponse, Http404
    from .models import AuthorizationLetterTemplate
    import os
    import mimetypes
    
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限预览模板文件')
        return redirect('business_pages:authorization_letter_template_list')
    
    template = get_object_or_404(AuthorizationLetterTemplate, id=template_id)
    
    if not template.template_file:
        raise Http404('模板文件不存在')
    
    try:
        # 获取文件名
        if template.template_file_name:
            filename = template.template_file_name
        else:
            filename = os.path.basename(template.template_file.name)
        
        # 根据文件扩展名确定 content_type
        ext = os.path.splitext(filename)[1].lower()
        content_type_map = {
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.xls': 'application/vnd.ms-excel',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.ppt': 'application/vnd.ms-powerpoint',
            '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
        }
        content_type = content_type_map.get(ext, 'application/octet-stream')
        
        response = FileResponse(
            template.template_file.open('rb'),
            content_type=content_type
        )
        # 设置文件名和内联显示
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    except Exception as e:
        logger.exception('预览模板文件失败: %s', str(e))
        messages.error(request, f'预览文件失败：{str(e)}')
        return redirect('business_pages:authorization_letter_template_edit', template_id=template_id)


@login_required
def authorization_letter_template_file_download(request, template_id):
    """下载业务委托书模板文件"""
    from django.http import FileResponse, Http404
    from .models import AuthorizationLetterTemplate
    
    permission_set = get_user_permission_codes(request.user)
    if not _permission_granted('customer_management.client.view', permission_set):
        messages.error(request, '您没有权限下载模板文件')
        return redirect('business_pages:authorization_letter_template_list')
    
    template = get_object_or_404(AuthorizationLetterTemplate, id=template_id)
    
    if not template.template_file:
        raise Http404('模板文件不存在')
    
    try:
        response = FileResponse(
            template.template_file.open('rb'),
            content_type='application/octet-stream'
        )
        # 设置下载文件名
        if template.template_file_name:
            response['Content-Disposition'] = f'attachment; filename="{template.template_file_name}"'
        else:
            import os
            filename = os.path.basename(template.template_file.name)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        logger.exception('下载模板文件失败: %s', str(e))
        messages.error(request, f'下载文件失败：{str(e)}')
        return redirect('business_pages:authorization_letter_template_edit', template_id=template_id)


# ==================== 商机导入功能 ====================

@login_required
def opportunity_import(request):
    """商机批量导入功能"""
    from django.http import HttpResponse
    from django.db import transaction
    from backend.apps.system_management.models import User
    
    permission_set = get_user_permission_codes(request.user)
    
    # 检查权限：需要商机管理权限
    if not _permission_granted('customer_management.opportunity.view', permission_set):
        messages.error(request, '您没有权限执行商机导入操作')
        return redirect('business_pages:opportunity_management')
    
    # 下载模板
    if request.GET.get('download') == 'template':
        service_type_sample_obj = ServiceType.objects.order_by('id').first()
        design_stage_sample_obj = DesignStage.objects.filter(is_active=True).order_by('order', 'id').first()
        design_stage_sample_label = design_stage_sample_obj.name if design_stage_sample_obj else ''
        status_label_map = dict(BusinessOpportunity.STATUS_CHOICES)
        status_sample_label = status_label_map.get('potential', '潜在客户')
        urgency_label_map = dict(BusinessOpportunity.URGENCY_CHOICES)
        urgency_sample_label = urgency_label_map.get('normal', '普通')
        opportunity_type_label_map = dict(BusinessOpportunity.OPPORTUNITY_TYPE_CHOICES)
        opportunity_type_sample_label = opportunity_type_label_map.get('project_cooperation', '项目合作')
        
        columns = [
            '商机编号（可留空自动生成）',
            '商机名称',
            '客户名称（必填）',
            '负责商务手机号（必填）',
            '商机类型',
            '服务类型（可填编码或名称）',
            '项目名称',
            '项目地址',
            '项目业态',
            '建筑面积（平方米）',
            '图纸阶段（可填编码或名称）',
            '预计金额（万元）',
            '成功概率（%）',
            '商机状态',
            '紧急程度',
            '预计签约时间（YYYY-MM-DD）',
            '商机描述',
            '备注',
        ]
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="opportunity_import_template.csv"'
        writer = csv.writer(response)
        writer.writerow(columns)
        writer.writerow([
            '',
            '锦城天府综合体一期商机',
            '成都锦城房地产开发有限公司',
            '13800000005',
            opportunity_type_sample_label,
            service_type_sample_obj.name if service_type_sample_obj else '',
            '锦城天府综合体一期',
            '成都市天府新区',
            '住宅',
            '50000',
            design_stage_sample_label,
            '500',
            '30',
            status_sample_label,
            urgency_sample_label,
            '2025-12-31',
            '这是一个示例商机',
            '备注信息',
        ])
        return response
    
    # 准备上下文数据
    design_stages = DesignStage.objects.filter(is_active=True).order_by('order', 'id')
    context = {
        'service_types': ServiceType.objects.order_by('order', 'id'),
        'design_stages': design_stages,
        'status_choices': BusinessOpportunity.STATUS_CHOICES,
        'urgency_choices': BusinessOpportunity.URGENCY_CHOICES,
        'opportunity_type_choices': BusinessOpportunity.OPPORTUNITY_TYPE_CHOICES,
        'import_results': None,
    }
    
    if request.method == 'POST':
        upload = request.FILES.get('import_file')
        if not upload:
            messages.error(request, '请上传 CSV 或 Excel 文件。')
        else:
            filename = upload.name.lower()
            is_excel = filename.endswith(('.xlsx', '.xls'))
            is_csv = filename.endswith('.csv')
            
            if not (is_csv or is_excel):
                messages.error(request, '仅支持 CSV 或 Excel 文件（.csv, .xlsx, .xls）。')
            elif upload.size > 10 * 1024 * 1024:  # 10MB
                messages.error(request, '文件过大，请控制在 10MB 以内。')
            else:
                try:
                    upload.seek(0)
                except Exception:
                    pass
                
                # 处理Excel文件
                if is_excel:
                    try:
                        import pandas as pd
                        # 尝试读取Excel文件
                        df = pd.read_excel(upload, engine='openpyxl' if filename.endswith('.xlsx') else None)
                        # 转换为CSV格式的字符串
                        csv_buffer = io.StringIO()
                        df.to_csv(csv_buffer, index=False, encoding='utf-8')
                        decoded_text = csv_buffer.getvalue()
                    except ImportError:
                        messages.error(request, '系统未安装 pandas 库，无法处理 Excel 文件。请使用 CSV 格式。')
                        decoded_text = None
                    except Exception as e:
                        messages.error(request, f'Excel 文件解析失败：{str(e)}')
                        decoded_text = None
                else:
                    # 处理CSV文件
                    raw_bytes = upload.read()
                    decoded_text = None
                    for enc in ('utf-8-sig', 'utf-8', 'gbk', 'gb2312'):
                        try:
                            decoded_text = raw_bytes.decode(enc)
                            break
                        except UnicodeDecodeError:
                            continue
                
                if decoded_text is None:
                    messages.error(request, '文件解析失败，请确认编码为 UTF-8 或 GBK（CSV），或使用标准 Excel 格式。')
                else:
                    text_io = io.StringIO(decoded_text)
                    reader = csv.DictReader(text_io)
                    
                    field_aliases = {
                        'opportunity_number': {'商机编号（可留空自动生成）', '商机编号', 'opportunity_number'},
                        'name': {'商机名称', 'name'},
                        'client_name': {'客户名称（必填）', '客户名称', 'client_name'},
                        'business_manager_phone': {'负责商务手机号（必填）', '负责商务手机号', '商务经理手机号', 'business_manager_phone'},
                        'opportunity_type': {'商机类型', 'opportunity_type'},
                        'service_type': {'服务类型（可填编码或名称）', '服务类型', 'service_type'},
                        'project_name': {'项目名称', 'project_name'},
                        'project_address': {'项目地址', 'project_address'},
                        'project_type': {'项目业态', 'project_type'},
                        'building_area': {'建筑面积（平方米）', '建筑面积', 'building_area'},
                        'drawing_stage': {'图纸阶段（可填编码或名称）', '图纸阶段', 'drawing_stage'},
                        'estimated_amount': {'预计金额（万元）', '预计金额', 'estimated_amount'},
                        'success_probability': {'成功概率（%）', '成功概率', 'success_probability'},
                        'status': {'商机状态', 'status'},
                        'urgency': {'紧急程度', 'urgency'},
                        'expected_sign_date': {'预计签约时间（YYYY-MM-DD）', '预计签约时间', 'expected_sign_date'},
                        'description': {'商机描述', 'description'},
                        'notes': {'备注', 'notes'},
                    }
                    
                    required_fields = {
                        'name',
                        'client_name',
                        'business_manager_phone',
                    }
                    
                    missing_labels = []
                    headers = set(reader.fieldnames or [])
                    for field in required_fields:
                        if not any(alias in headers for alias in field_aliases[field]):
                            missing_labels.append(next(iter(field_aliases[field])))
                    
                    if missing_labels:
                        messages.error(request, f'CSV 缺少必要字段：{", ".join(missing_labels)}。')
                    else:
                        def get_value(row, field):
                            for alias in field_aliases[field]:
                                if alias in row and row[alias] is not None:
                                    value = str(row.get(alias, '')).strip()
                                    if value:
                                        return value
                            return ''
                        
                        # 构建查找映射
                        service_type_lookup = {st.code: st for st in ServiceType.objects.all()}
                        service_type_name_lookup = {(st.name or '').strip(): st for st in ServiceType.objects.all()}
                        
                        design_stage_objects = DesignStage.objects.filter(is_active=True)
                        design_stage_id_map = {str(ds.id): ds for ds in design_stage_objects}
                        design_stage_code_map = {ds.code: ds for ds in design_stage_objects if ds.code}
                        design_stage_name_map = {ds.name: ds for ds in design_stage_objects}
                        
                        status_codes = {code for code, _ in BusinessOpportunity.STATUS_CHOICES}
                        status_label_map = {(label or '').strip(): code for code, label in BusinessOpportunity.STATUS_CHOICES}
                        
                        urgency_codes = {code for code, _ in BusinessOpportunity.URGENCY_CHOICES}
                        urgency_label_map = {(label or '').strip(): code for code, label in BusinessOpportunity.URGENCY_CHOICES}
                        
                        opportunity_type_codes = {code for code, _ in BusinessOpportunity.OPPORTUNITY_TYPE_CHOICES}
                        opportunity_type_label_map = {(label or '').strip(): code for code, label in BusinessOpportunity.OPPORTUNITY_TYPE_CHOICES}
                        
                        results = []
                        success_count = 0
                        failure_count = 0
                        
                        for row_index, row in enumerate(reader, start=2):
                            row_result = {'row': row_index, 'status': 'success', 'message': ''}
                            try:
                                with transaction.atomic():
                                    # 必填字段验证
                                    opportunity_name = get_value(row, 'name')
                                    if not opportunity_name:
                                        raise ValueError('商机名称不能为空')
                                    
                                    client_name = get_value(row, 'client_name')
                                    if not client_name:
                                        raise ValueError('客户名称不能为空')
                                    
                                    # 查找或创建客户
                                    client = Client.objects.filter(name=client_name).first()
                                    if not client:
                                        # 如果客户不存在，尝试创建（需要客户类型）
                                        client_type = ClientType.objects.first()
                                        if not client_type:
                                            raise ValueError(f'客户"{client_name}"不存在，且系统未配置客户类型，无法自动创建')
                                        client = Client.objects.create(
                                            name=client_name,
                                            client_type=client_type,
                                            created_by=request.user,
                                        )
                                    
                                    business_manager_phone = get_value(row, 'business_manager_phone')
                                    if not business_manager_phone:
                                        raise ValueError('负责商务手机号不能为空')
                                    business_manager = User.objects.filter(username=business_manager_phone).first()
                                    if not business_manager:
                                        raise ValueError(f'未找到对应的商务经理手机号：{business_manager_phone}')
                                    
                                    # 可选字段处理
                                    opportunity_number = get_value(row, 'opportunity_number')
                                    if opportunity_number and BusinessOpportunity.objects.filter(opportunity_number=opportunity_number).exists():
                                        raise ValueError(f'商机编号重复：{opportunity_number}')
                                    
                                    opportunity_type_raw = get_value(row, 'opportunity_type')
                                    opportunity_type = None
                                    if opportunity_type_raw:
                                        if opportunity_type_raw in opportunity_type_codes:
                                            opportunity_type = opportunity_type_raw
                                        else:
                                            opportunity_type = opportunity_type_label_map.get(opportunity_type_raw)
                                        if not opportunity_type:
                                            raise ValueError(f'商机类型取值无效：{opportunity_type_raw}')
                                    
                                    service_type_key = get_value(row, 'service_type')
                                    service_type = None
                                    if service_type_key:
                                        service_type = service_type_lookup.get(service_type_key)
                                        if not service_type:
                                            service_type = service_type_name_lookup.get(service_type_key)
                                        if not service_type:
                                            raise ValueError(f'服务类型取值无效：{service_type_key}')
                                    
                                    project_name = get_value(row, 'project_name') or None
                                    project_address = get_value(row, 'project_address') or None
                                    project_type = get_value(row, 'project_type') or None
                                    
                                    building_area_str = get_value(row, 'building_area')
                                    building_area = None
                                    if building_area_str:
                                        try:
                                            building_area = Decimal(building_area_str)
                                        except (ValueError, InvalidOperation):
                                            raise ValueError(f'建筑面积格式无效：{building_area_str}')
                                    
                                    drawing_stage_raw = get_value(row, 'drawing_stage')
                                    drawing_stage = None
                                    if drawing_stage_raw:
                                        if drawing_stage_raw in design_stage_id_map:
                                            drawing_stage = design_stage_id_map[drawing_stage_raw]
                                        elif drawing_stage_raw in design_stage_code_map:
                                            drawing_stage = design_stage_code_map[drawing_stage_raw]
                                        elif drawing_stage_raw in design_stage_name_map:
                                            drawing_stage = design_stage_name_map[drawing_stage_raw]
                                        if not drawing_stage:
                                            raise ValueError(f'图纸阶段取值无效：{drawing_stage_raw}')
                                    
                                    estimated_amount_str = get_value(row, 'estimated_amount')
                                    estimated_amount = Decimal('0')
                                    if estimated_amount_str:
                                        try:
                                            estimated_amount = Decimal(estimated_amount_str)
                                        except (ValueError, InvalidOperation):
                                            raise ValueError(f'预计金额格式无效：{estimated_amount_str}')
                                    
                                    success_probability_str = get_value(row, 'success_probability')
                                    success_probability = 10  # 默认值
                                    if success_probability_str:
                                        try:
                                            success_probability = int(success_probability_str)
                                            if success_probability not in [10, 30, 50, 70, 90]:
                                                raise ValueError(f'成功概率必须是 10、30、50、70 或 90，当前值：{success_probability}')
                                        except ValueError as e:
                                            if '必须是' in str(e):
                                                raise
                                            raise ValueError(f'成功概率格式无效：{success_probability_str}')
                                    
                                    status_raw = get_value(row, 'status') or 'potential'
                                    status = status_raw
                                    if status not in status_codes:
                                        status = status_label_map.get(status_raw)
                                    if not status or status not in status_codes:
                                        raise ValueError(f'商机状态取值无效：{status_raw}')
                                    
                                    urgency_raw = get_value(row, 'urgency') or 'normal'
                                    urgency = urgency_raw
                                    if urgency not in urgency_codes:
                                        urgency = urgency_label_map.get(urgency_raw)
                                    if not urgency or urgency not in urgency_codes:
                                        raise ValueError(f'紧急程度取值无效：{urgency_raw}')
                                    
                                    expected_sign_date_str = get_value(row, 'expected_sign_date')
                                    expected_sign_date = None
                                    if expected_sign_date_str:
                                        try:
                                            from datetime import datetime
                                            expected_sign_date = datetime.strptime(expected_sign_date_str, '%Y-%m-%d').date()
                                        except ValueError:
                                            raise ValueError(f'预计签约时间格式无效，应为 YYYY-MM-DD：{expected_sign_date_str}')
                                    
                                    description = get_value(row, 'description') or ''
                                    notes = get_value(row, 'notes') or ''
                                    
                                    # 创建商机
                                    opportunity = BusinessOpportunity(
                                        opportunity_number=opportunity_number or None,
                                        name=opportunity_name,
                                        client=client,
                                        business_manager=business_manager,
                                        opportunity_type=opportunity_type or '',
                                        service_type=service_type,
                                        project_name=project_name or '',
                                        project_address=project_address or '',
                                        project_type=project_type or '',
                                        building_area=building_area,
                                        drawing_stage=drawing_stage,
                                        estimated_amount=estimated_amount,
                                        success_probability=success_probability,
                                        status=status,
                                        urgency=urgency,
                                        expected_sign_date=expected_sign_date,
                                        description=description,
                                        notes=notes,
                                        created_by=request.user,
                                    )
                                    opportunity.save()
                                    
                                    success_count += 1
                                    row_result['message'] = f'导入成功，商机编号：{opportunity.opportunity_number}'
                            except Exception as exc:
                                failure_count += 1
                                row_result['status'] = 'failed'
                                row_result['message'] = str(exc)
                            results.append(row_result)
                        
                        context['import_results'] = {
                            'total': success_count + failure_count,
                            'success': success_count,
                            'failed': failure_count,
                            'rows': results,
                        }
                        if success_count:
                            messages.success(request, f'成功导入 {success_count} 条商机。')
                        if failure_count:
                            messages.warning(request, f'{failure_count} 条记录导入失败，请查看结果列表。')
    
    # 生成左侧菜单
    menu = _build_opportunity_management_menu(permission_set, 'opportunity_import')
    
    return render(
        request,
        'customer_management/opportunity_import.html',
        {
            **context,
            'menu': menu,
            'page_title': '商机批量导入',
            'page_description': '通过上传 CSV 或 Excel 文件批量导入商机数据',
        }
    )
